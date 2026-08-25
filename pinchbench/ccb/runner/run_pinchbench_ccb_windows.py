#!/usr/bin/env python3
# Runner revision: 2026-07-22-pinchbench-official-grader-api-ccb-windows-v1.4-stdin-session-fix
r"""
Run the local PinchBench task set with CCB (Claude Code Best) on native Windows.

Comparison target:
    PinchBench local tasks -> CCB default agent/tools -> OpenRouter
    -> deepseek/deepseek-v4-pro

Strict-alignment goals versus the paired OpenCode Windows v3 runner:
- Same checked-out PinchBench manifest and task order.
- Same four agreed integration-task exclusions.
- One worker / strictly serial task execution.
- Fresh CCB session per normal task.
- Multi-session task support via CCB --resume.
- Same task timeout policy, workspace staging, workspace reminder, and output schema.
- Raw CCB stream-json retained for audit.
- CCB events normalized to the transcript shape expected by PinchBench graders.
- Grading through the checked-out PinchBench scripts/lib_grading.py engine.
- Hybrid/llm_judge use that commit's DEFAULT_JUDGE_MODEL through direct API
  (judge_backend="api") and OPENROUTER_API_KEY.
- JSON, CSV, XLSX, progress.jsonl, and results.partial.json outputs.

CCB-specific, explicitly recorded adapter behavior:
- The tested model is selected with --model deepseek/deepseek-v4-pro while
  ANTHROPIC_BASE_URL points to OpenRouter's Anthropic-compatible endpoint.
- A fixed --allowedTools list auto-approves the ordinary headless benchmark tools.
  This does not restrict CCB's visible default tool set; it prevents interactive
  permission prompts from blocking a non-interactive -p benchmark run.
- --permission-mode acceptEdits is set explicitly for unattended workspace edits.
- Prompts are sent through stdin, with all CLI options placed before -p. This
  avoids Windows .cmd argument parsing turning stream-json runs into plain text.
- Every new CCB session receives an explicit UUID via --session-id; continuation
  turns use --resume with that exact UUID.
- CCB 2.8.4 does not expose a separate exact reasoning-token count in its result
  event, so reasoning_tokens remains null while total output_tokens is preserved.

Recommended location:
    C:\pinchbench-ccb\runner\run_pinchbench_ccb_windows.py

Examples:
    python run_pinchbench_ccb_windows.py --preflight --suite all
    python run_pinchbench_ccb_windows.py --suite task_sanity,task_email_triage,task_csv_stock_trend --keep-workspaces --verbose
    python run_pinchbench_ccb_windows.py --suite all --keep-workspaces --clear-judge-cache
"""
from __future__ import annotations

import argparse
import csv
import dataclasses
import datetime as dt
import hashlib
import json
import logging
import os
import platform
import queue
import re
import shutil
import signal
import subprocess
import sys
import threading
import time
import traceback
import uuid
from collections import Counter
from pathlib import Path
from typing import Any, Optional

os.environ.setdefault("PYTHONUTF8", "1")

try:
    import yaml  # type: ignore
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "缺少 PyYAML。请使用项目虚拟环境运行：\n"
        r"C:\pinchbench-ccb\.venv\Scripts\python.exe -m pip install pyyaml"
    ) from exc

LOGGER = logging.getLogger("pinchbench-ccb")

RUNNER_REVISION = "2026-07-22-pinchbench-official-grader-api-ccb-windows-v1.4-stdin-session-fix"
DEFAULT_MODEL = "deepseek/deepseek-v4-pro"
PINCHBENCH_DEFAULT_JUDGE_MODEL = "openrouter/anthropic/claude-opus-5"
EXPECTED_CCB_VERSION = "2.8.4"
DEFAULT_PERMISSION_MODE = "acceptEdits"
OPENROUTER_ANTHROPIC_BASE_URL = "https://openrouter.ai/api"
DEFAULT_ALLOWED_TOOLS = (
    "Read",
    "Write",
    "Edit",
    "Bash",
    "PowerShell",
    "Glob",
    "Grep",
    "WebFetch",
    "WebSearch",
    "NotebookEdit",
)

PERMISSION_WAIT_RE = re.compile(
    r"permission\s+is\s+pending|once\s+you\s+approve|approval\s+required|"
    r"waiting\s+for\s+approval|go\s+ahead\s+and\s+approve",
    re.IGNORECASE,
)

# Agreed exclusions for this independent project. They require external
# GitHub/Google Workspace integration tooling and credentials that are not
# part of the current Windows OpenCode test environment.
DEFAULT_SKIPPED_TASKS = {
    "task_gh_issue_triage",
    "task_gws_email_triage",
    "task_gws_cross_service",
    "task_gws_task_management",
}

# Explicitly known internet-dependent tasks from the historical runner.
# Newer task files can also mark network_required: true.
NETWORK_TASKS = {
    "task_earnings_analysis",
    "task_stock",
    "task_market_research",
    "task_polymarket_briefing",
    "task_executive_lookup",
    "task_events",
    "task_deep_research",
    "task_competitive_research",
    "task_oss_alternative_research",
    "task_pricing_research",
    "task_it_procurement",
    "task_eu_regulation_research",
    "task_byok_best_practices",
}

TEXT_FILE_SUFFIXES = {
    ".txt", ".md", ".py", ".js", ".ts", ".tsx", ".jsx", ".json", ".yaml", ".yml",
    ".csv", ".tsv", ".html", ".css", ".xml", ".toml", ".ini", ".cfg", ".log",
    ".ics", ".sh", ".ps1", ".bat", ".cmd", ".sql", ".java", ".go", ".rs", ".rb", ".php",
}

SKIP_WORKSPACE_NAMES = {
    "BOOTSTRAP.md", "SOUL.md", "USER.md", "IDENTITY.md", "HEARTBEAT.md",
    "TOOLS.md", "AGENTS.md", "CLAUDE.md",
}
SKIP_WORKSPACE_DIRS = {
    ".git", ".openclaw", ".opencode", "__pycache__", "node_modules", "skills",
}


@dataclasses.dataclass
class Task:
    task_id: str
    name: str
    category: str
    grading_type: str
    timeout_seconds: int
    workspace_files: list[dict[str, Any]]
    prompt: str
    expected_behavior: str
    grading_criteria: list[str]
    automated_checks: str
    llm_judge_rubric: str
    grading_weights: dict[str, float]
    metadata: dict[str, Any]
    file_path: Path
    multi_session: bool = False
    sessions: list[dict[str, Any]] = dataclasses.field(default_factory=list)


@dataclasses.dataclass
class GradeResult:
    score: Optional[float]
    grading_type: str
    breakdown: dict[str, float]
    notes: str = ""
    error: Optional[str] = None


# ---------------------------------------------------------------------------
# Task loading
# ---------------------------------------------------------------------------

def parse_sections(body: str) -> dict[str, str]:
    sections: dict[str, list[str]] = {}
    current: Optional[str] = None
    for line in body.splitlines():
        match = re.match(r"^##\s+(.+?)\s*$", line)
        if match:
            current = match.group(1).strip()
            sections.setdefault(current, [])
        elif current:
            sections[current].append(line)
    return {key: "\n".join(value).strip() for key, value in sections.items()}


def extract_grading_criteria(text: str) -> list[str]:
    criteria: list[str] = []
    for line in text.splitlines():
        match = re.match(r"^-\s+\[[ xX]\]\s+(.+)$", line.strip())
        if match:
            criteria.append(match.group(1).strip())
    return criteria


def load_task_file(path: Path, category_override: Optional[str] = None) -> Task:
    raw = path.read_text(encoding="utf-8")
    match = re.match(r"^---\s*\r?\n(.*?)\r?\n---\s*\r?\n?(.*)$", raw, re.DOTALL)
    if not match:
        raise ValueError(f"{path} 没有 YAML frontmatter")

    metadata = yaml.safe_load(match.group(1)) or {}
    if not isinstance(metadata, dict):
        raise ValueError(f"{path} 的 YAML frontmatter 不是对象")

    body = match.group(2)
    sections = parse_sections(body)
    task_id = str(metadata.get("id") or path.stem)

    weights = metadata.get("grading_weights") or {}
    if not isinstance(weights, dict):
        weights = {}
    normalized_weights: dict[str, float] = {}
    for key, value in weights.items():
        try:
            normalized_weights[str(key)] = float(value)
        except (TypeError, ValueError):
            pass

    raw_sessions = metadata.get("sessions") or []
    sessions = [item for item in raw_sessions if isinstance(item, dict)] if isinstance(raw_sessions, list) else []

    return Task(
        task_id=task_id,
        name=str(metadata.get("name") or task_id),
        category=str(category_override or metadata.get("category") or ""),
        grading_type=str(metadata.get("grading_type") or "automated"),
        timeout_seconds=int(metadata.get("timeout_seconds") or 120),
        workspace_files=list(metadata.get("workspace_files") or []),
        prompt=sections.get("Prompt", "").strip(),
        expected_behavior=sections.get("Expected Behavior", "").strip(),
        grading_criteria=extract_grading_criteria(sections.get("Grading Criteria", "")),
        automated_checks=sections.get("Automated Checks", "") or "",
        llm_judge_rubric=sections.get("LLM Judge Rubric", "") or "",
        grading_weights=normalized_weights,
        metadata=metadata,
        file_path=path,
        multi_session=bool(metadata.get("multi_session") or sessions),
        sessions=sessions,
    )


def parse_manifest(tasks_dir: Path) -> tuple[list[str], dict[str, str], list[str]]:
    manifest_path = tasks_dir / "manifest.yaml"
    if not manifest_path.exists():
        ordered = [
            p.stem for p in sorted(tasks_dir.glob("task_*.md"))
            if p.stem != "task_XX_name"
        ]
        return ordered, {}, []

    manifest = yaml.safe_load(manifest_path.read_text(encoding="utf-8")) or {}
    if not isinstance(manifest, dict):
        raise ValueError(f"{manifest_path} 不是有效 YAML 对象")

    category_map: dict[str, str] = {}
    core_tasks = [str(x) for x in (manifest.get("core") or [])]

    if "categories" in manifest:
        run_first = [str(x) for x in (manifest.get("run_first") or [])]
        categories = manifest.get("categories") or {}
        all_task_ids: list[str] = []
        if isinstance(categories, dict):
            for category, ids in categories.items():
                for task_id in ids or []:
                    task_id = str(task_id)
                    category_map[task_id] = str(category)
                    all_task_ids.append(task_id)

        seen: set[str] = set()
        ordered: list[str] = []
        for task_id in run_first + all_task_ids:
            if task_id not in seen:
                seen.add(task_id)
                ordered.append(task_id)
        return ordered, category_map, core_tasks

    ordered = [str(x) for x in (manifest.get("tasks") or [])]
    return ordered, category_map, core_tasks


def load_tasks(tasks_dir: Path) -> tuple[list[Task], list[str]]:
    """Load exactly the checked-out manifest when it exists.

    PinchBench ships template/example Markdown files alongside real tasks.
    Falling back to every task_*.md while a manifest exists can accidentally
    count a template as a benchmark task.  The manifest is therefore the
    source of truth for normal runs.
    """
    manifest_exists = (tasks_dir / "manifest.yaml").exists()
    ordered_ids, category_map, core_tasks = parse_manifest(tasks_dir)
    tasks: list[Task] = []

    for task_id in ordered_ids:
        if task_id == "task_XX_name":
            continue
        task_path = tasks_dir / f"{task_id}.md"
        if not task_path.exists():
            LOGGER.warning("manifest 引用了不存在的任务文件: %s", task_path)
            continue
        try:
            task = load_task_file(
                task_path,
                category_override=category_map.get(task_id),
            )
            if (
                task.task_id == "task_XX_name"
                or task.category.strip().lower() == "category_name"
            ):
                LOGGER.debug("忽略 PinchBench 模板任务: %s", task_path)
                continue
            tasks.append(task)
        except Exception as exc:
            LOGGER.warning("加载任务失败 %s: %s", task_path, exc)

    if manifest_exists:
        return tasks, core_tasks

    # Only use file discovery when no manifest is available.
    loaded_ids = {task.task_id for task in tasks}
    for task_path in sorted(tasks_dir.glob("task_*.md")):
        if task_path.stem in loaded_ids or task_path.stem == "task_XX_name":
            continue
        try:
            task = load_task_file(task_path)
            if (
                task.task_id == "task_XX_name"
                or task.category.strip().lower() == "category_name"
            ):
                continue
            tasks.append(task)
        except Exception as exc:
            LOGGER.warning("加载任务失败 %s: %s", task_path, exc)

    return tasks, core_tasks


def filter_tasks(
    tasks: list[Task],
    core_tasks: list[str],
    suite: str,
    limit: Optional[int],
) -> list[Task]:
    suite = suite.strip()
    task_by_id = {task.task_id: task for task in tasks}

    if suite == "all":
        selected = tasks
    elif suite == "core":
        core_set = set(core_tasks)
        selected = [task for task in tasks if task.task_id in core_set] if core_set else tasks[:25]
    elif suite == "automated-only":
        selected = [task for task in tasks if task.grading_type == "automated"]
    elif suite == "llm-judge-only":
        selected = [task for task in tasks if task.grading_type == "llm_judge"]
    elif suite == "hybrid-only":
        selected = [task for task in tasks if task.grading_type == "hybrid"]
    elif suite == "judge-required-only":
        selected = [task for task in tasks if task.grading_type in {"llm_judge", "hybrid"}]
    else:
        ids = [item.strip() for item in suite.split(",") if item.strip()]
        missing = [task_id for task_id in ids if task_id not in task_by_id]
        if missing:
            raise SystemExit(f"suite 中有未知任务: {', '.join(missing)}")
        selected = [task_by_id[task_id] for task_id in ids]

    if limit is not None:
        selected = selected[:limit]
    return selected


def is_network_task(task: Task) -> bool:
    return bool(
        task.task_id in NETWORK_TASKS
        or task.metadata.get("network_required") is True
        or task.metadata.get("requires_network") is True
    )


# ---------------------------------------------------------------------------
# Workspace and prerequisites
# ---------------------------------------------------------------------------

def safe_workspace_dest(dest: str) -> Path:
    rel = Path(dest)
    if rel.is_absolute() or ".." in rel.parts:
        raise ValueError(f"非法 workspace 文件路径: {dest}")
    return rel


def stage_workspace_files(
    task: Task,
    workspace: Path,
    skill_dir: Path,
) -> tuple[list[str], list[str]]:
    staged: list[str] = []
    missing: list[str] = []

    for wf in task.workspace_files:
        if not isinstance(wf, dict):
            missing.append(f"无法识别的 workspace_files 项: {wf!r}")
            continue

        dest_value = wf.get("dest") or wf.get("path") or wf.get("name")
        source_value = wf.get("source")
        content_value = wf.get("content")

        if not dest_value and source_value:
            dest_value = Path(str(source_value)).name
        if not dest_value:
            missing.append(f"workspace_files 项缺少 dest/path: {wf!r}")
            continue

        try:
            dest_rel = safe_workspace_dest(str(dest_value))
        except ValueError as exc:
            missing.append(str(exc))
            continue

        dest_path = workspace / dest_rel
        dest_path.parent.mkdir(parents=True, exist_ok=True)

        if content_value is not None:
            dest_path.write_text(str(content_value), encoding="utf-8")
            staged.append(str(dest_rel))
            continue

        if not source_value:
            missing.append(f"workspace_files 项缺少 source/content: {wf!r}")
            continue

        src_rel = Path(str(source_value))
        candidates = [
            skill_dir / "assets" / src_rel,
            skill_dir / src_rel,
            task.file_path.parent / src_rel,
            task.file_path.parent.parent / "assets" / src_rel,
        ]
        src_path = next((candidate for candidate in candidates if candidate.exists()), None)
        if src_path is None:
            missing.append(
                f"找不到预置文件 {source_value}; 尝试过: "
                + ", ".join(str(candidate) for candidate in candidates)
            )
            continue

        if src_path.is_dir():
            if dest_path.exists():
                shutil.rmtree(dest_path)
            shutil.copytree(src_path, dest_path)
        else:
            shutil.copy2(src_path, dest_path)
        staged.append(str(dest_rel))

    return staged, missing


def resolve_command(name: str) -> Optional[str]:
    candidates = [name]
    if sys.platform == "win32" and not Path(name).suffix:
        candidates = [f"{name}.cmd", f"{name}.exe", name]
    for candidate in candidates:
        resolved = shutil.which(candidate)
        if resolved:
            return resolved
    return None


def check_prerequisites(prereqs: list[str]) -> list[str]:
    missing: list[str] = []
    npm_command = resolve_command("npm") or "npm"

    for req_raw in prereqs:
        req = str(req_raw)
        if req.startswith("cli:"):
            command = req.split(":", 1)[1].strip()
            if not resolve_command(command):
                missing.append(req)
        elif req.startswith("npm:"):
            package = req.split(":", 1)[1].strip()
            try:
                result = subprocess.run(
                    [npm_command, "list", "-g", package, "--depth=0"],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=30,
                    check=False,
                )
                if result.returncode != 0:
                    missing.append(req)
            except Exception:
                missing.append(req)

    return missing


def choose_ccb_command(preferred: str) -> str:
    if preferred and preferred != "auto":
        preferred_path = Path(preferred).expanduser()
        if preferred_path.exists():
            return str(preferred_path.resolve())
        resolved = resolve_command(preferred)
        if resolved:
            return resolved
        raise SystemExit(f"找不到 CCB 命令 {preferred!r}")

    # Use the Node entry first.  The paired environment pins claude-code-best
    # 2.8.4 installed by npm; ccb-bun is intentionally only a fallback.
    commands = (
        ("ccb.cmd", "ccb.exe", "ccb", "ccb-bun.cmd", "ccb-bun.exe", "ccb-bun")
        if sys.platform == "win32"
        else ("ccb", "ccb-bun")
    )
    for command in commands:
        resolved = shutil.which(command)
        if resolved:
            return resolved

    raise SystemExit(
        "找不到 ccb/ccb-bun。请先运行 npm install -g claude-code-best@2.8.4，"
        "并确认 ccb.cmd 在 PATH 中。"
    )

def kill_proc_tree(proc: subprocess.Popen[str]) -> None:
    if proc.poll() is not None:
        return
    try:
        if sys.platform == "win32":
            subprocess.run(
                ["taskkill", "/F", "/T", "/PID", str(proc.pid)],
                capture_output=True,
                check=False,
            )
        else:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
    except Exception:
        try:
            proc.kill()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# CCB stream-json event handling
# ---------------------------------------------------------------------------

def _content_to_text(content: Any) -> str:
    if content is None:
        return ""
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        chunks: list[str] = []
        for item in content:
            if isinstance(item, str):
                chunks.append(item)
            elif isinstance(item, dict):
                item_type = str(item.get("type") or "")
                if item_type in {"text", "output_text"}:
                    chunks.append(str(item.get("text") or ""))
                elif item_type == "tool_result":
                    chunks.append(_content_to_text(item.get("content")))
        return "".join(chunks)
    if isinstance(content, dict):
        return _content_to_text(content.get("text") or content.get("content"))
    return str(content)


def ccb_assistant_text(event: dict[str, Any]) -> str:
    if event.get("type") != "assistant":
        return ""
    message = event.get("message")
    if not isinstance(message, dict):
        return ""
    content = message.get("content")
    if isinstance(content, str):
        return content
    if not isinstance(content, list):
        return ""
    chunks: list[str] = []
    for item in content:
        if isinstance(item, dict) and item.get("type") in {"text", "output_text"}:
            chunks.append(str(item.get("text") or ""))
    return "".join(chunks)


def ccb_thinking_text(event: dict[str, Any]) -> str:
    if event.get("type") != "assistant":
        return ""
    message = event.get("message")
    if not isinstance(message, dict):
        return ""
    content = message.get("content")
    if not isinstance(content, list):
        return ""
    chunks: list[str] = []
    for item in content:
        if isinstance(item, dict) and item.get("type") in {"thinking", "reasoning"}:
            chunks.append(str(item.get("thinking") or item.get("reasoning") or item.get("text") or ""))
    return "".join(chunks)


def extract_ccb_event_error(event: dict[str, Any]) -> str:
    event_type = str(event.get("type") or "")
    if event_type == "error":
        error = event.get("error") or event.get("message") or event.get("detail")
        if isinstance(error, dict):
            return str(error.get("message") or json.dumps(error, ensure_ascii=False))
        return str(error or "CCB emitted an error event")
    if event_type == "result" and event.get("is_error"):
        return str(event.get("result") or event.get("error") or "CCB returned is_error=true")
    return ""


def _to_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _to_float(value: Any) -> float:
    try:
        return float(value or 0.0)
    except (TypeError, ValueError):
        return 0.0


def parse_ccb_result(event: dict[str, Any]) -> Optional[dict[str, Any]]:
    if event.get("type") != "result":
        return None
    usage = event.get("usage") or {}
    if not isinstance(usage, dict):
        usage = {}
    permission_denials = event.get("permission_denials") or []
    if not isinstance(permission_denials, list):
        permission_denials = [permission_denials]
    model_usage = event.get("modelUsage") or {}
    if not isinstance(model_usage, dict):
        model_usage = {}
    return {
        "input_tokens": _to_int(usage.get("input_tokens")),
        "output_tokens": _to_int(usage.get("output_tokens")),
        "reasoning_tokens": None,
        "cache_read_tokens": _to_int(usage.get("cache_read_input_tokens")),
        "cache_write_tokens": _to_int(usage.get("cache_creation_input_tokens")),
        "cost_usd": _to_float(event.get("total_cost_usd")),
        "step_count": _to_int(event.get("num_turns")),
        "finish_reason": str(event.get("stop_reason") or event.get("subtype") or ""),
        "final_result": str(event.get("result") or ""),
        "is_error": bool(event.get("is_error")),
        "permission_denials": permission_denials,
        "model_usage": model_usage,
    }


def _ccb_message_provider(event: dict[str, Any]) -> str:
    message = event.get("message")
    if isinstance(message, dict):
        return str(message.get("provider") or "")
    return ""


def _ccb_message_model(event: dict[str, Any]) -> str:
    message = event.get("message")
    if isinstance(message, dict):
        return str(message.get("model") or "")
    return ""


def normalize_ccb_event(event: dict[str, Any]) -> list[dict[str, Any]]:
    """Convert completed CCB message blocks to PinchBench's generic transcript."""
    event_type = str(event.get("type") or "")
    normalized: list[dict[str, Any]] = []

    if event_type == "assistant":
        message = event.get("message")
        if not isinstance(message, dict):
            return normalized
        content = message.get("content")
        if not isinstance(content, list):
            if isinstance(content, str) and content:
                content = [{"type": "text", "text": content}]
            else:
                return normalized

        converted: list[dict[str, Any]] = []
        for item in content:
            if not isinstance(item, dict):
                continue
            item_type = str(item.get("type") or "")
            if item_type in {"text", "output_text"}:
                text = str(item.get("text") or "")
                if text:
                    converted.append({"type": "text", "text": text})
            elif item_type in {"tool_use", "toolCall", "tool_call"}:
                call_id = str(item.get("id") or item.get("tool_use_id") or "")
                tool_name = str(item.get("name") or item.get("tool") or "")
                tool_input = item.get("input") or item.get("arguments") or {}
                if not isinstance(tool_input, dict):
                    tool_input = {"value": tool_input}
                converted.append({
                    "type": "toolCall",
                    "id": call_id,
                    "name": tool_name,
                    "arguments": tool_input,
                    "input": tool_input,
                })
            # Thinking is intentionally omitted from the grader transcript.  The
            # paired OpenCode normalizer also excludes private reasoning blocks.

        if converted:
            normalized.append({
                "type": "message",
                "message": {"role": "assistant", "content": converted},
                "_adapter": {
                    "source": "ccb",
                    "event_type": event_type,
                    "provider": str(message.get("provider") or ""),
                    "model": str(message.get("model") or ""),
                },
            })
        return normalized

    if event_type == "user":
        message = event.get("message")
        if not isinstance(message, dict):
            return normalized
        content = message.get("content")
        if not isinstance(content, list):
            return normalized
        for item in content:
            if not isinstance(item, dict) or item.get("type") != "tool_result":
                continue
            call_id = str(item.get("tool_use_id") or item.get("toolCallId") or item.get("id") or "")
            result_text = _content_to_text(item.get("content"))
            normalized.append({
                "type": "message",
                "message": {
                    "role": "toolResult",
                    "content": [{
                        "type": "toolResult",
                        "toolCallId": call_id,
                        "name": str(item.get("name") or ""),
                        "content": result_text,
                        "isError": bool(item.get("is_error")),
                    }],
                },
                "_adapter": {"source": "ccb", "event_type": event_type},
            })
        return normalized

    error_text = extract_ccb_event_error(event)
    if error_text:
        normalized.append({
            "type": "message",
            "message": {
                "role": "assistant",
                "content": [{"type": "text", "text": f"[CCB error] {error_text}"}],
            },
            "_adapter": {"source": "ccb", "event_type": event_type},
        })
    return normalized


def make_user_transcript_event(prompt: str, turn_index: int, session_label: str) -> dict[str, Any]:
    return {
        "type": "message",
        "message": {
            "role": "user",
            "content": [{"type": "text", "text": prompt}],
        },
        "_adapter": {
            "source": "pinchbench-ccb-runner",
            "turn": turn_index,
            "session_label": session_label,
        },
    }


def _stream_reader(
    stream: Any,
    kind: str,
    output_queue: "queue.Queue[tuple[str, Optional[str]]]",
) -> None:
    try:
        for line in iter(stream.readline, ""):
            output_queue.put((kind, line))
    finally:
        output_queue.put((kind, None))
        try:
            stream.close()
        except Exception:
            pass


def _stdin_writer(stream: Any, prompt: str, errors: list[str]) -> None:
    """Write one prompt to CCB stdin without exposing it to Windows .cmd parsing."""
    try:
        stream.write(prompt)
        if not prompt.endswith("\n"):
            stream.write("\n")
        stream.flush()
    except Exception as exc:
        errors.append(f"写入 CCB stdin 失败: {exc}")
    finally:
        try:
            stream.close()
        except Exception:
            pass


def build_ccb_command(
    *,
    ccb_command: str,
    model: str,
    permission_mode: str,
    allowed_tools: list[str],
    session_id: str,
    resume: bool,
) -> list[str]:
    """Build a Windows-safe CCB command.

    Important invariants:
    - Prompt is NOT part of argv; run_ccb_streaming sends it through stdin.
    - All options are placed before the final -p flag.
    - --allowedTools receives one comma-separated value, rather than a variadic
      tail that can swallow following options in Windows batch-file parsing.
    - New sessions use an explicit UUID; continuation turns use --resume.
    """
    command = [ccb_command]
    if resume:
        command.extend(["--resume", session_id])
    else:
        command.extend(["--session-id", session_id])

    command.extend([
        "--model", model,
        "--input-format", "text",
        "--output-format", "stream-json",
        "--permission-mode", permission_mode,
        "--verbose",
        "--include-partial-messages",
    ])
    if allowed_tools:
        command.extend(["--allowedTools", ",".join(allowed_tools)])
    command.append("-p")
    return command


def run_ccb_streaming(
    cmd: list[str],
    cwd: Path,
    timeout: float,
    raw_stdout_path: Path,
    stderr_path: Path,
    prompt: str,
    expected_session_id: str,
) -> dict[str, Any]:
    """Run one CCB print-mode turn with stdin prompt transport.

    CCB is required to emit stream-json with a terminal result event. Plain
    text stdout is retained for diagnosis and normalized as a fallback message,
    but it is treated as a protocol error so an invalid benchmark cannot pass.
    """
    monotonic_start = time.monotonic()
    deadline = monotonic_start + timeout

    raw_stdout_path.parent.mkdir(parents=True, exist_ok=True)
    stderr_path.parent.mkdir(parents=True, exist_ok=True)

    stdout_lines: list[str] = []
    stderr_lines: list[str] = []
    raw_events: list[dict[str, Any]] = []
    normalized_events: list[dict[str, Any]] = []
    output_chunks: list[str] = []
    plain_stdout_chunks: list[str] = []
    event_counts: Counter[str] = Counter()

    session_id: Optional[str] = None
    ttft: Optional[float] = None
    error_messages: list[str] = []
    stdin_errors: list[str] = []
    status = "success"
    timed_out = False

    usage_seen = False
    result_event_seen = False
    assistant_text_seen = False
    input_tokens: Optional[int] = None
    output_tokens: Optional[int] = None
    reasoning_tokens: Optional[int] = None
    cache_read_tokens: Optional[int] = None
    cache_write_tokens: Optional[int] = None
    cost_usd: Optional[float] = None
    step_count = 0
    tool_errors = 0
    finish_reasons: list[str] = []
    providers: set[str] = set()
    models: set[str] = set()
    permission_denials: list[Any] = []
    model_usage: dict[str, Any] = {}
    thinking_char_count = 0
    final_result = ""

    env = os.environ.copy()
    env.setdefault("PYTHONUTF8", "1")
    env.setdefault("PYTHONIOENCODING", "utf-8")
    env.setdefault("NO_COLOR", "1")

    creationflags = 0
    if sys.platform == "win32":
        creationflags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)

    try:
        proc = subprocess.Popen(
            cmd,
            cwd=str(cwd),
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
            errors="replace",
            bufsize=1,
            env=env,
            start_new_session=(sys.platform != "win32"),
            creationflags=creationflags,
        )
    except Exception as exc:
        return {
            "status": "error", "success": False, "returncode": None,
            "output": "", "error": f"无法启动 CCB: {exc}", "stderr": "",
            "elapsed": time.monotonic() - monotonic_start, "ttft": None,
            "input_tokens": None, "output_tokens": None, "reasoning_tokens": None,
            "cache_read_tokens": None, "cache_write_tokens": None, "cost_usd": None,
            "usage_complete": False, "step_count": 0, "tool_errors": 0,
            "finish_reasons": [], "session_id": expected_session_id, "raw_events": [],
            "normalized_events": [], "event_counts": {}, "providers": [],
            "models": [], "permission_denials": [], "model_usage": {},
            "thinking_char_count": 0, "stream_json_ok": False,
            "protocol_mode": "launch_error",
        }

    if proc.stdin is None or proc.stdout is None or proc.stderr is None:
        kill_proc_tree(proc)
        raise RuntimeError("CCB stdin/stdout/stderr pipe creation failed")

    output_queue: "queue.Queue[tuple[str, Optional[str]]]" = queue.Queue()
    reader_threads = [
        threading.Thread(target=_stream_reader, args=(proc.stdout, "stdout", output_queue), daemon=True),
        threading.Thread(target=_stream_reader, args=(proc.stderr, "stderr", output_queue), daemon=True),
    ]
    for thread in reader_threads:
        thread.start()

    stdin_thread = threading.Thread(
        target=_stdin_writer,
        args=(proc.stdin, prompt, stdin_errors),
        daemon=True,
    )
    stdin_thread.start()

    open_streams = {"stdout", "stderr"}
    with (
        raw_stdout_path.open("w", encoding="utf-8", newline="") as raw_file,
        stderr_path.open("w", encoding="utf-8", newline="") as err_file,
    ):
        while open_streams or proc.poll() is None:
            if time.monotonic() > deadline:
                timed_out = True
                status = "timeout"
                error_messages.append(f"超时 ({timeout:.0f}s)")
                kill_proc_tree(proc)
                break

            try:
                kind, line = output_queue.get(timeout=0.2)
            except queue.Empty:
                continue

            if line is None:
                open_streams.discard(kind)
                continue

            if kind == "stderr":
                stderr_lines.append(line)
                err_file.write(line)
                err_file.flush()
                continue

            stdout_lines.append(line)
            raw_file.write(line)
            raw_file.flush()
            stripped = line.strip()
            if not stripped:
                continue

            try:
                event = json.loads(stripped)
            except json.JSONDecodeError:
                if ttft is None:
                    ttft = time.monotonic() - monotonic_start
                plain_stdout_chunks.append(stripped)
                output_chunks.append(stripped)
                event_counts["non_json_stdout"] += 1
                continue

            if not isinstance(event, dict):
                event_counts["non_object_json"] += 1
                continue

            raw_events.append(event)
            event_type = str(event.get("type") or "unknown")
            subtype = str(event.get("subtype") or "")
            event_counts[event_type] += 1
            if subtype:
                event_counts[f"{event_type}.{subtype}"] += 1

            candidate_session = event.get("session_id") or event.get("sessionID")
            if candidate_session:
                session_id = str(candidate_session)

            provider = _ccb_message_provider(event)
            model = _ccb_message_model(event)
            if provider:
                providers.add(provider)
            if model:
                models.add(model)

            text = ccb_assistant_text(event)
            if text:
                assistant_text_seen = True
                if ttft is None:
                    ttft = time.monotonic() - monotonic_start
                output_chunks.append(text)

            thinking_char_count += len(ccb_thinking_text(event))
            normalized_events.extend(normalize_ccb_event(event))

            if event_type == "user":
                message = event.get("message")
                content = message.get("content") if isinstance(message, dict) else None
                if isinstance(content, list):
                    for item in content:
                        if isinstance(item, dict) and item.get("type") == "tool_result" and item.get("is_error"):
                            tool_errors += 1

            parsed_result = parse_ccb_result(event)
            if parsed_result is not None:
                result_event_seen = True
                usage_seen = True
                input_tokens = parsed_result["input_tokens"]
                output_tokens = parsed_result["output_tokens"]
                reasoning_tokens = parsed_result["reasoning_tokens"]
                cache_read_tokens = parsed_result["cache_read_tokens"]
                cache_write_tokens = parsed_result["cache_write_tokens"]
                cost_usd = parsed_result["cost_usd"]
                step_count = parsed_result["step_count"]
                if parsed_result["finish_reason"]:
                    finish_reasons.append(parsed_result["finish_reason"])
                final_result = parsed_result["final_result"]
                permission_denials = parsed_result["permission_denials"]
                model_usage = parsed_result["model_usage"]
                if parsed_result["is_error"]:
                    status = "error"

            event_error = extract_ccb_event_error(event)
            if event_error:
                error_messages.append(event_error)
                status = "error"

    try:
        proc.wait(timeout=15)
    except subprocess.TimeoutExpired:
        kill_proc_tree(proc)
        status = "timeout"
        timed_out = True
        error_messages.append("进程未正常退出，已强制终止")

    stdin_thread.join(timeout=2)
    for thread in reader_threads:
        thread.join(timeout=1)

    if stdin_errors:
        error_messages.extend(stdin_errors)
        if status == "success":
            status = "stdin_error"

    returncode = proc.returncode
    stderr = "".join(stderr_lines).strip()
    if returncode not in (0, None) and status == "success":
        status = "error"
        error_messages.append(stderr or f"CCB 退出码 {returncode}")

    stream_json_ok = bool(raw_events and result_event_seen)
    protocol_mode = "stream-json" if stream_json_ok else "plain-text-fallback"

    # Preserve diagnostic value and make automated text graders see the answer,
    # but never let a plain-text fallback count as a valid benchmark execution.
    fallback_text = "\n".join(plain_stdout_chunks).strip() or final_result.strip()
    if not assistant_text_seen and fallback_text:
        normalized_events.append({
            "type": "message",
            "message": {
                "role": "assistant",
                "content": [{"type": "text", "text": fallback_text}],
            },
            "_adapter": {
                "source": "ccb",
                "event_type": "plain_text_fallback",
                "protocol_error": True,
            },
        })

    if status == "success" and not stream_json_ok:
        status = "protocol_error"
        error_messages.append(
            "CCB 未按要求返回 stream-json/result 事件；已保存纯文本用于诊断，"
            "本次执行不计为有效 benchmark 结果"
        )

    effective_session_id = session_id or expected_session_id
    if stream_json_ok and session_id != expected_session_id:
        status = "session_mismatch"
        error_messages.append(
            f"CCB 返回 session_id={session_id or '(missing)'}，"
            f"与请求的 {expected_session_id} 不一致"
        )

    output = "\n".join(chunk.strip() for chunk in output_chunks if chunk.strip()).strip()
    if not output:
        output = final_result.strip()

    if PERMISSION_WAIT_RE.search(output):
        status = "permission_blocked"
        error_messages.append("CCB 输出显示仍在等待人工权限批准")

    success = (
        status == "success"
        and not timed_out
        and returncode in (0, None)
        and stream_json_ok
    )

    return {
        "status": status,
        "success": success,
        "returncode": returncode,
        "output": output,
        "error": " | ".join(dict.fromkeys(message for message in error_messages if message)),
        "stderr": stderr,
        "elapsed": time.monotonic() - monotonic_start,
        "ttft": ttft,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "reasoning_tokens": reasoning_tokens,
        "cache_read_tokens": cache_read_tokens,
        "cache_write_tokens": cache_write_tokens,
        "cost_usd": cost_usd,
        "usage_complete": usage_seen,
        "step_count": step_count,
        "tool_errors": tool_errors,
        "finish_reasons": finish_reasons,
        "session_id": effective_session_id,
        "raw_events": raw_events,
        "normalized_events": normalized_events,
        "event_counts": dict(event_counts),
        "stdout_lines": stdout_lines,
        "providers": sorted(providers),
        "models": sorted(models),
        "permission_denials": permission_denials,
        "model_usage": model_usage,
        "thinking_char_count": thinking_char_count,
        "stream_json_ok": stream_json_ok,
        "protocol_mode": protocol_mode,
    }

def load_pinchbench_grading(skill_dir: Path) -> tuple[Optional[Any], Optional[str]]:
    """Load the grading engine from the exact checked-out PinchBench commit."""
    scripts_dir = (skill_dir / "scripts").resolve()
    grading_path = scripts_dir / "lib_grading.py"
    if not grading_path.exists():
        return None, f"找不到 PinchBench grading engine: {grading_path}"

    scripts_value = str(scripts_dir)
    if scripts_value not in sys.path:
        sys.path.insert(0, scripts_value)

    try:
        import lib_grading  # type: ignore
    except Exception as exc:
        return None, (
            "无法导入 PinchBench scripts/lib_grading.py。"
            "请安装本地仓库依赖，例如："
            r"C:\pinchbench-ccb\.venv\Scripts\python.exe -m pip install -e "
            f'"{skill_dir}"。原始错误: {exc}'
        )

    return lib_grading, None


def grade_with_pinchbench_default(
    *,
    grader_module: Any,
    task: Task,
    execution_result: dict[str, Any],
    workspace: Path,
    skill_dir: Path,
    judge_timeout: float,
    verbose: bool,
) -> GradeResult:
    """Use PinchBench's own grader with its checked-out default judge model.

    The judge is called directly through the API backend.  For the current
    commit, DEFAULT_JUDGE_MODEL is openrouter/anthropic/claude-opus-5 and
    lib_agent.py reads OPENROUTER_API_KEY.
    """
    grading_execution = dict(execution_result)
    grading_execution["workspace"] = str(workspace)

    try:
        result = grader_module.grade_task(
            task=task,
            execution_result=grading_execution,
            skill_dir=skill_dir,
            judge_model=getattr(
                grader_module,
                "DEFAULT_JUDGE_MODEL",
                PINCHBENCH_DEFAULT_JUDGE_MODEL,
            ),
            judge_timeout_seconds=judge_timeout,
            judge_backend="api",
            verbose=verbose,
        )

        raw_score = getattr(result, "score", None)
        max_score = getattr(result, "max_score", 1.0)
        score: Optional[float]
        if raw_score is None:
            score = None
        else:
            score = float(raw_score)
            max_score_float = float(max_score or 1.0)
            if max_score_float != 1.0:
                score = score / max_score_float
            score = max(0.0, min(1.0, score))

        notes = str(getattr(result, "notes", "") or "")
        grade_error: Optional[str] = None
        if task.grading_type in {"llm_judge", "hybrid"}:
            lower_notes = notes.lower()
            failure_markers = (
                "llm judge failed",
                "no parseable response",
                "response parsed but no score",
                "openrouter_api_key not set",
                "judge api call failed",
            )
            if any(marker in lower_notes for marker in failure_markers):
                grade_error = notes or "LLM judge failed"

        return GradeResult(
            score=score,
            grading_type=str(getattr(result, "grading_type", task.grading_type)),
            breakdown=dict(getattr(result, "breakdown", {}) or {}),
            notes=notes,
            error=grade_error,
        )
    except Exception as exc:
        return GradeResult(
            score=None,
            grading_type=task.grading_type,
            breakdown={},
            notes="",
            error=(
                "PinchBench 官方评分失败: "
                f"{exc}\n{traceback.format_exc(limit=8)}"
            ),
        )


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def truncate_excel(value: Any, limit: int = 32000) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    return text if len(text) <= limit else text[:limit] + "...[truncated]"


def save_csv(results: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "task_id", "name", "category", "grading_type", "network_task",
        "multi_session", "session_count", "success", "status", "returncode",
        "score", "elapsed", "ttft", "input_tokens", "output_tokens",
        "reasoning_tokens", "cache_read_tokens", "cache_write_tokens",
        "cost_usd", "usage_complete", "stream_json_complete", "protocol_modes",
        "step_count", "tool_errors", "providers", "unexpected_models", "permission_denials_count", "thinking_char_count",
        "workspace", "transcript", "error", "grade_error", "grade_notes",
    ]
    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for row in results:
            export = dict(row)
            export["providers"] = ",".join(row.get("providers") or [])
            export["unexpected_models"] = ",".join(row.get("unexpected_models") or [])
            export["permission_denials_count"] = len(row.get("permission_denials") or [])
            export["protocol_modes"] = ",".join(row.get("protocol_modes") or [])
            writer.writerow({key: truncate_excel(export.get(key), 10000) for key in fields})

def save_xlsx(
    results: list[dict[str, Any]],
    summary: dict[str, Any],
    output_path: Path,
) -> None:
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
    except ImportError:
        LOGGER.warning("未安装 openpyxl，跳过 XLSX。可运行: python -m pip install openpyxl")
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "详细结果"
    headers = [
        "任务ID", "名称", "类别", "打分类型", "联网题", "多轮任务",
        "会话调用数", "成功", "状态", "退出码", "分数", "耗时(s)",
        "TTFT估计(s)", "输入Token", "输出Token", "推理Token",
        "缓存读取Token", "缓存写入Token", "费用USD", "Usage完整",
        "StreamJSON完整", "协议模式", "Step数", "工具错误数", "上游Provider", "非目标模型", "权限拒绝数", "Thinking字符数",
        "Workspace", "Transcript", "运行错误", "打分错误", "打分备注",
    ]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(results, start=2):
        values = [
            row.get("task_id"), row.get("name"), row.get("category"), row.get("grading_type"),
            "是" if row.get("network_task") else "否",
            "是" if row.get("multi_session") else "否",
            row.get("session_count"), "✓" if row.get("success") else "✗",
            row.get("status"), row.get("returncode") if row.get("returncode") is not None else "",
            row.get("score") if row.get("score") is not None else "", row.get("elapsed"),
            row.get("ttft") if row.get("ttft") is not None else "",
            row.get("input_tokens") if row.get("input_tokens") is not None else "",
            row.get("output_tokens") if row.get("output_tokens") is not None else "",
            row.get("reasoning_tokens") if row.get("reasoning_tokens") is not None else "",
            row.get("cache_read_tokens") if row.get("cache_read_tokens") is not None else "",
            row.get("cache_write_tokens") if row.get("cache_write_tokens") is not None else "",
            row.get("cost_usd") if row.get("cost_usd") is not None else "",
            "是" if row.get("usage_complete") else "否",
            "是" if row.get("stream_json_complete") else "否",
            ",".join(row.get("protocol_modes") or []),
            row.get("step_count"), row.get("tool_errors"),
            ",".join(row.get("providers") or []), ",".join(row.get("unexpected_models") or []),
            len(row.get("permission_denials") or []), row.get("thinking_char_count") or 0, row.get("workspace"), row.get("transcript"),
            row.get("error") or "", row.get("grade_error") or "", row.get("grade_notes") or "",
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column, value=truncate_excel(value))

    for cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in cells)
        worksheet.column_dimensions[cells[0].column_letter].width = min(max_length + 3, 60)

    summary_sheet = workbook.create_sheet("汇总")
    for row_index, (key, value) in enumerate(summary.items(), start=1):
        summary_sheet.cell(row=row_index, column=1, value=key).font = Font(bold=True)
        if isinstance(value, (dict, list, tuple, set)):
            value = json.dumps(value, ensure_ascii=False)
        summary_sheet.cell(row=row_index, column=2, value=value)
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 40
    workbook.save(output_path)

def optional_sum(results: list[dict[str, Any]], field: str) -> Optional[float]:
    values = [row.get(field) for row in results if row.get(field) is not None]
    if not values:
        return None
    return float(sum(float(value) for value in values))


def print_summary(
    results: list[dict[str, Any]],
    total_elapsed: float,
) -> dict[str, Any]:
    succeeded = sum(1 for row in results if row.get("success"))
    failed = len(results) - succeeded
    scores = [float(row["score"]) for row in results if row.get("score") is not None]
    ttfts = [float(row["ttft"]) for row in results if row.get("ttft") is not None]
    average_elapsed = (
        sum(float(row.get("elapsed") or 0.0) for row in results) / len(results)
        if results else 0.0
    )
    provider_counts: Counter[str] = Counter()
    for row in results:
        provider_counts.update(row.get("providers") or [])

    summary = {
        "任务总数": len(results),
        "成功": succeeded,
        "失败": failed,
        "成功率": round(succeeded / len(results), 4) if results else 0.0,
        "有分数任务数": len(scores),
        "打分失败任务数": sum(1 for row in results if row.get("grade_error")),
        "平均分数": round(sum(scores) / len(scores), 4) if scores else None,
        "总耗时(s)": round(total_elapsed, 2),
        "平均耗时/任务(s)": round(average_elapsed, 2),
        "平均TTFT估计(s)": round(sum(ttfts) / len(ttfts), 4) if ttfts else None,
        "总输入Token": int(optional_sum(results, "input_tokens") or 0)
            if any(row.get("input_tokens") is not None for row in results) else None,
        "总输出Token": int(optional_sum(results, "output_tokens") or 0)
            if any(row.get("output_tokens") is not None for row in results) else None,
        "总推理Token": int(optional_sum(results, "reasoning_tokens") or 0)
            if any(row.get("reasoning_tokens") is not None for row in results) else None,
        "总费用USD": round(optional_sum(results, "cost_usd") or 0.0, 6)
            if any(row.get("cost_usd") is not None for row in results) else None,
        "Usage缺失任务数": sum(1 for row in results if not row.get("usage_complete")),
        "StreamJSON失败任务数": sum(1 for row in results if not row.get("stream_json_complete")),
        "联网任务数": sum(1 for row in results if row.get("network_task")),
        "多轮任务数": sum(1 for row in results if row.get("multi_session")),
        "权限拒绝任务数": sum(1 for row in results if row.get("permission_denials")),
        "上游Provider计数": dict(provider_counts),
    }

    print("\n" + "=" * 100)
    print("汇总")
    print("=" * 100)
    for key, value in summary.items():
        print(f"{key:<22}: {value}")
    print("-" * 100)
    for row in results:
        marker = "✓" if row.get("success") else "✗"
        score = f"{row['score']:.3f}" if row.get("score") is not None else "N/A"
        ttft = f"{row['ttft']:.2f}s" if row.get("ttft") is not None else "N/A"
        error = row.get("error") or row.get("grade_error") or ""
        error_suffix = f"  [{str(error)[:70]}]" if error else ""
        print(
            f"{marker} {row['task_id']:<42} score={score:<6} "
            f"elapsed={row['elapsed']:.1f}s ttft={ttft:<8}{error_suffix}"
        )
    print("=" * 100)
    return summary

# ---------------------------------------------------------------------------
# Preflight and run metadata
# ---------------------------------------------------------------------------

def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def command_output(command: list[str], cwd: Optional[Path] = None) -> str:
    try:
        result = subprocess.run(
            command,
            cwd=str(cwd) if cwd else None,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
            check=False,
        )
        return (result.stdout or result.stderr).strip()
    except Exception as exc:
        return f"unavailable: {exc}"


def write_run_config(
    output_path: Path,
    args: argparse.Namespace,
    skill_dir: Path,
    tasks_dir: Path,
    ccb_command: str,
    selected: list[Task],
    judge_model: str,
) -> None:
    manifest_path = tasks_dir / "manifest.yaml"
    git_command = resolve_command("git")
    commit = command_output([git_command, "rev-parse", "HEAD"], cwd=skill_dir) if git_command else ""
    agent_key = os.environ.get("ANTHROPIC_AUTH_TOKEN") or ""
    judge_key = os.environ.get("OPENROUTER_API_KEY") or ""
    config = {
        "runner_revision": RUNNER_REVISION,
        "created_at": dt.datetime.now(dt.timezone.utc).astimezone().isoformat(),
        "platform": platform.platform(),
        "python": sys.version,
        "ccb_command": ccb_command,
        "ccb_version": command_output([ccb_command, "--version"]),
        "expected_ccb_version": args.expected_ccb_version,
        "model": args.model,
        "model_environment": {
            key: os.environ.get(key)
            for key in (
                "ANTHROPIC_MODEL",
                "ANTHROPIC_DEFAULT_HAIKU_MODEL",
                "ANTHROPIC_DEFAULT_SONNET_MODEL",
                "ANTHROPIC_DEFAULT_OPUS_MODEL",
                "ANTHROPIC_REASONING_MODEL",
                "CLAUDE_CODE_SUBAGENT_MODEL",
            )
        },
        "agent": "CCB default (not explicitly set)",
        "variant": "default (not explicitly set)",
        "permission_mode": args.permission_mode,
        "headless_auto_approved_tools": [x.strip() for x in args.allowed_tools.split(",") if x.strip()],
        "tool_visibility": "CCB default; --allowedTools receives one comma-separated auto-approval value",
        "prompt_transport": "stdin",
        "output_protocol": "required stream-json with terminal result event",
        "session_strategy": "explicit --session-id UUID for new sessions; --resume UUID for continuation",
        "auto_approval": args.permission_mode == "acceptEdits",
        "openrouter_routing": "default/dynamic provider",
        "anthropic_base_url": os.environ.get("ANTHROPIC_BASE_URL"),
        "claude_config_dir": os.environ.get("CLAUDE_CONFIG_DIR"),
        "agent_key_env": "ANTHROPIC_AUTH_TOKEN",
        "agent_key_present": bool(agent_key),
        "agent_judge_key_same": bool(agent_key and judge_key and agent_key == judge_key),
        "worker_count": 1,
        "task_concurrency": 1,
        "skill_dir": str(skill_dir),
        "tasks_dir": str(tasks_dir),
        "pinchbench_commit": commit,
        "manifest_sha256": sha256_file(manifest_path) if manifest_path.exists() else None,
        "suite": args.suite,
        "limit": args.limit,
        "default_skipped_tasks": sorted(DEFAULT_SKIPPED_TASKS),
        "additional_skip": args.skip,
        "skip_network": args.skip_network,
        "task_count": len(selected),
        "task_ids": [task.task_id for task in selected],
        "timeout_multiplier": args.timeout_multiplier,
        "network_timeout": args.network_timeout,
        "workspace_instruction": not args.no_workspace_instruction,
        "grading_enabled": not args.no_grade,
        "grading_engine": str(skill_dir / "scripts" / "lib_grading.py"),
        "judge_backend": "api",
        "judge_model": judge_model,
        "judge_key_env": "OPENROUTER_API_KEY",
        "judge_key_present": bool(judge_key),
        "judge_concurrency": 1,
        "judge_is_separate_from_tested_model": True,
        "reasoning_token_note": "CCB result events do not expose an exact separate reasoning-token count; field remains null.",
        "environment_proxy_present": {
            key: bool(os.environ.get(key))
            for key in ("HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "NO_PROXY")
        },
    }
    output_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")

def print_preflight(
    selected: list[Task],
    skill_dir: Path,
    tasks_dir: Path,
    ccb_command: str,
    grader_module: Optional[Any],
    grader_error: Optional[str],
    args: argparse.Namespace,
) -> tuple[dict[str, list[str]], list[str], list[str]]:
    prerequisite_failures: dict[str, list[str]] = {}
    fixture_failures: list[str] = []
    environment_failures: list[str] = []

    needs_judge = (
        not args.no_grade
        and any(task.grading_type in {"llm_judge", "hybrid"} for task in selected)
    )
    judge_model = str(getattr(grader_module, "DEFAULT_JUDGE_MODEL", PINCHBENCH_DEFAULT_JUDGE_MODEL)) if grader_module is not None else PINCHBENCH_DEFAULT_JUDGE_MODEL
    ccb_version = command_output([ccb_command, "--version"])
    agent_key = os.environ.get("ANTHROPIC_AUTH_TOKEN") or ""
    judge_key = os.environ.get("OPENROUTER_API_KEY") or ""
    base_url = (os.environ.get("ANTHROPIC_BASE_URL") or "").rstrip("/")
    configured_model = os.environ.get("ANTHROPIC_MODEL") or ""
    config_dir = os.environ.get("CLAUDE_CONFIG_DIR") or ""
    api_key_empty = not bool(os.environ.get("ANTHROPIC_API_KEY"))
    keys_same = bool(agent_key and judge_key and agent_key == judge_key)

    if args.expected_ccb_version and args.expected_ccb_version not in ccb_version and not args.allow_version_mismatch:
        environment_failures.append(f"CCB 版本不是固定版本 {args.expected_ccb_version}: {ccb_version}")
    if base_url != OPENROUTER_ANTHROPIC_BASE_URL:
        environment_failures.append(f"ANTHROPIC_BASE_URL 应为 {OPENROUTER_ANTHROPIC_BASE_URL}，当前为 {base_url or '(missing)'}")
    if not agent_key:
        environment_failures.append("ANTHROPIC_AUTH_TOKEN 未设置，CCB 无法调用 OpenRouter")
    if not api_key_empty:
        environment_failures.append("ANTHROPIC_API_KEY 必须为空，避免与 ANTHROPIC_AUTH_TOKEN 冲突")
    if args.permission_mode != DEFAULT_PERMISSION_MODE:
        environment_failures.append(
            f"无人值守对照要求 --permission-mode {DEFAULT_PERMISSION_MODE}，当前为 {args.permission_mode}"
        )
    model_env_names = (
        "ANTHROPIC_MODEL",
        "ANTHROPIC_DEFAULT_HAIKU_MODEL",
        "ANTHROPIC_DEFAULT_SONNET_MODEL",
        "ANTHROPIC_DEFAULT_OPUS_MODEL",
        "ANTHROPIC_REASONING_MODEL",
        "CLAUDE_CODE_SUBAGENT_MODEL",
    )
    for env_name in model_env_names:
        env_value = os.environ.get(env_name) or ""
        if env_value != args.model:
            environment_failures.append(
                f"{env_name} 必须等于 {args.model}，当前为 {env_value or '(missing)'}"
            )
    if not config_dir:
        environment_failures.append("CLAUDE_CONFIG_DIR 未设置，无法保证独立 CCB profile")
    if needs_judge and not judge_key:
        environment_failures.append("OPENROUTER_API_KEY 未设置，PinchBench Judge 无法评分")
    if needs_judge and agent_key and judge_key and not keys_same and not args.allow_key_mismatch:
        environment_failures.append("ANTHROPIC_AUTH_TOKEN 与 OPENROUTER_API_KEY 不同；严格对照要求使用同一 OpenRouter Key")

    print("=" * 100)
    print("PinchBench CCB Windows runner preflight")
    print("=" * 100)
    print(f"Runner revision      : {RUNNER_REVISION}")
    print(f"Skill dir            : {skill_dir}")
    print(f"Tasks dir            : {tasks_dir}")
    print(f"CCB command          : {ccb_command}")
    print(f"CCB version          : {ccb_version}")
    print(f"Expected CCB version : {args.expected_ccb_version}")
    print(f"Model                : {args.model}")
    print(f"Suite                : {args.suite}")
    print(f"Selected tasks       : {len(selected)}")
    print(f"Default skipped      : {len(DEFAULT_SKIPPED_TASKS)}")
    for task_id in sorted(DEFAULT_SKIPPED_TASKS):
        print(f"  - {task_id}")
    print("Worker/concurrency   : 1 / 1")
    print("Judge concurrency    : 1 (synchronous)")
    print("Default CCB agent    : yes")
    print(f"Headless allowedTools: {args.allowed_tools or '(none)'}")
    print(f"Permission mode      : {args.permission_mode}")
    print("Prompt transport     : stdin")
    print("Output protocol      : stream-json required")
    print("Session strategy     : explicit UUID / --resume")
    print(f"ANTHROPIC_BASE_URL   : {base_url or 'missing'}")
    print(f"CLAUDE_CONFIG_DIR    : {config_dir or 'missing'}")
    print(f"Agent key            : {'set' if agent_key else 'missing'}")
    print(f"Judge key            : {'set' if judge_key else 'missing'}")
    print(f"Agent/Judge key same : {'yes' if keys_same else 'no'}")
    print(f"ANTHROPIC_API_KEY    : {'empty' if api_key_empty else 'NOT EMPTY'}")
    print(f"Grading engine       : {skill_dir / 'scripts' / 'lib_grading.py'}")
    print("Judge backend        : api")
    print(f"Judge model          : {judge_model}")
    print("OpenClaw required    : no")
    print(f"Grader import        : {'ok' if grader_error is None else 'FAILED'}")
    if grader_error:
        print(f"  {grader_error}")
    print()

    category_counts = Counter(task.category or "(uncategorized)" for task in selected)
    grading_counts = Counter(task.grading_type for task in selected)
    print("Grading types:")
    for key, value in sorted(grading_counts.items()):
        print(f"  {key:<24} {value}")
    print("Categories:")
    for key, value in sorted(category_counts.items()):
        print(f"  {key:<24} {value}")
    print(f"Network-marked tasks : {sum(1 for task in selected if is_network_task(task))}")
    print(f"Multi-session tasks  : {sum(1 for task in selected if task.multi_session)}")
    print(f"Default judge needed : {'yes' if needs_judge else 'no'}")
    print()

    for task in selected:
        prerequisites = task.metadata.get("prerequisites") or []
        missing = check_prerequisites(list(prerequisites)) if prerequisites else []
        if missing:
            prerequisite_failures[task.task_id] = missing
        for workspace_file in task.workspace_files:
            if not isinstance(workspace_file, dict):
                fixture_failures.append(f"{task.task_id}: invalid workspace file entry")
                continue
            if workspace_file.get("content") is not None:
                continue
            source = workspace_file.get("source")
            if not source:
                continue
            src_rel = Path(str(source))
            candidates = [
                skill_dir / "assets" / src_rel,
                skill_dir / src_rel,
                task.file_path.parent / src_rel,
                task.file_path.parent.parent / "assets" / src_rel,
            ]
            if not any(candidate.exists() for candidate in candidates):
                fixture_failures.append(f"{task.task_id}: missing fixture {source}")

    if prerequisite_failures:
        print("Missing declared prerequisites:")
        for task_id, missing in prerequisite_failures.items():
            print(f"  {task_id}: {', '.join(missing)}")
    else:
        print("Missing declared prerequisites: none")
    if fixture_failures:
        print("Missing workspace fixtures:")
        for item in fixture_failures:
            print(f"  {item}")
    else:
        print("Missing workspace fixtures: none")
    if environment_failures:
        print("Environment alignment failures:")
        for item in environment_failures:
            print(f"  - {item}")
    else:
        print("Environment alignment failures: none")
    print("=" * 100)
    return prerequisite_failures, fixture_failures, environment_failures

# ---------------------------------------------------------------------------
# Main execution
# ---------------------------------------------------------------------------

def build_turns(task: Task) -> list[dict[str, Any]]:
    if task.multi_session and task.sessions:
        turns: list[dict[str, Any]] = []
        for index, item in enumerate(task.sessions, start=1):
            prompt = str(item.get("prompt") or "").strip()
            if not prompt:
                continue
            turns.append({
                "id": str(item.get("id") or f"turn_{index}"),
                "prompt": prompt,
                "new_session": bool(item.get("new_session")),
            })
        if turns:
            return turns

    return [{
        "id": "single",
        "prompt": task.prompt,
        "new_session": True,
    }]


def aggregate_optional_int(
    current: Optional[int],
    value: Optional[int],
) -> Optional[int]:
    if value is None:
        return current
    return (current or 0) + int(value)


def aggregate_optional_float(
    current: Optional[float],
    value: Optional[float],
) -> Optional[float]:
    if value is None:
        return current
    return (current or 0.0) + float(value)


def _empty_execution_result(
    task: Task,
    workspace: Path,
    transcript_dir: Path,
    status: str,
    error: str,
) -> dict[str, Any]:
    return {
        "task_id": task.task_id, "name": task.name, "category": task.category,
        "grading_type": task.grading_type, "network_task": is_network_task(task),
        "multi_session": task.multi_session, "session_count": 0, "success": False,
        "status": status, "returncode": None, "elapsed": 0.0, "ttft": None,
        "input_tokens": None, "output_tokens": None, "reasoning_tokens": None,
        "cache_read_tokens": None, "cache_write_tokens": None, "cost_usd": None,
        "usage_complete": False, "stream_json_complete": False, "protocol_modes": [],
        "step_count": 0, "tool_errors": 0,
        "event_counts": {}, "providers": [], "models": [], "unexpected_models": [],
        "permission_denials": [], "model_usage": {}, "thinking_char_count": 0,
        "output": "", "error": error,
        "stderr": "", "score": None, "breakdown": {}, "grade_notes": "",
        "grade_error": error, "workspace": str(workspace), "transcript": str(transcript_dir),
        "normalized_transcript": [], "turn_results": [],
    }


def execute_task(
    task: Task,
    task_index: int,
    task_count: int,
    skill_dir: Path,
    workspace: Path,
    transcript_dir: Path,
    ccb_command: str,
    args: argparse.Namespace,
) -> dict[str, Any]:
    print(f"[{task_index}/{task_count}] {task.task_id} — {task.name}")
    prompt_preview = re.sub(r"\s+", " ", task.prompt)[:150]
    print(f"  prompt: {prompt_preview}{'...' if len(task.prompt) > 150 else ''}")

    prerequisites = task.metadata.get("prerequisites") or []
    missing_prerequisites = check_prerequisites(list(prerequisites)) if prerequisites else []
    if missing_prerequisites:
        error = f"缺少依赖: {', '.join(missing_prerequisites)}"
        print(f"  跳过: {error}\n")
        return _empty_execution_result(task, workspace, transcript_dir, "missing_prerequisite", error)

    if workspace.exists():
        shutil.rmtree(workspace)
    workspace.mkdir(parents=True, exist_ok=True)
    transcript_dir.mkdir(parents=True, exist_ok=True)

    staged, fixture_errors = stage_workspace_files(task, workspace, skill_dir)
    if staged:
        LOGGER.info("  staged files: %s", ", ".join(staged[:8]) + (" ..." if len(staged) > 8 else ""))
    if fixture_errors:
        error = "缺少或无效的 workspace fixture: " + " | ".join(fixture_errors)
        print(f"  失败: {error}\n")
        return _empty_execution_result(task, workspace, transcript_dir, "missing_fixture", error)

    turns = build_turns(task)
    total_timeout = args.network_timeout if is_network_task(task) else max(1.0, task.timeout_seconds * args.timeout_multiplier)
    task_deadline = time.monotonic() + total_timeout

    normalized_transcript: list[dict[str, Any]] = []
    outputs: list[str] = []
    errors: list[str] = []
    stderr_chunks: list[str] = []
    turn_results: list[dict[str, Any]] = []
    current_session_id: Optional[str] = None
    task_status = "success"
    task_success = True
    returncode: Optional[int] = 0
    total_elapsed = 0.0
    task_ttft: Optional[float] = None
    input_tokens: Optional[int] = None
    output_tokens: Optional[int] = None
    reasoning_tokens: Optional[int] = None
    cache_read_tokens: Optional[int] = None
    cache_write_tokens: Optional[int] = None
    cost_usd: Optional[float] = None
    usage_complete = True
    stream_json_complete = True
    protocol_modes: list[str] = []
    step_count = 0
    tool_errors = 0
    event_counts: Counter[str] = Counter()
    providers: set[str] = set()
    models: set[str] = set()
    permission_denials: list[Any] = []
    model_usage: dict[str, Any] = {}
    thinking_char_count = 0

    allowed_tools = [tool.strip() for tool in args.allowed_tools.split(",") if tool.strip()]

    for turn_index, turn in enumerate(turns, start=1):
        remaining = task_deadline - time.monotonic()
        if remaining <= 0:
            task_success = False
            task_status = "timeout"
            errors.append(f"任务总超时 ({total_timeout:.0f}s)")
            break

        turn_id = str(turn["id"])
        turn_prompt = str(turn["prompt"])
        metadata_new_session = bool(turn.get("new_session"))
        starts_new_session = current_session_id is None or metadata_new_session
        requested_session_id = str(uuid.uuid4()) if starts_new_session else str(current_session_id)
        resume_requested = not starts_new_session

        normalized_transcript.append(make_user_transcript_event(turn_prompt, turn_index, turn_id))
        prompt = turn_prompt
        if not args.no_workspace_instruction:
            prompt += (
                "\n\nIMPORTANT: You are running in an isolated workspace. "
                "Read, write, and edit files only in the current working directory. "
                "Complete the requested work now; do not wait for interactive approval."
            )

        command = build_ccb_command(
            ccb_command=ccb_command,
            model=args.model,
            permission_mode=args.permission_mode,
            allowed_tools=allowed_tools,
            session_id=requested_session_id,
            resume=resume_requested,
        )

        raw_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.jsonl"
        stderr_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.stderr.txt"
        turn_result = run_ccb_streaming(
            command,
            cwd=workspace,
            timeout=max(1.0, remaining),
            raw_stdout_path=raw_path,
            stderr_path=stderr_path,
            prompt=prompt,
            expected_session_id=requested_session_id,
        )

        normalized_transcript.extend(turn_result["normalized_events"])
        if turn_result.get("output"):
            outputs.append(f"## Turn {turn_index} ({turn_id})\n{turn_result['output']}")
        if turn_result.get("error"):
            errors.append(f"turn {turn_index} ({turn_id}): {turn_result['error']}")
        if turn_result.get("stderr"):
            stderr_chunks.append(f"## Turn {turn_index} ({turn_id})\n{turn_result['stderr']}")

        total_elapsed += float(turn_result.get("elapsed") or 0.0)
        if task_ttft is None and turn_result.get("ttft") is not None:
            task_ttft = float(turn_result["ttft"])
        input_tokens = aggregate_optional_int(input_tokens, turn_result.get("input_tokens"))
        output_tokens = aggregate_optional_int(output_tokens, turn_result.get("output_tokens"))
        reasoning_tokens = aggregate_optional_int(reasoning_tokens, turn_result.get("reasoning_tokens"))
        cache_read_tokens = aggregate_optional_int(cache_read_tokens, turn_result.get("cache_read_tokens"))
        cache_write_tokens = aggregate_optional_int(cache_write_tokens, turn_result.get("cache_write_tokens"))
        cost_usd = aggregate_optional_float(cost_usd, turn_result.get("cost_usd"))
        usage_complete = usage_complete and bool(turn_result.get("usage_complete"))
        stream_json_complete = stream_json_complete and bool(turn_result.get("stream_json_ok"))
        protocol_modes.append(str(turn_result.get("protocol_mode") or "unknown"))
        step_count += int(turn_result.get("step_count") or 0)
        tool_errors += int(turn_result.get("tool_errors") or 0)
        event_counts.update(turn_result.get("event_counts") or {})
        providers.update(turn_result.get("providers") or [])
        models.update(turn_result.get("models") or [])
        permission_denials.extend(turn_result.get("permission_denials") or [])
        model_usage.update(turn_result.get("model_usage") or {})
        thinking_char_count += int(turn_result.get("thinking_char_count") or 0)

        returned_session_id = str(turn_result.get("session_id") or requested_session_id)
        current_session_id = returned_session_id

        turn_results.append({
            "turn": turn_index,
            "turn_id": turn_id,
            "new_session": metadata_new_session,
            "started_new_session": starts_new_session,
            "resume_requested": resume_requested,
            "requested_session_id": requested_session_id,
            "session_id": returned_session_id,
            "success": turn_result.get("success"),
            "status": turn_result.get("status"),
            "returncode": turn_result.get("returncode"),
            "elapsed": turn_result.get("elapsed"),
            "ttft": turn_result.get("ttft"),
            "input_tokens": turn_result.get("input_tokens"),
            "output_tokens": turn_result.get("output_tokens"),
            "reasoning_tokens": turn_result.get("reasoning_tokens"),
            "cost_usd": turn_result.get("cost_usd"),
            "usage_complete": turn_result.get("usage_complete"),
            "stream_json_ok": turn_result.get("stream_json_ok"),
            "protocol_mode": turn_result.get("protocol_mode"),
            "step_count": turn_result.get("step_count"),
            "tool_errors": turn_result.get("tool_errors"),
            "event_counts": turn_result.get("event_counts"),
            "providers": turn_result.get("providers"),
            "models": turn_result.get("models"),
            "permission_denials": turn_result.get("permission_denials"),
            "model_usage": turn_result.get("model_usage"),
            "thinking_char_count": turn_result.get("thinking_char_count"),
            "raw_transcript": str(raw_path),
            "stderr_path": str(stderr_path),
            "error": turn_result.get("error"),
        })

        if not turn_result.get("success"):
            task_success = False
            task_status = str(turn_result.get("status") or "error")
            returncode = turn_result.get("returncode")
            break
        returncode = turn_result.get("returncode")

    observed_models = set(models) | {str(key) for key in model_usage.keys()}
    unexpected_models = sorted(model for model in observed_models if model and model != args.model)
    if unexpected_models:
        task_success = False
        task_status = "model_mismatch"
        errors.append("检测到非目标模型参与执行: " + ", ".join(unexpected_models))

    normalized_path = transcript_dir / "normalized.jsonl"
    with normalized_path.open("w", encoding="utf-8") as file:
        for event in normalized_transcript:
            file.write(json.dumps(event, ensure_ascii=False) + "\n")
    turn_results_path = transcript_dir / "turn_results.json"
    turn_results_path.write_text(json.dumps(turn_results, ensure_ascii=False, indent=2), encoding="utf-8")

    return {
        "task_id": task.task_id, "name": task.name, "category": task.category,
        "grading_type": task.grading_type, "network_task": is_network_task(task),
        "multi_session": task.multi_session, "session_count": len(turn_results),
        "success": task_success, "status": task_status, "returncode": returncode,
        "elapsed": total_elapsed, "ttft": task_ttft, "input_tokens": input_tokens,
        "output_tokens": output_tokens, "reasoning_tokens": reasoning_tokens,
        "cache_read_tokens": cache_read_tokens, "cache_write_tokens": cache_write_tokens,
        "cost_usd": cost_usd, "usage_complete": usage_complete and bool(turn_results),
        "stream_json_complete": stream_json_complete and bool(turn_results),
        "protocol_modes": protocol_modes,
        "step_count": step_count, "tool_errors": tool_errors, "event_counts": dict(event_counts),
        "providers": sorted(providers), "models": sorted(models),
        "unexpected_models": unexpected_models,
        "permission_denials": permission_denials, "model_usage": model_usage,
        "thinking_char_count": thinking_char_count,
        "output": "\n\n".join(outputs), "error": " | ".join(errors),
        "stderr": "\n\n".join(stderr_chunks), "score": None, "breakdown": {},
        "grade_notes": "", "grade_error": None, "workspace": str(workspace),
        "transcript": str(transcript_dir), "normalized_transcript_path": str(normalized_path),
        "transcript_data": normalized_transcript, "turn_results": turn_results,
    }

def build_args() -> argparse.Namespace:
    script_path = Path(__file__).resolve()
    inferred_root = script_path.parent.parent
    inferred_skill = inferred_root / "skill"
    parser = argparse.ArgumentParser(
        description="Run local PinchBench tasks with CCB on native Windows, one task at a time.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--skill-dir", default=str(inferred_skill if inferred_skill.exists() else Path.cwd()), help="PinchBench skill 仓库根目录")
    parser.add_argument("--tasks-dir", default=None, help="任务目录；默认是 <skill-dir>/tasks")
    parser.add_argument("--suite", default="core", help="all、core、automated-only、llm-judge-only、hybrid-only、judge-required-only，或逗号分隔任务 ID")
    parser.add_argument("--limit", type=int, default=None, help="最多运行多少个任务，仅用于测试 runner")
    parser.add_argument("--skip", default="", help="额外跳过的任务 ID；四个已约定 integration 任务默认跳过")
    parser.add_argument("--skip-network", action="store_true", help="跳过明确标记的联网任务；正式全量测试通常不应使用")
    parser.add_argument("--ccb", default="auto", help="CCB 命令：auto、ccb.cmd、ccb-bun.cmd 或绝对路径")
    parser.add_argument("--model", default=DEFAULT_MODEL, help="通过 OpenRouter Anthropic-compatible endpoint 使用的模型 ID")
    parser.add_argument("--expected-ccb-version", default=EXPECTED_CCB_VERSION, help="严格对照所固定的 CCB 版本")
    parser.add_argument("--allow-version-mismatch", action="store_true", help="允许 CCB 版本与固定版本不一致；正式对照不建议使用")
    parser.add_argument("--allow-key-mismatch", action="store_true", help="允许 Agent 和 Judge 使用不同 OpenRouter Key；正式对照不建议使用")
    parser.add_argument("--allowed-tools", default=",".join(DEFAULT_ALLOWED_TOOLS), help="CCB 非交互模式下自动批准的工具；不限制其他默认工具的可见性")
    parser.add_argument("--timeout-multiplier", type=float, default=3.0, help="非联网任务 timeout_seconds 的倍数，与 OpenCode v3 runner 对齐")
    parser.add_argument("--network-timeout", type=float, default=300.0, help="明确标记的联网任务总超时秒数，与 OpenCode v3 runner 对齐")
    parser.add_argument("--judge-timeout", type=float, default=300.0, help="PinchBench 默认 LLM judge 单次请求超时秒数")
    parser.add_argument("--results-dir", default=None, help="运行结果根目录；默认是 <skill-dir>/../runs")
    parser.add_argument("--keep-workspaces", action="store_true", help="保留成功任务 workspace；正式评测建议启用")
    parser.add_argument("--no-grade", action="store_true", help="只执行任务，不打分")
    parser.add_argument("--no-xlsx", action="store_true", help="不输出 XLSX")
    parser.add_argument(
        "--permission-mode",
        default=DEFAULT_PERMISSION_MODE,
        help=(
            "CCB permission mode。严格无人值守对照固定为 acceptEdits；"
            "不要使用 bypassPermissions。"
        ),
    )
    parser.add_argument("--no-workspace-instruction", action="store_true", help="不在 prompt 后追加隔离 workspace 提醒")
    parser.add_argument("--preflight", action="store_true", help="只做本地预检，不调用模型")
    parser.add_argument("--no-judge-cache", action="store_true", help="禁用 PinchBench Judge 缓存")
    parser.add_argument("--clear-judge-cache", action="store_true", help="运行前清空 PinchBench Judge 缓存")
    parser.add_argument("--verbose", action="store_true", help="打印更详细日志")
    return parser.parse_args()

def main() -> int:
    args = build_args()
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-8s %(message)s",
        datefmt="%H:%M:%S",
    )

    skill_dir = Path(args.skill_dir).expanduser().resolve()
    tasks_dir = Path(args.tasks_dir).expanduser().resolve() if args.tasks_dir else skill_dir / "tasks"
    if not skill_dir.exists():
        raise SystemExit(f"找不到 PinchBench 仓库目录: {skill_dir}")
    if not tasks_dir.exists():
        raise SystemExit(f"找不到 tasks 目录: {tasks_dir}\n请确认 --skill-dir 指向本地 pinchbench/skill 仓库。")
    if not (tasks_dir / "manifest.yaml").exists():
        LOGGER.warning("tasks/manifest.yaml 不存在，将按 task_*.md 文件回退加载")

    ccb_command = choose_ccb_command(args.ccb)
    grader_module, grader_error = load_pinchbench_grading(skill_dir)
    all_tasks, core_tasks = load_tasks(tasks_dir)
    selected = filter_tasks(all_tasks, core_tasks, args.suite, args.limit)

    skip_ids = set(DEFAULT_SKIPPED_TASKS)
    skip_ids.update(item.strip() for item in args.skip.split(",") if item.strip())
    if args.skip_network:
        skip_ids.update(task.task_id for task in selected if is_network_task(task))
    selected = [task for task in selected if task.task_id not in skip_ids]
    if not selected:
        raise SystemExit("没有可运行任务。请检查 --suite/--limit/--skip。")

    needs_judge = (
        not args.no_grade
        and any(task.grading_type in {"llm_judge", "hybrid"} for task in selected)
    )
    prerequisite_failures, fixture_failures, environment_failures = print_preflight(
        selected, skill_dir, tasks_dir, ccb_command, grader_module, grader_error, args,
    )

    if args.preflight:
        if fixture_failures: return 2
        if prerequisite_failures: return 3
        if grader_error: return 4
        if environment_failures: return 6
        return 0

    if grader_error and not args.no_grade:
        raise SystemExit(grader_error)
    if environment_failures:
        raise SystemExit("CCB/OpenRouter 环境未达到严格对照条件：\n- " + "\n- ".join(environment_failures))

    results_root = Path(args.results_dir).expanduser().resolve() if args.results_dir else skill_dir.parent / "runs"
    run_id = dt.datetime.now().strftime("ccb_%Y%m%d_%H%M%S")
    results_dir = results_root / run_id
    workspaces_dir = results_dir / "workspaces"
    transcripts_dir = results_dir / "transcripts"
    results_dir.mkdir(parents=True, exist_ok=True)
    workspaces_dir.mkdir(parents=True, exist_ok=True)
    transcripts_dir.mkdir(parents=True, exist_ok=True)
    progress_path = results_dir / "progress.jsonl"
    partial_json_path = results_dir / "results.partial.json"

    judge_model = str(getattr(grader_module, "DEFAULT_JUDGE_MODEL", PINCHBENCH_DEFAULT_JUDGE_MODEL)) if grader_module is not None else PINCHBENCH_DEFAULT_JUDGE_MODEL
    write_run_config(results_dir / "run_config.json", args, skill_dir, tasks_dir, ccb_command, selected, judge_model)

    judge_cache_dir = results_root / ".judge_cache"
    if grader_module is not None and hasattr(grader_module, "set_judge_cache_dir") and not args.no_judge_cache:
        grader_module.set_judge_cache_dir(judge_cache_dir)
        if args.clear_judge_cache and hasattr(grader_module, "clear_judge_cache"):
            grader_module.clear_judge_cache()

    print()
    print(f"Results dir          : {results_dir}")
    print("Execution backend    : CCB default agent via OpenRouter Anthropic endpoint")
    print(f"Permission mode      : {args.permission_mode}")
    print("Prompt transport     : stdin")
    print("Output protocol      : stream-json required")
    print("Session strategy     : explicit UUID / --resume")
    print("Grading engine       : PinchBench scripts/lib_grading.py")
    print("Judge backend        : api")
    print(f"Judge model          : {judge_model}")
    print(f"Judge cache          : {'disabled' if args.no_judge_cache else judge_cache_dir}")
    print()

    results: list[dict[str, Any]] = []
    total_start = time.monotonic()
    for index, task in enumerate(selected, start=1):
        workspace = workspaces_dir / task.task_id
        transcript_dir = transcripts_dir / task.task_id
        execution = execute_task(
            task, index, len(selected), skill_dir, workspace, transcript_dir,
            ccb_command, args,
        )

        grade_result = GradeResult(score=None, grading_type=task.grading_type, breakdown={})
        if execution["status"] not in {"missing_prerequisite", "missing_fixture"} and not args.no_grade:
            grading_execution = dict(execution)
            grading_execution["transcript"] = execution.get("transcript_data", [])
            grade_result = grade_with_pinchbench_default(
                grader_module=grader_module, task=task,
                execution_result=grading_execution, workspace=workspace,
                skill_dir=skill_dir, judge_timeout=args.judge_timeout,
                verbose=args.verbose,
            )

        execution["score"] = round(float(grade_result.score), 4) if grade_result.score is not None else None
        execution["breakdown"] = grade_result.breakdown
        execution["grade_notes"] = grade_result.notes
        execution["grade_error"] = grade_result.error
        execution.pop("transcript_data", None)
        results.append(execution)

        with progress_path.open("a", encoding="utf-8") as progress_file:
            progress_file.write(json.dumps(execution, ensure_ascii=False) + "\n")
            progress_file.flush()
        partial_json_path.write_text(
            json.dumps({"completed": len(results), "results": results}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        marker = "✓" if execution["success"] else "✗"
        score_text = f"score={execution['score']:.3f}" if execution["score"] is not None else "score=N/A"
        ttft_text = f"ttft={execution['ttft']:.2f}s" if execution["ttft"] is not None else "ttft=N/A"
        token_text = f"{execution['input_tokens']}+{execution['output_tokens']}" if execution["usage_complete"] else "N/A"
        provider_text = ",".join(execution.get("providers") or []) or "N/A"
        print(f"  {marker} elapsed={execution['elapsed']:.1f}s {ttft_text} {score_text} tokens={token_text} provider={provider_text}")
        if execution.get("error"):
            print(f"  运行错误: {str(execution['error'])[:500]}")
        if execution.get("grade_error"):
            print(f"  打分错误: {str(execution['grade_error'])[:500]}")
        if execution.get("permission_denials"):
            print(f"  权限拒绝: {len(execution['permission_denials'])}")
        if execution.get("output"):
            preview = re.sub(r"\s+", " ", str(execution["output"]))[:260]
            print(f"  输出预览: {preview}{'...' if len(str(execution['output'])) > 260 else ''}")
        print()

        if not args.keep_workspaces and execution["success"] and not execution.get("grade_error"):
            shutil.rmtree(workspace, ignore_errors=True)
            execution["workspace"] = "已清理；使用 --keep-workspaces 可保留"

    total_elapsed = time.monotonic() - total_start
    summary = print_summary(results, total_elapsed)
    payload = {"summary": summary, "results": results}
    json_path = results_dir / "results.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    csv_path = results_dir / "results.csv"
    save_csv(results, csv_path)
    xlsx_path = results_dir / "results.xlsx"
    if not args.no_xlsx:
        save_xlsx(results, summary, xlsx_path)

    if grader_module is not None and hasattr(grader_module, "get_judge_cache_stats") and not args.no_judge_cache:
        try:
            stats = grader_module.get_judge_cache_stats()
            print("Judge cache stats   : " f"entries={stats.get('entries', 0)} " f"hits={stats.get('hits', 0)} " f"misses={stats.get('misses', 0)}")
        except Exception as exc:
            LOGGER.warning("读取 Judge cache stats 失败: %s", exc)

    print("\n结果文件:")
    print(f"  Config : {results_dir / 'run_config.json'}")
    print(f"  JSON   : {json_path}")
    print(f"  CSV    : {csv_path}")
    if not args.no_xlsx:
        print(f"  XLSX   : {xlsx_path}")
    print(f"  Logs   : {transcripts_dir}")
    print(f"  Progress JSONL : {progress_path}")
    print(f"  Partial JSON   : {partial_json_path}")
    return 0



if __name__ == "__main__":
    raise SystemExit(main())
