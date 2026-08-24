#!/usr/bin/env python3
# Runner revision: 2026-08-10-our-framework-qwen-kimi-hybrid-v1
"""
Run the local PinchBench task set with 我们的框架 (LocalClaw/LocalCoding source bundle) on native Windows.

Formal modes:
- qwen: fixed qwen/qwen3.6-27b
- kimi: fixed moonshotai/kimi-k3
- hybrid: framework-native Smart Hybrid, Qwen default -> Kimi K3 upgrade

Benchmark invariants:
- Native Windows, one worker, strictly serial.
- Same checked-out PinchBench manifest, fixtures, grader and task order.
- Fresh framework session for normal tasks; multi-session tasks preserve only their defined task-local session.
- Cross-task projects memory is deleted between tasks.
- Extra MCP / Skills / global benchmark CLAUDE.md injections are rejected/cleaned.
- Smart Hybrid's own framework-native injection is kept only in hybrid mode as the treatment.
- Raw websocket events and raw Claude Agent SDK stream messages are retained.
- Agent time and grading time are recorded separately.
- Agent usage is taken from Claude Agent SDK result usage/modelUsage when present.
- Judge is explicitly pinned to openrouter/anthropic/claude-opus-5 and uses OPENROUTER_API_KEY.
"""
from __future__ import annotations

import argparse
import csv
import dataclasses
import datetime as dt
import hashlib
import json
import logging
import os
import platform
import queue
import re
import shutil
import signal
import subprocess
import sys
import threading
import time
import traceback
import urllib.error
import urllib.request
from collections import Counter
from pathlib import Path
from typing import Any, Optional

os.environ.setdefault("PYTHONUTF8", "1")
os.environ.setdefault("OPENCODE_DISABLE_AUTOUPDATE", "true")

try:
    import yaml  # type: ignore
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "缺少 PyYAML。请使用项目虚拟环境运行：\n"
        r"C:\pinchbench-opencode\.venv\Scripts\python.exe -m pip install pyyaml"
    ) from exc

LOGGER = logging.getLogger("pinchbench-opencode-kimi")

RUNNER_REVISION = "2026-08-13-our-framework-qwen-kimi-hybrid-v4.4-infra-backoff"
DEFAULT_MODEL = "moonshotai/kimi-k3"
DEFAULT_JUDGE_MODEL = "openrouter/anthropic/claude-opus-5"
GRADER_COMPAT_REVISION = "2026-08-20-opus5-anthropic-pin-plain-semantic-v4_1"

# Agreed exclusions for this independent project. They require external
# GitHub/Google Workspace integration tooling and credentials that are not
# part of the current Windows OpenCode test environment.
DEFAULT_SKIPPED_TASKS = {
    "task_gh_issue_triage",
    "task_gws_email_triage",
    "task_gws_cross_service",
    "task_gws_task_management",
}

# Explicitly known internet-dependent tasks from the historical runner.
# Newer task files can also mark network_required: true.
NETWORK_TASKS = {
    "task_earnings_analysis",
    "task_stock",
    "task_market_research",
    "task_polymarket_briefing",
    "task_executive_lookup",
    "task_events",
    "task_deep_research",
    "task_competitive_research",
    "task_oss_alternative_research",
    "task_pricing_research",
    "task_it_procurement",
    "task_eu_regulation_research",
    "task_byok_best_practices",
}

TEXT_FILE_SUFFIXES = {
    ".txt", ".md", ".py", ".js", ".ts", ".tsx", ".jsx", ".json", ".yaml", ".yml",
    ".csv", ".tsv", ".html", ".css", ".xml", ".toml", ".ini", ".cfg", ".log",
    ".ics", ".sh", ".ps1", ".bat", ".cmd", ".sql", ".java", ".go", ".rs", ".rb", ".php",
}

SKIP_WORKSPACE_NAMES = {
    "BOOTSTRAP.md", "SOUL.md", "USER.md", "IDENTITY.md", "HEARTBEAT.md",
    "TOOLS.md", "AGENTS.md", "CLAUDE.md",
}
SKIP_WORKSPACE_DIRS = {
    ".git", ".openclaw", ".opencode", "__pycache__", "node_modules", "skills",
}


@dataclasses.dataclass
class Task:
    task_id: str
    name: str
    category: str
    grading_type: str
    timeout_seconds: int
    workspace_files: list[dict[str, Any]]
    prompt: str
    expected_behavior: str
    grading_criteria: list[str]
    automated_checks: str
    llm_judge_rubric: str
    grading_weights: dict[str, float]
    metadata: dict[str, Any]
    file_path: Path
    multi_session: bool = False
    sessions: list[dict[str, Any]] = dataclasses.field(default_factory=list)


@dataclasses.dataclass
class GradeResult:
    score: Optional[float]
    grading_type: str
    breakdown: dict[str, float]
    notes: str = ""
    error: Optional[str] = None


# ---------------------------------------------------------------------------
# Task loading
# ---------------------------------------------------------------------------

def parse_sections(body: str) -> dict[str, str]:
    sections: dict[str, list[str]] = {}
    current: Optional[str] = None
    for line in body.splitlines():
        match = re.match(r"^##\s+(.+?)\s*$", line)
        if match:
            current = match.group(1).strip()
            sections.setdefault(current, [])
        elif current:
            sections[current].append(line)
    return {key: "\n".join(value).strip() for key, value in sections.items()}


def extract_grading_criteria(text: str) -> list[str]:
    criteria: list[str] = []
    for line in text.splitlines():
        match = re.match(r"^-\s+\[[ xX]\]\s+(.+)$", line.strip())
        if match:
            criteria.append(match.group(1).strip())
    return criteria


def load_task_file(path: Path, category_override: Optional[str] = None) -> Task:
    raw = path.read_text(encoding="utf-8")
    match = re.match(r"^---\s*\r?\n(.*?)\r?\n---\s*\r?\n?(.*)$", raw, re.DOTALL)
    if not match:
        raise ValueError(f"{path} 没有 YAML frontmatter")

    metadata = yaml.safe_load(match.group(1)) or {}
    if not isinstance(metadata, dict):
        raise ValueError(f"{path} 的 YAML frontmatter 不是对象")

    body = match.group(2)
    sections = parse_sections(body)
    task_id = str(metadata.get("id") or path.stem)

    weights = metadata.get("grading_weights") or {}
    if not isinstance(weights, dict):
        weights = {}
    normalized_weights: dict[str, float] = {}
    for key, value in weights.items():
        try:
            normalized_weights[str(key)] = float(value)
        except (TypeError, ValueError):
            pass

    raw_sessions = metadata.get("sessions") or []
    sessions = [item for item in raw_sessions if isinstance(item, dict)] if isinstance(raw_sessions, list) else []

    return Task(
        task_id=task_id,
        name=str(metadata.get("name") or task_id),
        category=str(category_override or metadata.get("category") or ""),
        grading_type=str(metadata.get("grading_type") or "automated"),
        timeout_seconds=int(metadata.get("timeout_seconds") or 120),
        workspace_files=list(metadata.get("workspace_files") or []),
        prompt=sections.get("Prompt", "").strip(),
        expected_behavior=sections.get("Expected Behavior", "").strip(),
        grading_criteria=extract_grading_criteria(sections.get("Grading Criteria", "")),
        automated_checks=sections.get("Automated Checks", "") or "",
        llm_judge_rubric=sections.get("LLM Judge Rubric", "") or "",
        grading_weights=normalized_weights,
        metadata=metadata,
        file_path=path,
        multi_session=bool(metadata.get("multi_session") or sessions),
        sessions=sessions,
    )


def parse_manifest(tasks_dir: Path) -> tuple[list[str], dict[str, str], list[str]]:
    manifest_path = tasks_dir / "manifest.yaml"
    if not manifest_path.exists():
        ordered = [
            p.stem for p in sorted(tasks_dir.glob("task_*.md"))
            if p.stem != "task_XX_name"
        ]
        return ordered, {}, []

    manifest = yaml.safe_load(manifest_path.read_text(encoding="utf-8")) or {}
    if not isinstance(manifest, dict):
        raise ValueError(f"{manifest_path} 不是有效 YAML 对象")

    category_map: dict[str, str] = {}
    core_tasks = [str(x) for x in (manifest.get("core") or [])]

    if "categories" in manifest:
        run_first = [str(x) for x in (manifest.get("run_first") or [])]
        categories = manifest.get("categories") or {}
        all_task_ids: list[str] = []
        if isinstance(categories, dict):
            for category, ids in categories.items():
                for task_id in ids or []:
                    task_id = str(task_id)
                    category_map[task_id] = str(category)
                    all_task_ids.append(task_id)

        seen: set[str] = set()
        ordered: list[str] = []
        for task_id in run_first + all_task_ids:
            if task_id not in seen:
                seen.add(task_id)
                ordered.append(task_id)
        return ordered, category_map, core_tasks

    ordered = [str(x) for x in (manifest.get("tasks") or [])]
    return ordered, category_map, core_tasks


def load_tasks(tasks_dir: Path) -> tuple[list[Task], list[str]]:
    """Load exactly the checked-out manifest when it exists.

    PinchBench ships template/example Markdown files alongside real tasks.
    Falling back to every task_*.md while a manifest exists can accidentally
    count a template as a benchmark task.  The manifest is therefore the
    source of truth for normal runs.
    """
    manifest_exists = (tasks_dir / "manifest.yaml").exists()
    ordered_ids, category_map, core_tasks = parse_manifest(tasks_dir)
    tasks: list[Task] = []

    for task_id in ordered_ids:
        if task_id == "task_XX_name":
            continue
        task_path = tasks_dir / f"{task_id}.md"
        if not task_path.exists():
            LOGGER.warning("manifest 引用了不存在的任务文件: %s", task_path)
            continue
        try:
            task = load_task_file(
                task_path,
                category_override=category_map.get(task_id),
            )
            if (
                task.task_id == "task_XX_name"
                or task.category.strip().lower() == "category_name"
            ):
                LOGGER.debug("忽略 PinchBench 模板任务: %s", task_path)
                continue
            tasks.append(task)
        except Exception as exc:
            LOGGER.warning("加载任务失败 %s: %s", task_path, exc)

    if manifest_exists:
        return tasks, core_tasks

    # Only use file discovery when no manifest is available.
    loaded_ids = {task.task_id for task in tasks}
    for task_path in sorted(tasks_dir.glob("task_*.md")):
        if task_path.stem in loaded_ids or task_path.stem == "task_XX_name":
            continue
        try:
            task = load_task_file(task_path)
            if (
                task.task_id == "task_XX_name"
                or task.category.strip().lower() == "category_name"
            ):
                continue
            tasks.append(task)
        except Exception as exc:
            LOGGER.warning("加载任务失败 %s: %s", task_path, exc)

    return tasks, core_tasks


def filter_tasks(
    tasks: list[Task],
    core_tasks: list[str],
    suite: str,
    limit: Optional[int],
) -> list[Task]:
    suite = suite.strip()
    task_by_id = {task.task_id: task for task in tasks}

    if suite == "all":
        selected = tasks
    elif suite == "core":
        core_set = set(core_tasks)
        selected = [task for task in tasks if task.task_id in core_set] if core_set else tasks[:25]
    elif suite == "automated-only":
        selected = [task for task in tasks if task.grading_type == "automated"]
    elif suite == "llm-judge-only":
        selected = [task for task in tasks if task.grading_type == "llm_judge"]
    elif suite == "hybrid-only":
        selected = [task for task in tasks if task.grading_type == "hybrid"]
    elif suite == "judge-required-only":
        selected = [task for task in tasks if task.grading_type in {"llm_judge", "hybrid"}]
    else:
        ids = [item.strip() for item in suite.split(",") if item.strip()]
        missing = [task_id for task_id in ids if task_id not in task_by_id]
        if missing:
            raise SystemExit(f"suite 中有未知任务: {', '.join(missing)}")
        selected = [task_by_id[task_id] for task_id in ids]

    if limit is not None:
        selected = selected[:limit]
    return selected


def is_network_task(task: Task) -> bool:
    return bool(
        task.task_id in NETWORK_TASKS
        or task.metadata.get("network_required") is True
        or task.metadata.get("requires_network") is True
    )


# ---------------------------------------------------------------------------
# Workspace and prerequisites
# ---------------------------------------------------------------------------

def safe_workspace_dest(dest: str) -> Path:
    rel = Path(dest)
    if rel.is_absolute() or ".." in rel.parts:
        raise ValueError(f"非法 workspace 文件路径: {dest}")
    return rel


def stage_workspace_files(
    task: Task,
    workspace: Path,
    skill_dir: Path,
) -> tuple[list[str], list[str]]:
    staged: list[str] = []
    missing: list[str] = []

    for wf in task.workspace_files:
        if not isinstance(wf, dict):
            missing.append(f"无法识别的 workspace_files 项: {wf!r}")
            continue

        dest_value = wf.get("dest") or wf.get("path") or wf.get("name")
        source_value = wf.get("source")
        content_value = wf.get("content")

        if not dest_value and source_value:
            dest_value = Path(str(source_value)).name
        if not dest_value:
            missing.append(f"workspace_files 项缺少 dest/path: {wf!r}")
            continue

        try:
            dest_rel = safe_workspace_dest(str(dest_value))
        except ValueError as exc:
            missing.append(str(exc))
            continue

        dest_path = workspace / dest_rel
        dest_path.parent.mkdir(parents=True, exist_ok=True)

        if content_value is not None:
            dest_path.write_text(str(content_value), encoding="utf-8")
            staged.append(str(dest_rel))
            continue

        if not source_value:
            missing.append(f"workspace_files 项缺少 source/content: {wf!r}")
            continue

        src_rel = Path(str(source_value))
        candidates = [
            skill_dir / "assets" / src_rel,
            skill_dir / src_rel,
            task.file_path.parent / src_rel,
            task.file_path.parent.parent / "assets" / src_rel,
        ]
        src_path = next((candidate for candidate in candidates if candidate.exists()), None)
        if src_path is None:
            missing.append(
                f"找不到预置文件 {source_value}; 尝试过: "
                + ", ".join(str(candidate) for candidate in candidates)
            )
            continue

        if src_path.is_dir():
            if dest_path.exists():
                shutil.rmtree(dest_path)
            shutil.copytree(src_path, dest_path)
        else:
            shutil.copy2(src_path, dest_path)
        staged.append(str(dest_rel))

    return staged, missing


def resolve_command(name: str) -> Optional[str]:
    candidates = [name]
    if sys.platform == "win32" and not Path(name).suffix:
        candidates = [f"{name}.cmd", f"{name}.exe", name]
    for candidate in candidates:
        resolved = shutil.which(candidate)
        if resolved:
            return resolved
    return None


def check_prerequisites(prereqs: list[str]) -> list[str]:
    missing: list[str] = []
    npm_command = resolve_command("npm") or "npm"

    for req_raw in prereqs:
        req = str(req_raw)
        if req.startswith("cli:"):
            command = req.split(":", 1)[1].strip()
            if not resolve_command(command):
                missing.append(req)
        elif req.startswith("npm:"):
            package = req.split(":", 1)[1].strip()
            try:
                result = subprocess.run(
                    [npm_command, "list", "-g", package, "--depth=0"],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=30,
                    check=False,
                )
                if result.returncode != 0:
                    missing.append(req)
            except Exception:
                missing.append(req)

    return missing


def choose_opencode_command(preferred: str) -> str:
    if preferred and preferred != "auto":
        preferred_path = Path(preferred).expanduser()
        if preferred_path.exists():
            return str(preferred_path.resolve())
        resolved = resolve_command(preferred)
        if resolved:
            return resolved
        raise SystemExit(f"找不到 OpenCode 命令 {preferred!r}")

    for command in (("opencode.cmd", "opencode.exe", "opencode") if sys.platform == "win32" else ("opencode",)):
        resolved = shutil.which(command)
        if resolved:
            return resolved

    raise SystemExit(
        "找不到 OpenCode。请确认已安装 opencode-ai，且 opencode.cmd 在 PATH 中。"
    )


def kill_proc_tree(proc: subprocess.Popen[str]) -> None:
    if proc.poll() is not None:
        return
    try:
        if sys.platform == "win32":
            subprocess.run(
                ["taskkill", "/F", "/T", "/PID", str(proc.pid)],
                capture_output=True,
                check=False,
            )
        else:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
    except Exception:
        try:
            proc.kill()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# OpenCode event handling
# ---------------------------------------------------------------------------

def event_text(event: dict[str, Any]) -> str:
    event_type = str(event.get("type") or "")
    part = event.get("part")

    if event_type == "text" and isinstance(part, dict):
        return str(part.get("text") or "")

    # Compatibility fallbacks for older/newer event shapes.
    if isinstance(part, dict) and part.get("type") == "text":
        return str(part.get("text") or "")
    if isinstance(event.get("text"), str):
        return str(event.get("text") or "")

    message = event.get("message")
    if isinstance(message, dict):
        content = message.get("content")
        if isinstance(content, str):
            return content
        if isinstance(content, list):
            chunks: list[str] = []
            for item in content:
                if isinstance(item, dict) and item.get("type") in {"text", "output_text"}:
                    chunks.append(str(item.get("text") or ""))
            return "".join(chunks)

    return ""


def extract_event_error(event: dict[str, Any]) -> str:
    if event.get("type") != "error":
        return ""
    error = event.get("error")
    if isinstance(error, str):
        return error
    if isinstance(error, dict):
        data = error.get("data")
        if isinstance(data, dict) and data.get("message"):
            return str(data.get("message"))
        if error.get("message"):
            return str(error.get("message"))
        if error.get("name"):
            return str(error.get("name"))
        return json.dumps(error, ensure_ascii=False)
    return str(error or "OpenCode emitted an error event")


def parse_step_finish(event: dict[str, Any]) -> Optional[dict[str, Any]]:
    if event.get("type") != "step_finish":
        return None
    part = event.get("part")
    if not isinstance(part, dict):
        return None

    tokens = part.get("tokens") or {}
    if not isinstance(tokens, dict):
        tokens = {}
    cache = tokens.get("cache") or {}
    if not isinstance(cache, dict):
        cache = {}

    def to_int(value: Any) -> int:
        try:
            return int(value or 0)
        except (TypeError, ValueError):
            return 0

    def to_float(value: Any) -> float:
        try:
            return float(value or 0.0)
        except (TypeError, ValueError):
            return 0.0

    return {
        "input_tokens": to_int(tokens.get("input")),
        "output_tokens": to_int(tokens.get("output")),
        "reasoning_tokens": to_int(tokens.get("reasoning")),
        "cache_read_tokens": to_int(cache.get("read")),
        "cache_write_tokens": to_int(cache.get("write")),
        "cost_usd": to_float(part.get("cost")),
        "finish_reason": str(part.get("reason") or ""),
    }


def normalize_opencode_event(event: dict[str, Any]) -> list[dict[str, Any]]:
    """
    Convert OpenCode raw events to the generic PinchBench transcript shape.

    PinchBench graders expect events shaped like:
      {"type": "message", "message": {"role": ..., "content": [...]}}
    """
    event_type = str(event.get("type") or "")
    normalized: list[dict[str, Any]] = []

    if event_type == "text":
        text = event_text(event)
        if text:
            normalized.append({
                "type": "message",
                "message": {
                    "role": "assistant",
                    "content": [{"type": "text", "text": text}],
                },
                "_adapter": {"source": "opencode", "event_type": event_type},
            })
        return normalized

    if event_type == "tool_use":
        part = event.get("part")
        if not isinstance(part, dict):
            return normalized

        state = part.get("state")
        if not isinstance(state, dict):
            state = {}

        tool_name = str(part.get("tool") or "")
        call_id = str(part.get("callID") or part.get("id") or "")
        tool_input = state.get("input")
        if not isinstance(tool_input, dict):
            tool_input = {}

        normalized.append({
            "type": "message",
            "message": {
                "role": "assistant",
                "content": [{
                    "type": "toolCall",
                    "id": call_id,
                    "name": tool_name,
                    "arguments": tool_input,
                    "input": tool_input,
                }],
            },
            "_adapter": {
                "source": "opencode",
                "event_type": event_type,
                "status": state.get("status"),
            },
        })

        status = str(state.get("status") or "")
        output = state.get("output")
        error = state.get("error")
        result_text = str(output if output is not None else error if error is not None else "")
        normalized.append({
            "type": "message",
            "message": {
                "role": "toolResult",
                "content": [{
                    "type": "toolResult",
                    "toolCallId": call_id,
                    "name": tool_name,
                    "content": result_text,
                    "isError": status == "error",
                }],
            },
            "_adapter": {
                "source": "opencode",
                "event_type": event_type,
                "status": status,
            },
        })
        return normalized

    if event_type == "error":
        error_text = extract_event_error(event)
        normalized.append({
            "type": "message",
            "message": {
                "role": "assistant",
                "content": [{"type": "text", "text": f"[OpenCode error] {error_text}"}],
            },
            "_adapter": {"source": "opencode", "event_type": event_type},
        })

    return normalized


def make_user_transcript_event(prompt: str, turn_index: int, session_label: str) -> dict[str, Any]:
    return {
        "type": "message",
        "message": {
            "role": "user",
            "content": [{"type": "text", "text": prompt}],
        },
        "_adapter": {
            "source": "pinchbench-opencode-runner",
            "turn": turn_index,
            "session_label": session_label,
        },
    }


def _stream_reader(
    stream: Any,
    kind: str,
    output_queue: "queue.Queue[tuple[str, Optional[str]]]",
) -> None:
    try:
        for line in iter(stream.readline, ""):
            output_queue.put((kind, line))
    finally:
        output_queue.put((kind, None))
        try:
            stream.close()
        except Exception:
            pass


def run_opencode_streaming(
    cmd: list[str],
    cwd: Path,
    timeout: float,
    raw_stdout_path: Path,
    stderr_path: Path,
    stdin_path: Path,
) -> dict[str, Any]:
    """
    Run one OpenCode turn.

    The prompt is supplied through ``stdin_path`` rather than as a positional
    command-line argument.  This is intentional: on native Windows an npm
    ``opencode.cmd`` shim is parsed by ``cmd.exe`` and multiline prompt
    arguments can be truncated or altered.  OpenCode 1.18.4 reads non-TTY
    stdin for ``opencode run``, so a UTF-8 file handle preserves the prompt
    exactly and also leaves an auditable copy beside the raw transcript.

    Uses reader threads rather than selectors because Windows selectors do not
    reliably support ordinary subprocess pipes.
    """
    monotonic_start = time.monotonic()
    wall_start_ms = time.time() * 1000.0
    deadline = monotonic_start + timeout

    raw_stdout_path.parent.mkdir(parents=True, exist_ok=True)
    stderr_path.parent.mkdir(parents=True, exist_ok=True)

    stdout_lines: list[str] = []
    stderr_lines: list[str] = []
    raw_events: list[dict[str, Any]] = []
    normalized_events: list[dict[str, Any]] = []
    output_chunks: list[str] = []
    event_counts: Counter[str] = Counter()

    session_id: Optional[str] = None
    ttft: Optional[float] = None
    error_messages: list[str] = []
    status = "success"
    timed_out = False

    usage_seen = False
    input_tokens = 0
    output_tokens = 0
    reasoning_tokens = 0
    cache_read_tokens = 0
    cache_write_tokens = 0
    cost_usd = 0.0
    step_count = 0
    tool_errors = 0
    finish_reasons: list[str] = []

    env = os.environ.copy()
    env.setdefault("PYTHONUTF8", "1")
    env.setdefault("OPENCODE_DISABLE_AUTOUPDATE", "true")
    env.setdefault("NO_COLOR", "1")

    creationflags = 0
    if sys.platform == "win32":
        creationflags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)

    try:
        with stdin_path.open("rb") as stdin_file:
            proc = subprocess.Popen(
                cmd,
                cwd=str(cwd),
                stdin=stdin_file,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace",
                bufsize=1,
                env=env,
                start_new_session=(sys.platform != "win32"),
                creationflags=creationflags,
            )
    except Exception as exc:
        return {
            "status": "error",
            "success": False,
            "returncode": None,
            "output": "",
            "error": f"无法启动 OpenCode: {exc}",
            "stderr": "",
            "elapsed": time.monotonic() - monotonic_start,
            "ttft": None,
            "input_tokens": None,
            "output_tokens": None,
            "reasoning_tokens": None,
            "cache_read_tokens": None,
            "cache_write_tokens": None,
            "cost_usd": None,
            "total_tokens": None,
            "token_source": "opencode_step_finish",
            "token_coverage_complete": False,
            "token_verified_against_openrouter": False,
            "usage_complete": False,
            "step_count": 0,
            "tool_errors": 0,
            "finish_reasons": [],
            "session_id": None,
            "raw_events": [],
            "normalized_events": [],
            "event_counts": {},
            "prompt_transport": "stdin_file",
            "prompt_path": str(stdin_path),
        }

    if proc.stdout is None or proc.stderr is None:
        kill_proc_tree(proc)
        raise RuntimeError("OpenCode stdout/stderr pipe creation failed")

    output_queue: "queue.Queue[tuple[str, Optional[str]]]" = queue.Queue()
    threads = [
        threading.Thread(
            target=_stream_reader,
            args=(proc.stdout, "stdout", output_queue),
            daemon=True,
        ),
        threading.Thread(
            target=_stream_reader,
            args=(proc.stderr, "stderr", output_queue),
            daemon=True,
        ),
    ]
    for thread in threads:
        thread.start()

    open_streams = {"stdout", "stderr"}

    with (
        raw_stdout_path.open("w", encoding="utf-8", newline="") as raw_file,
        stderr_path.open("w", encoding="utf-8", newline="") as err_file,
    ):
        while open_streams or proc.poll() is None:
            if time.monotonic() > deadline:
                timed_out = True
                status = "timeout"
                error_messages.append(f"超时 ({timeout:.0f}s)")
                kill_proc_tree(proc)
                break

            try:
                kind, line = output_queue.get(timeout=0.2)
            except queue.Empty:
                continue

            if line is None:
                open_streams.discard(kind)
                continue

            if kind == "stderr":
                stderr_lines.append(line)
                err_file.write(line)
                err_file.flush()
                continue

            stdout_lines.append(line)
            raw_file.write(line)
            raw_file.flush()

            stripped = line.strip()
            if not stripped:
                continue

            try:
                event = json.loads(stripped)
            except json.JSONDecodeError:
                if ttft is None:
                    ttft = time.monotonic() - monotonic_start
                output_chunks.append(stripped)
                event_counts["non_json_stdout"] += 1
                continue

            if not isinstance(event, dict):
                event_counts["non_object_json"] += 1
                continue

            raw_events.append(event)
            event_type = str(event.get("type") or "unknown")
            event_counts[event_type] += 1

            if not session_id and event.get("sessionID"):
                session_id = str(event.get("sessionID"))

            text = event_text(event)
            if text:
                if ttft is None:
                    part = event.get("part")
                    candidate: Optional[float] = None
                    if isinstance(part, dict):
                        part_time = part.get("time")
                        if isinstance(part_time, dict):
                            try:
                                start_ms = float(part_time.get("start"))
                                estimated = (start_ms - wall_start_ms) / 1000.0
                                if 0.0 <= estimated <= (time.monotonic() - monotonic_start + 5.0):
                                    candidate = estimated
                            except (TypeError, ValueError):
                                pass
                    ttft = candidate if candidate is not None else time.monotonic() - monotonic_start
                output_chunks.append(text)

            normalized_events.extend(normalize_opencode_event(event))

            usage = parse_step_finish(event)
            if usage is not None:
                usage_seen = True
                step_count += 1
                input_tokens += usage["input_tokens"]
                output_tokens += usage["output_tokens"]
                reasoning_tokens += usage["reasoning_tokens"]
                cache_read_tokens += usage["cache_read_tokens"]
                cache_write_tokens += usage["cache_write_tokens"]
                cost_usd += usage["cost_usd"]
                if usage["finish_reason"]:
                    finish_reasons.append(usage["finish_reason"])

            if event_type == "tool_use":
                part = event.get("part")
                if isinstance(part, dict):
                    state = part.get("state")
                    if isinstance(state, dict) and state.get("status") == "error":
                        tool_errors += 1

            event_error = extract_event_error(event)
            if event_error:
                error_messages.append(event_error)
                status = "error"

    try:
        proc.wait(timeout=15)
    except subprocess.TimeoutExpired:
        kill_proc_tree(proc)
        status = "timeout"
        timed_out = True
        error_messages.append("进程未正常退出，已强制终止")

    # Drain anything already queued after process exit.
    drain_deadline = time.monotonic() + 2.0
    while time.monotonic() < drain_deadline:
        try:
            kind, line = output_queue.get_nowait()
        except queue.Empty:
            break
        if line is None:
            continue
        if kind == "stderr":
            stderr_lines.append(line)
        else:
            stdout_lines.append(line)

    for thread in threads:
        thread.join(timeout=1)

    returncode = proc.returncode
    stderr = "".join(stderr_lines).strip()

    if returncode not in (0, None) and status == "success":
        status = "error"
        error_messages.append(stderr or f"OpenCode 退出码 {returncode}")

    # A non-zero exit or timeout is failure. Missing final text alone is not.
    success = status == "success" and not timed_out and returncode in (0, None)

    output = "\n".join(chunk.strip() for chunk in output_chunks if chunk.strip()).strip()
    elapsed = time.monotonic() - monotonic_start

    return {
        "status": status,
        "success": success,
        "returncode": returncode,
        "output": output,
        "error": " | ".join(dict.fromkeys(message for message in error_messages if message)),
        "stderr": stderr,
        "elapsed": elapsed,
        "ttft": ttft,
        "input_tokens": input_tokens if usage_seen else None,
        "output_tokens": output_tokens if usage_seen else None,
        "reasoning_tokens": reasoning_tokens if usage_seen else None,
        "cache_read_tokens": cache_read_tokens if usage_seen else None,
        "cache_write_tokens": cache_write_tokens if usage_seen else None,
        "cost_usd": cost_usd if usage_seen else None,
        "usage_complete": usage_seen,
        "step_count": step_count,
        "tool_errors": tool_errors,
        "finish_reasons": finish_reasons,
        "session_id": session_id,
        "raw_events": raw_events,
        "normalized_events": normalized_events,
        "event_counts": dict(event_counts),
        "stdout_lines": stdout_lines,
        "prompt_transport": "stdin_file",
        "prompt_path": str(stdin_path),
    }



def _normalize_judge_content(content: Any) -> str:
    if content is None:
        return ""
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts: list[str] = []
        for item in content:
            if isinstance(item, str):
                parts.append(item)
            elif isinstance(item, dict):
                value = item.get("text")
                if value is None:
                    value = item.get("content")
                if value is not None:
                    parts.append(str(value))
        return "\n".join(parts)
    return str(content)


def _resolve_git_bash() -> Optional[str]:
    candidates = [
        Path(r"C:\Program Files\Git\bin\bash.exe"),
        Path(r"C:\Program Files\Git\usr\bin\bash.exe"),
        Path(r"C:\Program Files (x86)\Git\bin\bash.exe"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return str(candidate)
    found = shutil.which("bash")
    if found and "\\git\\" in str(found).lower():
        return str(Path(found).resolve())
    return None


def install_pinchbench_grader_compat(grader_module: Any) -> dict[str, Any]:
    """Apply deterministic Windows/Judge transport compatibility only.

    This does not change task rubrics, automated checks, workspace outputs, or
    the selected Judge model. It prevents known Windows Bash failures and
    treats empty/truncated/unparseable OpenRouter Judge responses as retryable
    transport failures instead of valid grades.
    """
    if getattr(grader_module, "_our_framework_grader_compat_v3", False):
        return dict(getattr(grader_module, "_our_framework_grader_compat_meta", {}) or {})

    meta: dict[str, Any] = {
        "revision": GRADER_COMPAT_REVISION,
        "judge_none_content_guard": False,
        "judge_parse_retry": False,
        "judge_json_schema": False,
        "judge_max_completion_tokens": 8192,
        "windows_git_bash_compat": False,
        "bash_path": "",
    }

    original_parse = getattr(grader_module, "_parse_judge_text", None)
    if callable(original_parse):
        def safe_parse(raw_text: Any) -> dict[str, Any]:
            normalized = _normalize_judge_content(raw_text)
            if not normalized.strip():
                return {}
            return original_parse(normalized)
        grader_module._parse_judge_text = safe_parse
        meta["judge_none_content_guard"] = True

    original_extract = getattr(grader_module, "_extract_grading_code", None)
    bash_path = _resolve_git_bash()
    if callable(original_extract):
        def patched_extract(task: Any) -> str:
            code = original_extract(task)
            if os.name == "nt" and '/bin/bash' in code:
                if not bash_path:
                    raise RuntimeError(
                        "PinchBench Windows grader compatibility requires Git Bash for "
                        "task_git_rescue_recovery, but bash.exe was not found."
                    )
                code = code.replace(
                    "return subprocess.run(\n                cmd,\n                cwd=repo,\n                shell=True,\n                executable=\"/bin/bash\",",
                    f"return subprocess.run(\n                [{bash_path!r}, '-c', cmd],\n                cwd=repo,\n                shell=False,",
                )
                code = code.replace(
                    "execution = subprocess.run(\n            script,\n            cwd=repo,\n            shell=True,\n            executable=\"/bin/bash\",",
                    f"execution = subprocess.run(\n            [{bash_path!r}, '-c', script],\n            cwd=repo,\n            shell=False,",
                )
                # Generic fallback for grading snippets that invoke /bin/bash in a
                # shape not covered by the two historical transformations above.
                # Replacing the literal executable path preserves the grading script
                # semantics while making it runnable on native Windows Git Bash.
                if '/bin/bash' in code:
                    code = code.replace('/bin/bash', bash_path.replace('\\', '/'))
            return code
        grader_module._extract_grading_code = patched_extract
        meta["windows_git_bash_compat"] = bool(bash_path)
        meta["bash_path"] = bash_path or ""

    if callable(getattr(grader_module, "call_judge_api", None)):
        def robust_call_judge_api(*, prompt: str, model: str, timeout_seconds: float = 300.0) -> dict[str, Any]:
            key = os.environ.get("OPENROUTER_API_KEY", "")
            if not key:
                return {"status": "error", "text": "", "error": "OPENROUTER_API_KEY not set"}

            bare_model = model.removeprefix("openrouter/")
            endpoint = "https://openrouter.ai/api/v1/chat/completions"
            system_msg = (
                "You are a strict grading function. "
                "Respond with ONLY a JSON object, no prose, no markdown fences, no extra text. "
                "You MUST provide at least one criterion score and a non-empty concise note."
            )

            def semantic_payload(parsed: dict[str, Any]) -> dict[str, Any] | None:
                normalize = getattr(grader_module, "_normalize_judge_response", None)
                normalized = normalize(parsed) if callable(normalize) else parsed
                if not isinstance(normalized, dict):
                    return None
                scores = normalized.get("scores") or {}
                total = normalized.get("total")
                notes = str(normalized.get("notes") or "").strip()
                if not isinstance(scores, dict) or not scores:
                    return None
                clean_scores: dict[str, float] = {}
                for name, value in scores.items():
                    if isinstance(value, bool) or not isinstance(value, (int, float)):
                        return None
                    fv = float(value)
                    if not 0.0 <= fv <= 1.0:
                        return None
                    clean_scores[str(name)] = fv
                if isinstance(total, bool) or not isinstance(total, (int, float)):
                    return None
                ft = float(total)
                if not 0.0 <= ft <= 1.0 or not notes:
                    return None
                # Enforce the prompt's arithmetic rule instead of trusting a malformed total.
                mean_score = sum(clean_scores.values()) / len(clean_scores)
                if abs(ft - mean_score) > 0.02:
                    ft = mean_score
                return {"scores": clean_scores, "total": ft, "notes": notes}

            last_error = ""
            for json_mode in (False,):
                body: dict[str, Any] = {
                    "model": bare_model,
                    "messages": [
                        {"role": "system", "content": system_msg},
                        {"role": "user", "content": prompt},
                    ],
                    "temperature": 0.0,
                    "max_completion_tokens": 8192,
                    "provider": {
                        "order": ["anthropic"],
                        "only": ["anthropic"],
                        "allow_fallbacks": False,
                    },
                }
                if json_mode:
                    # Avoid the previous dynamic json_schema.  Its schema allowed
                    # scores={} / notes="" and some providers could return that
                    # technically-valid but semantically-empty object.
                    body["response_format"] = {"type": "json_object"}

                req = urllib.request.Request(
                    endpoint,
                    data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
                    headers={
                        "Authorization": f"Bearer {key}",
                        "Content-Type": "application/json",
                        "HTTP-Referer": "https://pinchbench.com",
                        "X-Title": "PinchBench-Judge",
                    },
                    method="POST",
                )
                try:
                    with urllib.request.urlopen(req, timeout=timeout_seconds) as response:
                        data = json.loads(response.read().decode("utf-8"))
                except urllib.error.HTTPError as exc:
                    try:
                        detail = exc.read().decode("utf-8", errors="replace")[:1000]
                    except Exception:
                        detail = ""
                    last_error = f"HTTP {exc.code}: {detail}"
                    if json_mode and exc.code in (400, 404, 422):
                        continue
                    return {"status": "error", "text": "", "error": last_error}
                except urllib.error.URLError as exc:
                    return {"status": "error", "text": "", "error": str(exc)}
                except TimeoutError:
                    return {"status": "timeout", "text": "", "error": "Request timed out"}
                except Exception as exc:
                    return {"status": "error", "text": "", "error": f"{type(exc).__name__}: {exc}"}

                choices = data.get("choices") or []
                if not choices:
                    last_error = "No choices in response"
                    if json_mode:
                        continue
                    return {"status": "error", "text": "", "error": last_error}

                choice = choices[0] or {}
                message = choice.get("message") or {}
                judge_text = _normalize_judge_content(message.get("content"))
                finish_reason = str(choice.get("finish_reason") or "")
                choice_error = choice.get("error") or data.get("error")
                provider_name = str(data.get("provider") or data.get("provider_name") or "anthropic")

                if choice_error:
                    last_error = f"Judge provider error: {choice_error}"
                    if json_mode:
                        continue
                    return {"status": "error", "text": "", "error": last_error}
                if finish_reason.lower() in {"length", "error"}:
                    last_error = f"Judge response truncated/errored (finish_reason={finish_reason})"
                    if json_mode:
                        continue
                    return {"status": "error", "text": "", "error": last_error}

                parsed = safe_parse(judge_text) if callable(original_parse) else {}
                canonical = semantic_payload(parsed)
                if canonical is not None:
                    LOGGER.info(
                        "Judge accepted: model=%s provider=%s mode=%s scores=%d total=%.4f notes_chars=%d",
                        bare_model, provider_name, "json_object" if json_mode else "plain",
                        len(canonical["scores"]), canonical["total"], len(canonical["notes"]),
                    )
                    return {"status": "success", "text": json.dumps(canonical, ensure_ascii=False)}

                last_error = (
                    "Judge response was parseable but semantically incomplete "
                    f"(provider={provider_name}, mode={'json_object' if json_mode else 'plain'}, "
                    f"chars={len(judge_text)})"
                )
                if json_mode:
                    continue
                # Important: return empty text on failure so lib_grading cannot
                # accidentally parse a failed/partial payload into a real score.
                return {"status": "error", "text": "", "error": last_error}

            return {"status": "error", "text": "", "error": last_error or "Judge call failed"}

        grader_module.call_judge_api = robust_call_judge_api
        meta["judge_parse_retry"] = True
        meta["judge_json_schema"] = False

    grader_module._our_framework_grader_compat_v3 = True
    grader_module._our_framework_grader_compat_meta = dict(meta)
    return meta


# ---------------------------------------------------------------------------
# PinchBench official/default grading
# ---------------------------------------------------------------------------

def load_pinchbench_grading(skill_dir: Path) -> tuple[Optional[Any], Optional[str]]:
    """Load the grading engine from the exact checked-out PinchBench commit."""
    scripts_dir = (skill_dir / "scripts").resolve()
    grading_path = scripts_dir / "lib_grading.py"
    if not grading_path.exists():
        return None, f"找不到 PinchBench grading engine: {grading_path}"

    scripts_value = str(scripts_dir)
    if scripts_value not in sys.path:
        sys.path.insert(0, scripts_value)

    try:
        import lib_grading  # type: ignore
        install_pinchbench_grader_compat(lib_grading)
    except Exception as exc:
        return None, (
            "无法导入 PinchBench scripts/lib_grading.py。"
            "请安装本地仓库依赖，例如："
            r"C:\pinchbench-opencode\.venv\Scripts\python.exe -m pip install -e "
            f'"{skill_dir}"。原始错误: {exc}'
        )

    return lib_grading, None


def grade_with_pinchbench_default(
    *,
    grader_module: Any,
    task: Task,
    execution_result: dict[str, Any],
    workspace: Path,
    skill_dir: Path,
    judge_timeout: float,
    judge_model: str,
    verbose: bool,
) -> GradeResult:
    """Use PinchBench's own grader with its explicitly pinned judge model.

    The judge is called directly through the API backend.  For the current
    commit, DEFAULT_JUDGE_MODEL is openrouter/anthropic/claude-opus-5 and
    lib_agent.py reads OPENROUTER_API_KEY.
    """
    grading_execution = dict(execution_result)
    grading_execution["workspace"] = str(workspace)

    try:
        result = grader_module.grade_task(
            task=task,
            execution_result=grading_execution,
            skill_dir=skill_dir,
            judge_model=judge_model,
            judge_timeout_seconds=judge_timeout,
            judge_backend="api",
            verbose=verbose,
        )

        raw_score = getattr(result, "score", None)
        max_score = getattr(result, "max_score", 1.0)
        score: Optional[float]
        if raw_score is None:
            score = None
        else:
            score = float(raw_score)
            max_score_float = float(max_score or 1.0)
            if max_score_float != 1.0:
                score = score / max_score_float
            score = max(0.0, min(1.0, score))

        notes = str(getattr(result, "notes", "") or "")
        grade_error: Optional[str] = None
        if task.grading_type in {"llm_judge", "hybrid"}:
            lower_notes = notes.lower()
            failure_markers = (
                "llm judge failed",
                "no parseable response",
                "response parsed but no score",
                "openrouter_api_key not set",
                "judge api call failed",
            )
            if any(marker in lower_notes for marker in failure_markers):
                grade_error = notes or "LLM judge failed"

        return GradeResult(
            score=score,
            grading_type=str(getattr(result, "grading_type", task.grading_type)),
            breakdown=dict(getattr(result, "breakdown", {}) or {}),
            notes=notes,
            error=grade_error,
        )
    except Exception as exc:
        return GradeResult(
            score=None,
            grading_type=task.grading_type,
            breakdown={},
            notes="",
            error=(
                "PinchBench 官方评分失败: "
                f"{exc}\n{traceback.format_exc(limit=8)}"
            ),
        )


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def truncate_excel(value: Any, limit: int = 32000) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    return text if len(text) <= limit else text[:limit] + "...[truncated]"


def save_csv(results: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "task_id", "name", "category", "grading_type", "model", "network_task",
        "multi_session", "session_count", "success", "status", "returncode",
        "score", "elapsed", "agent_elapsed", "grading_elapsed", "end_to_end_elapsed",
        "ttft", "input_tokens", "output_tokens", "reasoning_tokens",
        "cache_read_tokens", "cache_write_tokens", "total_tokens", "cost_usd",
        "token_source", "token_coverage_complete", "token_verified_against_openrouter",
        "usage_complete", "step_count", "tool_errors",
        "workspace", "transcript", "error", "grade_error", "grade_notes",
    ]

    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for row in results:
            writer.writerow({
                key: truncate_excel(row.get(key), 10000)
                for key in fields
            })


def save_xlsx(
    results: list[dict[str, Any]],
    summary: dict[str, Any],
    output_path: Path,
) -> None:
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
    except ImportError:
        LOGGER.warning(
            "未安装 openpyxl，跳过 XLSX。可运行: "
            "python -m pip install openpyxl"
        )
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "详细结果"

    headers = [
        "任务ID", "名称", "类别", "打分类型", "模型", "联网题", "多轮任务",
        "会话调用数", "成功", "状态", "退出码", "分数", "旧兼容耗时(s)",
        "Agent耗时(s)", "Grading耗时(s)", "端到端耗时(s)", "TTFT估计(s)",
        "输入Token", "输出Token", "推理Token", "缓存读取Token", "缓存写入Token",
        "派生总Token", "费用USD(OpenCode)", "Token来源", "Token覆盖完整",
        "已与OpenRouter核验", "Usage完整", "Step数", "工具错误数",
        "Workspace", "Transcript", "运行错误", "打分错误", "打分备注",
    ]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(results, start=2):
        values = [
            row.get("task_id"),
            row.get("name"),
            row.get("category"),
            row.get("grading_type"),
            row.get("model") or "",
            "是" if row.get("network_task") else "否",
            "是" if row.get("multi_session") else "否",
            row.get("session_count"),
            "✓" if row.get("success") else "✗",
            row.get("status"),
            row.get("returncode") if row.get("returncode") is not None else "",
            row.get("score") if row.get("score") is not None else "",
            row.get("elapsed"),
            row.get("agent_elapsed") if row.get("agent_elapsed") is not None else "",
            row.get("grading_elapsed") if row.get("grading_elapsed") is not None else "",
            row.get("end_to_end_elapsed") if row.get("end_to_end_elapsed") is not None else "",
            row.get("ttft") if row.get("ttft") is not None else "",
            row.get("input_tokens") if row.get("input_tokens") is not None else "",
            row.get("output_tokens") if row.get("output_tokens") is not None else "",
            row.get("reasoning_tokens") if row.get("reasoning_tokens") is not None else "",
            row.get("cache_read_tokens") if row.get("cache_read_tokens") is not None else "",
            row.get("cache_write_tokens") if row.get("cache_write_tokens") is not None else "",
            row.get("total_tokens") if row.get("total_tokens") is not None else "",
            row.get("cost_usd") if row.get("cost_usd") is not None else "",
            row.get("token_source") or "",
            "是" if row.get("token_coverage_complete") else "否",
            "是" if row.get("token_verified_against_openrouter") else "否",
            "是" if row.get("usage_complete") else "否",
            row.get("step_count"),
            row.get("tool_errors"),
            row.get("workspace"),
            row.get("transcript"),
            row.get("error") or "",
            row.get("grade_error") or "",
            row.get("grade_notes") or "",
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(
                row=row_index,
                column=column,
                value=truncate_excel(value),
            )

    for cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in cells)
        worksheet.column_dimensions[cells[0].column_letter].width = min(
            max_length + 3,
            60,
        )

    summary_sheet = workbook.create_sheet("汇总")
    for row_index, (key, value) in enumerate(summary.items(), start=1):
        summary_sheet.cell(row=row_index, column=1, value=key).font = Font(bold=True)
        summary_sheet.cell(row=row_index, column=2, value=value)
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 30

    workbook.save(output_path)


def optional_sum(results: list[dict[str, Any]], field: str) -> Optional[float]:
    values = [row.get(field) for row in results if row.get(field) is not None]
    if not values:
        return None
    return float(sum(float(value) for value in values))


def print_summary(
    results: list[dict[str, Any]],
    total_elapsed: float,
) -> dict[str, Any]:
    succeeded = sum(1 for row in results if row.get("success"))
    failed = len(results) - succeeded
    scores = [
        float(row["score"])
        for row in results
        if row.get("score") is not None
    ]
    ttfts = [
        float(row["ttft"])
        for row in results
        if row.get("ttft") is not None
    ]
    average_elapsed = (
        sum(float(row.get("agent_elapsed") or row.get("elapsed") or 0.0) for row in results) / len(results)
        if results else 0.0
    )
    total_agent_elapsed = sum(float(row.get("agent_elapsed") or row.get("elapsed") or 0.0) for row in results)
    total_grading_elapsed = sum(float(row.get("grading_elapsed") or 0.0) for row in results)
    total_end_to_end_elapsed = sum(float(row.get("end_to_end_elapsed") or 0.0) for row in results)

    summary = {
        "任务总数": len(results),
        "成功": succeeded,
        "失败": failed,
        "成功率": round(succeeded / len(results), 4) if results else 0.0,
        "有分数任务数": len(scores),
        "打分失败任务数": sum(1 for row in results if row.get("grade_error")),
        "平均分数": round(sum(scores) / len(scores), 4) if scores else None,
        "整批墙钟总耗时(s)": round(total_elapsed, 2),
        "Agent累计耗时(s)": round(total_agent_elapsed, 2),
        "Grading累计耗时(s)": round(total_grading_elapsed, 2),
        "任务端到端累计耗时(s)": round(total_end_to_end_elapsed, 2),
        "平均Agent耗时/任务(s)": round(average_elapsed, 2),
        "平均TTFT估计(s)": round(sum(ttfts) / len(ttfts), 4) if ttfts else None,
        "总输入Token": int(optional_sum(results, "input_tokens") or 0)
            if any(row.get("input_tokens") is not None for row in results) else None,
        "总输出Token": int(optional_sum(results, "output_tokens") or 0)
            if any(row.get("output_tokens") is not None for row in results) else None,
        "总推理Token": int(optional_sum(results, "reasoning_tokens") or 0)
            if any(row.get("reasoning_tokens") is not None for row in results) else None,
        "总缓存读取Token": int(optional_sum(results, "cache_read_tokens") or 0)
            if any(row.get("cache_read_tokens") is not None for row in results) else None,
        "总缓存写入Token": int(optional_sum(results, "cache_write_tokens") or 0)
            if any(row.get("cache_write_tokens") is not None for row in results) else None,
        "派生总Token": int(optional_sum(results, "total_tokens") or 0)
            if any(row.get("total_tokens") is not None for row in results) else None,
        "Token来源": "OpenCode step_finish JSONL (逐 step 累加)",
        "已与OpenRouter逐请求核验": False,
        "总费用USD(OpenCode上报)": round(optional_sum(results, "cost_usd") or 0.0, 6)
            if any(row.get("cost_usd") is not None for row in results) else None,
        "Usage缺失任务数": sum(
            1 for row in results if not row.get("usage_complete")
        ),
        "联网任务数": sum(1 for row in results if row.get("network_task")),
        "多轮任务数": sum(1 for row in results if row.get("multi_session")),
    }

    print("\n" + "=" * 100)
    print("汇总")
    print("=" * 100)
    for key, value in summary.items():
        print(f"{key:<22}: {value}")

    print("-" * 100)
    for row in results:
        marker = "✓" if row.get("success") else "✗"
        score = (
            f"{row['score']:.3f}"
            if row.get("score") is not None
            else "N/A"
        )
        ttft = (
            f"{row['ttft']:.2f}s"
            if row.get("ttft") is not None
            else "N/A"
        )
        error = row.get("error") or row.get("grade_error") or ""
        error_suffix = f"  [{str(error)[:70]}]" if error else ""
        print(
            f"{marker} {row['task_id']:<42} "
            f"score={score:<6} elapsed={row['elapsed']:.1f}s "
            f"ttft={ttft:<8}{error_suffix}"
        )
    print("=" * 100)
    return summary


# ---------------------------------------------------------------------------
# Preflight and run metadata
# ---------------------------------------------------------------------------

def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def command_output(command: list[str], cwd: Optional[Path] = None) -> str:
    try:
        result = subprocess.run(
            command,
            cwd=str(cwd) if cwd else None,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
            check=False,
        )
        return (result.stdout or result.stderr).strip()
    except Exception as exc:
        return f"unavailable: {exc}"


def write_run_config(
    output_path: Path,
    args: argparse.Namespace,
    skill_dir: Path,
    tasks_dir: Path,
    opencode_command: str,
    selected: list[Task],
    judge_model: str,
) -> None:
    manifest_path = tasks_dir / "manifest.yaml"
    git_command = resolve_command("git")

    commit = ""
    if git_command:
        commit = command_output(
            [git_command, "rev-parse", "HEAD"],
            cwd=skill_dir,
        )

    config = {
        "runner_revision": RUNNER_REVISION,
        "created_at": dt.datetime.now(dt.timezone.utc).astimezone().isoformat(),
        "platform": platform.platform(),
        "python": sys.version,
        "opencode_command": opencode_command,
        "opencode_version": command_output([opencode_command, "--version"]),
        "model": args.model,
        "agent": "OpenCode default (not explicitly set)",
        "variant": "default (not explicitly set)",
        "auto_approval": False,
        "prompt_transport": "stdin_file_utf8",
        "prompt_in_argv": False,
        "openrouter_routing": "default",
        "worker_count": 1,
        "task_concurrency": 1,
        "skill_dir": str(skill_dir),
        "tasks_dir": str(tasks_dir),
        "pinchbench_commit": commit,
        "manifest_sha256": sha256_file(manifest_path) if manifest_path.exists() else None,
        "suite": args.suite,
        "limit": args.limit,
        "default_skipped_tasks": sorted(DEFAULT_SKIPPED_TASKS),
        "additional_skip": args.skip,
        "skip_network": args.skip_network,
        "task_count": len(selected),
        "task_ids": [task.task_id for task in selected],
        "timeout_multiplier": args.timeout_multiplier,
        "network_timeout": args.network_timeout,
        "workspace_instruction": not args.no_workspace_instruction,
        "grading_enabled": not args.no_grade,
        "grading_engine": str(skill_dir / "scripts" / "lib_grading.py"),
        "grader_compat_revision": GRADER_COMPAT_REVISION,
        "judge_transport_policy": "anthropic_provider_pinned; plain_json_text; semantic-empty responses retry/fail; failed calls never parsed",
        "judge_max_completion_tokens": 8192,
        "judge_backend": "api",
        "judge_model": judge_model,
        "judge_key_env": "OPENROUTER_API_KEY",
        "judge_key_present": bool(os.environ.get("OPENROUTER_API_KEY")),
        "judge_concurrency": 1,
        "judge_is_separate_from_tested_model": True,
        "token_accounting": {
            "source": "OpenCode --format json step_finish events",
            "aggregation": "sum every step_finish across all turns",
            "derived_total_formula": "input + output + reasoning + cache_read + cache_write",
            "openrouter_reconciled": False,
            "cost_source": "OpenCode step_finish cost",
        },
        "environment_proxy_present": {
            key: bool(os.environ.get(key))
            for key in ("HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "NO_PROXY")
        },
    }
    output_path.write_text(
        json.dumps(config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def print_preflight(
    selected: list[Task],
    skill_dir: Path,
    tasks_dir: Path,
    opencode_command: str,
    grader_module: Optional[Any],
    grader_error: Optional[str],
    args: argparse.Namespace,
) -> tuple[dict[str, list[str]], list[str]]:
    prerequisite_failures: dict[str, list[str]] = {}
    fixture_failures: list[str] = []

    needs_judge = (
        not args.no_grade
        and any(task.grading_type in {"llm_judge", "hybrid"} for task in selected)
    )
    judge_model = str(args.judge_model)
    judge_key_present = bool(os.environ.get("OPENROUTER_API_KEY"))

    print("=" * 100)
    print("PinchBench OpenCode + Kimi K3 Windows runner preflight")
    print("=" * 100)
    print(f"Runner revision      : {RUNNER_REVISION}")
    print(f"Skill dir            : {skill_dir}")
    print(f"Tasks dir            : {tasks_dir}")
    print(f"OpenCode command     : {opencode_command}")
    print(f"OpenCode version     : {command_output([opencode_command, '--version'])}")
    print(f"Model                : {args.model}")
    print(f"Suite                : {args.suite}")
    print(f"Selected tasks       : {len(selected)}")
    print(f"Default skipped      : {len(DEFAULT_SKIPPED_TASKS)}")
    for task_id in sorted(DEFAULT_SKIPPED_TASKS):
        print(f"  - {task_id}")
    print("Worker/concurrency   : 1 / 1")
    print("Judge concurrency    : 1 (synchronous)")
    print("Default agent/tools  : yes (no --agent, no tool list)")
    print("Auto approval        : no")
    print("Prompt transport     : UTF-8 stdin file (never argv)")
    print("Reasoning variant    : default")
    print(f"Grading engine       : {skill_dir / 'scripts' / 'lib_grading.py'}")
    print("Judge backend        : api")
    print(f"Judge model          : {judge_model}")
    print(f"OPENROUTER_API_KEY   : {'set' if judge_key_present else 'missing'}")
    print("OpenClaw required    : no")
    print(f"Grader import        : {'ok' if grader_error is None else 'FAILED'}")
    if grader_error:
        print(f"  {grader_error}")
    print()

    category_counts = Counter(task.category or "(uncategorized)" for task in selected)
    grading_counts = Counter(task.grading_type for task in selected)
    print("Grading types:")
    for key, value in sorted(grading_counts.items()):
        print(f"  {key:<24} {value}")
    print("Categories:")
    for key, value in sorted(category_counts.items()):
        print(f"  {key:<24} {value}")
    print(f"Network-marked tasks : {sum(1 for task in selected if is_network_task(task))}")
    print(f"Multi-session tasks  : {sum(1 for task in selected if task.multi_session)}")
    print(f"Default judge needed : {'yes' if needs_judge else 'no'}")
    if needs_judge and not judge_key_present:
        print("Judge readiness      : FAILED (OPENROUTER_API_KEY missing)")
    elif needs_judge:
        print("Judge readiness      : ok")
    else:
        print("Judge readiness      : not required")
    print()

    for task in selected:
        prerequisites = task.metadata.get("prerequisites") or []
        missing = check_prerequisites(list(prerequisites)) if prerequisites else []
        if missing:
            prerequisite_failures[task.task_id] = missing

        for workspace_file in task.workspace_files:
            if not isinstance(workspace_file, dict):
                fixture_failures.append(f"{task.task_id}: invalid workspace file entry")
                continue
            if workspace_file.get("content") is not None:
                continue
            source = workspace_file.get("source")
            if not source:
                continue
            src_rel = Path(str(source))
            candidates = [
                skill_dir / "assets" / src_rel,
                skill_dir / src_rel,
                task.file_path.parent / src_rel,
                task.file_path.parent.parent / "assets" / src_rel,
            ]
            if not any(candidate.exists() for candidate in candidates):
                fixture_failures.append(f"{task.task_id}: missing fixture {source}")

    if prerequisite_failures:
        print("Missing declared prerequisites:")
        for task_id, missing in prerequisite_failures.items():
            print(f"  {task_id}: {', '.join(missing)}")
    else:
        print("Missing declared prerequisites: none")

    if fixture_failures:
        print("Missing workspace fixtures:")
        for item in fixture_failures:
            print(f"  {item}")
    else:
        print("Missing workspace fixtures: none")

    print("=" * 100)
    return prerequisite_failures, fixture_failures



# ---------------------------------------------------------------------------
# Our Framework benchmark execution (LocalClaw/LocalCoding source bundle)
# ---------------------------------------------------------------------------

OUR_FRAMEWORK_RUNNER_REVISION = "2026-08-13-our-framework-qwen-kimi-hybrid-v4.4-infra-backoff"
OUR_FRAMEWORK_ENDPOINT_ID = "openrouter-benchmark"
QWEN_MODEL = "qwen/qwen3.6-27b"
KIMI_MODEL = "moonshotai/kimi-k3"
OUR_FRAMEWORK_JUDGE_MODEL = "openrouter/anthropic/claude-opus-5"
GATEWAY_TOKEN = "localclaw-internal"


def build_turns(task: Task) -> list[dict[str, Any]]:
    if task.multi_session and task.sessions:
        turns: list[dict[str, Any]] = []
        for index, item in enumerate(task.sessions, start=1):
            prompt = str(item.get("prompt") or "").strip()
            if not prompt:
                continue
            turns.append({
                "id": str(item.get("id") or f"turn_{index}"),
                "prompt": prompt,
                "new_session": bool(item.get("new_session")),
            })
        if turns:
            return turns
    return [{"id": "single", "prompt": task.prompt, "new_session": True}]


def safe_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def pick(mapping: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in mapping and mapping[key] is not None:
            return mapping[key]
    return None


def usage_from_mapping(usage: Any) -> dict[str, Optional[float]]:
    if not isinstance(usage, dict):
        return {
            "input_tokens": None,
            "output_tokens": None,
            "reasoning_tokens": None,
            "cache_read_tokens": None,
            "cache_write_tokens": None,
            "cost_usd": None,
        }
    out_details = pick(usage, "output_tokens_details", "outputTokensDetails")
    reasoning = None
    if isinstance(out_details, dict):
        reasoning = safe_int(pick(out_details, "thinking_tokens", "thinkingTokens"))
    return {
        "input_tokens": safe_int(pick(usage, "input_tokens", "inputTokens")),
        "output_tokens": safe_int(pick(usage, "output_tokens", "outputTokens")),
        "reasoning_tokens": reasoning,
        "cache_read_tokens": safe_int(pick(usage, "cache_read_input_tokens", "cacheReadInputTokens", "cache_read_tokens", "cacheReadTokens")),
        "cache_write_tokens": safe_int(pick(usage, "cache_creation_input_tokens", "cacheCreationInputTokens", "cache_write_tokens", "cacheWriteTokens")),
        "cost_usd": safe_float(pick(usage, "cost_usd", "costUSD", "cost")),
    }


def billed_total_tokens(
    input_tokens: Optional[int],
    output_tokens: Optional[int],
    cache_read_tokens: Optional[int],
    cache_write_tokens: Optional[int],
) -> Optional[int]:
    # reasoning/thinking is a subset of output_tokens for Claude-style usage and
    # is therefore intentionally NOT added again.
    values = (input_tokens, output_tokens, cache_read_tokens, cache_write_tokens)
    if all(v is None for v in values):
        return None
    return int(sum(int(v or 0) for v in values))


def merge_usage(target: dict[str, Any], item: dict[str, Any]) -> None:
    for key in ("input_tokens", "output_tokens", "reasoning_tokens", "cache_read_tokens", "cache_write_tokens"):
        value = item.get(key)
        if value is not None:
            target[key] = int(target.get(key) or 0) + int(value)
    if item.get("cost_usd") is not None:
        target["cost_usd"] = float(target.get("cost_usd") or 0.0) + float(item["cost_usd"])


def normalize_sdk_message(msg: dict[str, Any]) -> list[dict[str, Any]]:
    msg_type = str(msg.get("type") or "")
    inner = msg.get("message")
    if not isinstance(inner, dict):
        inner = {}
    content = inner.get("content")
    if not isinstance(content, list):
        content = []

    normalized: list[dict[str, Any]] = []

    if msg_type == "assistant":
        blocks: list[dict[str, Any]] = []
        for block in content:
            if not isinstance(block, dict):
                continue
            btype = str(block.get("type") or "")
            if btype in {"text", "output_text"}:
                text = str(block.get("text") or "")
                if text:
                    blocks.append({"type": "text", "text": text})
            elif btype in {"tool_use", "toolCall"}:
                tool_input = block.get("input")
                if not isinstance(tool_input, dict):
                    tool_input = block.get("arguments") if isinstance(block.get("arguments"), dict) else {}
                blocks.append({
                    "type": "toolCall",
                    "id": str(block.get("id") or block.get("tool_use_id") or ""),
                    "name": str(block.get("name") or block.get("tool") or ""),
                    "arguments": tool_input,
                    "input": tool_input,
                })
        if blocks:
            normalized.append({
                "type": "message",
                "message": {"role": "assistant", "content": blocks},
                "_adapter": {"source": "our-framework", "sdk_type": msg_type},
            })
        return normalized

    if msg_type == "user":
        if msg.get("isReplay") is True:
            # Model-switch breadcrumb/replay noise is kept in raw logs and
            # escalation.status, but not duplicated into grader transcript.
            return normalized
        text_blocks: list[dict[str, Any]] = []
        for block in content:
            if not isinstance(block, dict):
                continue
            btype = str(block.get("type") or "")
            if btype == "tool_result":
                raw_content = block.get("content")
                if isinstance(raw_content, list):
                    parts: list[str] = []
                    for part in raw_content:
                        if isinstance(part, dict) and part.get("type") == "text":
                            parts.append(str(part.get("text") or ""))
                        elif isinstance(part, str):
                            parts.append(part)
                    result_text = "\n".join(x for x in parts if x)
                else:
                    result_text = str(raw_content or "")
                normalized.append({
                    "type": "message",
                    "message": {
                        "role": "toolResult",
                        "content": [{
                            "type": "toolResult",
                            "toolCallId": str(block.get("tool_use_id") or block.get("toolCallId") or ""),
                            "name": str(block.get("name") or ""),
                            "content": result_text,
                            "isError": bool(block.get("is_error") or block.get("isError")),
                        }],
                    },
                    "_adapter": {"source": "our-framework", "sdk_type": msg_type},
                })
            elif btype == "text":
                text = str(block.get("text") or "")
                if text:
                    text_blocks.append({"type": "text", "text": text})
        if text_blocks:
            normalized.append({
                "type": "message",
                "message": {"role": "user", "content": text_blocks},
                "_adapter": {"source": "our-framework", "sdk_type": msg_type},
            })
    return normalized


def extract_assistant_text(msg: dict[str, Any]) -> str:
    if msg.get("type") != "assistant":
        return ""
    inner = msg.get("message")
    if not isinstance(inner, dict):
        return ""
    content = inner.get("content")
    if not isinstance(content, list):
        return ""
    chunks: list[str] = []
    for block in content:
        if isinstance(block, dict) and block.get("type") in {"text", "output_text"}:
            text = str(block.get("text") or "")
            if text:
                chunks.append(text)
    return "\n".join(chunks)


def extract_result_usage(msg: dict[str, Any]) -> dict[str, Any]:
    usage = usage_from_mapping(msg.get("usage"))
    if usage.get("cost_usd") is None:
        usage["cost_usd"] = safe_float(pick(msg, "total_cost_usd", "totalCostUsd", "totalCostUSD"))

    raw_model_usage = pick(msg, "modelUsage", "model_usage")
    model_usage: dict[str, dict[str, Any]] = {}
    if isinstance(raw_model_usage, dict):
        for model_name, raw in raw_model_usage.items():
            if not isinstance(raw, dict):
                continue
            parsed = usage_from_mapping(raw)
            if parsed.get("cost_usd") is None:
                parsed["cost_usd"] = safe_float(pick(raw, "costUSD", "cost_usd", "cost"))
            parsed["total_tokens"] = billed_total_tokens(
                parsed.get("input_tokens"),
                parsed.get("output_tokens"),
                parsed.get("cache_read_tokens"),
                parsed.get("cache_write_tokens"),
            )
            model_usage[str(model_name)] = parsed

    usage["total_tokens"] = billed_total_tokens(
        usage.get("input_tokens"),
        usage.get("output_tokens"),
        usage.get("cache_read_tokens"),
        usage.get("cache_write_tokens"),
    )
    usage["model_usage"] = model_usage
    usage["num_turns"] = safe_int(pick(msg, "num_turns", "numTurns"))
    usage["duration_ms"] = safe_float(pick(msg, "duration_ms", "durationMs"))
    usage["duration_api_ms"] = safe_float(pick(msg, "duration_api_ms", "durationApiMs"))
    return usage


def classify_model_name(model_name: str, args: argparse.Namespace) -> str:
    name = model_name.lower()
    if args.qwen_model.lower() in name or name in {args.qwen_model.lower(), "qwen"}:
        return "qwen"
    if args.kimi_model.lower() in name or name in {args.kimi_model.lower(), "kimi"}:
        return "kimi"
    if "qwen" in name and "3.6" in name:
        return "qwen"
    if "kimi" in name and "k3" in name:
        return "kimi"
    return "other"


def clean_cross_task_state(config_dir: Path) -> None:
    projects_dir = config_dir / "projects"
    if projects_dir.exists():
        shutil.rmtree(projects_dir, ignore_errors=True)

    claude_json = config_dir / ".claude.json"
    payload: dict[str, Any] = {}
    if claude_json.exists():
        try:
            raw = json.loads(claude_json.read_text(encoding="utf-8-sig"))
            if isinstance(raw, dict):
                payload = raw
        except Exception:
            payload = {}
    payload["mcpServers"] = {}
    payload["mcpServersManaged"] = []
    payload["hasCompletedOnboarding"] = True
    payload["bypassPermissionsModeAccepted"] = True
    payload["hasTrustDialogAccepted"] = True
    claude_json.parent.mkdir(parents=True, exist_ok=True)
    claude_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def audit_benchmark_isolation(config_dir: Path) -> list[str]:
    problems: list[str] = []
    claude_json = config_dir / ".claude.json"
    if not claude_json.exists():
        problems.append(f"missing {claude_json}")
    else:
        try:
            payload = json.loads(claude_json.read_text(encoding="utf-8-sig"))
            servers = payload.get("mcpServers") if isinstance(payload, dict) else None
            if isinstance(servers, dict) and servers:
                problems.append("MCP is not empty: " + ", ".join(sorted(str(k) for k in servers)))
        except Exception as exc:
            problems.append(f"invalid .claude.json: {exc}")

    global_claude = config_dir / "CLAUDE.md"
    if global_claude.exists():
        text = global_claude.read_text(encoding="utf-8", errors="replace")
        forbidden = ["local-claw:secrets", "local-claw:cron-guard"]
        for item in forbidden:
            if item in text:
                problems.append(f"native extra instruction still present in config/CLAUDE.md: {item}")

    skills_dir = config_dir / "skills"
    if skills_dir.exists() and any(p.is_file() for p in skills_dir.rglob("*")):
        problems.append(f"skills directory is not empty: {skills_dir}")
    return problems


def http_json(url: str, *, method: str = "GET", body: Optional[dict[str, Any]] = None, headers: Optional[dict[str, str]] = None, timeout: float = 10.0) -> Any:
    import urllib.request
    data = None if body is None else json.dumps(body).encode("utf-8")
    req = urllib.request.Request(url, data=data, method=method)
    for key, value in (headers or {}).items():
        req.add_header(key, value)
    if data is not None:
        req.add_header("Content-Type", "application/json")
    # Local framework endpoints must never go through the external proxy.
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    with opener.open(req, timeout=timeout) as response:
        raw = response.read().decode("utf-8", errors="replace")
        if not raw.strip():
            return {}
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return raw.strip()


def server_health(server_url: str) -> tuple[bool, str]:
    try:
        data = http_json(server_url.rstrip("/") + "/api/health", timeout=3.0)
        return True, json.dumps(data, ensure_ascii=False)
    except Exception as exc:
        return False, str(exc)


def get_endpoints(server_url: str) -> list[dict[str, Any]]:
    data = http_json(server_url.rstrip("/") + "/api/endpoints", timeout=5.0)
    return data if isinstance(data, list) else []


def make_session_payload(prompt: str, workspace: Path, task_title: str, args: argparse.Namespace) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "title": task_title,
        "prompt": prompt,
        "cwd": str(workspace),
        "permissionMode": args.permission_mode,
    }
    if args.mode == "qwen":
        payload["model"] = args.qwen_model
        payload["endpointId"] = args.endpoint_id
    elif args.mode == "kimi":
        payload["model"] = args.kimi_model
        payload["endpointId"] = args.endpoint_id
    else:
        payload["smartHybrid"] = {
            "defaultModel": {"endpointId": args.endpoint_id, "model": args.qwen_model},
            "upgradeModel": {"endpointId": args.endpoint_id, "model": args.kimi_model},
        }
    return payload


def run_framework_turn(
    *,
    prompt: str,
    workspace: Path,
    task_title: str,
    timeout: float,
    transcript_dir: Path,
    turn_index: int,
    turn_id: str,
    args: argparse.Namespace,
    existing_ws: Any = None,
    existing_session_id: Optional[str] = None,
    new_session: bool = True,
) -> dict[str, Any]:
    try:
        import websocket  # type: ignore
    except ImportError as exc:
        raise RuntimeError("missing websocket-client; install with: python -m pip install websocket-client") from exc

    ws = existing_ws
    owns_ws = ws is None
    if ws is None:
        ws_url = args.server_url.replace("http://", "ws://").replace("https://", "wss://").rstrip("/") + "/ws"
        ws = websocket.create_connection(ws_url, timeout=1.0, http_proxy_host=None, http_proxy_port=None)

    raw_events_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.server-events.jsonl"
    sdk_messages_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.sdk-messages.jsonl"
    raw_events_file = raw_events_path.open("w", encoding="utf-8", newline="\n")
    sdk_file = sdk_messages_path.open("w", encoding="utf-8", newline="\n")

    monotonic_start = time.monotonic()
    deadline = monotonic_start + timeout
    session_id = existing_session_id
    status = "running"
    errors: list[str] = []
    output_chunks: list[str] = []
    normalized_events: list[dict[str, Any]] = []
    ttft: Optional[float] = None
    result_message: Optional[dict[str, Any]] = None
    routing_decisions: list[dict[str, Any]] = []
    escalation_events: list[dict[str, Any]] = []
    permission_requests = 0
    retry_count = 0
    tool_errors = 0
    assistant_messages = 0
    critical_ids: set[str] = set()
    model_call_counts: Counter[str] = Counter()

    # Timeout-safe usage telemetry. Claude Agent SDK may emit the same assistant
    # message multiple times as thinking/tool blocks are assembled. Deduplicate
    # by message.id and retain the latest usage snapshot for that message.
    assistant_usage_by_id: dict[str, dict[str, Any]] = {}
    assistant_model_by_id: dict[str, str] = {}

    try:
        if not new_session and session_id:
            send_event = {
                "type": "session.continue",
                "payload": {
                    "sessionId": session_id,
                    "prompt": prompt,
                    "permissionMode": args.permission_mode,
                },
            }
            if args.mode == "qwen":
                send_event["payload"]["model"] = args.qwen_model
                send_event["payload"]["endpointId"] = args.endpoint_id
            elif args.mode == "kimi":
                send_event["payload"]["model"] = args.kimi_model
                send_event["payload"]["endpointId"] = args.endpoint_id
            else:
                send_event["payload"]["smartHybrid"] = {
                    "defaultModel": {"endpointId": args.endpoint_id, "model": args.qwen_model},
                    "upgradeModel": {"endpointId": args.endpoint_id, "model": args.kimi_model},
                }
        else:
            send_event = {
                "type": "session.start",
                "payload": make_session_payload(prompt, workspace, task_title, args),
            }
        ws.send(json.dumps(send_event, ensure_ascii=False))

        while True:
            if time.monotonic() > deadline:
                status = "timeout"
                errors.append(f"turn timeout ({timeout:.0f}s)")
                if session_id:
                    try:
                        ws.send(json.dumps({"type": "session.stop", "payload": {"sessionId": session_id}}))
                    except Exception:
                        pass
                break
            try:
                ws.settimeout(min(1.0, max(0.1, deadline - time.monotonic())))
                raw = ws.recv()
            except websocket.WebSocketTimeoutException:
                continue
            except Exception as exc:
                status = "error"
                errors.append(f"websocket receive failed: {exc}")
                break
            if raw is None:
                status = "error"
                errors.append("websocket closed")
                break
            if isinstance(raw, bytes):
                raw = raw.decode("utf-8", errors="replace")
            try:
                event = json.loads(str(raw))
            except json.JSONDecodeError:
                continue
            if not isinstance(event, dict):
                continue
            raw_events_file.write(json.dumps(event, ensure_ascii=False) + "\n")
            raw_events_file.flush()
            etype = str(event.get("type") or "")
            payload = event.get("payload") if isinstance(event.get("payload"), dict) else {}

            if etype == "ping":
                continue

            candidate_sid = payload.get("sessionId")
            if not session_id and candidate_sid and etype == "session.status" and payload.get("status") == "running":
                session_id = str(candidate_sid)
            if session_id and candidate_sid and str(candidate_sid) != session_id:
                continue

            if etype == "permission.request":
                permission_requests += 1
                if args.auto_approve_permissions and session_id:
                    ws.send(json.dumps({
                        "type": "permission.response",
                        "payload": {
                            "sessionId": session_id,
                            "toolUseId": str(payload.get("toolUseId") or ""),
                            "result": {
                                "behavior": "allow",
                                "updatedInput": payload.get("input"),
                            },
                        },
                    }, ensure_ascii=False))
                else:
                    errors.append("permission request received but auto-approve is disabled")
                    status = "error"
                    break
                continue

            if etype == "routing.decision":
                decision = payload.get("decision")
                if isinstance(decision, dict):
                    routing_decisions.append(decision)
                continue

            if etype == "escalation.status":
                escalation_events.append(dict(payload))
                continue

            if etype == "session.retry":
                retry_count += 1
                continue

            if etype == "tasks.snapshot":
                tasks = payload.get("tasks")
                if isinstance(tasks, list):
                    for item in tasks:
                        if isinstance(item, dict) and item.get("critical") is True and item.get("id"):
                            critical_ids.add(str(item["id"]))
                continue

            if etype == "runner.error":
                errors.append(str(payload.get("message") or "runner.error"))
                status = "error"
                if not session_id or not candidate_sid or str(candidate_sid) == session_id:
                    break
                continue

            if etype == "stream.message":
                msg = payload.get("message")
                if not isinstance(msg, dict):
                    continue
                sdk_file.write(json.dumps(msg, ensure_ascii=False) + "\n")
                sdk_file.flush()
                normalized_events.extend(normalize_sdk_message(msg))

                text = extract_assistant_text(msg)
                if text:
                    if ttft is None:
                        ttft = time.monotonic() - monotonic_start
                    output_chunks.append(text)
                if msg.get("type") == "assistant":
                    assistant_messages += 1
                    inner = msg.get("message")
                    if isinstance(inner, dict):
                        message_id = str(inner.get("id") or "")
                        message_model = str(inner.get("model") or "")
                        if message_id and isinstance(inner.get("usage"), dict):
                            parsed_assistant_usage = usage_from_mapping(inner.get("usage"))
                            parsed_assistant_usage["total_tokens"] = billed_total_tokens(
                                parsed_assistant_usage.get("input_tokens"),
                                parsed_assistant_usage.get("output_tokens"),
                                parsed_assistant_usage.get("cache_read_tokens"),
                                parsed_assistant_usage.get("cache_write_tokens"),
                            )
                            assistant_usage_by_id[message_id] = parsed_assistant_usage
                            if message_model:
                                assistant_model_by_id[message_id] = message_model
                        blocks = inner.get("content")
                        if isinstance(blocks, list):
                            for block in blocks:
                                if isinstance(block, dict) and block.get("type") == "tool_use":
                                    pass
                if msg.get("type") == "user":
                    inner = msg.get("message")
                    if isinstance(inner, dict) and isinstance(inner.get("content"), list):
                        for block in inner["content"]:
                            if isinstance(block, dict) and block.get("type") == "tool_result" and block.get("is_error") is True:
                                tool_errors += 1
                if msg.get("type") == "stream_event":
                    sev = msg.get("event")
                    if isinstance(sev, dict) and sev.get("type") == "message_start":
                        message = sev.get("message")
                        if isinstance(message, dict) and message.get("model"):
                            model_call_counts[str(message["model"])] += 1
                if msg.get("type") == "result":
                    result_message = msg
                continue

            if etype == "session.status" and session_id:
                s = str(payload.get("status") or "")
                if s in {"completed", "error"}:
                    status = s
                    if s == "error" and payload.get("error"):
                        errors.append(str(payload.get("error")))
                    break

        elapsed = time.monotonic() - monotonic_start

        # Prefer the SDK's final result usage when present. On timeout/session
        # abort there is often no final result message, even though every
        # completed model call already emitted assistant.message.usage. In that
        # case, recover exact token telemetry from unique assistant message IDs.
        result_usage = extract_result_usage(result_message or {})
        result_usage_complete = (
            result_usage.get("input_tokens") is not None
            and result_usage.get("output_tokens") is not None
        )
        usage_source = "claude_agent_sdk_result_usage"
        if result_usage_complete:
            usage = result_usage
        elif assistant_usage_by_id:
            usage = {}
            fallback_model_usage: dict[str, dict[str, Any]] = {}
            for message_id, stats in assistant_usage_by_id.items():
                merge_usage(usage, stats)
                model_name = assistant_model_by_id.get(message_id)
                if model_name:
                    bucket = fallback_model_usage.setdefault(model_name, {})
                    merge_usage(bucket, stats)
            usage["total_tokens"] = billed_total_tokens(
                usage.get("input_tokens"),
                usage.get("output_tokens"),
                usage.get("cache_read_tokens"),
                usage.get("cache_write_tokens"),
            )
            for stats in fallback_model_usage.values():
                stats["total_tokens"] = billed_total_tokens(
                    stats.get("input_tokens"),
                    stats.get("output_tokens"),
                    stats.get("cache_read_tokens"),
                    stats.get("cache_write_tokens"),
                )
            usage["model_usage"] = fallback_model_usage
            usage["num_turns"] = len(assistant_usage_by_id)
            usage["duration_ms"] = result_usage.get("duration_ms")
            usage["duration_api_ms"] = result_usage.get("duration_api_ms")
            usage_source = "claude_agent_sdk_assistant_usage_dedup"
        else:
            usage = result_usage

        result_subtype = str((result_message or {}).get("subtype") or "")
        result_is_error = bool((result_message or {}).get("is_error") is True)
        success = status == "completed" and not result_is_error and (not result_subtype or result_subtype == "success")
        if status == "completed" and result_message is None:
            errors.append("session completed without SDK result message; token usage unavailable")

        return {
            "ws": ws,
            "owns_ws": owns_ws,
            "session_id": session_id,
            "success": success,
            "status": "success" if success else status,
            "returncode": 0 if success else None,
            "elapsed": elapsed,
            "ttft": ttft,
            "output": "\n".join(x for x in output_chunks if x).strip(),
            "error": " | ".join(dict.fromkeys(x for x in errors if x)),
            "normalized_events": normalized_events,
            "raw_events_path": str(raw_events_path),
            "sdk_messages_path": str(sdk_messages_path),
            "result_message": result_message,
            "usage": usage,
            "usage_source": usage_source,
            "usage_complete": usage.get("input_tokens") is not None and usage.get("output_tokens") is not None,
            "routing_decisions": routing_decisions,
            "escalation_events": escalation_events,
            "permission_requests": permission_requests,
            "retry_count": retry_count,
            "tool_errors": tool_errors,
            "assistant_messages": assistant_messages,
            "critical_task_ids": sorted(critical_ids),
            "model_call_counts": dict(model_call_counts),
        }
    except Exception as exc:
        # LOCALCLAW_PER_TASK_CRASH_GUARD_20260819
        # Convert an unexpected SDK/WebSocket/tool-processing exception into a
        # task-level error instead of terminating the entire PinchBench batch.
        import traceback
        elapsed = time.monotonic() - monotonic_start
        crash = f"run_framework_turn exception: {type(exc).__name__}: {exc}"
        try:
            sdk_file.write(json.dumps({"type": "runner_exception", "error": crash, "traceback": traceback.format_exc()}, ensure_ascii=False) + "\n")
            sdk_file.flush()
        except Exception:
            pass
        return {
            "ws": ws,
            "owns_ws": owns_ws,
            "session_id": session_id,
            "success": False,
            "status": "runner_exception",
            "returncode": None,
            "elapsed": elapsed,
            "ttft": ttft,
            "output": "\n".join(x for x in output_chunks if x).strip(),
            "error": crash,
            "normalized_events": normalized_events,
            "raw_events_path": str(raw_events_path),
            "sdk_messages_path": str(sdk_messages_path),
            "result_message": result_message,
            "usage": {},
            "usage_source": "runner_exception",
            "usage_complete": False,
            "routing_decisions": routing_decisions,
            "escalation_events": escalation_events,
            "permission_requests": permission_requests,
            "retry_count": retry_count,
            "tool_errors": tool_errors,
            "assistant_messages": assistant_messages,
            "critical_task_ids": sorted(critical_ids),
            "model_call_counts": dict(model_call_counts),
        }
    finally:
        raw_events_file.close()
        sdk_file.close()


def close_framework_session(ws: Any, session_id: Optional[str]) -> None:
    if not ws or not session_id:
        return
    try:
        ws.send(json.dumps({"type": "session.delete", "payload": {"sessionId": session_id}}))
        deadline = time.monotonic() + 3.0
        while time.monotonic() < deadline:
            try:
                ws.settimeout(0.5)
                raw = ws.recv()
            except Exception:
                break
            if isinstance(raw, bytes):
                raw = raw.decode("utf-8", errors="replace")
            try:
                event = json.loads(str(raw))
            except Exception:
                continue
            if isinstance(event, dict) and event.get("type") == "session.deleted":
                payload = event.get("payload")
                if isinstance(payload, dict) and str(payload.get("sessionId") or "") == session_id:
                    break
    except Exception:
        pass



INFRA_FAILURE_OPENROUTER_504 = "openrouter_upstream_504"
INFRA_FAILURE_WEBSEARCH_429 = "websearch_429"
INFRA_FAILURE_PROVIDER_429 = "openrouter_provider_429"  # OPENROUTER_PROVIDER_429_RETRY_20260820


def inspect_sdk_infrastructure_failures(transcript_dir: Path) -> dict[str, Any]:
    """Inspect frozen SDK JSONL for narrowly defined external infrastructure failures.

    Only narrowly defined external infrastructure failures are eligible for automatic task-level retry:
    - OpenRouter/provider upstream idle timeout with HTTP/code 504.
    - WebSearch tool_result explicitly failing with HTTP status 429.
    - The selected model provider itself returning HTTP 429 before a usable model response.

    Normal PinchBench hard timeouts, WebFetch slowness, low scores, tool loops,
    and ordinary tool errors are deliberately NOT retryable here.
    """
    upstream_504 = False
    websearch_429 = False
    provider_429 = False
    evidence: list[str] = []
    tool_names: dict[str, str] = {}

    for sdk_path in sorted(transcript_dir.glob("*.sdk-messages.jsonl")):
        try:
            lines = sdk_path.read_text(encoding="utf-8", errors="replace").splitlines()
        except Exception:
            continue
        for raw in lines:
            if "Upstream idle timeout exceeded" in raw and (
                '"code": 504' in raw or '"code":504' in raw or '"error_type": "timeout"' in raw
            ):
                upstream_504 = True
                if len(evidence) < 10:
                    evidence.append(f"{sdk_path.name}: OpenRouter upstream idle timeout / 504")
            try:
                msg = json.loads(raw)
            except Exception:
                continue
            if not isinstance(msg, dict):
                continue
            if msg.get("type") == "assistant":
                # OPENROUTER_PROVIDER_429_RETRY_20260820:
                # Provider-side model failures arrive as synthetic assistant messages
                # with a top-level error object (for example Alibaba shared-pool 429).
                top_error = msg.get("error")
                if isinstance(top_error, dict):
                    status = top_error.get("status")
                    nested = top_error.get("error")
                    nested_message = ""
                    provider_name = ""
                    provider_error_code = ""
                    limit_source = ""
                    if isinstance(nested, dict):
                        nested_message = str(nested.get("message") or "")
                        metadata = nested.get("metadata")
                        if isinstance(metadata, dict):
                            provider_name = str(metadata.get("provider_name") or "")
                            provider_error_code = str(metadata.get("provider_error_code") or "")
                            limit_source = str(metadata.get("limit_source") or "")
                    if (
                        str(status) == "429"
                        or top_error.get("code") == 429
                        or "429" in nested_message
                    ) and (
                        provider_name
                        or provider_error_code
                        or "Provider returned error" in nested_message
                        or "rate-limit" in nested_message.lower()
                        or "rate limit" in nested_message.lower()
                    ):
                        provider_429 = True
                        if len(evidence) < 10:
                            detail = f" provider={provider_name}" if provider_name else ""
                            if provider_error_code:
                                detail += f" code={provider_error_code}"
                            if limit_source:
                                detail += f" source={limit_source}"
                            evidence.append(f"{sdk_path.name}: OpenRouter/model-provider HTTP 429{detail}")
                inner = msg.get("message")
                if isinstance(inner, dict):
                    blocks = inner.get("content")
                    if isinstance(blocks, list):
                        for block in blocks:
                            if isinstance(block, dict) and block.get("type") == "tool_use" and block.get("id"):
                                tool_names[str(block["id"])] = str(block.get("name") or "")
            elif msg.get("type") == "user":
                inner = msg.get("message")
                blocks = inner.get("content") if isinstance(inner, dict) else None
                if isinstance(blocks, list):
                    for block in blocks:
                        if not isinstance(block, dict) or block.get("type") != "tool_result":
                            continue
                        tool_id = str(block.get("tool_use_id") or "")
                        if tool_names.get(tool_id) != "WebSearch" or block.get("is_error") is not True:
                            continue
                        content = block.get("content")
                        text = content if isinstance(content, str) else json.dumps(content, ensure_ascii=False, default=str)
                        if re.search(r"(?:status\s+code\s+429|HTTP\s*429|Too\s+Many\s+Requests)", text, re.IGNORECASE):
                            websearch_429 = True
                            if len(evidence) < 10:
                                evidence.append(f"{sdk_path.name}: WebSearch HTTP 429")

    reasons: list[str] = []
    if upstream_504:
        reasons.append(INFRA_FAILURE_OPENROUTER_504)
    if websearch_429:
        reasons.append(INFRA_FAILURE_WEBSEARCH_429)
    if provider_429:
        reasons.append(INFRA_FAILURE_PROVIDER_429)
    return {
        "retryable": bool(reasons),
        "reasons": reasons,
        "evidence": evidence,
    }


def archive_infra_attempt(
    *,
    results_dir: Path,
    task_id: str,
    attempt_number: int,
    workspace: Path,
    transcript_dir: Path,
    execution: dict[str, Any],
    diagnosis: dict[str, Any],
) -> dict[str, Any]:
    root = results_dir / "infra_attempts" / task_id / f"attempt_{attempt_number:02d}"
    root.mkdir(parents=True, exist_ok=True)
    archived_transcript = root / "transcript"
    archived_workspace = root / "workspace"
    if transcript_dir.exists():
        if archived_transcript.exists():
            shutil.rmtree(archived_transcript, ignore_errors=True)
        shutil.move(str(transcript_dir), str(archived_transcript))
    if workspace.exists():
        if archived_workspace.exists():
            shutil.rmtree(archived_workspace, ignore_errors=True)
        shutil.move(str(workspace), str(archived_workspace))
    metadata = {
        "task_id": task_id,
        "attempt_number": attempt_number,
        "diagnosis": diagnosis,
        "status": execution.get("status"),
        "success": execution.get("success"),
        "elapsed": execution.get("elapsed"),
        "total_tokens": execution.get("total_tokens"),
        "qwen_calls": execution.get("qwen_calls"),
        "kimi_calls": execution.get("kimi_calls"),
        "error": execution.get("error"),
        "transcript": str(archived_transcript),
        "workspace": str(archived_workspace),
    }
    (root / "attempt.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    return metadata


def execute_task(
    task: Task,
    task_index: int,
    task_count: int,
    skill_dir: Path,
    workspace: Path,
    transcript_dir: Path,
    args: argparse.Namespace,
) -> dict[str, Any]:
    print(f"[{task_index}/{task_count}] {task.task_id} — {task.name}")
    prompt_preview = re.sub(r"\s+", " ", task.prompt)[:150]
    print(f"  prompt: {prompt_preview}{'...' if len(task.prompt) > 150 else ''}")

    prerequisites = task.metadata.get("prerequisites") or []
    missing_prerequisites = check_prerequisites(list(prerequisites)) if prerequisites else []
    if missing_prerequisites:
        error = f"缺少依赖: {', '.join(missing_prerequisites)}"
        print(f"  跳过: {error}\n")
        return base_empty_execution(task, args, "missing_prerequisite", error, workspace, transcript_dir)

    if workspace.exists():
        shutil.rmtree(workspace)
    workspace.mkdir(parents=True, exist_ok=True)
    staged, missing = stage_workspace_files(task, workspace, skill_dir)
    if missing:
        error = " | ".join(missing)
        print(f"  fixture 错误: {error}\n")
        return base_empty_execution(task, args, "missing_fixture", error, workspace, transcript_dir)

    transcript_dir.mkdir(parents=True, exist_ok=True)
    (transcript_dir / "staged_files.json").write_text(json.dumps(staged, ensure_ascii=False, indent=2), encoding="utf-8")

    clean_cross_task_state(Path(args.config_dir))
    isolation = audit_benchmark_isolation(Path(args.config_dir))
    if isolation:
        error = "benchmark isolation failed: " + " | ".join(isolation)
        print(f"  隔离检查失败: {error}\n")
        return base_empty_execution(task, args, "isolation_error", error, workspace, transcript_dir)

    total_timeout = args.network_timeout if is_network_task(task) else max(1.0, task.timeout_seconds * args.timeout_multiplier)
    task_deadline = time.monotonic() + total_timeout
    turns = build_turns(task)

    normalized_transcript: list[dict[str, Any]] = []
    turn_results: list[dict[str, Any]] = []
    outputs: list[str] = []
    errors: list[str] = []
    current_session_id: Optional[str] = None
    ws = None
    task_success = True
    task_status = "success"
    total_elapsed = 0.0
    task_ttft: Optional[float] = None
    aggregate_usage: dict[str, Any] = {}
    aggregate_model_usage: dict[str, dict[str, Any]] = {}
    aggregate_calls: Counter[str] = Counter()
    escalation_events: list[dict[str, Any]] = []
    routing_decisions: list[dict[str, Any]] = []
    permission_requests = 0
    retry_count = 0
    tool_errors = 0
    step_count = 0
    critical_ids: set[str] = set()
    usage_complete = True
    token_sources: set[str] = set()

    try:
        for turn_index, turn in enumerate(turns, start=1):
            remaining = task_deadline - time.monotonic()
            if remaining <= 0:
                task_success = False
                task_status = "timeout"
                errors.append(f"任务总超时 ({total_timeout:.0f}s)")
                break

            turn_id = str(turn["id"])
            new_session = bool(turn.get("new_session"))
            turn_prompt = str(turn["prompt"])
            normalized_transcript.append(make_user_transcript_event(turn_prompt, turn_index, turn_id))

            prompt = turn_prompt
            if not args.no_workspace_instruction:
                prompt += "\n\nIMPORTANT: You are running in an isolated workspace. Read, write, and edit files only in the current working directory."
            prompt_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.prompt.txt"
            prompt_path.write_text(prompt + ("" if prompt.endswith("\n") else "\n"), encoding="utf-8", newline="\n")

            if new_session and current_session_id and ws:
                close_framework_session(ws, current_session_id)
                current_session_id = None

            turn_result = run_framework_turn(
                prompt=prompt,
                workspace=workspace,
                task_title=f"PB {task.task_id} turn {turn_index}",
                timeout=max(1.0, remaining),
                transcript_dir=transcript_dir,
                turn_index=turn_index,
                turn_id=turn_id,
                args=args,
                existing_ws=ws,
                existing_session_id=current_session_id,
                new_session=new_session or not current_session_id,
            )
            ws = turn_result["ws"]
            current_session_id = turn_result.get("session_id") or current_session_id
            normalized_transcript.extend(turn_result["normalized_events"])
            if turn_result.get("output"):
                outputs.append(f"## Turn {turn_index} ({turn_id})\n{turn_result['output']}")
            if turn_result.get("error"):
                errors.append(f"turn {turn_index} ({turn_id}): {turn_result['error']}")

            total_elapsed += float(turn_result.get("elapsed") or 0.0)
            if task_ttft is None and turn_result.get("ttft") is not None:
                task_ttft = float(turn_result["ttft"])
            usage = turn_result.get("usage") or {}
            merge_usage(aggregate_usage, usage)
            usage_complete = usage_complete and bool(turn_result.get("usage_complete"))
            if turn_result.get("usage_source"):
                token_sources.add(str(turn_result.get("usage_source")))
            for model_name, model_stats in (usage.get("model_usage") or {}).items():
                bucket = aggregate_model_usage.setdefault(model_name, {})
                merge_usage(bucket, model_stats)
            aggregate_calls.update(turn_result.get("model_call_counts") or {})
            escalation_events.extend(turn_result.get("escalation_events") or [])
            routing_decisions.extend(turn_result.get("routing_decisions") or [])
            permission_requests += int(turn_result.get("permission_requests") or 0)
            retry_count += int(turn_result.get("retry_count") or 0)
            tool_errors += int(turn_result.get("tool_errors") or 0)
            step_count += int((usage.get("num_turns") or turn_result.get("assistant_messages") or 0))
            critical_ids.update(turn_result.get("critical_task_ids") or [])

            turn_results.append({
                "turn": turn_index,
                "turn_id": turn_id,
                "new_session": new_session,
                "session_id": current_session_id,
                "success": turn_result.get("success"),
                "status": turn_result.get("status"),
                "elapsed": turn_result.get("elapsed"),
                "ttft": turn_result.get("ttft"),
                "usage": usage,
                "usage_source": turn_result.get("usage_source"),
                "model_call_counts": turn_result.get("model_call_counts"),
                "routing_decisions": turn_result.get("routing_decisions"),
                "escalation_events": turn_result.get("escalation_events"),
                "permission_requests": turn_result.get("permission_requests"),
                "retry_count": turn_result.get("retry_count"),
                "tool_errors": turn_result.get("tool_errors"),
                "critical_task_ids": turn_result.get("critical_task_ids"),
                "raw_server_events": turn_result.get("raw_events_path"),
                "raw_sdk_messages": turn_result.get("sdk_messages_path"),
                "prompt_path": str(prompt_path),
                "error": turn_result.get("error"),
            })

            if not turn_result.get("success"):
                task_success = False
                task_status = str(turn_result.get("status") or "error")
                break
    except Exception as exc:
        # LOCALCLAW_PER_TASK_CRASH_GUARD_20260819
        import traceback
        task_success = False
        task_status = "runner_exception"
        errors.append(f"execute_task exception: {type(exc).__name__}: {exc}")
        try:
            (transcript_dir / "runner_exception.txt").write_text(traceback.format_exc(), encoding="utf-8")
        except Exception:
            pass
    finally:
        if ws:
            close_framework_session(ws, current_session_id)
            try:
                ws.close()
            except Exception:
                pass
        clean_cross_task_state(Path(args.config_dir))

    normalized_path = transcript_dir / "normalized.jsonl"
    with normalized_path.open("w", encoding="utf-8") as file:
        for event in normalized_transcript:
            file.write(json.dumps(event, ensure_ascii=False) + "\n")
    (transcript_dir / "turn_results.json").write_text(json.dumps(turn_results, ensure_ascii=False, indent=2), encoding="utf-8")

    aggregate_usage["total_tokens"] = billed_total_tokens(
        aggregate_usage.get("input_tokens"), aggregate_usage.get("output_tokens"),
        aggregate_usage.get("cache_read_tokens"), aggregate_usage.get("cache_write_tokens"),
    )
    for stats in aggregate_model_usage.values():
        stats["total_tokens"] = billed_total_tokens(stats.get("input_tokens"), stats.get("output_tokens"), stats.get("cache_read_tokens"), stats.get("cache_write_tokens"))

    qwen_usage: dict[str, Any] = {}
    kimi_usage: dict[str, Any] = {}
    qwen_calls = 0
    kimi_calls = 0
    for model_name, stats in aggregate_model_usage.items():
        kind = classify_model_name(model_name, args)
        if kind == "qwen":
            merge_usage(qwen_usage, stats)
        elif kind == "kimi":
            merge_usage(kimi_usage, stats)
    for model_name, count in aggregate_calls.items():
        kind = classify_model_name(model_name, args)
        if kind == "qwen": qwen_calls += int(count)
        elif kind == "kimi": kimi_calls += int(count)
    qwen_usage["total_tokens"] = billed_total_tokens(qwen_usage.get("input_tokens"), qwen_usage.get("output_tokens"), qwen_usage.get("cache_read_tokens"), qwen_usage.get("cache_write_tokens"))
    kimi_usage["total_tokens"] = billed_total_tokens(kimi_usage.get("input_tokens"), kimi_usage.get("output_tokens"), kimi_usage.get("cache_read_tokens"), kimi_usage.get("cache_write_tokens"))

    active_escalations = [e for e in escalation_events if e.get("active") is True]
    deescalations = [e for e in escalation_events if e.get("active") is False]
    escalated_to_kimi = any(classify_model_name(str(e.get("model") or ""), args) == "kimi" for e in active_escalations)

    return {
        "task_id": task.task_id,
        "name": task.name,
        "category": task.category,
        "grading_type": task.grading_type,
        "mode": args.mode,
        "model": args.qwen_model if args.mode == "qwen" else args.kimi_model if args.mode == "kimi" else "smart-hybrid",
        "default_model": args.qwen_model if args.mode == "hybrid" else None,
        "upgrade_model": args.kimi_model if args.mode == "hybrid" else None,
        "endpoint_id": args.endpoint_id,
        "network_task": is_network_task(task),
        "multi_session": task.multi_session,
        "session_count": len(turn_results),
        "success": task_success,
        "status": task_status,
        "returncode": 0 if task_success else None,
        "elapsed": total_elapsed,
        "ttft": task_ttft,
        "input_tokens": aggregate_usage.get("input_tokens"),
        "output_tokens": aggregate_usage.get("output_tokens"),
        "reasoning_tokens": aggregate_usage.get("reasoning_tokens"),
        "cache_read_tokens": aggregate_usage.get("cache_read_tokens"),
        "cache_write_tokens": aggregate_usage.get("cache_write_tokens"),
        "total_tokens": aggregate_usage.get("total_tokens"),
        "cost_usd": aggregate_usage.get("cost_usd"),
        "token_source": (
            next(iter(token_sources))
            if len(token_sources) == 1
            else "+".join(sorted(token_sources))
            if token_sources
            else "claude_agent_sdk_result_usage"
        ),
        "token_coverage_complete": usage_complete and bool(turn_results),
        "token_verified_against_openrouter": False,
        "usage_complete": usage_complete and bool(turn_results),
        "step_count": step_count,
        "tool_errors": tool_errors,
        "permission_requests": permission_requests,
        "retry_count": retry_count,
        "escalation_count": len(active_escalations),
        "deescalation_count": len(deescalations),
        "escalated_to_kimi": escalated_to_kimi,
        "critical_task_count": len(critical_ids),
        "qwen_calls": qwen_calls,
        "kimi_calls": kimi_calls,
        "qwen_input_tokens": qwen_usage.get("input_tokens"),
        "qwen_output_tokens": qwen_usage.get("output_tokens"),
        "qwen_reasoning_tokens": qwen_usage.get("reasoning_tokens"),
        "qwen_cache_read_tokens": qwen_usage.get("cache_read_tokens"),
        "qwen_cache_write_tokens": qwen_usage.get("cache_write_tokens"),
        "qwen_total_tokens": qwen_usage.get("total_tokens"),
        "qwen_cost_usd": qwen_usage.get("cost_usd"),
        "kimi_input_tokens": kimi_usage.get("input_tokens"),
        "kimi_output_tokens": kimi_usage.get("output_tokens"),
        "kimi_reasoning_tokens": kimi_usage.get("reasoning_tokens"),
        "kimi_cache_read_tokens": kimi_usage.get("cache_read_tokens"),
        "kimi_cache_write_tokens": kimi_usage.get("cache_write_tokens"),
        "kimi_total_tokens": kimi_usage.get("total_tokens"),
        "kimi_cost_usd": kimi_usage.get("cost_usd"),
        "model_usage_json": json.dumps(aggregate_model_usage, ensure_ascii=False),
        "model_call_counts_json": json.dumps(dict(aggregate_calls), ensure_ascii=False),
        "routing_decisions_json": json.dumps(routing_decisions, ensure_ascii=False),
        "escalation_events_json": json.dumps(escalation_events, ensure_ascii=False),
        "critical_task_ids_json": json.dumps(sorted(critical_ids), ensure_ascii=False),
        "output": "\n\n".join(outputs),
        "error": " | ".join(errors),
        "stderr": "",
        "score": None,
        "breakdown": {},
        "grade_notes": "",
        "grade_error": None,
        "workspace": str(workspace),
        "transcript": str(transcript_dir),
        "normalized_transcript_path": str(normalized_path),
        "transcript_data": normalized_transcript,
        "turn_results": turn_results,
    }


def base_empty_execution(task: Task, args: argparse.Namespace, status: str, error: str, workspace: Path, transcript_dir: Path) -> dict[str, Any]:
    return {
        "task_id": task.task_id, "name": task.name, "category": task.category, "grading_type": task.grading_type,
        "mode": args.mode, "model": args.qwen_model if args.mode == "qwen" else args.kimi_model if args.mode == "kimi" else "smart-hybrid",
        "default_model": args.qwen_model if args.mode == "hybrid" else None, "upgrade_model": args.kimi_model if args.mode == "hybrid" else None,
        "endpoint_id": args.endpoint_id, "network_task": is_network_task(task), "multi_session": task.multi_session, "session_count": 0,
        "success": False, "status": status, "returncode": None, "elapsed": 0.0, "ttft": None,
        "input_tokens": None, "output_tokens": None, "reasoning_tokens": None, "cache_read_tokens": None, "cache_write_tokens": None,
        "total_tokens": None, "cost_usd": None, "token_source": "claude_agent_sdk_result_usage", "token_coverage_complete": False,
        "token_verified_against_openrouter": False, "usage_complete": False, "step_count": 0, "tool_errors": 0, "permission_requests": 0,
        "retry_count": 0, "escalation_count": 0, "deescalation_count": 0, "escalated_to_kimi": False, "critical_task_count": 0,
        "qwen_calls": 0, "kimi_calls": 0, "qwen_total_tokens": None, "kimi_total_tokens": None, "qwen_cost_usd": None, "kimi_cost_usd": None,
        "model_usage_json": "{}", "model_call_counts_json": "{}", "routing_decisions_json": "[]", "escalation_events_json": "[]",
        "output": "", "error": error, "stderr": "", "score": None, "breakdown": {}, "grade_notes": "", "grade_error": None,
        "workspace": str(workspace), "transcript": str(transcript_dir), "normalized_transcript_path": "", "transcript_data": [], "turn_results": [],
    }


def write_run_config(output_path: Path, args: argparse.Namespace, skill_dir: Path, tasks_dir: Path, selected: list[Task], judge_model: str) -> None:
    manifest_path = tasks_dir / "manifest.yaml"
    source_pkg = Path(args.source_dir) / "package.json"
    source_version = None
    if source_pkg.exists():
        try:
            source_version = json.loads(source_pkg.read_text(encoding="utf-8" )).get("version")
        except Exception:
            pass
    commit = command_output([resolve_command("git") or "git", "rev-parse", "HEAD"], cwd=skill_dir) if resolve_command("git") else ""
    config = {
        "runner_revision": OUR_FRAMEWORK_RUNNER_REVISION,
        "created_at": dt.datetime.now(dt.timezone.utc).astimezone().isoformat(),
        "platform": platform.platform(),
        "python": sys.version,
        "framework_alias": "我们的框架",
        "framework_source_dir": str(Path(args.source_dir)),
        "framework_version": source_version,
        "framework_package_json_sha256": sha256_file(source_pkg) if source_pkg.exists() else None,
        "server_url": args.server_url,
        "endpoint_id": args.endpoint_id,
        "mode": args.mode,
        "model": args.qwen_model if args.mode == "qwen" else args.kimi_model if args.mode == "kimi" else "smart-hybrid",
        "qwen_model": args.qwen_model,
        "kimi_model": args.kimi_model,
        "smart_hybrid": args.mode == "hybrid",
        "smart_hybrid_default_model": args.qwen_model if args.mode == "hybrid" else None,
        "smart_hybrid_upgrade_model": args.kimi_model if args.mode == "hybrid" else None,
        "local_gateway": True,
        "openrouter_routing": "default",
        "worker_count": 1,
        "task_concurrency": 1,
        "permission_mode": args.permission_mode,
        "permission_adapter": "immediate allow response to framework permission.request" if args.auto_approve_permissions else "none",
        "mcp_policy": "OFF: isolated .claude.json mcpServers must be empty",
        "skills_policy": "OFF: isolated config skills dir must be empty",
        "memory_policy": "cross-task OFF: config/projects deleted before and after every task",
        "native_language_constraint": "ON (framework-native harness behavior)",
        "smart_hybrid_claude_md_injection": "ON only in hybrid (framework-native treatment)",
        "custom_task_prompt": False,
        "workspace_instruction": not args.no_workspace_instruction,
        "skill_dir": str(skill_dir),
        "tasks_dir": str(tasks_dir),
        "pinchbench_commit": commit,
        "manifest_sha256": sha256_file(manifest_path) if manifest_path.exists() else None,
        "suite": args.suite,
        "limit": args.limit,
        "default_skipped_tasks": sorted(DEFAULT_SKIPPED_TASKS),
        "additional_skip": args.skip,
        "skip_network": args.skip_network,
        "task_count": len(selected),
        "task_ids": [t.task_id for t in selected],
        "timeout_multiplier": args.timeout_multiplier,
        "network_timeout": args.network_timeout,
        "infra_retry_policy": {
            "max_retries_per_task": args.infra_retries,
            "retryable_only": [INFRA_FAILURE_OPENROUTER_504, INFRA_FAILURE_WEBSEARCH_429, INFRA_FAILURE_PROVIDER_429],
            "normal_deadline_timeout_retry": False,
            "low_score_retry": False,
            "webfetch_slow_retry": False,
            "fresh_workspace_and_session": True,
        },
        "grading_enabled": not args.no_grade,
        "grading_engine": str(skill_dir / "scripts" / "lib_grading.py"),
        "grader_compat_revision": GRADER_COMPAT_REVISION,
        "judge_transport_policy": "anthropic_provider_pinned; plain_json_text; semantic-empty responses retry/fail; failed calls never parsed",
        "judge_max_completion_tokens": 8192,
        "judge_backend": "api",
        "judge_model": judge_model,
        "judge_key_env": "OPENROUTER_API_KEY",
        "judge_key_present": bool(os.environ.get("OPENROUTER_API_KEY")),
        "judge_is_separate_from_tested_agent": True,
        "token_source": "Claude Agent SDK result.usage/modelUsage; timeout fallback uses deduplicated assistant.message.usage by message.id",
        "token_total_definition": "input + output + cache_read + cache_creation; reasoning is subset of output and is not double-counted",
        "token_verified_against_openrouter": False,
    }
    output_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")


def print_preflight(selected: list[Task], skill_dir: Path, tasks_dir: Path, grader_module: Optional[Any], grader_error: Optional[str], args: argparse.Namespace) -> tuple[dict[str, list[str]], list[str]]:
    prerequisite_failures: dict[str, list[str]] = {}
    fixture_failures: list[str] = []
    needs_judge = not args.no_grade and any(t.grading_type in {"llm_judge", "hybrid"} for t in selected)
    health_ok, health_detail = server_health(args.server_url)
    isolation = audit_benchmark_isolation(Path(args.config_dir))
    endpoint_ok = False
    endpoint_detail = ""
    if health_ok:
        try:
            endpoints = get_endpoints(args.server_url)
            target = next((e for e in endpoints if e.get("id") == args.endpoint_id), None)
            if target:
                ids = {str(m.get("id")) for m in target.get("models", []) if isinstance(m, dict)}
                endpoint_ok = args.qwen_model in ids and args.kimi_model in ids and target.get("enabled") is True
                endpoint_detail = f"models={sorted(ids)} enabled={target.get('enabled')} apiType={target.get('apiType')}"
            else:
                endpoint_detail = f"endpoint {args.endpoint_id!r} not found"
        except Exception as exc:
            endpoint_detail = str(exc)

    print("=" * 100)
    print("PinchBench 我们的框架 Windows runner preflight")
    print("=" * 100)
    print(f"Runner revision      : {OUR_FRAMEWORK_RUNNER_REVISION}")
    print(f"Mode                 : {args.mode}")
    print(f"Qwen                 : {args.qwen_model}")
    print(f"Kimi                 : {args.kimi_model}")
    print(f"Endpoint             : {args.endpoint_id}")
    print(f"Server               : {'ok' if health_ok else 'FAILED'} — {health_detail}")
    print(f"Endpoint audit       : {'ok' if endpoint_ok else 'FAILED'} — {endpoint_detail}")
    print(f"Isolation            : {'ok' if not isolation else 'FAILED'}")
    for item in isolation:
        print(f"  - {item}")
    print(f"Permission mode      : {args.permission_mode}")
    print(f"Permission adapter   : {'auto-allow immediate' if args.auto_approve_permissions else 'disabled'}")
    print("Worker/concurrency   : 1 / 1")
    print("Gateway              : ON")
    print("Extra MCP/Skills     : OFF / OFF")
    print("Cross-task memory    : reset per task")
    print(f"Infra retries        : {args.infra_retries} (ONLY OpenRouter 504 / WebSearch 429 / model-provider 429; fresh task attempt)")
    print(f"Judge                : {args.judge_model}")
    print(f"OPENROUTER_API_KEY   : {'set' if os.environ.get('OPENROUTER_API_KEY') else 'missing'}")
    print(f"Grader import        : {'ok' if grader_error is None else 'FAILED'}")
    if grader_error: print(f"  {grader_error}")
    print(f"Selected tasks       : {len(selected)}")

    for task in selected:
        prereqs = task.metadata.get("prerequisites") or []
        missing = check_prerequisites(list(prereqs)) if prereqs else []
        if missing: prerequisite_failures[task.task_id] = missing
        for wf in task.workspace_files:
            if not isinstance(wf, dict):
                fixture_failures.append(f"{task.task_id}: invalid workspace file entry"); continue
            if wf.get("content") is not None: continue
            source = wf.get("source")
            if not source: continue
            src_rel = Path(str(source))
            candidates = [skill_dir / "assets" / src_rel, skill_dir / src_rel, task.file_path.parent / src_rel, task.file_path.parent.parent / "assets" / src_rel]
            if not any(c.exists() for c in candidates): fixture_failures.append(f"{task.task_id}: missing fixture {source}")

    print(f"Missing prerequisites: {sum(len(v) > 0 for v in prerequisite_failures.values())} task(s)")
    print(f"Missing fixtures      : {len(fixture_failures)}")
    print("=" * 100)
    args._health_ok = health_ok
    args._endpoint_ok = endpoint_ok
    args._isolation_ok = not isolation
    args._needs_judge = needs_judge
    return prerequisite_failures, fixture_failures


def save_csv(results: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "task_id","name","category","grading_type","mode","model","default_model","upgrade_model","success","status","score",
        "agent_elapsed","grading_elapsed","end_to_end_elapsed","ttft","input_tokens","output_tokens","reasoning_tokens","cache_read_tokens","cache_write_tokens","total_tokens","cost_usd","usage_complete",
        "qwen_calls","kimi_calls","qwen_total_tokens","kimi_total_tokens","qwen_cost_usd","kimi_cost_usd","escalation_count","deescalation_count","escalated_to_kimi","critical_task_count",
        "permission_requests","retry_count","infra_retry_count","infra_retry_reasons_json","infra_discarded_elapsed","infra_discarded_total_tokens","step_count","tool_errors","workspace","transcript","error","grade_error","grade_notes","model_usage_json","model_call_counts_json","escalation_events_json",
    ]
    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields); writer.writeheader()
        for row in results: writer.writerow({k: truncate_excel(row.get(k), 10000) for k in fields})


def save_xlsx(results: list[dict[str, Any]], summary: dict[str, Any], output_path: Path) -> None:
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
    except ImportError:
        LOGGER.warning("未安装 openpyxl，跳过 XLSX")
        return
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "详细结果"
    columns = [
        ("任务ID","task_id"),("名称","name"),("类别","category"),("打分类型","grading_type"),("模式","mode"),("模型","model"),("基础模型","default_model"),("升级模型","upgrade_model"),
        ("成功","success"),("状态","status"),("分数","score"),("Agent耗时(s)","agent_elapsed"),("Grading耗时(s)","grading_elapsed"),("端到端耗时(s)","end_to_end_elapsed"),("TTFT(s)","ttft"),
        ("输入Token","input_tokens"),("输出Token","output_tokens"),("推理Token(输出子集)","reasoning_tokens"),("缓存读取Token","cache_read_tokens"),("缓存创建Token","cache_write_tokens"),("总Token(不重复推理)","total_tokens"),("Agent费用USD","cost_usd"),
        ("Qwen调用","qwen_calls"),("Kimi调用","kimi_calls"),("Qwen Token","qwen_total_tokens"),("Kimi Token","kimi_total_tokens"),("Qwen费用USD","qwen_cost_usd"),("Kimi费用USD","kimi_cost_usd"),
        ("升级次数","escalation_count"),("降级次数","deescalation_count"),("是否升级Kimi","escalated_to_kimi"),("Critical任务数","critical_task_count"),("权限请求","permission_requests"),("会话API重试","retry_count"),("基础设施任务重试","infra_retry_count"),("基础设施重试原因","infra_retry_reasons_json"),("丢弃尝试耗时(s)","infra_discarded_elapsed"),("丢弃尝试Token","infra_discarded_total_tokens"),("Step数","step_count"),("工具错误","tool_errors"),
        ("Workspace","workspace"),("Transcript","transcript"),("运行错误","error"),("打分错误","grade_error"),("打分备注","grade_notes"),("Model Usage JSON","model_usage_json"),("Model Calls JSON","model_call_counts_json"),("Escalation JSON","escalation_events_json"),
    ]
    fill = PatternFill("solid", fgColor="4472C4"); font = Font(bold=True, color="FFFFFF")
    for i,(label,_) in enumerate(columns,1):
        c=ws.cell(1,i,label); c.fill=fill; c.font=font; c.alignment=Alignment(horizontal="center")
    for r,row in enumerate(results,2):
        for c,(_,key) in enumerate(columns,1):
            value=row.get(key)
            if key=="success": value="✓" if value else "✗"
            ws.cell(r,c,truncate_excel(value))
    for cells in ws.columns:
        max_len=max(len(str(c.value or "")) for c in cells); ws.column_dimensions[cells[0].column_letter].width=min(max_len+3,60)
    ss=wb.create_sheet("汇总")
    for r,(k,v) in enumerate(summary.items(),1): ss.cell(r,1,k).font=Font(bold=True); ss.cell(r,2,v)
    ss.column_dimensions["A"].width=34; ss.column_dimensions["B"].width=34
    wb.save(output_path)


def percentile(values: list[float], p: float) -> Optional[float]:
    if not values: return None
    xs=sorted(values); idx=(len(xs)-1)*p; lo=int(idx); hi=min(lo+1,len(xs)-1); frac=idx-lo
    return xs[lo]*(1-frac)+xs[hi]*frac


def print_summary(results: list[dict[str, Any]], total_elapsed: float) -> dict[str, Any]:
    scores=[float(r["score"]) for r in results if r.get("score") is not None]
    agent_times=[float(r.get("agent_elapsed") or 0.0) for r in results]
    total_tokens=sum(int(r.get("total_tokens") or 0) for r in results if r.get("total_tokens") is not None)
    summary={
        "任务总数":len(results),"执行成功":sum(1 for r in results if r.get("success")),"执行失败":sum(1 for r in results if not r.get("success")),
        "有分数任务数":len(scores),"平均分数":round(sum(scores)/len(scores),4) if scores else None,"打分失败任务数":sum(1 for r in results if r.get("grade_error")),
        "整次运行壁钟(s)":round(total_elapsed,2),"Agent累计耗时(s)":round(sum(agent_times),2),"Agent平均耗时(s)":round(sum(agent_times)/len(agent_times),2) if agent_times else None,
        "Agent中位耗时(s)":round(percentile(agent_times,0.5) or 0,2) if agent_times else None,"Agent P90耗时(s)":round(percentile(agent_times,0.9) or 0,2) if agent_times else None,
        "总输入Token":int(sum(int(r.get("input_tokens") or 0) for r in results if r.get("input_tokens") is not None)),
        "总输出Token":int(sum(int(r.get("output_tokens") or 0) for r in results if r.get("output_tokens") is not None)),
        "总推理Token(输出子集)":int(sum(int(r.get("reasoning_tokens") or 0) for r in results if r.get("reasoning_tokens") is not None)),
        "总Token(不重复推理)":int(total_tokens),"总Agent费用USD":round(sum(float(r.get("cost_usd") or 0) for r in results if r.get("cost_usd") is not None),6),
        "Usage缺失任务数":sum(1 for r in results if not r.get("usage_complete")),"Timeout数":sum(1 for r in results if r.get("status")=="timeout"),
        "权限请求数":sum(int(r.get("permission_requests") or 0) for r in results),"会话API重试数":sum(int(r.get("retry_count") or 0) for r in results),
        "基础设施任务重试数":sum(int(r.get("infra_retry_count") or 0) for r in results),
        "基础设施丢弃尝试耗时(s)":round(sum(float(r.get("infra_discarded_elapsed") or 0.0) for r in results),2),
        "基础设施丢弃尝试Token":int(sum(int(r.get("infra_discarded_total_tokens") or 0) for r in results)),
        "Hybrid升级任务数":sum(1 for r in results if r.get("escalated_to_kimi")),"Hybrid升级事件数":sum(int(r.get("escalation_count") or 0) for r in results),
        "Qwen调用数":sum(int(r.get("qwen_calls") or 0) for r in results),"Kimi调用数":sum(int(r.get("kimi_calls") or 0) for r in results),
        "Qwen Token":int(sum(int(r.get("qwen_total_tokens") or 0) for r in results if r.get("qwen_total_tokens") is not None)),"Kimi Token":int(sum(int(r.get("kimi_total_tokens") or 0) for r in results if r.get("kimi_total_tokens") is not None)),
        "Qwen费用USD":round(sum(float(r.get("qwen_cost_usd") or 0) for r in results if r.get("qwen_cost_usd") is not None),6),"Kimi费用USD":round(sum(float(r.get("kimi_cost_usd") or 0) for r in results if r.get("kimi_cost_usd") is not None),6),
    }
    print("\n"+"="*100); print("汇总"); print("="*100)
    for k,v in summary.items(): print(f"{k:<28}: {v}")
    print("="*100)
    return summary


def build_args() -> argparse.Namespace:
    script_path=Path(__file__).resolve(); root=script_path.parent.parent
    parser=argparse.ArgumentParser(description="Run PinchBench with 我们的框架 on native Windows", formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("--skill-dir",default=str(root/"skill")); parser.add_argument("--tasks-dir",default=None)
    parser.add_argument("--source-dir",default=str(root/"framework"/"localclaw-localcoding-dev")); parser.add_argument("--config-dir",default=str(root/"config"))
    parser.add_argument("--server-url",default="http://127.0.0.1:10086"); parser.add_argument("--endpoint-id",default=OUR_FRAMEWORK_ENDPOINT_ID)
    parser.add_argument("--mode",choices=["qwen","kimi","hybrid"],required=True); parser.add_argument("--qwen-model",default=QWEN_MODEL); parser.add_argument("--kimi-model",default=KIMI_MODEL)
    parser.add_argument("--permission-mode",choices=["plan","default","acceptEdits","bypassPermissions"],default="acceptEdits")
    parser.add_argument("--no-auto-approve-permissions",dest="auto_approve_permissions",action="store_false"); parser.set_defaults(auto_approve_permissions=True)
    parser.add_argument("--suite",default="core"); parser.add_argument("--limit",type=int,default=None); parser.add_argument("--skip",default=""); parser.add_argument("--skip-network",action="store_true")
    parser.add_argument("--timeout-multiplier",type=float,default=3.0); parser.add_argument("--network-timeout",type=float,default=300.0)
    parser.add_argument("--infra-retries",type=int,default=1,help="Retry a whole task only for explicit OpenRouter upstream 504 or WebSearch HTTP 429; fresh workspace/session")
    parser.add_argument("--judge-timeout",type=float,default=300.0); parser.add_argument("--judge-model",default=OUR_FRAMEWORK_JUDGE_MODEL)
    parser.add_argument("--results-dir",default=str(root/"runs")); parser.add_argument("--keep-workspaces",action="store_true"); parser.add_argument("--no-grade",action="store_true"); parser.add_argument("--no-xlsx",action="store_true")
    parser.add_argument("--no-workspace-instruction",action="store_true"); parser.add_argument("--preflight",action="store_true"); parser.add_argument("--no-judge-cache",action="store_true"); parser.add_argument("--clear-judge-cache",action="store_true"); parser.add_argument("--verbose",action="store_true")
    return parser.parse_args()


def main() -> int:
    args=build_args()
    if args.infra_retries < 0 or args.infra_retries > 3:
        raise SystemExit("--infra-retries must be between 0 and 3")
    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO,format="%(asctime)s %(levelname)-8s %(message)s",datefmt="%H:%M:%S")
    skill_dir=Path(args.skill_dir).expanduser().resolve(); tasks_dir=Path(args.tasks_dir).expanduser().resolve() if args.tasks_dir else skill_dir/"tasks"
    if not skill_dir.exists(): raise SystemExit(f"找不到 PinchBench 仓库目录: {skill_dir}")
    if not tasks_dir.exists(): raise SystemExit(f"找不到 tasks 目录: {tasks_dir}")
    grader_module,grader_error=load_pinchbench_grading(skill_dir)
    all_tasks,core_tasks=load_tasks(tasks_dir); selected=filter_tasks(all_tasks,core_tasks,args.suite,args.limit)
    skip_ids=set(DEFAULT_SKIPPED_TASKS); skip_ids.update(x.strip() for x in args.skip.split(",") if x.strip())
    if args.skip_network: skip_ids.update(t.task_id for t in selected if is_network_task(t))
    selected=[t for t in selected if t.task_id not in skip_ids]
    if not selected: raise SystemExit("没有可运行任务")
    prereq,fixtures=print_preflight(selected,skill_dir,tasks_dir,grader_module,grader_error,args)
    needs_judge=not args.no_grade and any(t.grading_type in {"llm_judge","hybrid"} for t in selected)
    if args.preflight:
        if not getattr(args,"_health_ok",False): return 10
        if not getattr(args,"_endpoint_ok",False): return 11
        if not getattr(args,"_isolation_ok",False): return 12
        if fixtures: return 2
        if prereq: return 3
        if grader_error: return 4
        if needs_judge and not os.environ.get("OPENROUTER_API_KEY"): return 5
        return 0
    if grader_error and not args.no_grade: raise SystemExit(grader_error)
    if needs_judge and not os.environ.get("OPENROUTER_API_KEY"): raise SystemExit("Judge 需要当前 PowerShell 的 OPENROUTER_API_KEY")
    if not getattr(args,"_health_ok",False) or not getattr(args,"_endpoint_ok",False) or not getattr(args,"_isolation_ok",False): raise SystemExit("Preflight failed; do not start formal run")

    results_root=Path(args.results_dir).expanduser().resolve(); run_id=dt.datetime.now().strftime(f"our_framework_{args.mode}_%Y%m%d_%H%M%S"); results_dir=results_root/run_id
    workspaces_dir=results_dir/"workspaces"; transcripts_dir=results_dir/"transcripts"; results_dir.mkdir(parents=True,exist_ok=True); workspaces_dir.mkdir(parents=True,exist_ok=True); transcripts_dir.mkdir(parents=True,exist_ok=True)
    progress_path=results_dir/"progress.jsonl"; partial_path=results_dir/"results.partial.json"; judge_model=str(args.judge_model)
    write_run_config(results_dir/"run_config.json",args,skill_dir,tasks_dir,selected,judge_model)
    judge_cache_dir=results_root/".judge_cache"
    if grader_module is not None and hasattr(grader_module,"set_judge_cache_dir") and not args.no_judge_cache:
        grader_module.set_judge_cache_dir(judge_cache_dir)
        if args.clear_judge_cache and hasattr(grader_module,"clear_judge_cache"): grader_module.clear_judge_cache()
    print(f"\nResults dir          : {results_dir}\nMode                 : {args.mode}\nJudge model          : {judge_model}\n")

    results:list[dict[str,Any]]=[]; total_start=time.monotonic()
    for index,task in enumerate(selected,1):
        task_e2e_start=time.monotonic(); workspace=workspaces_dir/task.task_id; transcript_dir=transcripts_dir/task.task_id
        discarded_attempts: list[dict[str, Any]] = []
        attempt_number = 1
        while True:
            execution=execute_task(task,index,len(selected),skill_dir,workspace,transcript_dir,args)
            diagnosis=inspect_sdk_infrastructure_failures(transcript_dir)
            raw_reasons = list(diagnosis.get("reasons") or [])
            retry_reasons = [
                reason for reason in raw_reasons
                if reason == INFRA_FAILURE_OPENROUTER_504
                or (reason == INFRA_FAILURE_WEBSEARCH_429 and not execution.get("success"))
                or (reason == INFRA_FAILURE_PROVIDER_429 and not execution.get("success"))
            ]
            diagnosis["retryable"] = bool(retry_reasons)
            diagnosis["reasons"] = retry_reasons
            if diagnosis.get("retryable") and len(discarded_attempts) < args.infra_retries:
                archived=archive_infra_attempt(
                    results_dir=results_dir, task_id=task.task_id, attempt_number=attempt_number,
                    workspace=workspace, transcript_dir=transcript_dir, execution=execution, diagnosis=diagnosis,
                )
                discarded_attempts.append(archived)
                retry_reasons_now = list(diagnosis.get("reasons") or [])
                retry_wait = 90 if (INFRA_FAILURE_WEBSEARCH_429 in retry_reasons_now or INFRA_FAILURE_PROVIDER_429 in retry_reasons_now) else 10
                print(f"  ↻ infrastructure retry {len(discarded_attempts)}/{args.infra_retries}: {','.join(retry_reasons_now)} — fresh workspace/session after {retry_wait}s cooldown")
                time.sleep(retry_wait)
                attempt_number += 1
                continue
            break
        execution["infra_retry_count"] = len(discarded_attempts)
        execution["infra_retry_reasons_json"] = json.dumps([a.get("diagnosis", {}).get("reasons", []) for a in discarded_attempts], ensure_ascii=False)
        execution["infra_discarded_attempts_json"] = json.dumps(discarded_attempts, ensure_ascii=False)
        execution["infra_discarded_elapsed"] = round(sum(float(a.get("elapsed") or 0.0) for a in discarded_attempts), 6)
        execution["infra_discarded_total_tokens"] = int(sum(int(a.get("total_tokens") or 0) for a in discarded_attempts))
        grade_result=GradeResult(score=None,grading_type=task.grading_type,breakdown={})
        grading_attempted=execution["status"] not in {"missing_prerequisite","missing_fixture","isolation_error"} and not args.no_grade
        grading_start=time.monotonic()
        if grading_attempted:
            grading_execution=dict(execution); grading_execution["transcript"]=execution.get("transcript_data",[])
            grade_result=grade_with_pinchbench_default(grader_module=grader_module,task=task,execution_result=grading_execution,workspace=workspace,skill_dir=skill_dir,judge_timeout=args.judge_timeout,judge_model=judge_model,verbose=args.verbose)
        execution["agent_elapsed"]=float(execution.get("elapsed") or 0.0); execution["grading_elapsed"]=time.monotonic()-grading_start if grading_attempted else 0.0; execution["end_to_end_elapsed"]=time.monotonic()-task_e2e_start
        execution["score"]=round(float(grade_result.score),4) if grade_result.score is not None else None; execution["breakdown"]=grade_result.breakdown; execution["grade_notes"]=grade_result.notes; execution["grade_error"]=grade_result.error
        execution.pop("transcript_data",None); results.append(execution)
        with progress_path.open("a",encoding="utf-8") as f: f.write(json.dumps(execution,ensure_ascii=False)+"\n"); f.flush()
        partial_path.write_text(json.dumps({"completed":len(results),"results":results},ensure_ascii=False,indent=2),encoding="utf-8")
        marker="✓" if execution["success"] else "✗"; score_text=f"{execution['score']:.3f}" if execution["score"] is not None else "N/A"
        print(f"  {marker} mode={args.mode} score={score_text} agent={execution['agent_elapsed']:.1f}s grade={execution['grading_elapsed']:.1f}s tokens={execution.get('total_tokens')} qwen_calls={execution.get('qwen_calls')} kimi_calls={execution.get('kimi_calls')} escalations={execution.get('escalation_count')} infra_retries={execution.get('infra_retry_count',0)}")
        if execution.get("error"): print(f"  运行错误: {str(execution['error'])[:500]}")
        if execution.get("grade_error"): print(f"  打分错误: {str(execution['grade_error'])[:500]}")
        print()
        if not args.keep_workspaces and execution["success"] and not execution.get("grade_error"): shutil.rmtree(workspace,ignore_errors=True)

    total_elapsed=time.monotonic()-total_start; summary=print_summary(results,total_elapsed); payload={"summary":summary,"results":results}
    json_path=results_dir/"results.json"; json_path.write_text(json.dumps(payload,ensure_ascii=False,indent=2),encoding="utf-8"); csv_path=results_dir/"results.csv"; save_csv(results,csv_path)
    xlsx_path=results_dir/"results.xlsx"
    if not args.no_xlsx: save_xlsx(results,summary,xlsx_path)
    print(f"\n结果文件:\n  Config : {results_dir/'run_config.json'}\n  JSON   : {json_path}\n  CSV    : {csv_path}\n  XLSX   : {xlsx_path if not args.no_xlsx else '(disabled)'}\n  Logs   : {transcripts_dir}\n  Progress JSONL : {progress_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
