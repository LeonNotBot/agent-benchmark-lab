#!/usr/bin/env python3
# Runner revision: 2026-07-27-pinchbench-codex-windows-v1-jsonl
r"""
Run the local PinchBench task set with Codex CLI on native Windows.

Target evaluation chain:
    PinchBench local tasks -> Codex built-in agent/tools -> OpenRouter
    -> deepseek/deepseek-v4-pro

Comparability contract with the validated Windows Qwen Code/OpenCode runs:
- Same checked-out PinchBench tasks, manifest order, fixtures, exclusions and commit.
- Same Python environment, official PinchBench grading engine, judge backend/model/cache.
- Same serial execution (one worker, one task at a time).
- Same task timeouts, network timeout, workspace staging and workspace instruction.
- Same UTF-8 stdin prompt transport; prompts are never placed in Windows .cmd argv.
- Same core JSON/CSV/XLSX result fields, with additional Codex audit fields.
- Fresh Codex thread for normal tasks; explicit `codex exec resume` for continuation turns.
- Raw stdout JSONL, stderr, prompt, final-message file and normalized transcript are retained.

Intentional agent-adapter differences:
- Codex CLI is pinned separately and invoked with `codex exec --json`.
- A dedicated CODEX_HOME supplies the tested OpenRouter provider configuration.
- `approval_policy=never` and `sandbox_mode=workspace-write` are required for unattended
  execution; Windows uses the configured `unelevated` Codex sandbox.
- `--ignore-rules` disables user/project exec-policy rules. No custom system prompt,
  AGENTS.md, memories, hooks, MCP servers or skills are added by this runner.
- Item-level command failures and Codex warning items are retained for audit, but do not
  automatically fail a task when the turn itself completes successfully.
- Codex JSONL does not reliably expose a runtime model identity field. The requested
  model is verified through both argv and the isolated config, and that limitation is
  recorded rather than pretending it was observed from the stream.

Recommended location:
    C:\pinchbench-codex\runner\run_pinchbench_codex_windows.py

Examples:
    python run_pinchbench_codex_windows.py --preflight --suite all
    python run_pinchbench_codex_windows.py --suite task_sanity,task_iterative_code_refine --keep-workspaces --verbose
    python run_pinchbench_codex_windows.py --suite all --keep-workspaces --verbose

Scoring:
- Automated tasks use PinchBench's checked-out automated grading code.
- Hybrid and llm_judge tasks use PinchBench's checked-out DEFAULT_JUDGE_MODEL
  through judge_backend="api" and OPENROUTER_API_KEY.
- The judge is separate from the tested Codex + DeepSeek V4 Pro agent.
"""
from __future__ import annotations

import argparse
import csv
import dataclasses
import datetime as dt
import hashlib
import json
import logging
import os
import platform
import queue
import re
import shutil
import signal
import subprocess
import sys
import threading
import time
import traceback
from collections import Counter
from pathlib import Path
from typing import Any, Optional

os.environ.setdefault("PYTHONUTF8", "1")
os.environ.setdefault("PYTHONIOENCODING", "utf-8")

try:
    import yaml  # type: ignore
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "缺少 PyYAML。请使用项目虚拟环境运行：\n"
        r"C:\pinchbench-codex\.venv\Scripts\python.exe -m pip install pyyaml"
    ) from exc

try:
    import tomllib  # type: ignore[attr-defined]
except ImportError:  # pragma: no cover - Python 3.10 fallback
    try:
        import tomli as tomllib  # type: ignore[no-redef]
    except ImportError as exc:
        raise SystemExit(
            "Python 3.10 需要 tomli：python -m pip install tomli"
        ) from exc

LOGGER = logging.getLogger("pinchbench-codex")

RUNNER_REVISION = "2026-07-27-codex-windows-v1.1-encoding-normalization+encoding-v3.2+transport-recovery-v1"
DEFAULT_MODEL = "deepseek/deepseek-v4-pro"
DEFAULT_CODEX_VERSION = "0.145.0"
DEFAULT_APPROVAL_POLICY = "never"
DEFAULT_SANDBOX_MODE = "workspace-write"
DEFAULT_WINDOWS_SANDBOX = "unelevated"
DEFAULT_PROVIDER = "openrouter"
DEFAULT_WIRE_API = "responses"
PINCHBENCH_DEFAULT_JUDGE_MODEL = "openrouter/anthropic/claude-haiku-4.5"

DEFAULT_SKIPPED_TASKS = {
    "task_gh_issue_triage",
    "task_gws_email_triage",
    "task_gws_cross_service",
    "task_gws_task_management",
}

NETWORK_TASKS = {
    "task_earnings_analysis",
    "task_stock",
    "task_market_research",
    "task_polymarket_briefing",
    "task_executive_lookup",
    "task_events",
    "task_deep_research",
    "task_competitive_research",
    "task_oss_alternative_research",
    "task_pricing_research",
    "task_it_procurement",
    "task_eu_regulation_research",
    "task_byok_best_practices",
}

TEXT_FILE_SUFFIXES = {
    ".txt", ".md", ".py", ".js", ".ts", ".tsx", ".jsx", ".json", ".yaml", ".yml",
    ".csv", ".tsv", ".html", ".css", ".xml", ".toml", ".ini", ".cfg", ".log",
    ".ics", ".sh", ".ps1", ".bat", ".cmd", ".sql", ".java", ".go", ".rs", ".rb", ".php",
}

SKIP_WORKSPACE_NAMES = {
    "BOOTSTRAP.md", "SOUL.md", "USER.md", "IDENTITY.md", "HEARTBEAT.md",
    "TOOLS.md", "AGENTS.md", "CLAUDE.md",
}
SKIP_WORKSPACE_DIRS = {
    ".git", ".openclaw", ".opencode", ".qwen", ".codex",
    "__pycache__", "node_modules", "skills",
}


@dataclasses.dataclass
class Task:
    task_id: str
    name: str
    category: str
    grading_type: str
    timeout_seconds: int
    workspace_files: list[dict[str, Any]]
    prompt: str
    expected_behavior: str
    grading_criteria: list[str]
    automated_checks: str
    llm_judge_rubric: str
    grading_weights: dict[str, float]
    metadata: dict[str, Any]
    file_path: Path
    multi_session: bool = False
    sessions: list[dict[str, Any]] = dataclasses.field(default_factory=list)


@dataclasses.dataclass
class GradeResult:
    score: Optional[float]
    grading_type: str
    breakdown: dict[str, float]
    notes: str = ""
    error: Optional[str] = None


# ---------------------------------------------------------------------------
# Task loading
# ---------------------------------------------------------------------------

def parse_sections(body: str) -> dict[str, str]:
    sections: dict[str, list[str]] = {}
    current: Optional[str] = None
    for line in body.splitlines():
        match = re.match(r"^##\s+(.+?)\s*$", line)
        if match:
            current = match.group(1).strip()
            sections.setdefault(current, [])
        elif current:
            sections[current].append(line)
    return {key: "\n".join(value).strip() for key, value in sections.items()}


def extract_grading_criteria(text: str) -> list[str]:
    criteria: list[str] = []
    for line in text.splitlines():
        match = re.match(r"^-\s+\[[ xX]\]\s+(.+)$", line.strip())
        if match:
            criteria.append(match.group(1).strip())
    return criteria


def load_task_file(path: Path, category_override: Optional[str] = None) -> Task:
    raw = path.read_text(encoding="utf-8")
    match = re.match(r"^---\s*\r?\n(.*?)\r?\n---\s*\r?\n?(.*)$", raw, re.DOTALL)
    if not match:
        raise ValueError(f"{path} 没有 YAML frontmatter")

    metadata = yaml.safe_load(match.group(1)) or {}
    if not isinstance(metadata, dict):
        raise ValueError(f"{path} 的 YAML frontmatter 不是对象")

    body = match.group(2)
    sections = parse_sections(body)
    task_id = str(metadata.get("id") or path.stem)

    weights = metadata.get("grading_weights") or {}
    if not isinstance(weights, dict):
        weights = {}
    normalized_weights: dict[str, float] = {}
    for key, value in weights.items():
        try:
            normalized_weights[str(key)] = float(value)
        except (TypeError, ValueError):
            pass

    raw_sessions = metadata.get("sessions") or []
    sessions = [item for item in raw_sessions if isinstance(item, dict)] if isinstance(raw_sessions, list) else []

    return Task(
        task_id=task_id,
        name=str(metadata.get("name") or task_id),
        category=str(category_override or metadata.get("category") or ""),
        grading_type=str(metadata.get("grading_type") or "automated"),
        timeout_seconds=int(metadata.get("timeout_seconds") or 120),
        workspace_files=list(metadata.get("workspace_files") or []),
        prompt=sections.get("Prompt", "").strip(),
        expected_behavior=sections.get("Expected Behavior", "").strip(),
        grading_criteria=extract_grading_criteria(sections.get("Grading Criteria", "")),
        automated_checks=sections.get("Automated Checks", "") or "",
        llm_judge_rubric=sections.get("LLM Judge Rubric", "") or "",
        grading_weights=normalized_weights,
        metadata=metadata,
        file_path=path,
        multi_session=bool(metadata.get("multi_session") or sessions),
        sessions=sessions,
    )


def parse_manifest(tasks_dir: Path) -> tuple[list[str], dict[str, str], list[str]]:
    manifest_path = tasks_dir / "manifest.yaml"
    if not manifest_path.exists():
        ordered = [
            p.stem for p in sorted(tasks_dir.glob("task_*.md"))
            if p.stem != "task_XX_name"
        ]
        return ordered, {}, []

    manifest = yaml.safe_load(manifest_path.read_text(encoding="utf-8")) or {}
    if not isinstance(manifest, dict):
        raise ValueError(f"{manifest_path} 不是有效 YAML 对象")

    category_map: dict[str, str] = {}
    core_tasks = [str(x) for x in (manifest.get("core") or [])]

    if "categories" in manifest:
        run_first = [str(x) for x in (manifest.get("run_first") or [])]
        categories = manifest.get("categories") or {}
        all_task_ids: list[str] = []
        if isinstance(categories, dict):
            for category, ids in categories.items():
                for task_id in ids or []:
                    task_id = str(task_id)
                    category_map[task_id] = str(category)
                    all_task_ids.append(task_id)

        seen: set[str] = set()
        ordered: list[str] = []
        for task_id in run_first + all_task_ids:
            if task_id not in seen:
                seen.add(task_id)
                ordered.append(task_id)
        return ordered, category_map, core_tasks

    ordered = [str(x) for x in (manifest.get("tasks") or [])]
    return ordered, category_map, core_tasks


def load_tasks(tasks_dir: Path) -> tuple[list[Task], list[str]]:
    """Load exactly the checked-out manifest when it exists.

    PinchBench ships template/example Markdown files alongside real tasks.
    Falling back to every task_*.md while a manifest exists can accidentally
    count a template as a benchmark task.  The manifest is therefore the
    source of truth for normal runs.
    """
    manifest_exists = (tasks_dir / "manifest.yaml").exists()
    ordered_ids, category_map, core_tasks = parse_manifest(tasks_dir)
    tasks: list[Task] = []

    for task_id in ordered_ids:
        if task_id == "task_XX_name":
            continue
        task_path = tasks_dir / f"{task_id}.md"
        if not task_path.exists():
            LOGGER.warning("manifest 引用了不存在的任务文件: %s", task_path)
            continue
        try:
            task = load_task_file(
                task_path,
                category_override=category_map.get(task_id),
            )
            if (
                task.task_id == "task_XX_name"
                or task.category.strip().lower() == "category_name"
            ):
                LOGGER.debug("忽略 PinchBench 模板任务: %s", task_path)
                continue
            tasks.append(task)
        except Exception as exc:
            LOGGER.warning("加载任务失败 %s: %s", task_path, exc)

    if manifest_exists:
        return tasks, core_tasks

    # Only use file discovery when no manifest is available.
    loaded_ids = {task.task_id for task in tasks}
    for task_path in sorted(tasks_dir.glob("task_*.md")):
        if task_path.stem in loaded_ids or task_path.stem == "task_XX_name":
            continue
        try:
            task = load_task_file(task_path)
            if (
                task.task_id == "task_XX_name"
                or task.category.strip().lower() == "category_name"
            ):
                continue
            tasks.append(task)
        except Exception as exc:
            LOGGER.warning("加载任务失败 %s: %s", task_path, exc)

    return tasks, core_tasks


def filter_tasks(
    tasks: list[Task],
    core_tasks: list[str],
    suite: str,
    limit: Optional[int],
) -> list[Task]:
    suite = suite.strip()
    task_by_id = {task.task_id: task for task in tasks}

    if suite == "all":
        selected = tasks
    elif suite == "core":
        core_set = set(core_tasks)
        selected = [task for task in tasks if task.task_id in core_set] if core_set else tasks[:25]
    elif suite == "automated-only":
        selected = [task for task in tasks if task.grading_type == "automated"]
    elif suite == "llm-judge-only":
        selected = [task for task in tasks if task.grading_type == "llm_judge"]
    elif suite == "hybrid-only":
        selected = [task for task in tasks if task.grading_type == "hybrid"]
    elif suite == "judge-required-only":
        selected = [task for task in tasks if task.grading_type in {"llm_judge", "hybrid"}]
    else:
        ids = [item.strip() for item in suite.split(",") if item.strip()]
        missing = [task_id for task_id in ids if task_id not in task_by_id]
        if missing:
            raise SystemExit(f"suite 中有未知任务: {', '.join(missing)}")
        selected = [task_by_id[task_id] for task_id in ids]

    if limit is not None:
        selected = selected[:limit]
    return selected


def is_network_task(task: Task) -> bool:
    return bool(
        task.task_id in NETWORK_TASKS
        or task.metadata.get("network_required") is True
        or task.metadata.get("requires_network") is True
    )


# ---------------------------------------------------------------------------
# Workspace and prerequisites
# ---------------------------------------------------------------------------

def safe_workspace_dest(dest: str) -> Path:
    rel = Path(dest)
    if rel.is_absolute() or ".." in rel.parts:
        raise ValueError(f"非法 workspace 文件路径: {dest}")
    return rel


def stage_workspace_files(
    task: Task,
    workspace: Path,
    skill_dir: Path,
) -> tuple[list[str], list[str]]:
    staged: list[str] = []
    missing: list[str] = []

    for wf in task.workspace_files:
        if not isinstance(wf, dict):
            missing.append(f"无法识别的 workspace_files 项: {wf!r}")
            continue

        dest_value = wf.get("dest") or wf.get("path") or wf.get("name")
        source_value = wf.get("source")
        content_value = wf.get("content")

        if not dest_value and source_value:
            dest_value = Path(str(source_value)).name
        if not dest_value:
            missing.append(f"workspace_files 项缺少 dest/path: {wf!r}")
            continue

        try:
            dest_rel = safe_workspace_dest(str(dest_value))
        except ValueError as exc:
            missing.append(str(exc))
            continue

        dest_path = workspace / dest_rel
        dest_path.parent.mkdir(parents=True, exist_ok=True)

        if content_value is not None:
            dest_path.write_text(str(content_value), encoding="utf-8")
            staged.append(str(dest_rel))
            continue

        if not source_value:
            missing.append(f"workspace_files 项缺少 source/content: {wf!r}")
            continue

        src_rel = Path(str(source_value))
        candidates = [
            skill_dir / "assets" / src_rel,
            skill_dir / src_rel,
            task.file_path.parent / src_rel,
            task.file_path.parent.parent / "assets" / src_rel,
        ]
        src_path = next((candidate for candidate in candidates if candidate.exists()), None)
        if src_path is None:
            missing.append(
                f"找不到预置文件 {source_value}; 尝试过: "
                + ", ".join(str(candidate) for candidate in candidates)
            )
            continue

        if src_path.is_dir():
            if dest_path.exists():
                shutil.rmtree(dest_path)
            shutil.copytree(src_path, dest_path)
        else:
            shutil.copy2(src_path, dest_path)
        staged.append(str(dest_rel))

    return staged, missing



def snapshot_workspace_files(workspace: Path) -> dict[str, str]:
    """Return SHA-256 hashes for files present before the agent runs."""
    snapshot: dict[str, str] = {}
    for path in sorted(workspace.rglob("*")):
        if not path.is_file():
            continue
        rel = path.relative_to(workspace)
        if any(part in SKIP_WORKSPACE_DIRS for part in rel.parts):
            continue
        try:
            snapshot[str(rel)] = sha256_file(path)
        except OSError:
            continue
    return snapshot


def normalize_workspace_text_encodings(
    workspace: Path,
    baseline_hashes: dict[str, str],
    audit_dir: Path,
) -> dict[str, Any]:
    """Normalize changed/new text artifacts to UTF-8 without BOM before grading.

    This Windows compatibility layer changes encoding only, never logical text
    content. It handles explicit UTF BOM encodings and BOM-less Windows text
    only when strict decoding and exact byte round-trip checks succeed.
    Original bytes are retained under the task transcript directory.
    """
    audit_dir.mkdir(parents=True, exist_ok=True)
    originals_dir = audit_dir / "originals"
    records: list[dict[str, Any]] = []
    errors: list[str] = []

    bom_decoders: list[tuple[bytes, str, str]] = [
        (b"\xff\xfe\x00\x00", "utf-32", "utf-32-le-bom"),
        (b"\x00\x00\xfe\xff", "utf-32", "utf-32-be-bom"),
        (b"\xef\xbb\xbf", "utf-8-sig", "utf-8-bom"),
        (b"\xff\xfe", "utf-16", "utf-16-le-bom"),
        (b"\xfe\xff", "utf-16", "utf-16-be-bom"),
    ]

    def acceptable_text(text: str) -> bool:
        if "\x00" in text:
            return False
        if not text:
            return True
        disallowed_controls = sum(
            1
            for char in text
            if ord(char) < 32 and char not in "\t\r\n\f"
        )
        return disallowed_controls / max(1, len(text)) <= 0.002

    def strict_roundtrip_decode(
        raw_bytes: bytes,
        codec: str,
    ) -> Optional[str]:
        try:
            decoded = raw_bytes.decode(codec, errors="strict")
            if not acceptable_text(decoded):
                return None
            if decoded.encode(codec, errors="strict") != raw_bytes:
                return None
            return decoded
        except (LookupError, UnicodeError):
            return None

    for path in sorted(workspace.rglob("*")):
        if not path.is_file():
            continue
        rel = path.relative_to(workspace)
        if any(part in SKIP_WORKSPACE_DIRS for part in rel.parts):
            continue
        if path.suffix.lower() not in TEXT_FILE_SUFFIXES:
            continue

        rel_text = str(rel)
        try:
            raw = path.read_bytes()
        except OSError as exc:
            errors.append(f"{rel_text}: read failed: {exc}")
            continue

        current_hash = hashlib.sha256(raw).hexdigest()
        if baseline_hashes.get(rel_text) == current_hash:
            continue

        text: Optional[str] = None
        detected = ""
        confidence = ""

        for prefix, decoder, label in bom_decoders:
            if not raw.startswith(prefix):
                continue
            try:
                candidate = raw.decode(decoder, errors="strict")
            except UnicodeError as exc:
                errors.append(f"{rel_text}: {label} decode failed: {exc}")
                candidate = None
            if candidate is not None and acceptable_text(candidate):
                text = candidate
                detected = label
                confidence = "explicit-bom"
            break

        if text is None:
            try:
                raw.decode("utf-8", errors="strict")
                continue
            except UnicodeDecodeError:
                pass

            fallback_codecs = (
                ("mbcs", "windows-ansi-mbcs"),
                ("cp936", "windows-cp936"),
                ("gb18030", "windows-gb18030"),
            )
            for codec, label in fallback_codecs:
                candidate = strict_roundtrip_decode(raw, codec)
                if candidate is None:
                    continue
                text = candidate
                detected = label
                confidence = "strict-roundtrip"
                break

        if text is None:
            errors.append(
                f"{rel_text}: invalid UTF-8 and no safe Windows text codec "
                "passed strict round-trip validation"
            )
            continue

        normalized = text.encode("utf-8")
        if normalized == raw:
            continue

        original_path = originals_dir / rel
        original_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            original_path.write_bytes(raw)
            path.write_bytes(normalized)
        except OSError as exc:
            errors.append(f"{rel_text}: normalization write failed: {exc}")
            continue

        records.append({
            "path": rel_text,
            "detected_encoding": detected,
            "confidence": confidence,
            "original_size": len(raw),
            "normalized_size": len(normalized),
            "original_sha256": current_hash,
            "normalized_sha256": hashlib.sha256(normalized).hexdigest(),
            "original_copy": str(original_path),
        })

    result = {
        "enabled": True,
        "policy": (
            "changed_or_new_text_files_only; explicit BOM encodings and "
            "strict-roundtrip BOM-less Windows ANSI/CP936/GB18030 converted "
            "to UTF-8 without BOM"
        ),
        "normalized_count": len(records),
        "normalized_files": [record["path"] for record in records],
        "records": records,
        "errors": errors,
        "audit_dir": str(audit_dir),
    }
    (audit_dir / "encoding_normalization.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return result

def resolve_command(name: str) -> Optional[str]:
    candidates = [name]
    if sys.platform == "win32" and not Path(name).suffix:
        candidates = [f"{name}.cmd", f"{name}.exe", name]
    for candidate in candidates:
        resolved = shutil.which(candidate)
        if resolved:
            return resolved
    return None


def check_prerequisites(prereqs: list[str]) -> list[str]:
    missing: list[str] = []
    npm_command = resolve_command("npm") or "npm"

    for req_raw in prereqs:
        req = str(req_raw)
        if req.startswith("cli:"):
            command = req.split(":", 1)[1].strip()
            if not resolve_command(command):
                missing.append(req)
        elif req.startswith("npm:"):
            package = req.split(":", 1)[1].strip()
            try:
                result = subprocess.run(
                    [npm_command, "list", "-g", package, "--depth=0"],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=30,
                    check=False,
                )
                if result.returncode != 0:
                    missing.append(req)
            except Exception:
                missing.append(req)

    return missing


def choose_codex_command(preferred: str, inferred_root: Optional[Path] = None) -> str:
    if preferred and preferred != "auto":
        preferred_path = Path(preferred).expanduser()
        if preferred_path.exists():
            return str(preferred_path.resolve())
        resolved = resolve_command(preferred)
        if resolved:
            return resolved
        raise SystemExit(f"找不到 Codex 命令 {preferred!r}")

    candidates: list[Path] = []
    if inferred_root is not None:
        candidates.extend([
            inferred_root / "codex-cli" / "node_modules" / ".bin" / "codex.cmd",
            inferred_root / "codex-cli" / "node_modules" / ".bin" / "codex.exe",
        ])
    for candidate in candidates:
        if candidate.exists():
            return str(candidate.resolve())

    for command in (("codex.cmd", "codex.exe", "codex") if sys.platform == "win32" else ("codex",)):
        resolved = shutil.which(command)
        if resolved:
            return resolved

    raise SystemExit(
        "找不到 Codex。建议固定安装到 "
        r"C:\pinchbench-codex\codex-cli\node_modules\.bin\codex.cmd，"
        "并通过 --codex 指定。"
    )


def kill_proc_tree(proc: subprocess.Popen[str]) -> None:
    if proc.poll() is not None:
        return
    try:
        if sys.platform == "win32":
            subprocess.run(
                ["taskkill", "/F", "/T", "/PID", str(proc.pid)],
                capture_output=True,
                check=False,
            )
        else:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
    except Exception:
        try:
            proc.kill()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Codex JSONL event handling
# ---------------------------------------------------------------------------

NON_TOOL_ITEM_TYPES = {"agent_message", "reasoning", "error"}


def _to_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    try:
        return json.dumps(value, ensure_ascii=False)
    except (TypeError, ValueError):
        return str(value)


def codex_item(event: dict[str, Any]) -> dict[str, Any]:
    item = event.get("item")
    return item if isinstance(item, dict) else {}


def codex_visible_text(event: dict[str, Any]) -> str:
    if event.get("type") != "item.completed":
        return ""
    item = codex_item(event)
    if item.get("type") != "agent_message":
        return ""
    return str(item.get("text") or "")


def codex_item_error(event: dict[str, Any]) -> str:
    if event.get("type") != "item.completed":
        return ""
    item = codex_item(event)
    if item.get("type") != "error":
        return ""
    return str(item.get("message") or item.get("text") or "Codex item error")


def codex_turn_failure(event: dict[str, Any]) -> str:
    event_type = str(event.get("type") or "")
    if event_type not in {"turn.failed", "error"}:
        return ""
    error = event.get("error")
    if isinstance(error, dict):
        return str(error.get("message") or _stringify(error))
    if error:
        return str(error)
    return str(event.get("message") or event_type)


def codex_usage(event: dict[str, Any]) -> Optional[dict[str, int]]:
    if event.get("type") != "turn.completed":
        return None
    usage = event.get("usage")
    if not isinstance(usage, dict):
        return {}
    return {
        "input_tokens": _to_int(usage.get("input_tokens")),
        "output_tokens": _to_int(usage.get("output_tokens")),
        "reasoning_tokens": _to_int(
            usage.get("reasoning_output_tokens")
            or usage.get("reasoning_tokens")
        ),
        "cache_read_tokens": _to_int(
            usage.get("cached_input_tokens")
            or usage.get("cache_read_input_tokens")
        ),
        "cache_write_tokens": _to_int(
            usage.get("cache_write_input_tokens")
            or usage.get("cache_write_tokens")
        ),
    }


def codex_tool_item(item: dict[str, Any]) -> bool:
    item_type = str(item.get("type") or "")
    return bool(item_type and item_type not in NON_TOOL_ITEM_TYPES)


def codex_tool_name(item: dict[str, Any]) -> str:
    item_type = str(item.get("type") or "tool")
    if item_type == "command_execution":
        return "shell"
    if item_type == "file_change":
        return "file_change"
    return item_type


def codex_tool_arguments(item: dict[str, Any]) -> dict[str, Any]:
    if item.get("type") == "command_execution":
        return {"command": str(item.get("command") or "")}
    excluded = {
        "id", "type", "status", "aggregated_output", "exit_code", "result", "output",
    }
    return {
        str(key): value
        for key, value in item.items()
        if key not in excluded
    }


def codex_tool_result_text(item: dict[str, Any]) -> str:
    for key in ("aggregated_output", "output", "result", "message"):
        value = item.get(key)
        if value not in (None, ""):
            return _stringify(value)
    remaining = {
        str(key): value
        for key, value in item.items()
        if key not in {"id", "type", "status"}
    }
    return _stringify(remaining)


def normalize_codex_event(event: dict[str, Any]) -> list[dict[str, Any]]:
    """Convert Codex JSONL items to PinchBench's generic transcript shape."""
    event_type = str(event.get("type") or "")
    item = codex_item(event)
    item_type = str(item.get("type") or "")
    normalized: list[dict[str, Any]] = []

    if event_type == "item.completed" and item_type == "agent_message":
        text = str(item.get("text") or "")
        if text:
            normalized.append({
                "type": "message",
                "message": {
                    "role": "assistant",
                    "content": [{"type": "text", "text": text}],
                },
                "_adapter": {
                    "source": "codex",
                    "event_type": event_type,
                    "item_id": str(item.get("id") or ""),
                },
            })
        return normalized

    if event_type == "item.started" and codex_tool_item(item):
        tool_input = codex_tool_arguments(item)
        normalized.append({
            "type": "message",
            "message": {
                "role": "assistant",
                "content": [{
                    "type": "toolCall",
                    "id": str(item.get("id") or ""),
                    "name": codex_tool_name(item),
                    "arguments": tool_input,
                    "input": tool_input,
                }],
            },
            "_adapter": {
                "source": "codex",
                "event_type": event_type,
                "item_type": item_type,
            },
        })
        return normalized

    if event_type == "item.completed" and codex_tool_item(item):
        status = str(item.get("status") or "")
        exit_code = item.get("exit_code")
        is_error = (
            status in {"failed", "declined", "error"}
            or (exit_code is not None and _to_int(exit_code) != 0)
        )
        normalized.append({
            "type": "message",
            "message": {
                "role": "toolResult",
                "content": [{
                    "type": "toolResult",
                    "toolCallId": str(item.get("id") or ""),
                    "name": codex_tool_name(item),
                    "content": codex_tool_result_text(item),
                    "isError": is_error,
                }],
            },
            "_adapter": {
                "source": "codex",
                "event_type": event_type,
                "item_type": item_type,
                "status": status,
                "exit_code": exit_code,
            },
        })
    return normalized


def make_user_transcript_event(prompt: str, turn_index: int, session_label: str) -> dict[str, Any]:
    return {
        "type": "message",
        "message": {
            "role": "user",
            "content": [{"type": "text", "text": prompt}],
        },
        "_adapter": {
            "source": "pinchbench-codex-runner",
            "turn": turn_index,
            "session_label": session_label,
        },
    }


def _stream_reader(
    stream: Any,
    kind: str,
    output_queue: "queue.Queue[tuple[str, Optional[str]]]",
) -> None:
    try:
        for line in iter(stream.readline, ""):
            output_queue.put((kind, line))
    finally:
        output_queue.put((kind, None))
        try:
            stream.close()
        except Exception:
            pass


def _spawn_command(logical_cmd: list[str]) -> list[str]:
    """Use cmd.exe for Windows .cmd launchers without placing the prompt in argv."""
    if (
        sys.platform == "win32"
        and logical_cmd
        and Path(logical_cmd[0]).suffix.lower() in {".cmd", ".bat"}
    ):
        comspec = os.environ.get("ComSpec") or os.environ.get("COMSPEC") or "cmd.exe"
        command_line = subprocess.list2cmdline(logical_cmd)
        return [comspec, "/d", "/s", "/c", command_line]
    return logical_cmd


def run_codex_streaming(
    logical_cmd: list[str],
    cwd: Path,
    timeout: float,
    raw_stdout_path: Path,
    stderr_path: Path,
    stdin_path: Path,
    final_message_path: Path,
    requested_model: str,
    configured_model: str,
    approval_policy: str,
    sandbox_mode: str,
    windows_sandbox: str,
) -> dict[str, Any]:
    """Run one Codex non-interactive turn with UTF-8 stdin and JSONL stdout."""
    monotonic_start = time.monotonic()
    deadline = monotonic_start + timeout

    raw_stdout_path.parent.mkdir(parents=True, exist_ok=True)
    stderr_path.parent.mkdir(parents=True, exist_ok=True)
    final_message_path.parent.mkdir(parents=True, exist_ok=True)

    stdout_lines: list[str] = []
    stderr_lines: list[str] = []
    raw_events: list[dict[str, Any]] = []
    normalized_events: list[dict[str, Any]] = []
    output_chunks: list[str] = []
    event_counts: Counter[str] = Counter()

    session_id: Optional[str] = None
    ttft: Optional[float] = None
    fatal_errors: list[str] = []
    item_errors: list[str] = []
    command_failures: list[dict[str, Any]] = []
    permission_denials: list[dict[str, Any]] = []
    status = "success"
    timed_out = False

    turn_started = False
    turn_completed = False
    usage_seen = False
    input_tokens = 0
    output_tokens = 0
    reasoning_tokens = 0
    cache_read_tokens = 0
    cache_write_tokens = 0
    step_count = 0
    tool_errors = 0
    tool_call_count = 0
    agent_message_count = 0
    model_metadata_warnings: list[str] = []

    env = os.environ.copy()
    env.setdefault("PYTHONUTF8", "1")
    env.setdefault("PYTHONIOENCODING", "utf-8")
    env.setdefault("NO_COLOR", "1")
    env.setdefault("FORCE_COLOR", "0")

    creationflags = 0
    if sys.platform == "win32":
        creationflags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)

    spawn_cmd = _spawn_command(logical_cmd)

    try:
        with stdin_path.open("rb") as stdin_file:
            proc = subprocess.Popen(
                spawn_cmd,
                cwd=str(cwd),
                stdin=stdin_file,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace",
                bufsize=1,
                env=env,
                start_new_session=(sys.platform != "win32"),
                creationflags=creationflags,
            )
    except Exception as exc:
        return {
            "status": "error",
            "success": False,
            "returncode": None,
            "output": "",
            "error": f"无法启动 Codex: {exc}",
            "stderr": "",
            "elapsed": time.monotonic() - monotonic_start,
            "ttft": None,
            "input_tokens": None,
            "output_tokens": None,
            "reasoning_tokens": None,
            "cache_read_tokens": None,
            "cache_write_tokens": None,
            "cost_usd": None,
            "total_tokens": None,
            "usage_complete": False,
            "step_count": 0,
            "tool_errors": 0,
            "tool_call_count": 0,
            "item_error_count": 0,
            "fatal_error_count": 1,
            "command_failures": [],
            "permission_denials": [],
            "session_id": None,
            "raw_events": [],
            "normalized_events": [],
            "event_counts": {},
            "prompt_transport": "stdin_file",
            "prompt_path": str(stdin_path),
            "final_message_path": str(final_message_path),
            "requested_model": requested_model,
            "configured_model": configured_model,
            "observed_models": [],
            "unexpected_models": [],
            "model_verification": "not_run",
            "model_metadata_warnings": [],
            "approval_policy": approval_policy,
            "sandbox_mode": sandbox_mode,
            "windows_sandbox": windows_sandbox,
        }

    if proc.stdout is None or proc.stderr is None:
        kill_proc_tree(proc)
        raise RuntimeError("Codex stdout/stderr pipe creation failed")

    output_queue: "queue.Queue[tuple[str, Optional[str]]]" = queue.Queue()
    threads = [
        threading.Thread(
            target=_stream_reader,
            args=(proc.stdout, "stdout", output_queue),
            daemon=True,
        ),
        threading.Thread(
            target=_stream_reader,
            args=(proc.stderr, "stderr", output_queue),
            daemon=True,
        ),
    ]
    for thread in threads:
        thread.start()

    open_streams = {"stdout", "stderr"}

    with (
        raw_stdout_path.open("w", encoding="utf-8", newline="") as raw_file,
        stderr_path.open("w", encoding="utf-8", newline="") as err_file,
    ):
        while open_streams or proc.poll() is None:
            if time.monotonic() > deadline:
                timed_out = True
                status = "timeout"
                fatal_errors.append(f"超时 ({timeout:.0f}s)")
                kill_proc_tree(proc)
                break

            try:
                kind, line = output_queue.get(timeout=0.2)
            except queue.Empty:
                continue

            if line is None:
                open_streams.discard(kind)
                continue

            if kind == "stderr":
                stderr_lines.append(line)
                err_file.write(line)
                err_file.flush()
                continue

            stdout_lines.append(line)
            raw_file.write(line)
            raw_file.flush()

            stripped = line.strip()
            if not stripped:
                continue

            try:
                event = json.loads(stripped)
            except json.JSONDecodeError:
                event_counts["non_json_stdout"] += 1
                continue

            if not isinstance(event, dict):
                event_counts["non_object_json"] += 1
                continue

            raw_events.append(event)
            event_type = str(event.get("type") or "unknown")
            event_counts[event_type] += 1

            if event_type == "thread.started" and event.get("thread_id"):
                session_id = str(event.get("thread_id"))
            elif event_type == "turn.started":
                turn_started = True
            elif event_type == "turn.completed":
                turn_completed = True

            normalized_events.extend(normalize_codex_event(event))

            visible_text = codex_visible_text(event)
            if visible_text:
                agent_message_count += 1
                step_count += 1
                if ttft is None:
                    ttft = time.monotonic() - monotonic_start
                output_chunks.append(visible_text)

            item = codex_item(event)
            item_type = str(item.get("type") or "")
            if event_type == "item.started" and codex_tool_item(item):
                tool_call_count += 1

            if event_type == "item.completed" and codex_tool_item(item):
                item_status = str(item.get("status") or "")
                exit_code = item.get("exit_code")
                failed = (
                    item_status in {"failed", "declined", "error"}
                    or (exit_code is not None and _to_int(exit_code) != 0)
                )
                if failed:
                    tool_errors += 1
                    failure = {
                        "item_id": str(item.get("id") or ""),
                        "item_type": item_type,
                        "status": item_status,
                        "exit_code": exit_code,
                        "command": str(item.get("command") or ""),
                        "output": codex_tool_result_text(item),
                    }
                    command_failures.append(failure)
                    if item_status == "declined":
                        permission_denials.append(failure)

            item_error = codex_item_error(event)
            if item_error:
                item_errors.append(item_error)
                if "Model metadata for" in item_error:
                    model_metadata_warnings.append(item_error)

            fatal = codex_turn_failure(event)
            if fatal:
                fatal_errors.append(fatal)
                status = "error"

            usage = codex_usage(event)
            if usage is not None:
                usage_seen = bool(usage)
                input_tokens = usage.get("input_tokens", 0)
                output_tokens = usage.get("output_tokens", 0)
                reasoning_tokens = usage.get("reasoning_tokens", 0)
                cache_read_tokens = usage.get("cache_read_tokens", 0)
                cache_write_tokens = usage.get("cache_write_tokens", 0)

    try:
        proc.wait(timeout=15)
    except subprocess.TimeoutExpired:
        kill_proc_tree(proc)
        status = "timeout"
        timed_out = True
        fatal_errors.append("进程未正常退出，已强制终止")

    for thread in threads:
        thread.join(timeout=1)

    returncode = proc.returncode
    stderr = "".join(stderr_lines).strip()

    if returncode not in (0, None):
        status = "error"
        fatal_errors.append(stderr or f"Codex 退出码 {returncode}")

    if status == "success" and not turn_started:
        status = "error"
        fatal_errors.append("Codex 未输出 turn.started 事件")
    if status == "success" and not turn_completed:
        status = "error"
        fatal_errors.append("Codex 未输出 turn.completed 事件")

    final_message = ""
    if final_message_path.exists():
        try:
            final_message = final_message_path.read_text(
                encoding="utf-8-sig",
                errors="replace",
            ).strip()
        except OSError:
            final_message = ""

    output = "\n".join(
        chunk.strip() for chunk in output_chunks if chunk.strip()
    ).strip()
    if final_message:
        output = final_message

    if status == "success" and not output:
        status = "error"
        fatal_errors.append("Codex turn 完成但没有最终 assistant message")

    unexpected_models: list[str] = []
    if configured_model and configured_model != requested_model:
        unexpected_models.append(configured_model)
        if status == "success":
            status = "model_mismatch"
            fatal_errors.append(
                f"配置模型 {configured_model} 与请求模型 {requested_model} 不一致"
            )

    success = (
        status == "success"
        and not timed_out
        and returncode in (0, None)
        and turn_completed
        and bool(output)
        and not unexpected_models
    )

    elapsed = time.monotonic() - monotonic_start
    return {
        "status": status,
        "success": success,
        "returncode": returncode,
        "output": output,
        "error": " | ".join(dict.fromkeys(x for x in fatal_errors if x)),
        "stderr": stderr,
        "elapsed": elapsed,
        "ttft": ttft,
        "input_tokens": input_tokens if usage_seen else None,
        "output_tokens": output_tokens if usage_seen else None,
        "reasoning_tokens": reasoning_tokens if usage_seen else None,
        "cache_read_tokens": cache_read_tokens if usage_seen else None,
        "cache_write_tokens": cache_write_tokens if usage_seen else None,
        "cost_usd": None,
        "total_tokens": (
            input_tokens + output_tokens if usage_seen else None
        ),
        "usage_complete": usage_seen,
        "step_count": step_count,
        "tool_errors": tool_errors,
        "tool_call_count": tool_call_count,
        "item_error_count": len(item_errors),
        "fatal_error_count": len(fatal_errors),
        "item_errors": item_errors,
        "command_failures": command_failures,
        "permission_denials": permission_denials,
        "session_id": session_id,
        "raw_events": raw_events,
        "normalized_events": normalized_events,
        "event_counts": dict(event_counts),
        "stdout_lines": stdout_lines,
        "prompt_transport": "stdin_file",
        "prompt_path": str(stdin_path),
        "final_message_path": str(final_message_path),
        "requested_model": requested_model,
        "configured_model": configured_model,
        "observed_models": [],
        "unexpected_models": unexpected_models,
        "model_verification": "argv+isolated_config",
        "model_metadata_warnings": model_metadata_warnings,
        "approval_policy": approval_policy,
        "sandbox_mode": sandbox_mode,
        "windows_sandbox": windows_sandbox,
        "agent_message_count": agent_message_count,
    }


# ---------------------------------------------------------------------------
# PinchBench official/default grading
# ---------------------------------------------------------------------------

def load_pinchbench_grading(skill_dir: Path) -> tuple[Optional[Any], Optional[str]]:
    """Load the grading engine from the exact checked-out PinchBench commit."""
    scripts_dir = (skill_dir / "scripts").resolve()
    grading_path = scripts_dir / "lib_grading.py"
    if not grading_path.exists():
        return None, f"找不到 PinchBench grading engine: {grading_path}"

    scripts_value = str(scripts_dir)
    if scripts_value not in sys.path:
        sys.path.insert(0, scripts_value)

    try:
        import lib_grading  # type: ignore
    except Exception as exc:
        return None, (
            "无法导入 PinchBench scripts/lib_grading.py。"
            "请安装本地仓库依赖，例如："
            r"C:\pinchbench-opencode\.venv\Scripts\python.exe -m pip install -e "
            f'"{skill_dir}"。原始错误: {exc}'
        )

    return lib_grading, None


def grade_with_pinchbench_default(
    *,
    grader_module: Any,
    task: Task,
    execution_result: dict[str, Any],
    workspace: Path,
    skill_dir: Path,
    judge_timeout: float,
    verbose: bool,
) -> GradeResult:
    """Use PinchBench's own grader with its checked-out default judge model.

    The judge is called directly through the API backend.  For the current
    commit, DEFAULT_JUDGE_MODEL is openrouter/anthropic/claude-haiku-4.5 and
    lib_agent.py reads OPENROUTER_API_KEY.
    """
    grading_execution = dict(execution_result)
    grading_execution["workspace"] = str(workspace)

    try:
        result = grader_module.grade_task(
            task=task,
            execution_result=grading_execution,
            skill_dir=skill_dir,
            judge_model=getattr(
                grader_module,
                "DEFAULT_JUDGE_MODEL",
                PINCHBENCH_DEFAULT_JUDGE_MODEL,
            ),
            judge_timeout_seconds=judge_timeout,
            judge_backend="api",
            verbose=verbose,
        )

        raw_score = getattr(result, "score", None)
        max_score = getattr(result, "max_score", 1.0)
        score: Optional[float]
        if raw_score is None:
            score = None
        else:
            score = float(raw_score)
            max_score_float = float(max_score or 1.0)
            if max_score_float != 1.0:
                score = score / max_score_float
            score = max(0.0, min(1.0, score))

        notes = str(getattr(result, "notes", "") or "")
        grade_error: Optional[str] = None
        if task.grading_type in {"llm_judge", "hybrid"}:
            lower_notes = notes.lower()
            failure_markers = (
                "llm judge failed",
                "no parseable response",
                "response parsed but no score",
                "openrouter_api_key not set",
                "judge api call failed",
            )
            if any(marker in lower_notes for marker in failure_markers):
                grade_error = notes or "LLM judge failed"

        return GradeResult(
            score=score,
            grading_type=str(getattr(result, "grading_type", task.grading_type)),
            breakdown=dict(getattr(result, "breakdown", {}) or {}),
            notes=notes,
            error=grade_error,
        )
    except Exception as exc:
        return GradeResult(
            score=None,
            grading_type=task.grading_type,
            breakdown={},
            notes="",
            error=(
                "PinchBench 官方评分失败: "
                f"{exc}\n{traceback.format_exc(limit=8)}"
            ),
        )


POST_COMPLETION_STREAM_DISCONNECT_RE = re.compile(
    r"^turn 1 \(single\): Reconnecting\.\.\. \d+/\d+ "
    r"\(stream disconnected before completion: "
    r"stream closed before response\.completed\)$"
)


def maybe_recover_post_completion_stream_disconnect(
    execution: dict[str, Any],
) -> bool:
    """Reclassify one narrowly defined post-completion SSE disconnect."""
    if execution.get("success") is not False:
        return False
    if execution.get("status") != "error":
        return False
    if int(execution.get("session_count") or 0) != 1:
        return False

    original_error = str(execution.get("error") or "").strip()
    if POST_COMPLETION_STREAM_DISCONNECT_RE.fullmatch(original_error) is None:
        return False

    if not str(execution.get("output") or "").strip():
        return False
    if not bool(execution.get("usage_complete")):
        return False
    if execution.get("unexpected_models"):
        return False
    if execution.get("model_metadata_warnings"):
        return False
    if execution.get("encoding_normalization_errors"):
        return False
    if str(execution.get("grade_error") or "").strip():
        return False

    score_value = execution.get("score")
    if score_value is None:
        return False
    try:
        score_number = float(score_value)
    except (TypeError, ValueError):
        return False
    if not 0.0 <= score_number <= 1.0:
        return False

    event_counts = execution.get("event_counts") or {}
    if int(event_counts.get("turn.started") or 0) < 1:
        return False
    if int(event_counts.get("turn.completed") or 0) != 0:
        return False

    turns = execution.get("turn_results") or []
    if len(turns) != 1:
        return False
    turn = turns[0]
    if turn.get("success") is not False or turn.get("status") != "error":
        return False
    if str(turn.get("error") or "").strip() not in original_error:
        return False

    final_message_path_text = str(
        turn.get("final_message_path") or ""
    ).strip()
    if not final_message_path_text:
        return False
    final_message_path = Path(final_message_path_text)
    try:
        if (
            not final_message_path.is_file()
            or final_message_path.stat().st_size <= 0
        ):
            return False
    except OSError:
        return False

    workspace_text = str(execution.get("workspace") or "").strip()
    changed_files = [
        str(item)
        for item in (execution.get("changed_workspace_files") or [])
        if str(item).strip()
    ]
    if not workspace_text or not changed_files:
        return False

    workspace = Path(workspace_text)
    artifact_files: list[str] = []
    for relative_text in changed_files:
        candidate = workspace / relative_text
        try:
            if candidate.is_file() and candidate.stat().st_size > 0:
                artifact_files.append(relative_text)
        except OSError:
            continue
    if not artifact_files:
        return False

    original_status = execution.get("status")
    original_success = execution.get("success")
    original_returncode = execution.get("returncode")
    original_fatal_error_count = int(
        execution.get("fatal_error_count") or 0
    )

    warning = {
        "type": "post_completion_stream_disconnect",
        "classification": "recovered_after_successful_grading",
        "original_status": original_status,
        "original_success": original_success,
        "original_returncode": original_returncode,
        "original_error": original_error,
        "original_fatal_error_count": original_fatal_error_count,
        "score": score_number,
        "final_message_path": final_message_path_text,
        "changed_workspace_files": changed_files,
        "verified_artifact_files": artifact_files,
        "policy": (
            "single-turn only; exact stream-disconnect signature; "
            "final assistant message present; changed non-empty artifact "
            "present; usage complete; official grading succeeded; no model, "
            "metadata, or encoding error"
        ),
    }

    execution["post_completion_transport_recovered"] = True
    execution["transport_warning_count"] = 1
    execution["transport_warnings"] = [warning]
    execution["original_status"] = original_status
    execution["original_success"] = original_success
    execution["original_returncode"] = original_returncode
    execution["original_error"] = original_error
    execution["original_fatal_error_count"] = (
        original_fatal_error_count
    )
    execution["status"] = "success"
    execution["success"] = True
    execution["error"] = ""
    execution["fatal_error_count"] = max(
        0,
        original_fatal_error_count - 1,
    )
    return True

# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def truncate_excel(value: Any, limit: int = 32000) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    return text if len(text) <= limit else text[:limit] + "...[truncated]"


def save_csv(results: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "task_id", "name", "category", "grading_type", "network_task",
        "multi_session", "session_count", "success", "status", "returncode",
        "score", "elapsed", "ttft", "input_tokens", "output_tokens",
        "reasoning_tokens", "cache_read_tokens", "cache_write_tokens",
        "cost_usd", "total_tokens", "usage_complete", "step_count", "tool_errors",
        "tool_call_count", "item_error_count", "fatal_error_count",
        "observed_models", "unexpected_models", "model_verification",
        "codex_versions", "approval_policies", "sandbox_modes", "windows_sandboxes",
        "permission_denials", "command_failures",
        "post_completion_transport_recovered", "transport_warning_count",
        "transport_warnings", "original_status", "original_success",
        "original_returncode", "original_error",
        "original_fatal_error_count", "changed_workspace_files",
        "encoding_normalized_count", "encoding_normalized_files",
        "encoding_normalization_errors", "encoding_normalization_audit",
        "workspace", "transcript", "error", "grade_error", "grade_notes",
    ]

    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for row in results:
            writer.writerow({
                key: truncate_excel(row.get(key), 10000)
                for key in fields
            })


def save_xlsx(
    results: list[dict[str, Any]],
    summary: dict[str, Any],
    output_path: Path,
) -> None:
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
    except ImportError:
        LOGGER.warning(
            "未安装 openpyxl，跳过 XLSX。可运行: "
            "python -m pip install openpyxl"
        )
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "详细结果"

    headers = [
        "任务ID", "名称", "类别", "打分类型", "联网题", "多轮任务",
        "会话调用数", "成功", "状态", "退出码", "分数", "耗时(s)",
        "TTFT估计(s)", "输入Token", "输出Token", "推理Token",
        "缓存读取Token", "缓存写入Token", "费用USD", "总Token", "Usage完整",
        "Step数", "工具错误数", "工具调用数", "Item错误数", "致命错误数",
        "实际模型", "非目标模型", "模型验证", "Codex版本", "审批策略",
        "沙箱模式", "Windows沙箱", "权限拒绝", "命令失败明细",
        "编码规范化文件数", "编码规范化文件", "编码规范化错误", "编码审计目录",
        "Workspace", "Transcript", "运行错误", "打分错误", "打分备注",
    ]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(results, start=2):
        values = [
            row.get("task_id"),
            row.get("name"),
            row.get("category"),
            row.get("grading_type"),
            "是" if row.get("network_task") else "否",
            "是" if row.get("multi_session") else "否",
            row.get("session_count"),
            "✓" if row.get("success") else "✗",
            row.get("status"),
            row.get("returncode") if row.get("returncode") is not None else "",
            row.get("score") if row.get("score") is not None else "",
            row.get("elapsed"),
            row.get("ttft") if row.get("ttft") is not None else "",
            row.get("input_tokens") if row.get("input_tokens") is not None else "",
            row.get("output_tokens") if row.get("output_tokens") is not None else "",
            row.get("reasoning_tokens") if row.get("reasoning_tokens") is not None else "",
            row.get("cache_read_tokens") if row.get("cache_read_tokens") is not None else "",
            row.get("cache_write_tokens") if row.get("cache_write_tokens") is not None else "",
            row.get("cost_usd") if row.get("cost_usd") is not None else "",
            row.get("total_tokens") if row.get("total_tokens") is not None else "",
            "是" if row.get("usage_complete") else "否",
            row.get("step_count"),
            row.get("tool_errors"),
            row.get("tool_call_count"),
            row.get("item_error_count"),
            row.get("fatal_error_count"),
            json.dumps(row.get("observed_models") or [], ensure_ascii=False),
            json.dumps(row.get("unexpected_models") or [], ensure_ascii=False),
            row.get("model_verification") or "",
            json.dumps(row.get("codex_versions") or [], ensure_ascii=False),
            json.dumps(row.get("approval_policies") or [], ensure_ascii=False),
            json.dumps(row.get("sandbox_modes") or [], ensure_ascii=False),
            json.dumps(row.get("windows_sandboxes") or [], ensure_ascii=False),
            json.dumps(row.get("permission_denials") or [], ensure_ascii=False),
            json.dumps(row.get("command_failures") or [], ensure_ascii=False),
            row.get("encoding_normalized_count") or 0,
            json.dumps(row.get("encoding_normalized_files") or [], ensure_ascii=False),
            json.dumps(row.get("encoding_normalization_errors") or [], ensure_ascii=False),
            row.get("encoding_normalization_audit") or "",
            row.get("workspace"),
            row.get("transcript"),
            row.get("error") or "",
            row.get("grade_error") or "",
            row.get("grade_notes") or "",
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(
                row=row_index,
                column=column,
                value=truncate_excel(value),
            )

    for cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in cells)
        worksheet.column_dimensions[cells[0].column_letter].width = min(
            max_length + 3,
            60,
        )

    summary_sheet = workbook.create_sheet("汇总")
    for row_index, (key, value) in enumerate(summary.items(), start=1):
        summary_sheet.cell(row=row_index, column=1, value=key).font = Font(bold=True)
        summary_sheet.cell(row=row_index, column=2, value=value)
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 30

    workbook.save(output_path)


def optional_sum(results: list[dict[str, Any]], field: str) -> Optional[float]:
    values = [row.get(field) for row in results if row.get(field) is not None]
    if not values:
        return None
    return float(sum(float(value) for value in values))


def print_summary(
    results: list[dict[str, Any]],
    total_elapsed: float,
) -> dict[str, Any]:
    succeeded = sum(1 for row in results if row.get("success"))
    failed = len(results) - succeeded
    scores = [
        float(row["score"])
        for row in results
        if row.get("score") is not None
    ]
    ttfts = [
        float(row["ttft"])
        for row in results
        if row.get("ttft") is not None
    ]
    average_elapsed = (
        sum(float(row.get("elapsed") or 0.0) for row in results) / len(results)
        if results else 0.0
    )

    summary = {
        "任务总数": len(results),
        "成功": succeeded,
        "失败": failed,
        "成功率": round(succeeded / len(results), 4) if results else 0.0,
        "有分数任务数": len(scores),
        "打分失败任务数": sum(1 for row in results if row.get("grade_error")),
        "平均分数": round(sum(scores) / len(scores), 4) if scores else None,
        "总耗时(s)": round(total_elapsed, 2),
        "平均耗时/任务(s)": round(average_elapsed, 2),
        "平均TTFT估计(s)": round(sum(ttfts) / len(ttfts), 4) if ttfts else None,
        "总输入Token": int(optional_sum(results, "input_tokens") or 0)
            if any(row.get("input_tokens") is not None for row in results) else None,
        "总输出Token": int(optional_sum(results, "output_tokens") or 0)
            if any(row.get("output_tokens") is not None for row in results) else None,
        "总推理Token": int(optional_sum(results, "reasoning_tokens") or 0)
            if any(row.get("reasoning_tokens") is not None for row in results) else None,
        "总费用USD": round(optional_sum(results, "cost_usd") or 0.0, 6)
            if any(row.get("cost_usd") is not None for row in results) else None,
        "Usage缺失任务数": sum(
            1 for row in results if not row.get("usage_complete")
        ),
        "模型不一致任务数": sum(
            1 for row in results if row.get("unexpected_models")
        ),
        "权限拒绝任务数": sum(
            1 for row in results if row.get("permission_denials")
        ),
        "有可恢复工具错误任务数": sum(
            1 for row in results if int(row.get("tool_errors") or 0) > 0
        ),
        "有Codex Item警告任务数": sum(
            1 for row in results if int(row.get("item_error_count") or 0) > 0
        ),
        "发生编码规范化任务数": sum(
            1 for row in results if int(row.get("encoding_normalized_count") or 0) > 0
        ),
        "完成后传输警告任务数": sum(
            1
            for row in results
            if row.get("post_completion_transport_recovered")
        ),
        "编码规范化文件总数": sum(
            int(row.get("encoding_normalized_count") or 0) for row in results
        ),
        "联网任务数": sum(1 for row in results if row.get("network_task")),
        "多轮任务数": sum(1 for row in results if row.get("multi_session")),
    }

    print("\n" + "=" * 100)
    print("汇总")
    print("=" * 100)
    for key, value in summary.items():
        print(f"{key:<22}: {value}")

    print("-" * 100)
    for row in results:
        marker = "✓" if row.get("success") else "✗"
        score = (
            f"{row['score']:.3f}"
            if row.get("score") is not None
            else "N/A"
        )
        ttft = (
            f"{row['ttft']:.2f}s"
            if row.get("ttft") is not None
            else "N/A"
        )
        error = row.get("error") or row.get("grade_error") or ""
        error_suffix = f"  [{str(error)[:70]}]" if error else ""
        print(
            f"{marker} {row['task_id']:<42} "
            f"score={score:<6} elapsed={row['elapsed']:.1f}s "
            f"ttft={ttft:<8}{error_suffix}"
        )
    print("=" * 100)
    return summary


# ---------------------------------------------------------------------------
# Preflight and run metadata
# ---------------------------------------------------------------------------

def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def command_output(command: list[str], cwd: Optional[Path] = None) -> str:
    try:
        result = subprocess.run(
            _spawn_command(command),
            cwd=str(cwd) if cwd else None,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
            check=False,
        )
        return (result.stdout or result.stderr).strip().replace("\x00", "")
    except Exception as exc:
        return f"unavailable: {exc}"


def load_toml_file(path: Path) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    try:
        with path.open("rb") as file:
            value = tomllib.load(file)
        if not isinstance(value, dict):
            return None, f"TOML 根节点不是对象: {path}"
        return value, None
    except Exception as exc:
        return None, f"无法解析 {path}: {exc}"


def nested_get(value: dict[str, Any], *keys: str) -> Any:
    current: Any = value
    for key in keys:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def codex_config_snapshot(config: dict[str, Any]) -> dict[str, Any]:
    provider = nested_get(config, "model_providers", DEFAULT_PROVIDER) or {}
    return {
        "model": config.get("model"),
        "model_provider": config.get("model_provider"),
        "approval_policy": config.get("approval_policy"),
        "sandbox_mode": config.get("sandbox_mode"),
        "web_search": config.get("web_search"),
        "sandbox_network_access": nested_get(config, "sandbox_workspace_write", "network_access"),
        "provider_name": provider.get("name") if isinstance(provider, dict) else None,
        "provider_base_url": provider.get("base_url") if isinstance(provider, dict) else None,
        "provider_env_key": provider.get("env_key") if isinstance(provider, dict) else None,
        "wire_api": provider.get("wire_api") if isinstance(provider, dict) else None,
        "supports_websockets": provider.get("supports_websockets") if isinstance(provider, dict) else None,
        "windows_sandbox": nested_get(config, "windows", "sandbox"),
        "history_persistence": nested_get(config, "history", "persistence"),
        "memories_generate": nested_get(config, "memories", "generate_memories"),
        "memories_use": nested_get(config, "memories", "use_memories"),
        "analytics_enabled": nested_get(config, "analytics", "enabled"),
        "feedback_enabled": nested_get(config, "feedback", "enabled"),
        "hooks_enabled": nested_get(config, "features", "hooks"),
        "remote_plugin_enabled": nested_get(config, "features", "remote_plugin"),
        "mcp_server_count": len(config.get("mcp_servers") or {}),
    }


def validate_codex_config(config: dict[str, Any], args: argparse.Namespace) -> list[str]:
    failures: list[str] = []
    snapshot = codex_config_snapshot(config)
    expected = {
        "model": args.model,
        "model_provider": DEFAULT_PROVIDER,
        "approval_policy": args.approval_policy,
        "sandbox_mode": args.sandbox_mode,
        "sandbox_network_access": True,
        "provider_env_key": "OPENROUTER_API_KEY",
        "wire_api": DEFAULT_WIRE_API,
        "windows_sandbox": args.windows_sandbox,
    }
    for key, expected_value in expected.items():
        actual = snapshot.get(key)
        if actual != expected_value:
            failures.append(f"config {key} 不一致：预期 {expected_value!r}，实际 {actual!r}")

    if str(snapshot.get("provider_base_url") or "").rstrip("/") != "https://openrouter.ai/api/v1":
        failures.append(f"OpenRouter base_url 不一致：{snapshot.get('provider_base_url')!r}")
    if snapshot.get("memories_generate") is not False:
        failures.append("memories.generate_memories 必须为 false")
    if snapshot.get("memories_use") is not False:
        failures.append("memories.use_memories 必须为 false")
    if snapshot.get("hooks_enabled") not in {False, None}:
        failures.append("features.hooks 必须关闭")
    if snapshot.get("remote_plugin_enabled") not in {False, None}:
        failures.append("features.remote_plugin 必须关闭")
    if int(snapshot.get("mcp_server_count") or 0) != 0:
        failures.append("正式可比测试不能加载 MCP servers")
    return failures


def write_run_config(
    output_path: Path,
    args: argparse.Namespace,
    skill_dir: Path,
    tasks_dir: Path,
    codex_command: str,
    codex_home: Path,
    codex_config_path: Path,
    codex_config: dict[str, Any],
    selected: list[Task],
    judge_model: str,
) -> None:
    manifest_path = tasks_dir / "manifest.yaml"
    git_command = resolve_command("git")
    commit = command_output([git_command, "rev-parse", "HEAD"], cwd=skill_dir) if git_command else ""
    config = {
        "runner_revision": RUNNER_REVISION,
        "created_at": dt.datetime.now(dt.timezone.utc).astimezone().isoformat(),
        "platform": platform.platform(),
        "python": sys.version,
        "codex_command": codex_command,
        "codex_version": command_output([codex_command, "--version"]),
        "expected_codex_version": args.expected_codex_version,
        "model": args.model,
        "model_provider": DEFAULT_PROVIDER,
        "wire_api": DEFAULT_WIRE_API,
        "agent": "Codex built-in default agent/tools with isolated CODEX_HOME",
        "variant": "default agent; no custom system prompt, rules, memory, MCP or skills",
        "approval_policy": args.approval_policy,
        "sandbox_mode": args.sandbox_mode,
        "windows_sandbox": args.windows_sandbox,
        "ignore_rules": not args.no_ignore_rules,
        "prompt_transport": "stdin_file_utf8",
        "prompt_in_argv": False,
        "output_format": "jsonl",
        "final_message_capture": "codex --output-last-message",
        "openrouter_routing": "default; upstream provider not pinned",
        "worker_count": 1,
        "task_concurrency": 1,
        "skill_dir": str(skill_dir),
        "tasks_dir": str(tasks_dir),
        "codex_home": str(codex_home),
        "codex_config_path": str(codex_config_path),
        "codex_config_sha256": sha256_file(codex_config_path),
        "codex_config_non_secret": codex_config_snapshot(codex_config),
        "pinchbench_commit": commit,
        "manifest_sha256": sha256_file(manifest_path) if manifest_path.exists() else None,
        "suite": args.suite,
        "limit": args.limit,
        "default_skipped_tasks": sorted(DEFAULT_SKIPPED_TASKS),
        "additional_skip": args.skip,
        "skip_network": args.skip_network,
        "task_count": len(selected),
        "task_ids": [task.task_id for task in selected],
        "timeout_multiplier": args.timeout_multiplier,
        "network_timeout": args.network_timeout,
        "workspace_instruction": not args.no_workspace_instruction,
        "encoding_normalization_enabled": not args.no_encoding_normalization,
        "encoding_normalization_policy": (
            "changed/new text artifacts only; BOM and strict-roundtrip Windows ANSI text converted to UTF-8 without BOM; "
            "original bytes retained in transcripts"
        ),
        "grading_enabled": not args.no_grade,
        "grading_engine": str(skill_dir / "scripts" / "lib_grading.py"),
        "judge_backend": "api",
        "judge_model": judge_model,
        "judge_key_env": "OPENROUTER_API_KEY",
        "judge_key_present": bool(os.environ.get("OPENROUTER_API_KEY")),
        "judge_concurrency": 1,
        "judge_is_separate_from_tested_model": True,
        "environment_proxy_present": {
            key: bool(os.environ.get(key))
            for key in ("HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "NO_PROXY")
        },
        "model_identity_note": (
            "Codex JSONL does not reliably expose a runtime model field. "
            "The model is pinned in both isolated config.toml and --model argv."
        ),
        "comparability_note": (
            "Tasks/grader/timeouts/task order/skips/prompt transport/workspace instruction "
            "are aligned with the validated Windows Qwen Code runner. Codex sandbox and "
            "approval settings are explicit adapter differences required for unattended "
            "native-Windows execution. The same pre-grading Windows text-encoding "
            "normalization baseline is applied as in the prior comparison runs."
        ),
    }
    output_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")


def print_preflight(
    selected: list[Task],
    skill_dir: Path,
    tasks_dir: Path,
    codex_command: str,
    codex_home: Path,
    codex_config_path: Path,
    codex_config: Optional[dict[str, Any]],
    codex_config_error: Optional[str],
    grader_module: Optional[Any],
    grader_error: Optional[str],
    args: argparse.Namespace,
) -> tuple[dict[str, list[str]], list[str], list[str]]:
    prerequisite_failures: dict[str, list[str]] = {}
    fixture_failures: list[str] = []
    codex_failures: list[str] = []
    needs_judge = not args.no_grade and any(t.grading_type in {"llm_judge", "hybrid"} for t in selected)
    judge_model = str(getattr(grader_module, "DEFAULT_JUDGE_MODEL", PINCHBENCH_DEFAULT_JUDGE_MODEL)) if grader_module is not None else PINCHBENCH_DEFAULT_JUDGE_MODEL
    judge_key_present = bool(os.environ.get("OPENROUTER_API_KEY"))

    version_output = command_output([codex_command, "--version"])
    match = re.search(r"\d+\.\d+\.\d+(?:[-+][^\s]+)?", version_output)
    version = match.group(0) if match else version_output.strip()
    if args.expected_codex_version and version != args.expected_codex_version:
        codex_failures.append(f"Codex 版本不一致：预期 {args.expected_codex_version}，实际 {version_output}")
    if not codex_home.exists():
        codex_failures.append(f"CODEX_HOME 不存在: {codex_home}")
    if codex_config_error:
        codex_failures.append(codex_config_error)
    elif codex_config is not None:
        codex_failures.extend(validate_codex_config(codex_config, args))
    for custom_name in ("AGENTS.md", ".rules"):
        custom_path = codex_home / custom_name
        if custom_path.exists():
            codex_failures.append(f"隔离 CODEX_HOME 中存在自定义文件，正式测试前移走: {custom_path}")

    print("=" * 100)
    print("PinchBench Codex Windows runner preflight")
    print("=" * 100)
    print(f"Runner revision      : {RUNNER_REVISION}")
    print(f"Skill dir            : {skill_dir}")
    print(f"Tasks dir            : {tasks_dir}")
    print(f"Codex command        : {codex_command}")
    print(f"Codex version        : {version_output}")
    print(f"Expected version     : {args.expected_codex_version}")
    print(f"Model                : {args.model}")
    print(f"Provider / wire API  : {DEFAULT_PROVIDER} / {DEFAULT_WIRE_API}")
    print(f"Approval policy      : {args.approval_policy}")
    print(f"Sandbox mode         : {args.sandbox_mode}")
    print(f"Windows sandbox      : {args.windows_sandbox}")
    print(f"Ignore rules         : {'yes' if not args.no_ignore_rules else 'no'}")
    print(f"CODEX_HOME           : {codex_home}")
    print(f"Config               : {codex_config_path}")
    print(f"Suite                : {args.suite}")
    print(f"Selected tasks       : {len(selected)}")
    print(f"Default skipped      : {len(DEFAULT_SKIPPED_TASKS)}")
    for task_id in sorted(DEFAULT_SKIPPED_TASKS):
        print(f"  - {task_id}")
    print("Worker/concurrency   : 1 / 1")
    print("Judge concurrency    : 1 (synchronous)")
    print("Built-in agent/tools : yes; no runner-added tools or system prompt")
    print("Prompt transport     : UTF-8 stdin file (never argv)")
    print(f"Encoding normalize   : {'disabled' if args.no_encoding_normalization else 'enabled (changed/new text artifacts only)'}")
    print("Output transport     : Codex JSONL stdout; stderr separate")
    print("Model verification   : --model argv + isolated config.toml")
    print(f"Grading engine       : {skill_dir / 'scripts' / 'lib_grading.py'}")
    print("Judge backend        : api")
    print(f"Judge model          : {judge_model}")
    print(f"OPENROUTER_API_KEY   : {'set' if judge_key_present else 'missing'}")
    print(f"Grader import        : {'ok' if grader_error is None else 'FAILED'}")
    if grader_error:
        print(f"  {grader_error}")
    print()

    category_counts = Counter(task.category or "(uncategorized)" for task in selected)
    grading_counts = Counter(task.grading_type for task in selected)
    print("Grading types:")
    for key, value in sorted(grading_counts.items()):
        print(f"  {key:<24} {value}")
    print("Categories:")
    for key, value in sorted(category_counts.items()):
        print(f"  {key:<24} {value}")
    print(f"Network-marked tasks : {sum(1 for task in selected if is_network_task(task))}")
    print(f"Multi-session tasks  : {sum(1 for task in selected if task.multi_session)}")
    print(f"Default judge needed : {'yes' if needs_judge else 'no'}")
    if needs_judge and not judge_key_present:
        print("Judge readiness      : FAILED (OPENROUTER_API_KEY missing)")
    elif needs_judge:
        print("Judge readiness      : ok")
    else:
        print("Judge readiness      : not required")
    print()

    for task in selected:
        prerequisites = task.metadata.get("prerequisites") or []
        missing = check_prerequisites(list(prerequisites)) if prerequisites else []
        if missing:
            prerequisite_failures[task.task_id] = missing
        for workspace_file in task.workspace_files:
            if not isinstance(workspace_file, dict):
                fixture_failures.append(f"{task.task_id}: invalid workspace file entry")
                continue
            if workspace_file.get("content") is not None:
                continue
            source = workspace_file.get("source")
            if not source:
                continue
            src_rel = Path(str(source))
            candidates = [
                skill_dir / "assets" / src_rel,
                skill_dir / src_rel,
                task.file_path.parent / src_rel,
                task.file_path.parent.parent / "assets" / src_rel,
            ]
            if not any(candidate.exists() for candidate in candidates):
                fixture_failures.append(f"{task.task_id}: missing fixture {source}")

    if prerequisite_failures:
        print("Missing declared prerequisites:")
        for task_id, missing in prerequisite_failures.items():
            print(f"  {task_id}: {', '.join(missing)}")
    else:
        print("Missing declared prerequisites: none")
    if fixture_failures:
        print("Missing workspace fixtures:")
        for item in fixture_failures:
            print(f"  {item}")
    else:
        print("Missing workspace fixtures: none")
    if codex_failures:
        print("Codex environment failures:")
        for item in codex_failures:
            print(f"  {item}")
    else:
        print("Codex environment failures: none")
    print("=" * 100)
    return prerequisite_failures, fixture_failures, codex_failures


# ---------------------------------------------------------------------------
# Main execution
# ---------------------------------------------------------------------------

def build_turns(task: Task) -> list[dict[str, Any]]:
    """Use the same fresh/resume semantics as the validated Qwen runner."""
    if task.multi_session and task.sessions:
        turns: list[dict[str, Any]] = []
        for index, item in enumerate(task.sessions, start=1):
            prompt = str(item.get("prompt") or "").strip()
            if not prompt:
                continue
            is_first_executable_turn = not turns
            new_session = True if is_first_executable_turn else bool(item.get("new_session"))
            turns.append({
                "id": str(item.get("id") or f"turn_{index}"),
                "prompt": prompt,
                "new_session": new_session,
            })
        if turns:
            return turns
    return [{"id": "single", "prompt": task.prompt, "new_session": True}]


def aggregate_optional_int(current: Optional[int], value: Optional[int]) -> Optional[int]:
    if value is None:
        return current
    return (current or 0) + int(value)


def execute_task(
    task: Task,
    task_index: int,
    task_count: int,
    skill_dir: Path,
    workspace: Path,
    transcript_dir: Path,
    codex_command: str,
    configured_model: str,
    args: argparse.Namespace,
) -> dict[str, Any]:
    print(f"[{task_index}/{task_count}] {task.task_id} — {task.name}")
    prompt_preview = re.sub(r"\s+", " ", task.prompt)[:150]
    print(f"  prompt: {prompt_preview}{'...' if len(task.prompt) > 150 else ''}")

    def failure_result(status: str, error: str) -> dict[str, Any]:
        return {
            "task_id": task.task_id, "name": task.name, "category": task.category,
            "grading_type": task.grading_type, "network_task": is_network_task(task),
            "multi_session": task.multi_session, "session_count": 0,
            "success": False, "status": status, "returncode": None, "score": None,
            "elapsed": 0.0, "ttft": None, "input_tokens": None, "output_tokens": None,
            "reasoning_tokens": None, "cache_read_tokens": None,
            "cache_write_tokens": None, "cost_usd": None, "total_tokens": None,
            "usage_complete": False, "step_count": 0, "tool_errors": 0,
            "tool_call_count": 0, "item_error_count": 0, "fatal_error_count": 1,
            "observed_models": [], "unexpected_models": [],
            "model_verification": "not_run", "codex_versions": [],
            "approval_policies": [args.approval_policy],
            "sandbox_modes": [args.sandbox_mode],
            "windows_sandboxes": [args.windows_sandbox],
            "permission_denials": [], "command_failures": [],
            "encoding_normalization_enabled": not args.no_encoding_normalization,
            "encoding_normalized_count": 0, "encoding_normalized_files": [],
            "encoding_normalization_errors": [], "encoding_normalization_audit": "",
            "output": "", "error": error, "stderr": "", "breakdown": {},
            "grade_notes": "", "grade_error": error, "workspace": str(workspace),
            "transcript": str(transcript_dir), "normalized_transcript_path": "",
            "turn_results": [],
        }

    prerequisites = task.metadata.get("prerequisites") or []
    missing_prerequisites = check_prerequisites(list(prerequisites)) if prerequisites else []
    if missing_prerequisites:
        error = f"缺少依赖: {', '.join(missing_prerequisites)}"
        print(f"  跳过: {error}\n")
        return failure_result("missing_prerequisite", error)

    if workspace.exists():
        shutil.rmtree(workspace)
    workspace.mkdir(parents=True, exist_ok=True)
    transcript_dir.mkdir(parents=True, exist_ok=True)

    staged, fixture_errors = stage_workspace_files(task, workspace, skill_dir)
    if staged:
        LOGGER.info("  staged files: %s", ", ".join(staged[:8]) + (" ..." if len(staged) > 8 else ""))
    if fixture_errors:
        error = "缺少或无效的 workspace fixture: " + " | ".join(fixture_errors)
        print(f"  失败: {error}\n")
        return failure_result("missing_fixture", error)

    baseline_hashes = snapshot_workspace_files(workspace)

    turns = build_turns(task)
    total_timeout = args.network_timeout if is_network_task(task) else max(1.0, task.timeout_seconds * args.timeout_multiplier)
    task_deadline = time.monotonic() + total_timeout

    normalized_transcript: list[dict[str, Any]] = []
    outputs: list[str] = []
    errors: list[str] = []
    stderr_chunks: list[str] = []
    turn_results: list[dict[str, Any]] = []

    current_session_id: Optional[str] = None
    task_status = "success"
    task_success = True
    returncode: Optional[int] = 0
    total_elapsed = 0.0
    task_ttft: Optional[float] = None
    input_tokens: Optional[int] = None
    output_tokens: Optional[int] = None
    reasoning_tokens: Optional[int] = None
    cache_read_tokens: Optional[int] = None
    cache_write_tokens: Optional[int] = None
    total_tokens: Optional[int] = None
    usage_complete = True
    step_count = tool_errors = tool_call_count = item_error_count = fatal_error_count = 0
    event_counts: Counter[str] = Counter()
    unexpected_models: set[str] = set()
    permission_denials: list[Any] = []
    command_failures: list[Any] = []
    model_metadata_warnings: list[str] = []
    model_verifications: set[str] = set()
    approval_policies: set[str] = set()
    sandbox_modes: set[str] = set()
    windows_sandboxes: set[str] = set()

    for turn_index, turn in enumerate(turns, start=1):
        remaining = task_deadline - time.monotonic()
        if remaining <= 0:
            task_success = False
            task_status = "timeout"
            errors.append(f"任务总超时 ({total_timeout:.0f}s)")
            fatal_error_count += 1
            break

        turn_id = str(turn["id"])
        turn_prompt = str(turn["prompt"])
        new_session = bool(turn.get("new_session"))
        resume_required = turn_index > 1 and not new_session
        if resume_required and not current_session_id:
            task_success = False
            task_status = "session_error"
            errors.append(f"turn {turn_index} ({turn_id}) 需要恢复会话，但前一轮没有 thread_id")
            fatal_error_count += 1
            break

        normalized_transcript.append(make_user_transcript_event(turn_prompt, turn_index, turn_id))
        prompt = turn_prompt
        if not args.no_workspace_instruction:
            prompt += (
                "\n\nIMPORTANT: You are running in an isolated workspace. "
                "Read, write, and edit files only in the current working directory."
            )

        raw_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.jsonl"
        stderr_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.stderr.txt"
        prompt_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.prompt.txt"
        final_message_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.final.txt"
        prompt_payload = prompt if prompt.endswith("\n") else prompt + "\n"
        prompt_path.write_text(prompt_payload, encoding="utf-8", newline="\n")
        prompt_bytes = prompt_payload.encode("utf-8")
        prompt_sha256 = hashlib.sha256(prompt_bytes).hexdigest()

        common_options = [
            "--json", "--skip-git-repo-check", "--model", args.model,
            "--output-last-message", str(final_message_path),
        ]
        if not args.no_ignore_rules:
            common_options.append("--ignore-rules")

        if resume_required and current_session_id:
            logical_command = [
                codex_command, "exec", "resume", *common_options,
                current_session_id, "-",
            ]
        else:
            logical_command = [
                codex_command, "exec", "--color", "never",
                "--sandbox", args.sandbox_mode, *common_options, "-",
            ]

        turn_result = run_codex_streaming(
            logical_command, cwd=workspace, timeout=max(1.0, remaining),
            raw_stdout_path=raw_path, stderr_path=stderr_path,
            stdin_path=prompt_path, final_message_path=final_message_path,
            requested_model=args.model, configured_model=configured_model,
            approval_policy=args.approval_policy, sandbox_mode=args.sandbox_mode,
            windows_sandbox=args.windows_sandbox,
        )

        normalized_transcript.extend(turn_result["normalized_events"])
        if turn_result.get("output"):
            outputs.append(f"## Turn {turn_index} ({turn_id})\n{turn_result['output']}")
        if turn_result.get("error"):
            errors.append(f"turn {turn_index} ({turn_id}): {turn_result['error']}")
        if turn_result.get("stderr"):
            stderr_chunks.append(f"## Turn {turn_index} ({turn_id})\n{turn_result['stderr']}")

        total_elapsed += float(turn_result.get("elapsed") or 0.0)
        if task_ttft is None and turn_result.get("ttft") is not None:
            task_ttft = float(turn_result["ttft"])
        input_tokens = aggregate_optional_int(input_tokens, turn_result.get("input_tokens"))
        output_tokens = aggregate_optional_int(output_tokens, turn_result.get("output_tokens"))
        reasoning_tokens = aggregate_optional_int(reasoning_tokens, turn_result.get("reasoning_tokens"))
        cache_read_tokens = aggregate_optional_int(cache_read_tokens, turn_result.get("cache_read_tokens"))
        cache_write_tokens = aggregate_optional_int(cache_write_tokens, turn_result.get("cache_write_tokens"))
        total_tokens = aggregate_optional_int(total_tokens, turn_result.get("total_tokens"))
        usage_complete = usage_complete and bool(turn_result.get("usage_complete"))
        step_count += int(turn_result.get("step_count") or 0)
        tool_errors += int(turn_result.get("tool_errors") or 0)
        tool_call_count += int(turn_result.get("tool_call_count") or 0)
        item_error_count += int(turn_result.get("item_error_count") or 0)
        fatal_error_count += int(turn_result.get("fatal_error_count") or 0)
        event_counts.update(turn_result.get("event_counts") or {})
        unexpected_models.update(str(x) for x in (turn_result.get("unexpected_models") or []))
        permission_denials.extend(turn_result.get("permission_denials") or [])
        command_failures.extend(turn_result.get("command_failures") or [])
        model_metadata_warnings.extend(str(x) for x in (turn_result.get("model_metadata_warnings") or []))
        for field, target in [
            ("model_verification", model_verifications),
            ("approval_policy", approval_policies),
            ("sandbox_mode", sandbox_modes),
            ("windows_sandbox", windows_sandboxes),
        ]:
            if turn_result.get(field):
                target.add(str(turn_result.get(field)))

        returned_session_id = turn_result.get("session_id")
        if returned_session_id:
            current_session_id = str(returned_session_id)

        turn_results.append({
            "turn": turn_index, "turn_id": turn_id, "new_session": new_session,
            "session_id": current_session_id, "success": turn_result.get("success"),
            "status": turn_result.get("status"), "returncode": turn_result.get("returncode"),
            "elapsed": turn_result.get("elapsed"), "ttft": turn_result.get("ttft"),
            "input_tokens": turn_result.get("input_tokens"),
            "output_tokens": turn_result.get("output_tokens"),
            "reasoning_tokens": turn_result.get("reasoning_tokens"),
            "cache_read_tokens": turn_result.get("cache_read_tokens"),
            "cache_write_tokens": turn_result.get("cache_write_tokens"),
            "total_tokens": turn_result.get("total_tokens"),
            "usage_complete": turn_result.get("usage_complete"),
            "step_count": turn_result.get("step_count"),
            "tool_errors": turn_result.get("tool_errors"),
            "tool_call_count": turn_result.get("tool_call_count"),
            "item_error_count": turn_result.get("item_error_count"),
            "fatal_error_count": turn_result.get("fatal_error_count"),
            "event_counts": turn_result.get("event_counts"),
            "model_verification": turn_result.get("model_verification"),
            "unexpected_models": turn_result.get("unexpected_models"),
            "model_metadata_warnings": turn_result.get("model_metadata_warnings"),
            "approval_policy": turn_result.get("approval_policy"),
            "sandbox_mode": turn_result.get("sandbox_mode"),
            "windows_sandbox": turn_result.get("windows_sandbox"),
            "permission_denials": turn_result.get("permission_denials"),
            "command_failures": turn_result.get("command_failures"),
            "raw_transcript": str(raw_path), "stderr_path": str(stderr_path),
            "final_message_path": str(final_message_path),
            "prompt_transport": turn_result.get("prompt_transport"),
            "prompt_path": str(prompt_path), "prompt_chars": len(prompt_payload),
            "prompt_bytes": len(prompt_bytes), "prompt_sha256": prompt_sha256,
            "error": turn_result.get("error"),
        })

        if not turn_result.get("success"):
            task_success = False
            task_status = str(turn_result.get("status") or "error")
            returncode = turn_result.get("returncode")
            break
        returncode = turn_result.get("returncode")

    normalized_path = transcript_dir / "normalized.jsonl"
    with normalized_path.open("w", encoding="utf-8") as file:
        for event in normalized_transcript:
            file.write(json.dumps(event, ensure_ascii=False) + "\n")
    (transcript_dir / "turn_results.json").write_text(
        json.dumps(turn_results, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    if args.no_encoding_normalization:
        encoding_normalization = {
            "enabled": False,
            "policy": "disabled by --no-encoding-normalization",
            "normalized_count": 0,
            "normalized_files": [],
            "records": [],
            "errors": [],
            "audit_dir": "",
        }
    else:
        encoding_normalization = normalize_workspace_text_encodings(
            workspace, baseline_hashes, transcript_dir / "encoding_normalization"
        )
        if encoding_normalization["normalized_count"]:
            LOGGER.info(
                "  normalized text encodings: %s",
                ", ".join(encoding_normalization["normalized_files"][:8])
                + (" ..." if encoding_normalization["normalized_count"] > 8 else ""),
            )
        if encoding_normalization["errors"]:
            LOGGER.warning(
                "  encoding normalization warnings: %s",
                " | ".join(encoding_normalization["errors"][:5]),
            )

    final_workspace_hashes = snapshot_workspace_files(workspace)
    changed_workspace_files = sorted(
        relative_path
        for relative_path in (
            set(baseline_hashes) | set(final_workspace_hashes)
        )
        if baseline_hashes.get(relative_path)
        != final_workspace_hashes.get(relative_path)
    )

    return {
        "task_id": task.task_id, "name": task.name, "category": task.category,
        "grading_type": task.grading_type, "network_task": is_network_task(task),
        "multi_session": task.multi_session, "session_count": len(turn_results),
        "success": task_success, "status": task_status, "returncode": returncode,
        "elapsed": total_elapsed, "ttft": task_ttft, "input_tokens": input_tokens,
        "output_tokens": output_tokens, "reasoning_tokens": reasoning_tokens,
        "cache_read_tokens": cache_read_tokens, "cache_write_tokens": cache_write_tokens,
        "cost_usd": None, "total_tokens": total_tokens,
        "usage_complete": usage_complete and bool(turn_results), "step_count": step_count,
        "tool_errors": tool_errors, "tool_call_count": tool_call_count,
        "item_error_count": item_error_count, "fatal_error_count": fatal_error_count,
        "event_counts": dict(event_counts), "observed_models": [],
        "unexpected_models": sorted(unexpected_models),
        "model_verification": ",".join(sorted(model_verifications)),
        "model_metadata_warnings": list(dict.fromkeys(model_metadata_warnings)),
        "codex_versions": [args.expected_codex_version],
        "approval_policies": sorted(approval_policies),
        "sandbox_modes": sorted(sandbox_modes),
        "windows_sandboxes": sorted(windows_sandboxes),
        "permission_denials": permission_denials, "command_failures": command_failures,
        "output": "\n\n".join(outputs), "error": " | ".join(errors),
        "stderr": "\n\n".join(stderr_chunks), "score": None, "breakdown": {},
        "grade_notes": "", "grade_error": None, "workspace": str(workspace),
        "transcript": str(transcript_dir),
        "normalized_transcript_path": str(normalized_path),
        "transcript_data": normalized_transcript, "turn_results": turn_results,
        "encoding_normalization_enabled": encoding_normalization["enabled"],
        "encoding_normalized_count": encoding_normalization["normalized_count"],
        "encoding_normalized_files": encoding_normalization["normalized_files"],
        "encoding_normalization_errors": encoding_normalization["errors"],
        "encoding_normalization_audit": encoding_normalization["audit_dir"],
        "changed_workspace_files": changed_workspace_files,
        "post_completion_transport_recovered": False,
        "transport_warning_count": 0,
        "transport_warnings": [],
        "original_status": None,
        "original_success": None,
        "original_returncode": None,
        "original_error": "",
        "original_fatal_error_count": None,
    }


def build_args() -> argparse.Namespace:
    script_path = Path(__file__).resolve()
    inferred_root = script_path.parent.parent
    inferred_skill = inferred_root / "skill"
    inferred_codex = inferred_root / "codex-cli" / "node_modules" / ".bin" / "codex.cmd"

    parser = argparse.ArgumentParser(
        description="Run local PinchBench tasks with Codex on native Windows, one task at a time.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--skill-dir", default=str(inferred_skill if inferred_skill.exists() else Path.cwd()), help="PinchBench skill 仓库根目录")
    parser.add_argument("--tasks-dir", default=None, help="任务目录；默认是 <skill-dir>/tasks")
    parser.add_argument("--suite", default="core", help="all、core、automated-only、llm-judge-only、hybrid-only、judge-required-only，或逗号分隔任务 ID")
    parser.add_argument("--limit", type=int, default=None, help="最多运行多少个任务，仅用于测试 runner")
    parser.add_argument("--skip", default="", help="额外跳过的任务 ID；四个已约定 integration 任务默认跳过")
    parser.add_argument("--skip-network", action="store_true", help="跳过明确标记的联网任务；正式全量测试通常不应使用")
    parser.add_argument("--codex", default=str(inferred_codex) if inferred_codex.exists() else "auto", help="Codex 命令：auto、codex.cmd 或固定安装的绝对路径")
    parser.add_argument("--expected-codex-version", default=DEFAULT_CODEX_VERSION, help="正式测试要求的 Codex 精确版本；不一致时预检失败")
    parser.add_argument("--model", default=DEFAULT_MODEL, help="Codex 模型 ID；OpenRouter provider 由隔离 config.toml 决定")
    parser.add_argument("--approval-policy", default=DEFAULT_APPROVAL_POLICY, choices=["never"], help="正式无人值守测试固定为 never")
    parser.add_argument("--sandbox-mode", default=DEFAULT_SANDBOX_MODE, choices=["workspace-write"], help="正式对比固定为 workspace-write")
    parser.add_argument("--windows-sandbox", default=DEFAULT_WINDOWS_SANDBOX, choices=["unelevated"], help="正式 Windows Codex 沙箱固定为 unelevated")
    parser.add_argument("--no-ignore-rules", action="store_true", help="调试时允许 exec-policy rules；正式可比测试不要使用")
    parser.add_argument("--timeout-multiplier", type=float, default=3.0, help="非联网任务 timeout_seconds 的倍数")
    parser.add_argument("--network-timeout", type=float, default=300.0, help="明确标记的联网任务总超时秒数")
    parser.add_argument("--judge-timeout", type=float, default=300.0, help="PinchBench 默认 LLM judge 单次请求超时秒数")
    parser.add_argument("--results-dir", default=None, help="运行结果根目录；默认是 <skill-dir>/../runs")
    parser.add_argument("--keep-workspaces", action="store_true", help="保留成功任务 workspace；正式评测建议启用")
    parser.add_argument("--no-grade", action="store_true", help="只执行任务，不打分")
    parser.add_argument("--no-xlsx", action="store_true", help="不输出 XLSX")
    parser.add_argument("--no-workspace-instruction", action="store_true", help="不在 prompt 后追加隔离 workspace 提醒")
    parser.add_argument("--no-encoding-normalization", action="store_true", help="禁用评分前的 Windows 文本产物编码规范化；正式可比测试不要使用")
    parser.add_argument("--preflight", action="store_true", help="只做本地预检，不调用模型")
    parser.add_argument("--no-judge-cache", action="store_true", help="禁用 PinchBench Judge 缓存")
    parser.add_argument("--clear-judge-cache", action="store_true", help="运行前清空 PinchBench Judge 缓存")
    parser.add_argument("--verbose", action="store_true", help="打印更详细日志")
    return parser.parse_args()


def main() -> int:
    args = build_args()
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-8s %(message)s",
        datefmt="%H:%M:%S",
    )

    script_path = Path(__file__).resolve()
    inferred_root = script_path.parent.parent
    os.environ.setdefault("CODEX_HOME", str(inferred_root / "codex-home"))
    os.environ.setdefault("PYTHONUTF8", "1")
    os.environ.setdefault("PYTHONIOENCODING", "utf-8")

    codex_home = Path(os.environ["CODEX_HOME"]).expanduser().resolve()
    codex_config_path = codex_home / "config.toml"
    skill_dir = Path(args.skill_dir).expanduser().resolve()
    tasks_dir = Path(args.tasks_dir).expanduser().resolve() if args.tasks_dir else skill_dir / "tasks"

    if not skill_dir.exists():
        raise SystemExit(f"找不到 PinchBench 仓库目录: {skill_dir}")
    if not tasks_dir.exists():
        raise SystemExit(f"找不到 tasks 目录: {tasks_dir}\n请确认 --skill-dir 指向本地 pinchbench/skill 仓库。")
    if not (tasks_dir / "manifest.yaml").exists():
        LOGGER.warning("tasks/manifest.yaml 不存在，将按 task_*.md 文件回退加载")

    codex_command = choose_codex_command(args.codex, inferred_root)
    codex_config, codex_config_error = load_toml_file(codex_config_path)
    grader_module, grader_error = load_pinchbench_grading(skill_dir)
    judge_key_present = bool(os.environ.get("OPENROUTER_API_KEY"))

    all_tasks, core_tasks = load_tasks(tasks_dir)
    selected = filter_tasks(all_tasks, core_tasks, args.suite, args.limit)
    skip_ids = set(DEFAULT_SKIPPED_TASKS)
    skip_ids.update(item.strip() for item in args.skip.split(",") if item.strip())
    if args.skip_network:
        skip_ids.update(task.task_id for task in selected if is_network_task(task))
    selected = [task for task in selected if task.task_id not in skip_ids]
    if not selected:
        raise SystemExit("没有可运行任务。请检查 --suite/--limit/--skip。")

    needs_judge = not args.no_grade and any(
        task.grading_type in {"llm_judge", "hybrid"} for task in selected
    )

    prerequisite_failures, fixture_failures, codex_failures = print_preflight(
        selected, skill_dir, tasks_dir, codex_command, codex_home,
        codex_config_path, codex_config, codex_config_error,
        grader_module, grader_error, args,
    )

    if args.preflight:
        if codex_failures:
            return 2
        if fixture_failures:
            return 3
        if prerequisite_failures:
            return 4
        if grader_error:
            return 5
        if needs_judge and not judge_key_present:
            return 6
        return 0

    if codex_failures:
        raise SystemExit("Codex 环境预检失败：\n- " + "\n- ".join(codex_failures))
    if args.no_ignore_rules:
        raise SystemExit("正式可比评测要求 --ignore-rules。请移除 --no-ignore-rules。")
    if grader_error and not args.no_grade:
        raise SystemExit(grader_error)
    if needs_judge and not judge_key_present:
        raise SystemExit(
            "所选任务包含 hybrid/llm_judge，但当前 PowerShell 未设置 "
            "OPENROUTER_API_KEY。为避免跑完后无法评分，已在正式执行前停止。"
        )
    if codex_config is None:
        raise SystemExit(codex_config_error or "Codex config 未加载")
    configured_model = str(codex_config.get("model") or "")

    results_root = Path(args.results_dir).expanduser().resolve() if args.results_dir else skill_dir.parent / "runs"
    run_id = dt.datetime.now().strftime("codex_%Y%m%d_%H%M%S")
    results_dir = results_root / run_id
    workspaces_dir = results_dir / "workspaces"
    transcripts_dir = results_dir / "transcripts"
    results_dir.mkdir(parents=True, exist_ok=True)
    workspaces_dir.mkdir(parents=True, exist_ok=True)
    transcripts_dir.mkdir(parents=True, exist_ok=True)
    progress_path = results_dir / "progress.jsonl"
    progress_path.touch(exist_ok=True)
    partial_json_path = results_dir / "results.partial.json"

    judge_model = str(getattr(grader_module, "DEFAULT_JUDGE_MODEL", PINCHBENCH_DEFAULT_JUDGE_MODEL)) if grader_module is not None else PINCHBENCH_DEFAULT_JUDGE_MODEL
    write_run_config(
        results_dir / "run_config.json", args, skill_dir, tasks_dir,
        codex_command, codex_home, codex_config_path, codex_config,
        selected, judge_model,
    )

    judge_cache_dir = results_root / ".judge_cache"
    if grader_module is not None and hasattr(grader_module, "set_judge_cache_dir") and not args.no_judge_cache:
        grader_module.set_judge_cache_dir(judge_cache_dir)
        if args.clear_judge_cache and hasattr(grader_module, "clear_judge_cache"):
            grader_module.clear_judge_cache()

    print()
    print(f"Results dir          : {results_dir}")
    print(f"Codex                : {command_output([codex_command, '--version'])}")
    print(f"Model                : {args.model}")
    print(f"Approval / sandbox   : {args.approval_policy} / {args.sandbox_mode}")
    print(f"Windows sandbox      : {args.windows_sandbox}")
    print("Grading engine       : PinchBench scripts/lib_grading.py")
    print("Judge backend        : api")
    print(f"Judge model          : {judge_model}")
    print(f"Judge key            : {'set' if judge_key_present else 'not required'}")
    print(f"Judge cache          : {'disabled' if args.no_judge_cache else judge_cache_dir}")
    print()

    results: list[dict[str, Any]] = []
    total_start = time.monotonic()

    for index, task in enumerate(selected, start=1):
        workspace = workspaces_dir / task.task_id
        transcript_dir = transcripts_dir / task.task_id
        execution = execute_task(
            task, index, len(selected), skill_dir, workspace, transcript_dir,
            codex_command, configured_model, args,
        )

        grade_result = GradeResult(score=None, grading_type=task.grading_type, breakdown={})
        if execution["status"] not in {"missing_prerequisite", "missing_fixture"} and not args.no_grade:
            grading_execution = dict(execution)
            grading_execution["transcript"] = execution.get("transcript_data", [])
            grade_result = grade_with_pinchbench_default(
                grader_module=grader_module, task=task,
                execution_result=grading_execution, workspace=workspace,
                skill_dir=skill_dir, judge_timeout=args.judge_timeout,
                verbose=args.verbose,
            )

        execution["score"] = round(float(grade_result.score), 4) if grade_result.score is not None else None
        execution["breakdown"] = grade_result.breakdown
        execution["grade_notes"] = grade_result.notes
        execution["grade_error"] = grade_result.error

        transport_recovered = (
            maybe_recover_post_completion_stream_disconnect(execution)
        )
        execution.pop("transcript_data", None)
        results.append(execution)

        with progress_path.open("a", encoding="utf-8") as progress_file:
            progress_file.write(json.dumps(execution, ensure_ascii=False) + "\n")
            progress_file.flush()
        partial_json_path.write_text(
            json.dumps({"completed": len(results), "results": results}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        marker = "✓" if execution["success"] else "✗"
        score_text = f"score={execution['score']:.3f}" if execution["score"] is not None else "score=N/A"
        ttft_text = f"ttft={execution['ttft']:.2f}s" if execution["ttft"] is not None else "ttft=N/A"
        token_text = f"{execution['input_tokens']}+{execution['output_tokens']}" if execution["usage_complete"] else "N/A"
        print(
            f"  {marker} elapsed={execution['elapsed']:.1f}s {ttft_text} "
            f"{score_text} tokens={token_text} tool_errors={execution['tool_errors']}"
        )
        if execution.get("unexpected_models"):
            print("  模型配置异常: " + ", ".join(str(x) for x in execution["unexpected_models"]))
        if execution.get("error"):
            print(f"  运行错误: {str(execution['error'])[:500]}")
        if execution.get("grade_error"):
            print(f"  打分错误: {str(execution['grade_error'])[:500]}")
        if execution.get("post_completion_transport_recovered"):
            warning = (execution.get("transport_warnings") or [{}])[0]
            print(
                "  传输警告: 产物与官方评分完成后缺少 "
                "response.completed；已按严格规则归类为成功。"
            )
            print(
                "  原始错误: "
                + str(warning.get("original_error") or "")[:500]
            )
        if execution.get("output"):
            preview = re.sub(r"\s+", " ", str(execution["output"]))[:260]
            print(f"  输出预览: {preview}{'...' if len(str(execution['output'])) > 260 else ''}")
        print()

        if not args.keep_workspaces and execution["success"] and not execution.get("grade_error"):
            shutil.rmtree(workspace, ignore_errors=True)
            execution["workspace"] = "已清理；使用 --keep-workspaces 可保留"

    total_elapsed = time.monotonic() - total_start
    summary = print_summary(results, total_elapsed)
    payload = {"summary": summary, "results": results}
    json_path = results_dir / "results.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    csv_path = results_dir / "results.csv"
    save_csv(results, csv_path)
    xlsx_path = results_dir / "results.xlsx"
    if not args.no_xlsx:
        save_xlsx(results, summary, xlsx_path)

    if grader_module is not None and hasattr(grader_module, "get_judge_cache_stats") and not args.no_judge_cache:
        try:
            stats = grader_module.get_judge_cache_stats()
            print(
                "Judge cache stats   : "
                f"entries={stats.get('entries', 0)} "
                f"hits={stats.get('hits', 0)} "
                f"misses={stats.get('misses', 0)}"
            )
        except Exception as exc:
            LOGGER.warning("读取 Judge cache stats 失败: %s", exc)

    print("\n结果文件:")
    print(f"  Config : {results_dir / 'run_config.json'}")
    print(f"  JSON   : {json_path}")
    print(f"  CSV    : {csv_path}")
    if not args.no_xlsx:
        print(f"  XLSX   : {xlsx_path}")
    print(f"  Logs   : {transcripts_dir}")
    print(f"  Progress JSONL : {progress_path}")
    print(f"  Partial JSON   : {partial_json_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
