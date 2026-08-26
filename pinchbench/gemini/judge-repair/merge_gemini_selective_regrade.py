#!/usr/bin/env python3
from __future__ import annotations

import argparse
import copy
import csv
import datetime as dt
import hashlib
import json
import sys
from pathlib import Path
from typing import Any

EXPECTED_TASKS = {
    'task_subway_navigation',
    'task_competitive_research',
    'task_blog',
    'task_csv_stock_trend',
    'task_csv_gdp_per_capita',
    'task_csv_pension_liability',
    'task_csv_pension_risk',
    'task_log_ssh_failed_logins',
    'task_log_hdfs_block_ops',
    'task_meeting_gov_speaker_summary',
    'task_meeting_gov_qa_extract',
    'task_meeting_gov_recommendations',
    'task_meeting_gov_data_sources',
    'task_csv_finance_report',
    'task_csv_stations_by_elevation',
    'task_csv_stations_coverage',
    'task_csv_stations_filter',
    'task_csv_iris_classify',
    'task_csv_cities_filter',
    'task_csv_cities_density',
    'task_csv_pension_ranking',
    'task_meeting_advisory_acronyms',
}
PRESERVED_AUTOMATED_NA = 'task_git_rescue_recovery'
MODEL_STATUS_TASK = 'task_market_research'
EXPECTED_JUDGE = 'openrouter/anthropic/claude-opus-5'
TARGET_MODEL = 'deepseek/deepseek-v4-pro'
ZERO_TOKEN_MODEL = 'gemini-3-flash-preview'


def read_json(path: Path) -> Any:
    with path.open('r', encoding='utf-8-sig') as handle:
        return json.load(handle)


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding='utf-8')


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open('rb') as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b''):
            digest.update(chunk)
    return digest.hexdigest()


def result_rows(document: Any) -> list[dict[str, Any]]:
    rows = document.get('results') if isinstance(document, dict) else document
    if not isinstance(rows, list):
        raise ValueError('results.json does not contain a results list')
    return rows


def optional_sum(rows: list[dict[str, Any]], field: str) -> float | None:
    values = [row.get(field) for row in rows if row.get(field) is not None]
    if not values:
        return None
    return float(sum(float(value) for value in values))


def recompute_summary(rows: list[dict[str, Any]], original_summary: dict[str, Any]) -> dict[str, Any]:
    succeeded = sum(1 for row in rows if row.get('success'))
    scores = [float(row['score']) for row in rows if row.get('score') is not None]
    ttfts = [float(row['ttft']) for row in rows if row.get('ttft') is not None]
    average_elapsed = sum(float(row.get('elapsed') or 0.0) for row in rows) / len(rows) if rows else 0.0

    summary = dict(original_summary)
    summary.update({
        '任务总数': len(rows),
        '成功': succeeded,
        '失败': len(rows) - succeeded,
        '成功率': round(succeeded / len(rows), 4) if rows else 0.0,
        '有分数任务数': len(scores),
        '打分失败任务数': sum(1 for row in rows if row.get('grade_error')),
        '平均分数': round(sum(scores) / len(scores), 4) if scores else None,
        '平均耗时/任务(s)': round(average_elapsed, 2),
        '平均TTFT估计(s)': round(sum(ttfts) / len(ttfts), 4) if ttfts else None,
        '总输入Token': int(optional_sum(rows, 'input_tokens') or 0)
            if any(row.get('input_tokens') is not None for row in rows) else None,
        '总输出Token': int(optional_sum(rows, 'output_tokens') or 0)
            if any(row.get('output_tokens') is not None for row in rows) else None,
        '总推理Token': int(optional_sum(rows, 'reasoning_tokens') or 0)
            if any(row.get('reasoning_tokens') is not None for row in rows) else None,
        '总费用USD': round(optional_sum(rows, 'cost_usd') or 0.0, 6)
            if any(row.get('cost_usd') is not None for row in rows) else None,
        'Usage缺失任务数': sum(1 for row in rows if not row.get('usage_complete')),
        '模型不一致任务数': sum(1 for row in rows if row.get('unexpected_models')),
        '权限拒绝任务数': sum(1 for row in rows if row.get('permission_denials')),
        '联网任务数': sum(1 for row in rows if row.get('network_task')),
        '多轮任务数': sum(1 for row in rows if row.get('multi_session')),
    })
    return summary


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fields: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen:
                seen.add(key)
                fields.append(key)
    with path.open('w', encoding='utf-8-sig', newline='') as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction='ignore')
        writer.writeheader()
        for row in rows:
            flat = {}
            for key, value in row.items():
                flat[key] = json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value
            writer.writerow(flat)


def write_xlsx(path: Path, rows: list[dict[str, Any]], summary: dict[str, Any], audit: dict[str, Any]) -> bool:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return False

    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = '汇总'
    summary_sheet['A1'] = 'Gemini PinchBench corrected frozen-output result'
    summary_sheet['A1'].font = Font(size=15, bold=True)
    summary_sheet.merge_cells('A1:B1')
    summary_sheet.append([])
    for key, value in summary.items():
        summary_sheet.append([key, value])
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 30

    details = workbook.create_sheet('任务结果')
    columns = [
        'task_id', 'name', 'category', 'grading_type', 'success', 'status',
        'score', 'elapsed', 'ttft', 'grade_error', 'grade_notes',
        'input_tokens', 'output_tokens', 'unexpected_models', 'workspace',
    ]
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    for index, name in enumerate(columns, start=1):
        cell = details.cell(row=1, column=index, value=name)
        cell.fill = header_fill
        cell.font = header_font
    for r_index, row in enumerate(rows, start=2):
        for c_index, name in enumerate(columns, start=1):
            value = row.get(name)
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            details.cell(row=r_index, column=c_index, value=value)
    details.freeze_panes = 'A2'
    details.auto_filter.ref = details.dimensions
    for col, width in {'A': 40, 'B': 32, 'C': 20, 'D': 18, 'E': 10, 'F': 18,
                       'G': 12, 'H': 14, 'I': 14, 'J': 55, 'K': 80,
                       'L': 16, 'M': 16, 'N': 35, 'O': 80}.items():
        details.column_dimensions[col].width = width

    audit_sheet = workbook.create_sheet('修复审计')
    audit_sheet.append(['Key', 'Value'])
    for cell in audit_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for key, value in audit.items():
        audit_sheet.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    audit_sheet.column_dimensions['A'].width = 38
    audit_sheet.column_dimensions['B'].width = 110

    workbook.save(path)
    return True


def zero_token_model_evidence(run_dir: Path, row: dict[str, Any]) -> tuple[bool, dict[str, Any]]:
    turn_results = row.get('turn_results') or []
    if not turn_results:
        return False, {'reason': 'turn_results missing'}
    raw = str(turn_results[0].get('raw_transcript') or '')
    raw_path = Path(raw)
    if not raw_path.is_file():
        fallback = run_dir / 'transcripts' / MODEL_STATUS_TASK / 'turn_01_single.jsonl'
        raw_path = fallback
    if not raw_path.is_file():
        return False, {'reason': f'raw transcript missing: {raw_path}'}

    result_event = None
    with raw_path.open('r', encoding='utf-8-sig') as handle:
        for line in handle:
            try:
                event = json.loads(line)
            except json.JSONDecodeError:
                continue
            if event.get('type') == 'result':
                result_event = event
    if not isinstance(result_event, dict):
        return False, {'reason': 'result event missing'}
    models = ((result_event.get('stats') or {}).get('models') or {})
    target = models.get(TARGET_MODEL) or {}
    zero = models.get(ZERO_TOKEN_MODEL) or {}
    target_total = int(target.get('total_tokens') or 0)
    zero_total = int(zero.get('total_tokens') or 0)
    valid = target_total > 0 and zero_total == 0
    return valid, {
        'raw_transcript': str(raw_path),
        'target_model': TARGET_MODEL,
        'target_total_tokens': target_total,
        'zero_token_model': ZERO_TOKEN_MODEL,
        'zero_token_model_total_tokens': zero_total,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument('--original-run', required=True)
    parser.add_argument('--regrade-run', required=True)
    parser.add_argument('--output-root', required=True)
    args = parser.parse_args()

    original_run = Path(args.original_run).resolve()
    regrade_run = Path(args.regrade_run).resolve()
    output_root = Path(args.output_root).resolve()
    original_results_path = original_run / 'results.json'
    source_snapshot_path = regrade_run / 'source_runs_snapshot.json'
    regrade_results_path = regrade_run / 'results.json'

    for path in (original_results_path, source_snapshot_path, regrade_results_path):
        if not path.is_file():
            raise FileNotFoundError(path)

    source_snapshot = read_json(source_snapshot_path)
    if not isinstance(source_snapshot, list) or len(source_snapshot) != 1:
        raise RuntimeError('Expected exactly one source run in the selective regrade')
    snapshot = source_snapshot[0]
    if Path(str(snapshot.get('run_path'))).resolve() != original_run:
        raise RuntimeError('Regrade source run does not match the requested original run')
    actual_hash = sha256_file(original_results_path)
    if snapshot.get('results_sha256') != actual_hash:
        raise RuntimeError('Original results.json changed after the regrade snapshot was created')

    regrade_doc = read_json(regrade_results_path)
    metadata = regrade_doc.get('metadata') or {}
    if metadata.get('judge_model') != EXPECTED_JUDGE:
        raise RuntimeError(f"Unexpected Judge model: {metadata.get('judge_model')!r}")
    regrade_rows = result_rows(regrade_doc)
    regrade_map = {str(row.get('task_id')): row for row in regrade_rows}
    if set(regrade_map) != EXPECTED_TASKS:
        missing = sorted(EXPECTED_TASKS - set(regrade_map))
        extra = sorted(set(regrade_map) - EXPECTED_TASKS)
        raise RuntimeError(f'Regrade task set mismatch. missing={missing}, extra={extra}')

    bad: list[str] = []
    for task_id, row in regrade_map.items():
        if row.get('status') != 'completed':
            bad.append(f"{task_id}: status={row.get('status')}")
        if row.get('worker_status') != 'completed':
            bad.append(f"{task_id}: worker_status={row.get('worker_status')}")
        if row.get('new_score') is None:
            bad.append(f'{task_id}: new_score is null')
        if str(row.get('new_grade_error') or '').strip():
            bad.append(f"{task_id}: {row.get('new_grade_error')}")
        compat = row.get('judge_transport_compatibility') or {}
        if not compat.get('enabled'):
            bad.append(f'{task_id}: OpenRouter compatibility shim was not enabled')
    if bad:
        raise RuntimeError('Selective regrade is not clean; do not merge:\n' + '\n'.join(bad))

    original_doc = read_json(original_results_path)
    corrected_doc = copy.deepcopy(original_doc)
    corrected_rows = result_rows(corrected_doc)
    original_map = {str(row.get('task_id')): row for row in result_rows(original_doc)}
    corrected_map = {str(row.get('task_id')): row for row in corrected_rows}
    if len(corrected_map) != 143:
        raise RuntimeError(f'Expected 143 original tasks, found {len(corrected_map)}')
    if PRESERVED_AUTOMATED_NA not in corrected_map:
        raise RuntimeError(f'Missing preserved automated task: {PRESERVED_AUTOMATED_NA}')

    changes: list[dict[str, Any]] = []
    now = dt.datetime.now(dt.timezone.utc).isoformat()
    for task_id in sorted(EXPECTED_TASKS):
        target = corrected_map[task_id]
        source = regrade_map[task_id]
        old = original_map[task_id]
        old_score = old.get('score')
        target['score'] = float(source['new_score'])
        target['breakdown'] = copy.deepcopy(source.get('new_breakdown') or {})
        target['grade_notes'] = str(source.get('new_grade_notes') or '')
        target['grade_error'] = None
        target['regrade'] = {
            'kind': 'frozen_output_judge_repair',
            'judge_model': EXPECTED_JUDGE,
            'agent_execution_rerun': False,
            'source_regrade_run': str(regrade_run),
            'original_score': old_score,
            'new_score': target['score'],
            'regraded_at': now,
            'judge_transport_compatibility': copy.deepcopy(source.get('judge_transport_compatibility') or {}),
        }
        changes.append({
            'task_id': task_id,
            'original_score': old_score,
            'new_score': target['score'],
            'original_grade_error': old.get('grade_error'),
            'new_grade_error': None,
        })

    preserved_before = copy.deepcopy(corrected_map[PRESERVED_AUTOMATED_NA])

    market_row = corrected_map[MODEL_STATUS_TASK]
    evidence_ok, model_evidence = zero_token_model_evidence(original_run, market_row)
    if not evidence_ok:
        raise RuntimeError('Refusing to correct task_market_research without zero-token model evidence: ' + json.dumps(model_evidence, ensure_ascii=False))
    if market_row.get('status') != 'model_mismatch' or ZERO_TOKEN_MODEL not in (market_row.get('unexpected_models') or []):
        raise RuntimeError('task_market_research no longer matches the expected zero-token mismatch state')
    market_row['success'] = True
    market_row['status'] = 'success'
    market_row['error'] = ''
    market_row['unexpected_models'] = []
    for turn in market_row.get('turn_results') or []:
        if turn.get('status') == 'model_mismatch':
            turn['success'] = True
            turn['status'] = 'success'
            turn['error'] = ''
            turn['unexpected_models'] = []
    market_row['status_correction'] = {
        'kind': 'zero_token_observed_model_false_positive',
        'agent_execution_rerun': False,
        'corrected_at': now,
        'evidence': model_evidence,
    }

    if corrected_map[PRESERVED_AUTOMATED_NA] != preserved_before:
        raise RuntimeError('Automated N/A task was modified unexpectedly')

    original_summary = original_doc.get('summary') if isinstance(original_doc, dict) else {}
    corrected_summary = recompute_summary(corrected_rows, original_summary or {})
    if isinstance(corrected_doc, dict):
        corrected_doc['summary'] = corrected_summary
        corrected_doc['regrade_metadata'] = {
            'schema_version': 1,
            'created_at': now,
            'original_run': str(original_run),
            'original_results_sha256': actual_hash,
            'regrade_run': str(regrade_run),
            'judge_model': EXPECTED_JUDGE,
            'agent_execution_rerun': False,
            'regraded_task_count': len(EXPECTED_TASKS),
            'regraded_tasks': sorted(EXPECTED_TASKS),
            'preserved_automated_na': PRESERVED_AUTOMATED_NA,
            'model_status_correction_task': MODEL_STATUS_TASK,
        }

    stamp = dt.datetime.now().strftime('%Y%m%d-%H%M%S')
    output_dir = output_root / f'gemini_20260731_173523_opus5_repaired_{stamp}'
    output_dir.mkdir(parents=True, exist_ok=False)
    write_json(output_dir / 'results.json', corrected_doc)
    write_csv(output_dir / 'results.csv', corrected_rows)

    audit = {
        'created_at': now,
        'original_run': str(original_run),
        'original_results_sha256': actual_hash,
        'regrade_run': str(regrade_run),
        'judge_model': EXPECTED_JUDGE,
        'agent_execution_rerun': False,
        'regraded_task_count': len(EXPECTED_TASKS),
        'changes': changes,
        'preserved_automated_na': PRESERVED_AUTOMATED_NA,
        'preserved_automated_na_score': corrected_map[PRESERVED_AUTOMATED_NA].get('score'),
        'market_status_correction': model_evidence,
        'original_summary': original_summary,
        'corrected_summary': corrected_summary,
    }
    write_json(output_dir / 'audit.json', audit)
    write_json(output_dir / 'run_config.json', {
        'schema_version': 1,
        'kind': 'frozen_output_selective_regrade_merge',
        'created_at': now,
        'source_agent_run': str(original_run),
        'source_regrade_run': str(regrade_run),
        'judge_model': EXPECTED_JUDGE,
        'agent_execution_rerun': False,
        'source_agent_results_sha256': actual_hash,
        'regraded_tasks': sorted(EXPECTED_TASKS),
        'preserved_automated_na': PRESERVED_AUTOMATED_NA,
    })
    xlsx_written = write_xlsx(output_dir / 'results.xlsx', corrected_rows, corrected_summary, audit)

    print(f'OUTPUT_DIR={output_dir}')
    print(f"SCORED={corrected_summary.get('有分数任务数')}")
    print(f"GRADE_FAILURES={corrected_summary.get('打分失败任务数')}")
    print(f"MEAN_SCORE={corrected_summary.get('平均分数')}")
    print(f"SUCCESS={corrected_summary.get('成功')}/{corrected_summary.get('任务总数')}")
    print(f'XLSX_WRITTEN={xlsx_written}')
    return 0


if __name__ == '__main__':
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f'ERROR: {exc}', file=sys.stderr)
        raise
