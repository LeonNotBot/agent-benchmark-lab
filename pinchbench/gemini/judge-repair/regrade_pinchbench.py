#!/usr/bin/env python3
"""
PinchBench frozen-output regrader.

Contract:
- Reads original runs and PinchBench skill checkout without modifying them.
- Loads the same Task parser from the supplied formal Codex Runner.
- Calls the checked-out scripts/lib_grading.py grade_task API.
- Keeps task/rubric/automated checks/hybrid weights unchanged.
- The only scoring input intentionally changed is judge_model.
- Removes original score, breakdown, grade notes and grade error before grading.
- Copies each source workspace to scratch before calling the grader.
- Serial execution, one grading job at a time.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import http.client as http_client
import importlib
import inspect
import json
import os
import platform
import shutil
import signal
import sqlite3
import subprocess
import sys
import time
import threading
import traceback
from urllib import error as urllib_error
from urllib import request as urllib_request
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable, Optional

SCRIPT_REVISION = "2026-07-29-pinchbench-frozen-regrader-v1.6-image-gen-context-bounding"
DEFAULT_CONFIG = Path(__file__).with_name("regrade_config.json")
REMOVE_BEFORE_GRADING = {
    "score",
    "breakdown",
    "grade_notes",
    "grade_error",
    "original_score",
    "new_score",
    "score_delta",
    "regrade",
    "regrade_result",
}
RETRYABLE_ERROR_MARKERS = (
    "429",
    "rate limit",
    "temporarily unavailable",
    "timeout",
    "timed out",
    "502",
    "503",
    "504",
    "connection reset",
    "connection aborted",
)
INVALID_FILENAME = '<>:"/\\|?*'
RECOVERED_HYBRID_BREAKDOWNS_FILE = Path(__file__).with_name(
    "recovered_hybrid_automated_breakdowns.json"
)
_RECOVERED_HYBRID_BREAKDOWNS_CACHE: Optional[dict[str, dict[str, float]]] = None


def utc_now() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat()


def local_stamp() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def safe_slug(value: str) -> str:
    text = value.strip()
    for char in INVALID_FILENAME:
        text = text.replace(char, "_")
    text = "_".join(text.split())
    return text.strip("._") or "item"


def read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8-sig") as handle:
        return json.load(handle)


def write_json_atomic(path: Path, value: Any) -> None:
    """Atomically write JSON with Windows sharing-violation retries.

    Windows may temporarily deny os.replace() while antivirus, indexing,
    PowerShell Get-Content, or another monitor has the destination open.
    The payload is written to a process-specific temporary file, fsynced,
    and replace is retried with bounded exponential backoff.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(
        path.name
        + f".{os.getpid()}.{threading.get_ident()}.tmp"
    )
    payload = json.dumps(
        value,
        ensure_ascii=False,
        indent=2,
    )

    with temp.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write(payload)
        handle.flush()
        os.fsync(handle.fileno())

    delays = (
        0.05,
        0.10,
        0.20,
        0.40,
        0.80,
        1.00,
        1.00,
        1.00,
        1.00,
        1.00,
    )

    try:
        for attempt, delay in enumerate(delays, start=1):
            try:
                os.replace(temp, path)
                return
            except PermissionError:
                if attempt == len(delays):
                    raise
                time.sleep(delay)
    finally:
        try:
            if temp.exists():
                temp.unlink()
        except OSError:
            pass


def append_jsonl(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(value, ensure_ascii=False) + "\n")
        handle.flush()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def tree_sha256(root: Path) -> tuple[str, int, int]:
    digest = hashlib.sha256()
    file_count = 0
    total_bytes = 0
    for path in sorted(
        (item for item in root.rglob("*") if item.is_file()),
        key=lambda item: item.as_posix().lower(),
    ):
        relative = path.relative_to(root).as_posix()
        stat = path.stat()
        file_digest = sha256_file(path)
        digest.update(relative.encode("utf-8"))
        digest.update(b"\0")
        digest.update(str(stat.st_size).encode("ascii"))
        digest.update(b"\0")
        digest.update(file_digest.encode("ascii"))
        digest.update(b"\n")
        file_count += 1
        total_bytes += stat.st_size
    return digest.hexdigest(), file_count, total_bytes


def load_config(path: Path) -> dict[str, Any]:
    data = read_json(path)
    required = (
        "expected_pinchbench_commit",
        "expected_manifest_sha256",
        "skill_dir",
        "output_root",
        "judge_model",
        "runs",
    )
    missing = [key for key in required if key not in data]
    if missing:
        raise ValueError("Config keys missing: " + ", ".join(missing))
    return data


def load_reference_runner() -> Any:
    return importlib.import_module("reference_runner_contract")


def load_grader(skill_dir: Path) -> Any:
    scripts_dir = (skill_dir / "scripts").resolve()
    grading_path = scripts_dir / "lib_grading.py"
    if not grading_path.is_file():
        raise FileNotFoundError(
            f"PinchBench grading engine not found: {grading_path}"
        )
    scripts_text = str(scripts_dir)
    if scripts_text not in sys.path:
        sys.path.insert(0, scripts_text)
    if "lib_grading" in sys.modules:
        del sys.modules["lib_grading"]
    return importlib.import_module("lib_grading")


def results_rows(path: Path) -> list[dict[str, Any]]:
    data = read_json(path)
    if isinstance(data, dict):
        rows = data.get("results")
    else:
        rows = data
    if not isinstance(rows, list):
        raise ValueError(f"No results list in {path}")
    return [row for row in rows if isinstance(row, dict)]


def result_by_task(path: Path) -> dict[str, dict[str, Any]]:
    rows = results_rows(path)
    mapped: dict[str, dict[str, Any]] = {}
    for row in rows:
        task_id = str(row.get("task_id") or "").strip()
        if not task_id:
            continue
        if task_id in mapped:
            raise ValueError(f"Duplicate task id in {path}: {task_id}")
        mapped[task_id] = row
    return mapped


def source_workspace(run_path: Path, row: dict[str, Any], task_id: str) -> Path:
    explicit = str(row.get("workspace") or "").strip()
    candidates = []
    if explicit and explicit not in {
        "已清理；使用 --keep-workspaces 可保留",
        "cleaned",
    }:
        candidate = Path(explicit)
        if not candidate.is_absolute():
            candidate = run_path / candidate
        candidates.append(candidate)
    candidates.extend(
        (
            run_path / "workspaces" / task_id,
            run_path / "workspace" / task_id,
        )
    )
    for candidate in candidates:
        if candidate.is_dir():
            return candidate.resolve()
    raise FileNotFoundError(
        f"Workspace not found for {task_id}: "
        + " | ".join(str(item) for item in candidates)
    )


def transcript_events(
    run_path: Path,
    row: dict[str, Any],
    task_id: str,
) -> tuple[list[Any], str]:
    candidates: list[Path] = []
    explicit = str(row.get("normalized_transcript_path") or "").strip()
    if explicit:
        candidate = Path(explicit)
        if not candidate.is_absolute():
            candidate = run_path / candidate
        candidates.append(candidate)

    transcript_value = str(row.get("transcript") or "").strip()
    if transcript_value:
        candidate = Path(transcript_value)
        if not candidate.is_absolute():
            candidate = run_path / candidate
        if candidate.is_file():
            candidates.append(candidate)
        else:
            candidates.extend(
                (
                    candidate / "normalized.jsonl",
                    candidate / "normalized_transcript.jsonl",
                )
            )

    candidates.extend(
        (
            run_path / "transcripts" / task_id / "normalized.jsonl",
            run_path / "transcripts" / task_id / "normalized_transcript.jsonl",
        )
    )

    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate).lower()
        if key in seen:
            continue
        seen.add(key)
        if not candidate.is_file():
            continue
        events: list[Any] = []
        with candidate.open("r", encoding="utf-8-sig", errors="replace") as handle:
            for line in handle:
                if not line.strip():
                    continue
                try:
                    events.append(json.loads(line))
                except json.JSONDecodeError:
                    events.append({"raw": line.rstrip("\r\n")})
        return events, str(candidate.resolve())
    return [], ""


def normalize_score(result: Any) -> Optional[float]:
    raw = getattr(result, "score", None)
    if raw is None:
        return None
    max_score = float(getattr(result, "max_score", 1.0) or 1.0)
    score = float(raw)
    if max_score != 1.0:
        score /= max_score
    return max(0.0, min(1.0, score))


def grading_error(result: Any, grading_type: str) -> Optional[str]:
    notes = str(getattr(result, "notes", "") or "")
    if grading_type not in {"llm_judge", "hybrid"}:
        return None
    lower_notes = notes.lower()
    markers = (
        "llm judge failed",
        "no parseable response",
        "response parsed but no score",
        "openrouter_api_key not set",
        "judge api call failed",
    )
    if any(marker in lower_notes for marker in markers):
        return notes or "LLM judge failed"
    return None


def prepare_execution(
    original: dict[str, Any],
    workspace: Path,
    transcript: list[Any],
) -> dict[str, Any]:
    execution = {
        key: value
        for key, value in original.items()
        if key not in REMOVE_BEFORE_GRADING
        and not key.startswith("regrade_")
    }
    execution["workspace"] = str(workspace)
    execution["transcript"] = transcript
    return execution


def cache_stats(grader: Any) -> dict[str, Any]:
    if hasattr(grader, "get_judge_cache_stats"):
        try:
            stats = grader.get_judge_cache_stats()
            if isinstance(stats, dict):
                return stats
        except Exception:
            pass
    return {}


def configure_cache(grader: Any, cache_dir: Path) -> None:
    cache_dir.mkdir(parents=True, exist_ok=True)
    if hasattr(grader, "set_judge_cache_dir"):
        grader.set_judge_cache_dir(cache_dir)



def _openrouter_message_text(message: Any) -> str:
    """Extract visible assistant text from OpenRouter chat-completions."""
    if not isinstance(message, dict):
        return ""

    content = message.get("content")
    if isinstance(content, str):
        return content

    if isinstance(content, list):
        pieces: list[str] = []
        for block in content:
            if isinstance(block, str):
                pieces.append(block)
                continue
            if not isinstance(block, dict):
                continue
            text = block.get("text")
            if isinstance(text, str):
                pieces.append(text)
                continue
            nested = block.get("content")
            if isinstance(nested, str):
                pieces.append(nested)
        return "".join(pieces)

    return ""


def install_openrouter_judge_compatibility(
    grader: Any,
    *,
    agent: str,
    task_id: str,
    worker_result_path: Path,
) -> dict[str, Any]:
    """Install a transport-only OpenRouter compatibility shim."""
    context: dict[str, Any] = {
        "enabled": False,
        "api_attempts": 0,
        "empty_content_retries": 0,
        "length_retries": 0,
        "network_failures": 0,
        "raw_response_files": [],
        "completion_budgets": [],
        "last_finish_reason": None,
        "last_usage": {},
    }

    original_call = getattr(grader, "call_judge_api", None)
    if not callable(original_call):
        return context

    module = sys.modules.get(getattr(original_call, "__module__", ""))
    system_message = getattr(module, "_JUDGE_SYSTEM_MSG", None)
    if not isinstance(system_message, str) or not system_message:
        return context

    raw_root = (
        worker_result_path.parents[2]
        / "judge_raw_responses"
        / safe_slug(agent)
        / safe_slug(task_id)
    )
    raw_root.mkdir(parents=True, exist_ok=True)
    context["enabled"] = True

    def save_raw(
        attempt_number: int,
        payload: dict[str, Any],
        *,
        status_code: Optional[int],
        response_data: Any,
        error_text: str = "",
    ) -> None:
        record = {
            "schema_version": 1,
            "at": utc_now(),
            "agent": agent,
            "task_id": task_id,
            "attempt": attempt_number,
            "model": payload.get("model"),
            "max_completion_tokens": payload.get(
                "max_completion_tokens"
            ),
            "temperature": payload.get("temperature"),
            "prompt_sha256": hashlib.sha256(
                str(payload["messages"][1]["content"]).encode("utf-8")
            ).hexdigest(),
            "http_status": status_code,
            "response": response_data,
            "error": error_text,
        }
        path = raw_root / f"attempt_{attempt_number:02d}.json"
        write_json_atomic(path, record)
        context["raw_response_files"].append(str(path))

    def compatible_call(
        *,
        prompt: str,
        model: str,
        timeout_seconds: float = 120.0,
    ) -> dict[str, Any]:
        if not model.startswith("openrouter/"):
            return original_call(
                prompt=prompt,
                model=model,
                timeout_seconds=timeout_seconds,
            )

        api_key = os.environ.get("OPENROUTER_API_KEY")
        if not api_key:
            return {
                "status": "error",
                "text": "",
                "error": "OPENROUTER_API_KEY not set",
            }

        bare_model = model.removeprefix("openrouter/")
        budgets = (2048, 4096, 8192, 16384)
        last_description = "empty content"

        for local_index, budget in enumerate(budgets, start=1):
            context["api_attempts"] += 1
            attempt_number = int(context["api_attempts"])
            context["completion_budgets"].append(budget)

            payload_object = {
                "model": bare_model,
                "messages": [
                    {
                        "role": "system",
                        "content": system_message,
                    },
                    {
                        "role": "user",
                        "content": prompt,
                    },
                ],
                "temperature": 0.0,
                "max_completion_tokens": budget,
            }
            payload = json.dumps(
                payload_object,
                ensure_ascii=False,
            ).encode("utf-8")
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
                "HTTP-Referer": "https://pinchbench.com",
                "X-Title": "PinchBench-Judge",
            }
            req = urllib_request.Request(
                "https://openrouter.ai/api/v1/chat/completions",
                data=payload,
                headers=headers,
                method="POST",
            )

            try:
                with urllib_request.urlopen(
                    req,
                    timeout=timeout_seconds,
                ) as response:
                    status_code = getattr(response, "status", 200)
                    response_data = json.loads(
                        response.read().decode("utf-8")
                    )
            except urllib_error.HTTPError as exc:
                body = ""
                try:
                    body = exc.read().decode(
                        "utf-8",
                        errors="replace",
                    )
                except Exception:
                    pass
                save_raw(
                    attempt_number,
                    payload_object,
                    status_code=exc.code,
                    response_data=None,
                    error_text=body[:4000],
                )
                return {
                    "status": "error",
                    "text": "",
                    "error": f"HTTP {exc.code}: {body[:500]}",
                }
            except (
                http_client.IncompleteRead,
                http_client.RemoteDisconnected,
                ConnectionResetError,
                BrokenPipeError,
            ) as exc:
                context["network_failures"] += 1
                save_raw(
                    attempt_number,
                    payload_object,
                    status_code=None,
                    response_data=None,
                    error_text=repr(exc),
                )
                return {
                    "status": "error",
                    "text": "",
                    "error": repr(exc),
                }
            except urllib_error.URLError as exc:
                context["network_failures"] += 1
                save_raw(
                    attempt_number,
                    payload_object,
                    status_code=None,
                    response_data=None,
                    error_text=str(exc),
                )
                return {
                    "status": "error",
                    "text": "",
                    "error": str(exc),
                }
            except (TimeoutError, OSError) as exc:
                context["network_failures"] += 1
                save_raw(
                    attempt_number,
                    payload_object,
                    status_code=None,
                    response_data=None,
                    error_text=str(exc),
                )
                return {
                    "status": "timeout",
                    "text": "",
                    "error": str(exc) or "Request timed out",
                }

            save_raw(
                attempt_number,
                payload_object,
                status_code=status_code,
                response_data=response_data,
            )

            choices = (
                response_data.get("choices", [])
                if isinstance(response_data, dict)
                else []
            )
            if not choices:
                return {
                    "status": "error",
                    "text": "",
                    "error": "No choices in response",
                }

            choice = choices[0] if isinstance(choices[0], dict) else {}
            message = choice.get("message", {})
            text = _openrouter_message_text(message)
            finish_reason = choice.get("finish_reason")
            usage = (
                response_data.get("usage", {})
                if isinstance(response_data, dict)
                else {}
            )
            context["last_finish_reason"] = finish_reason
            context["last_usage"] = usage

            has_visible_text = (
                isinstance(text, str)
                and bool(text.strip())
            )
            normalized_finish_reason = str(
                finish_reason or ""
            ).strip().lower()
            was_truncated = normalized_finish_reason in {
                "length",
                "max_tokens",
            }

            if has_visible_text and not was_truncated:
                return {
                    "status": "success",
                    "text": text,
                }

            if was_truncated:
                context["length_retries"] += 1
                last_description = (
                    "OpenRouter truncated the Judge response; "
                    f"finish_reason={finish_reason!r}; "
                    f"visible_text_length={len(text)}; "
                    f"budget={budget}; usage={usage}"
                )
            else:
                context["empty_content_retries"] += 1
                last_description = (
                    "OpenRouter returned null/empty message.content; "
                    f"finish_reason={finish_reason!r}; "
                    f"budget={budget}; usage={usage}"
                )

            if local_index < len(budgets):
                time.sleep(min(4, 2 ** (local_index - 1)))

        raise RuntimeError(
            last_description
            + "; compatibility retries exhausted; raw responses: "
            + " | ".join(context["raw_response_files"])
        )

    grader.call_judge_api = compatible_call
    return context



def should_preserve_original_automated_na(
    task: Any,
    original: dict[str, Any],
) -> bool:
    """Preserve pre-existing automated-grader N/A in judge-only regrades."""
    return (
        str(getattr(task, "grading_type", "") or "") == "automated"
        and original.get("score") is None
        and bool(str(original.get("grade_error") or "").strip())
    )


def preserved_automated_na_payload(
    *,
    job: dict[str, Any],
    task: Any,
    original: dict[str, Any],
    judge_model: str,
    source: Path,
    source_hash: str,
    source_files: int,
    source_bytes: int,
    scratch: Path,
    transcript_path: str,
    transcript_event_count: int,
    started: float,
    judge_compat: dict[str, Any],
) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "agent": job["agent"],
        "task_id": str(job["task_id"]),
        "task_name": task.name,
        "category": task.category,
        "grading_type": task.grading_type,
        "judge_required": False,
        "judge_model": judge_model,
        "judge_backend": "api",
        "regrade_scope": "judge_model_only",
        "preserved_original_na": True,
        "preservation_reason": (
            "pre_existing_automated_grader_failure"
        ),
        "original_score": None,
        "new_score": None,
        "score_delta": None,
        "original_breakdown": original.get("breakdown") or {},
        "new_breakdown": original.get("breakdown") or {},
        "original_grade_notes": str(
            original.get("grade_notes") or ""
        ),
        "new_grade_notes": (
            "Preserved as N/A because this automated task already "
            "failed in the original formal run. Regrading is scoped "
            "to changing only the Judge model."
        ),
        "original_grade_error": str(
            original.get("grade_error") or ""
        ),
        "new_grade_error": None,
        "original_status": original.get("status"),
        "original_success": original.get("success"),
        "source_workspace": str(source),
        "source_workspace_sha256": source_hash,
        "source_workspace_file_count": source_files,
        "source_workspace_bytes": source_bytes,
        "scratch_workspace": str(scratch),
        "transcript_path": transcript_path,
        "transcript_event_count": transcript_event_count,
        "grading_elapsed_seconds": 0.0,
        "worker_elapsed_seconds": round(
            time.monotonic() - started,
            3,
        ),
        "cache_stats": {},
        "judge_transport_compatibility": judge_compat,
        "completed_at": utc_now(),
        "worker_status": "skipped",
        "error": "",
    }



def _score_dict_mean(values: dict[str, Any]) -> float:
    numbers = [
        float(value)
        for value in values.values()
        if isinstance(value, (int, float))
        and not isinstance(value, bool)
    ]
    return sum(numbers) / len(numbers) if numbers else 0.0


def _load_recovered_hybrid_breakdowns() -> dict[str, dict[str, float]]:
    global _RECOVERED_HYBRID_BREAKDOWNS_CACHE
    if _RECOVERED_HYBRID_BREAKDOWNS_CACHE is not None:
        return _RECOVERED_HYBRID_BREAKDOWNS_CACHE

    recovered: dict[str, dict[str, float]] = {}
    if RECOVERED_HYBRID_BREAKDOWNS_FILE.is_file():
        payload = read_json(RECOVERED_HYBRID_BREAKDOWNS_FILE)
        raw_tasks = payload.get("tasks", {}) if isinstance(payload, dict) else {}
        if not isinstance(raw_tasks, dict):
            raise ValueError(
                "Invalid recovered hybrid breakdown file: tasks must be an object"
            )
        for task_id, raw_breakdown in raw_tasks.items():
            if not isinstance(raw_breakdown, dict):
                continue
            clean: dict[str, float] = {}
            for key, value in raw_breakdown.items():
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    clean[str(key)] = float(value)
            if clean:
                recovered[str(task_id)] = clean

    _RECOVERED_HYBRID_BREAKDOWNS_CACHE = recovered
    return recovered


def original_automated_breakdown(
    task: Any,
    original: dict[str, Any],
) -> dict[str, float]:
    raw = original.get("breakdown") or {}
    result: dict[str, float] = {}

    if str(getattr(task, "grading_type", "") or "") == "hybrid":
        for key, value in raw.items():
            key_text = str(key)
            if not key_text.startswith("automated."):
                continue
            if isinstance(value, (int, float)) and not isinstance(
                value,
                bool,
            ):
                result[key_text.split(".", 1)[1]] = float(value)
        if result:
            return result

        task_id = str(getattr(task, "task_id", "") or "")
        recovered = _load_recovered_hybrid_breakdowns().get(task_id)
        if recovered:
            return dict(recovered)
        return {}

    for key, value in raw.items():
        if isinstance(value, (int, float)) and not isinstance(
            value,
            bool,
        ):
            result[str(key)] = float(value)
    return result


def preserved_original_automated_payload(
    *,
    job: dict[str, Any],
    task: Any,
    original: dict[str, Any],
    judge_model: str,
    source: Path,
    source_hash: str,
    source_files: int,
    source_bytes: int,
    scratch: Path,
    transcript_path: str,
    transcript_event_count: int,
    started: float,
    judge_compat: dict[str, Any],
) -> dict[str, Any]:
    old_score = original.get("score")
    old_score_number = (
        float(old_score) if old_score is not None else None
    )
    return {
        "schema_version": 1,
        "agent": job["agent"],
        "task_id": str(job["task_id"]),
        "task_name": task.name,
        "category": task.category,
        "grading_type": task.grading_type,
        "judge_required": False,
        "judge_model": judge_model,
        "judge_backend": "api",
        "regrade_scope": "judge_model_only",
        "preserved_original_automated": True,
        "preservation_reason": (
            "automated_result_outside_judge_change_scope"
        ),
        "original_score": old_score_number,
        "new_score": old_score_number,
        "score_delta": (
            0.0 if old_score_number is not None else None
        ),
        "original_breakdown": original.get("breakdown") or {},
        "new_breakdown": original.get("breakdown") or {},
        "original_grade_notes": str(
            original.get("grade_notes") or ""
        ),
        "new_grade_notes": (
            "Original automated result preserved exactly because "
            "this regrade changes only the Judge model."
        ),
        "original_grade_error": original.get("grade_error"),
        "new_grade_error": None,
        "original_status": original.get("status"),
        "original_success": original.get("success"),
        "source_workspace": str(source),
        "source_workspace_sha256": source_hash,
        "source_workspace_file_count": source_files,
        "source_workspace_bytes": source_bytes,
        "scratch_workspace": str(scratch),
        "transcript_path": transcript_path,
        "transcript_event_count": transcript_event_count,
        "grading_elapsed_seconds": 0.0,
        "worker_elapsed_seconds": round(
            time.monotonic() - started,
            3,
        ),
        "cache_stats": {},
        "judge_transport_compatibility": judge_compat,
        "completed_at": utc_now(),
        "worker_status": "skipped",
        "error": "",
    }


def freeze_hybrid_automated_component(
    grader: Any,
    task: Any,
    original: dict[str, Any],
) -> Any:
    """Replace only the hybrid automated pass with its original result."""
    breakdown = original_automated_breakdown(task, original)
    if not breakdown:
        raise RuntimeError(
            "Original hybrid automated breakdown is missing for "
            + str(getattr(task, "task_id", "unknown"))
        )

    original_function = getattr(grader, "_grade_automated")
    frozen_score = _score_dict_mean(breakdown)

    def frozen_automated(
        task: Any,
        execution_result: dict[str, Any],
        skill_dir: Optional[Path] = None,
        verbose: bool = False,
    ) -> Any:
        return grader.GradeResult(
            task_id=task.task_id,
            score=frozen_score,
            max_score=1.0,
            grading_type="automated",
            breakdown=dict(breakdown),
            notes="",
        )

    grader._grade_automated = frozen_automated
    return original_function



def install_bounded_workspace_evidence(
    grader: Any,
    *,
    task_id: str,
) -> dict[str, Any]:
    """Bound irrelevant workspace text for the image-generation Judge.

    The frozen grader recursively concatenates every UTF-8-readable workspace
    file. For task_image_gen, dependency environments can dwarf the actual
    task evidence and exceed the model context. This compatibility layer keeps
    the task prompt, transcript summary, rubric, Judge model, and score parser
    unchanged. It only prevents dependency/build trees from being treated as
    agent-authored deliverables and caps workspace text evidence.
    """
    context: dict[str, Any] = {
        "enabled": False,
        "task_id": task_id,
        "policy": "official_default",
        "workspace": "",
        "expected_output": "robot_cafe.png",
        "expected_output_present": False,
        "expected_output_bytes": None,
        "expected_output_sha256": None,
        "files_seen": 0,
        "files_included": 0,
        "files_skipped_dependency": 0,
        "files_skipped_binary_or_type": 0,
        "files_skipped_size_or_budget": 0,
        "included_characters": 0,
        "character_budget": 120000,
        "per_file_character_budget": 32000,
        "included_files": [],
    }

    if task_id != "task_image_gen":
        return context

    original_reader = getattr(
        grader,
        "_read_workspace_files",
        None,
    )
    if not callable(original_reader):
        return context

    skip_parts = {
        ".git",
        ".openclaw",
        "__pycache__",
        "node_modules",
        "skills",
        ".venv",
        "venv",
        "env",
        "site-packages",
        "dist-packages",
        ".pytest_cache",
        ".mypy_cache",
        ".ruff_cache",
        ".cache",
    }
    allowed_suffixes = {
        ".txt",
        ".md",
        ".py",
        ".json",
        ".jsonl",
        ".yaml",
        ".yml",
        ".toml",
        ".ini",
        ".cfg",
        ".ps1",
        ".bat",
        ".cmd",
        ".js",
        ".ts",
        ".html",
        ".css",
        ".csv",
        ".log",
    }
    max_files = 40
    max_file_bytes = 256 * 1024
    total_budget = int(context["character_budget"])
    per_file_budget = int(
        context["per_file_character_budget"]
    )

    def bounded_reader(workspace_path: str) -> str:
        workspace = Path(workspace_path)
        context["workspace"] = str(workspace)

        if not workspace.is_dir():
            return (
                "### Workspace Evidence Summary\n"
                "Workspace directory is missing."
            )

        output_path = workspace / "robot_cafe.png"
        context["expected_output_present"] = output_path.is_file()
        if output_path.is_file():
            stat = output_path.stat()
            context["expected_output_bytes"] = stat.st_size
            context["expected_output_sha256"] = sha256_file(
                output_path
            )

        sections: list[str] = [
            "### Workspace Evidence Summary",
            (
                "Expected output robot_cafe.png: "
                + (
                    "PRESENT"
                    if context["expected_output_present"]
                    else "MISSING"
                )
            ),
        ]

        if context["expected_output_present"]:
            sections.append(
                "robot_cafe.png size: "
                + str(context["expected_output_bytes"])
                + " bytes"
            )
            sections.append(
                "robot_cafe.png sha256: "
                + str(context["expected_output_sha256"])
            )

        remaining = total_budget
        for path in sorted(
            workspace.rglob("*"),
            key=lambda item: item.as_posix().lower(),
        ):
            if not path.is_file():
                continue

            context["files_seen"] += 1
            relative = path.relative_to(workspace)
            lower_parts = {
                part.lower()
                for part in relative.parts
            }

            if lower_parts & skip_parts:
                context["files_skipped_dependency"] += 1
                continue

            if path.name.lower() == "robot_cafe.png":
                continue

            if (
                path.suffix.lower() not in allowed_suffixes
                or path.stat().st_size > max_file_bytes
            ):
                context[
                    "files_skipped_binary_or_type"
                ] += 1
                continue

            if (
                context["files_included"] >= max_files
                or remaining <= 0
            ):
                context[
                    "files_skipped_size_or_budget"
                ] += 1
                continue

            try:
                content = path.read_text(
                    encoding="utf-8",
                    errors="strict",
                )
            except (OSError, UnicodeDecodeError):
                context[
                    "files_skipped_binary_or_type"
                ] += 1
                continue

            content = content[:per_file_budget]
            if len(content) > remaining:
                content = content[:remaining]

            if not content.strip():
                continue

            sections.append(
                "### File: "
                + relative.as_posix()
                + "\n"
                + content
            )
            context["files_included"] += 1
            context["included_files"].append(
                relative.as_posix()
            )
            context["included_characters"] += len(content)
            remaining -= len(content)

        sections.insert(
            2,
            (
                "Bounded workspace evidence policy: dependency and "
                "environment directories excluded; user-authored text "
                "limited to "
                + str(total_budget)
                + " characters."
            ),
        )
        return "\n\n".join(sections)

    grader._read_workspace_files = bounded_reader
    context["enabled"] = True
    context["policy"] = (
        "task_image_gen_dependency_exclusion_and_120k_char_cap"
    )
    return context


def worker_job(job_file: Path) -> int:
    job = read_json(job_file)
    started = time.monotonic()
    output_path = Path(job["worker_result_path"])
    scratch = Path(job["scratch_workspace"])
    source = Path(job["source_workspace"])
    skill_dir = Path(job["skill_dir"])
    task_id = str(job["task_id"])
    original = job["original_result"]
    judge_model = str(job["judge_model"])
    judge_timeout = float(job["judge_timeout_seconds"])
    keep_scratch = bool(job.get("keep_scratch_workspace"))

    result_payload: dict[str, Any]
    workspace_evidence_compat: dict[str, Any] = {
        "enabled": False,
        "task_id": task_id,
        "policy": "official_default",
    }
    judge_compat: dict[str, Any] = {
        "enabled": False,
        "api_attempts": 0,
        "empty_content_retries": 0,
        "length_retries": 0,
        "network_failures": 0,
        "raw_response_files": [],
        "completion_budgets": [],
        "last_finish_reason": None,
        "last_usage": {},
    }
    try:
        if scratch.exists():
            shutil.rmtree(scratch, ignore_errors=True)
        scratch.parent.mkdir(parents=True, exist_ok=True)

        source_hash, source_files, source_bytes = tree_sha256(source)
        shutil.copytree(source, scratch, copy_function=shutil.copy2)
        copied_hash, copied_files, copied_bytes = tree_sha256(scratch)
        if (
            source_hash != copied_hash
            or source_files != copied_files
            or source_bytes != copied_bytes
        ):
            raise RuntimeError("Scratch workspace copy verification failed.")

        reference = load_reference_runner()
        tasks, _ = reference.load_tasks(skill_dir / "tasks")
        task_map = {task.task_id: task for task in tasks}
        if task_id not in task_map:
            raise KeyError(f"Task not loaded from manifest: {task_id}")
        task = task_map[task_id]

        grader = load_grader(skill_dir)
        configure_cache(grader, Path(job["judge_cache_dir"]))
        judge_compat = install_openrouter_judge_compatibility(
            grader,
            agent=str(job["agent"]),
            task_id=task_id,
            worker_result_path=output_path,
        )
        workspace_evidence_compat = (
            install_bounded_workspace_evidence(
                grader,
                task_id=task_id,
            )
        )

        transcript, transcript_path = transcript_events(
            Path(job["run_path"]),
            original,
            task_id,
        )
        execution = prepare_execution(original, scratch, transcript)

        if should_preserve_original_automated_na(task, original):
            result_payload = preserved_automated_na_payload(
                job=job,
                task=task,
                original=original,
                judge_model=judge_model,
                source=source,
                source_hash=source_hash,
                source_files=source_files,
                source_bytes=source_bytes,
                scratch=scratch,
                transcript_path=transcript_path,
                transcript_event_count=len(transcript),
                started=started,
                judge_compat=judge_compat,
            )
            write_json_atomic(output_path, result_payload)
            return 0

        if (
            str(getattr(task, "grading_type", "") or "")
            == "automated"
        ):
            result_payload = preserved_original_automated_payload(
                job=job,
                task=task,
                original=original,
                judge_model=judge_model,
                source=source,
                source_hash=source_hash,
                source_files=source_files,
                source_bytes=source_bytes,
                scratch=scratch,
                transcript_path=transcript_path,
                transcript_event_count=len(transcript),
                started=started,
                judge_compat=judge_compat,
            )
            write_json_atomic(output_path, result_payload)
            return 0

        if str(original.get("status") or "") in {
            "missing_prerequisite",
            "missing_fixture",
        }:
            result_payload = {
                "schema_version": 1,
                "agent": job["agent"],
                "task_id": task_id,
                "task_name": task.name,
                "category": task.category,
                "grading_type": task.grading_type,
                "judge_required": task.grading_type
                in {"hybrid", "llm_judge"},
                "judge_model": judge_model,
                "judge_backend": "api",
                "original_score": original.get("score"),
                "new_score": None,
                "score_delta": None,
                "original_breakdown": original.get("breakdown") or {},
                "new_breakdown": {},
                "original_grade_notes": str(
                    original.get("grade_notes") or ""
                ),
                "new_grade_notes": (
                    "Skipped exactly as the formal runner: "
                    f"status={original.get('status')}"
                ),
                "original_grade_error": original.get("grade_error"),
                "new_grade_error": None,
                "original_status": original.get("status"),
                "original_success": original.get("success"),
                "source_workspace": str(source),
                "source_workspace_sha256": source_hash,
                "source_workspace_file_count": source_files,
                "source_workspace_bytes": source_bytes,
                "scratch_workspace": str(scratch),
                "transcript_path": transcript_path,
                "transcript_event_count": len(transcript),
                "grading_elapsed_seconds": 0.0,
                "worker_elapsed_seconds": round(
                    time.monotonic() - started,
                    3,
                ),
                "cache_stats": cache_stats(grader),
                "judge_transport_compatibility": judge_compat,
                "workspace_evidence_compatibility": (
                    workspace_evidence_compat
                ),
                "completed_at": utc_now(),
                "worker_status": "skipped",
                "error": "",
            }
            write_json_atomic(output_path, result_payload)
            return 0

        frozen_auto_restore: Any = None
        if str(getattr(task, "grading_type", "") or "") == "hybrid":
            frozen_auto_restore = freeze_hybrid_automated_component(
                grader,
                task,
                original,
            )

        grade_started = time.monotonic()
        try:
            official_result = grader.grade_task(
                task=task,
                execution_result=execution,
                skill_dir=skill_dir,
                judge_model=judge_model,
                judge_timeout_seconds=judge_timeout,
                judge_backend="api",
                verbose=bool(job.get("verbose")),
            )
        finally:
            if frozen_auto_restore is not None:
                grader._grade_automated = frozen_auto_restore
        grade_elapsed = time.monotonic() - grade_started
        new_score = normalize_score(official_result)
        old_score = original.get("score")
        old_score_number = (
            float(old_score) if old_score is not None else None
        )
        delta = (
            new_score - old_score_number
            if new_score is not None and old_score_number is not None
            else None
        )
        notes = str(getattr(official_result, "notes", "") or "")
        breakdown = dict(getattr(official_result, "breakdown", {}) or {})
        grade_error = grading_error(official_result, task.grading_type)

        result_payload = {
            "schema_version": 1,
            "agent": job["agent"],
            "task_id": task_id,
            "task_name": task.name,
            "category": task.category,
            "grading_type": task.grading_type,
            "judge_required": task.grading_type in {"hybrid", "llm_judge"},
            "judge_model": judge_model,
            "judge_backend": "api",
            "original_score": old_score_number,
            "new_score": (
                round(new_score, 4)
                if new_score is not None
                else None
            ),
            "score_delta": (
                round(delta, 4)
                if delta is not None
                else None
            ),
            "original_breakdown": original.get("breakdown") or {},
            "new_breakdown": breakdown,
            "original_grade_notes": str(
                original.get("grade_notes") or ""
            ),
            "new_grade_notes": notes,
            "original_grade_error": original.get("grade_error"),
            "new_grade_error": grade_error,
            "original_status": original.get("status"),
            "original_success": original.get("success"),
            "source_workspace": str(source),
            "source_workspace_sha256": source_hash,
            "source_workspace_file_count": source_files,
            "source_workspace_bytes": source_bytes,
            "scratch_workspace": str(scratch),
            "transcript_path": transcript_path,
            "transcript_event_count": len(transcript),
            "grading_elapsed_seconds": round(grade_elapsed, 3),
            "worker_elapsed_seconds": round(
                time.monotonic() - started,
                3,
            ),
            "cache_stats": cache_stats(grader),
            "judge_transport_compatibility": judge_compat,
            "workspace_evidence_compatibility": (
                workspace_evidence_compat
            ),
            "completed_at": utc_now(),
            "worker_status": (
                "completed"
                if grade_error is None and new_score is not None
                else "grade_error"
            ),
            "error": "",
        }
    except Exception as exc:
        result_payload = {
            "schema_version": 1,
            "agent": job.get("agent"),
            "task_id": task_id,
            "judge_model": judge_model,
            "worker_status": "failed",
            "new_score": None,
            "new_breakdown": {},
            "new_grade_notes": "",
            "new_grade_error": str(exc),
            "judge_transport_compatibility": judge_compat,
            "workspace_evidence_compatibility": (
                workspace_evidence_compat
            ),
            "worker_elapsed_seconds": round(
                time.monotonic() - started,
                3,
            ),
            "completed_at": utc_now(),
            "error": (
                f"{exc}\n{traceback.format_exc(limit=14)}"
            ),
        }
    finally:
        if not keep_scratch:
            shutil.rmtree(scratch, ignore_errors=True)

    write_json_atomic(output_path, result_payload)
    return 0 if result_payload["worker_status"] != "failed" else 1


def db_connect(path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(path)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA journal_mode=WAL")
    connection.execute("PRAGMA synchronous=FULL")
    return connection


def initialize_db(path: Path) -> None:
    with db_connect(path) as db:
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_order INTEGER NOT NULL,
                agent TEXT NOT NULL,
                agent_slug TEXT NOT NULL,
                run_path TEXT NOT NULL,
                task_id TEXT NOT NULL,
                grading_type TEXT NOT NULL,
                judge_required INTEGER NOT NULL,
                source_workspace TEXT NOT NULL,
                original_score REAL,
                original_status TEXT,
                status TEXT NOT NULL DEFAULT 'pending',
                attempts INTEGER NOT NULL DEFAULT 0,
                started_at TEXT,
                ended_at TEXT,
                heartbeat_at TEXT,
                worker_pid INTEGER,
                elapsed_seconds REAL,
                new_score REAL,
                score_delta REAL,
                grade_error TEXT,
                worker_result_path TEXT,
                stdout_path TEXT,
                stderr_path TEXT,
                UNIQUE(agent, task_id)
            );

            CREATE TABLE IF NOT EXISTS metadata (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            """
        )


def db_set_meta(db: sqlite3.Connection, key: str, value: Any) -> None:
    db.execute(
        """
        INSERT INTO metadata(key, value)
        VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value=excluded.value
        """,
        (key, json.dumps(value, ensure_ascii=False)),
    )


def db_get_meta(db: sqlite3.Connection, key: str) -> Any:
    row = db.execute(
        "SELECT value FROM metadata WHERE key=?",
        (key,),
    ).fetchone()
    return json.loads(row["value"]) if row else None


def run_counts(db: sqlite3.Connection) -> dict[str, int]:
    counts = Counter()
    for row in db.execute(
        "SELECT status, COUNT(*) AS n FROM jobs GROUP BY status"
    ):
        counts[str(row["status"])] = int(row["n"])
    counts["total"] = int(
        db.execute("SELECT COUNT(*) AS n FROM jobs").fetchone()["n"]
    )
    return dict(counts)


def latest_run_dir(output_root: Path, kind: Optional[str] = None) -> Path:
    patterns = ["regrade_*", "smoke_*"] if kind is None else [kind + "_*"]
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(
            path
            for path in output_root.glob(pattern)
            if path.is_dir() and (path / "state.sqlite").is_file()
        )
    if not candidates:
        raise FileNotFoundError(f"No regrade run under {output_root}")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def create_run(
    config_path: Path,
    suite: str,
    judge_model_override: Optional[str],
) -> Path:
    config = load_config(config_path)
    judge_model = (
        judge_model_override.strip()
        if judge_model_override
        else str(config["judge_model"])
    )
    output_root = Path(config["output_root"])
    output_root.mkdir(parents=True, exist_ok=True)
    prefix = "smoke" if suite == "smoke" else "regrade"
    run_dir = output_root / f"{prefix}_{local_stamp()}"
    run_dir.mkdir(parents=False, exist_ok=False)
    for name in (
        "task_results",
        "worker_logs",
        "scratch",
        "job_inputs",
        "exports",
    ):
        (run_dir / name).mkdir()

    skill_dir = Path(config["skill_dir"]).resolve()
    reference = load_reference_runner()
    tasks, _ = reference.load_tasks(skill_dir / "tasks")
    task_map = {task.task_id: task for task in tasks}
    if suite == "smoke":
        selected_ids = [str(item) for item in config["smoke_tasks"]]
    elif suite == "full":
        selected_ids = [
            task.task_id
            for task in tasks
            if task.task_id not in reference.DEFAULT_SKIPPED_TASKS
        ]
    else:
        raise ValueError(f"Unsupported suite: {suite}")

    missing = [task_id for task_id in selected_ids if task_id not in task_map]
    if missing:
        raise ValueError("Smoke/full task ids missing: " + ", ".join(missing))

    initialize_db(run_dir / "state.sqlite")
    source_snapshot: list[dict[str, Any]] = []
    job_order = 0

    with db_connect(run_dir / "state.sqlite") as db:
        for run_entry in config["runs"]:
            agent = str(run_entry["agent"])
            agent_slug = safe_slug(agent)
            run_path = Path(run_entry["path"]).resolve()
            run_config = read_json(run_path / "run_config.json")
            rows = result_by_task(run_path / "results.json")
            source_snapshot.append(
                {
                    "agent": agent,
                    "run_path": str(run_path),
                    "results_sha256": sha256_file(
                        run_path / "results.json"
                    ),
                    "run_config_sha256": sha256_file(
                        run_path / "run_config.json"
                    ),
                    "original_judge_model": run_config.get(
                        "judge_model"
                    ),
                    "pinchbench_commit": run_config.get(
                        "pinchbench_commit"
                    ),
                }
            )

            for task_id in selected_ids:
                original = rows[task_id]
                task = task_map[task_id]
                workspace = source_workspace(
                    run_path,
                    original,
                    task_id,
                )
                job_order += 1
                result_path = (
                    run_dir
                    / "task_results"
                    / agent_slug
                    / f"{task_id}.json"
                )
                stdout_path = (
                    run_dir
                    / "worker_logs"
                    / agent_slug
                    / f"{task_id}.stdout.txt"
                )
                stderr_path = (
                    run_dir
                    / "worker_logs"
                    / agent_slug
                    / f"{task_id}.stderr.txt"
                )
                db.execute(
                    """
                    INSERT INTO jobs(
                        job_order, agent, agent_slug, run_path,
                        task_id, grading_type, judge_required,
                        source_workspace, original_score,
                        original_status, status,
                        worker_result_path, stdout_path, stderr_path
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?, ?, ?)
                    """,
                    (
                        job_order,
                        agent,
                        agent_slug,
                        str(run_path),
                        task_id,
                        task.grading_type,
                        int(
                            task.grading_type
                            in {"hybrid", "llm_judge"}
                        ),
                        str(workspace),
                        original.get("score"),
                        original.get("status"),
                        str(result_path),
                        str(stdout_path),
                        str(stderr_path),
                    ),
                )

        metadata = {
            "script_revision": SCRIPT_REVISION,
            "created_at": utc_now(),
            "suite": suite,
            "judge_model": judge_model,
            "judge_backend": "api",
            "skill_dir": str(skill_dir),
            "expected_pinchbench_commit": config[
                "expected_pinchbench_commit"
            ],
            "expected_manifest_sha256": config[
                "expected_manifest_sha256"
            ],
            "judge_timeout_seconds": float(
                config["judge_timeout_seconds"]
            ),
            "job_hard_timeout_seconds": float(
                config["job_hard_timeout_seconds"]
            ),
            "source_runs": source_snapshot,
            "task_count_per_agent": len(selected_ids),
            "agent_count": len(config["runs"]),
            "job_count": job_order,
            "judge_required_job_count": int(
                db.execute(
                    "SELECT COUNT(*) AS n FROM jobs "
                    "WHERE judge_required=1"
                ).fetchone()["n"]
            ),
            "openrouter_key_present_at_creation": bool(
                os.environ.get("OPENROUTER_API_KEY")
            ),
            "platform": platform.platform(),
            "python": sys.version,
            "config_source": str(config_path.resolve()),
            "source_mutation_policy": "read-only; scratch copy per job",
        }
        for key, value in metadata.items():
            db_set_meta(db, key, value)
        db.commit()

    write_json_atomic(run_dir / "run_config.json", metadata)
    write_json_atomic(
        run_dir / "source_runs_snapshot.json",
        source_snapshot,
    )
    (run_dir / "progress.jsonl").touch()
    return run_dir


def kill_process_tree(process: subprocess.Popen[Any]) -> None:
    if process.poll() is not None:
        return
    if os.name == "nt":
        subprocess.run(
            [
                os.path.join(
                    os.environ.get("SystemRoot", r"C:\Windows"),
                    "System32",
                    "taskkill.exe",
                ),
                "/PID",
                str(process.pid),
                "/T",
                "/F",
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
    else:
        try:
            os.killpg(process.pid, signal.SIGTERM)
            time.sleep(2)
            if process.poll() is None:
                os.killpg(process.pid, signal.SIGKILL)
        except Exception:
            process.kill()


def heartbeat_payload(
    run_dir: Path,
    db: sqlite3.Connection,
    current: Optional[sqlite3.Row],
    state: str,
    current_elapsed: float = 0.0,
) -> dict[str, Any]:
    counts = run_counts(db)
    completed = counts.get("completed", 0)
    failed = counts.get("failed", 0)
    total = counts.get("total", 0)
    done = completed + failed
    elapsed_rows = db.execute(
        """
        SELECT elapsed_seconds FROM jobs
        WHERE status IN ('completed', 'failed')
        AND elapsed_seconds IS NOT NULL
        """
    ).fetchall()
    average = (
        sum(float(row["elapsed_seconds"]) for row in elapsed_rows)
        / len(elapsed_rows)
        if elapsed_rows
        else None
    )
    remaining = max(0, total - done)
    eta = average * remaining if average is not None else None
    scores = [
        float(row["new_score"])
        for row in db.execute(
            "SELECT new_score FROM jobs "
            "WHERE new_score IS NOT NULL"
        )
    ]
    return {
        "updated_at": utc_now(),
        "run_dir": str(run_dir),
        "state": state,
        "counts": counts,
        "completed_or_failed": done,
        "progress_fraction": done / total if total else 0.0,
        "mean_new_score": (
            sum(scores) / len(scores) if scores else None
        ),
        "average_job_seconds": average,
        "eta_seconds": eta,
        "current_job": (
            {
                "id": int(current["id"]),
                "agent": current["agent"],
                "task_id": current["task_id"],
                "grading_type": current["grading_type"],
                "judge_required": bool(current["judge_required"]),
                "worker_pid": current["worker_pid"],
                "elapsed_seconds": current_elapsed,
            }
            if current is not None
            else None
        ),
    }


def worker_command(
    script_path: Path,
    job_file: Path,
) -> tuple[list[str], dict[str, Any]]:
    command = [
        sys.executable,
        "-X",
        "utf8",
        str(script_path),
        "worker",
        "--job-file",
        str(job_file),
    ]
    kwargs: dict[str, Any] = {}
    if os.name != "nt":
        kwargs["start_new_session"] = True
    return command, kwargs


def reset_interrupted(db: sqlite3.Connection) -> int:
    cursor = db.execute(
        """
        UPDATE jobs
        SET status='pending',
            worker_pid=NULL,
            heartbeat_at=NULL,
            grade_error=CASE
                WHEN grade_error IS NULL OR grade_error=''
                THEN 'Recovered from interrupted run.'
                ELSE grade_error
            END
        WHERE status='running'
        """
    )
    db.commit()
    return cursor.rowcount


def run_queue(
    config_path: Path,
    run_dir: Path,
    retry_failed: bool,
    verbose: bool,
) -> int:
    config = load_config(config_path)
    run_dir = run_dir.resolve()
    state_path = run_dir / "state.sqlite"
    if not state_path.is_file():
        raise FileNotFoundError(state_path)
    if not os.environ.get("OPENROUTER_API_KEY"):
        raise RuntimeError(
            "OPENROUTER_API_KEY is missing. One OpenRouter key is "
            "required for the new Judge."
        )

    stop_path = run_dir / "STOP_REQUESTED"
    stop_path.unlink(missing_ok=True)
    script_path = Path(__file__).resolve()
    progress_path = run_dir / "progress.jsonl"
    heartbeat_path = run_dir / "heartbeat.json"
    job_timeout = float(config["job_hard_timeout_seconds"])
    poll_seconds = max(1.0, float(config["poll_seconds"]))
    keep_scratch = bool(config.get("keep_scratch_workspaces"))

    with db_connect(state_path) as db:
        reset_interrupted(db)
        if retry_failed:
            db.execute(
                """
                UPDATE jobs
                SET status='pending',
                    worker_pid=NULL,
                    heartbeat_at=NULL
                WHERE status='failed'
                """
            )
            db.commit()

        judge_model = str(db_get_meta(db, "judge_model"))
        skill_dir = str(db_get_meta(db, "skill_dir"))
        cache_dir = (
            Path(config["output_root"])
            / ".judge_cache"
            / safe_slug(judge_model)
        )
        cache_dir.mkdir(parents=True, exist_ok=True)
        db_set_meta(db, "judge_cache_dir", str(cache_dir))
        db_set_meta(db, "run_started_at", utc_now())
        db.commit()

        while True:
            if stop_path.exists():
                payload = heartbeat_payload(
                    run_dir,
                    db,
                    None,
                    "stopped",
                )
                write_json_atomic(heartbeat_path, payload)
                append_jsonl(
                    progress_path,
                    {
                        "event": "stop_acknowledged",
                        "at": utc_now(),
                    },
                )
                print("STOP_REQUESTED found. Queue stopped.")
                return 130

            job = db.execute(
                """
                SELECT * FROM jobs
                WHERE status='pending'
                ORDER BY job_order
                LIMIT 1
                """
            ).fetchone()
            if job is None:
                break

            result_path = Path(job["worker_result_path"])
            stdout_path = Path(job["stdout_path"])
            stderr_path = Path(job["stderr_path"])
            for path in (result_path, stdout_path, stderr_path):
                path.parent.mkdir(parents=True, exist_ok=True)
                path.unlink(missing_ok=True)

            run_path = Path(job["run_path"])
            originals = result_by_task(run_path / "results.json")
            original = originals[str(job["task_id"])]
            scratch = (
                run_dir
                / "scratch"
                / str(job["agent_slug"])
                / str(job["task_id"])
            )
            job_input = {
                "agent": job["agent"],
                "run_path": job["run_path"],
                "task_id": job["task_id"],
                "source_workspace": job["source_workspace"],
                "scratch_workspace": str(scratch),
                "worker_result_path": str(result_path),
                "skill_dir": skill_dir,
                "judge_model": judge_model,
                "judge_timeout_seconds": float(
                    config["judge_timeout_seconds"]
                ),
                "judge_cache_dir": str(cache_dir),
                "original_result": original,
                "keep_scratch_workspace": keep_scratch,
                "verbose": verbose,
            }
            job_file = (
                run_dir
                / "job_inputs"
                / f"{int(job['id']):04d}_"
                f"{job['agent_slug']}_{job['task_id']}.json"
            )
            write_json_atomic(job_file, job_input)

            command, popen_kwargs = worker_command(
                script_path,
                job_file,
            )
            started_monotonic = time.monotonic()
            started_at = utc_now()
            with stdout_path.open("wb") as stdout_handle, stderr_path.open(
                "wb"
            ) as stderr_handle:
                process = subprocess.Popen(
                    command,
                    cwd=str(Path(__file__).parent),
                    stdout=stdout_handle,
                    stderr=stderr_handle,
                    **popen_kwargs,
                )

                db.execute(
                    """
                    UPDATE jobs
                    SET status='running',
                        attempts=attempts+1,
                        started_at=?,
                        heartbeat_at=?,
                        worker_pid=?,
                        grade_error=NULL
                    WHERE id=?
                    """,
                    (
                        started_at,
                        started_at,
                        process.pid,
                        int(job["id"]),
                    ),
                )
                db.commit()
                append_jsonl(
                    progress_path,
                    {
                        "event": "job_started",
                        "at": started_at,
                        "job_id": int(job["id"]),
                        "agent": job["agent"],
                        "task_id": job["task_id"],
                        "grading_type": job["grading_type"],
                        "judge_required": bool(
                            job["judge_required"]
                        ),
                        "worker_pid": process.pid,
                    },
                )

                timed_out = False
                while process.poll() is None:
                    elapsed = time.monotonic() - started_monotonic
                    now = utc_now()
                    db.execute(
                        """
                        UPDATE jobs
                        SET heartbeat_at=?,
                            elapsed_seconds=?
                        WHERE id=?
                        """,
                        (now, elapsed, int(job["id"])),
                    )
                    db.commit()
                    refreshed = db.execute(
                        "SELECT * FROM jobs WHERE id=?",
                        (int(job["id"]),),
                    ).fetchone()
                    write_json_atomic(
                        heartbeat_path,
                        heartbeat_payload(
                            run_dir,
                            db,
                            refreshed,
                            (
                                "stopping_after_current"
                                if stop_path.exists()
                                else "running"
                            ),
                            elapsed,
                        ),
                    )
                    if elapsed >= job_timeout:
                        timed_out = True
                        kill_process_tree(process)
                        break
                    time.sleep(poll_seconds)

                try:
                    return_code = process.wait(timeout=15)
                except subprocess.TimeoutExpired:
                    kill_process_tree(process)
                    return_code = -9

            elapsed = time.monotonic() - started_monotonic
            ended_at = utc_now()
            worker_result: dict[str, Any]
            if timed_out:
                worker_result = {
                    "worker_status": "failed",
                    "new_score": None,
                    "new_grade_error": (
                        f"Hard timeout after {job_timeout:.0f}s"
                    ),
                    "error": (
                        f"Hard timeout after {job_timeout:.0f}s"
                    ),
                }
            elif result_path.is_file():
                try:
                    worker_result = read_json(result_path)
                except Exception as exc:
                    worker_result = {
                        "worker_status": "failed",
                        "new_score": None,
                        "new_grade_error": str(exc),
                        "error": (
                            "Worker result parse failed: "
                            f"{exc}"
                        ),
                    }
            else:
                worker_result = {
                    "worker_status": "failed",
                    "new_score": None,
                    "new_grade_error": (
                        "Worker did not create a result file."
                    ),
                    "error": (
                        "Worker did not create a result file. "
                        f"Exit code: {return_code}"
                    ),
                }

            worker_status = str(
                worker_result.get("worker_status") or "failed"
            )
            complete = (
                (
                    worker_status in {"completed", "grade_error"}
                    and worker_result.get("new_score") is not None
                )
                or worker_status == "skipped"
            )
            final_status = "completed" if complete else "failed"
            grade_error = str(
                worker_result.get("new_grade_error")
                or worker_result.get("error")
                or ""
            )
            new_score = worker_result.get("new_score")
            score_delta = worker_result.get("score_delta")

            db.execute(
                """
                UPDATE jobs
                SET status=?,
                    ended_at=?,
                    heartbeat_at=?,
                    worker_pid=NULL,
                    elapsed_seconds=?,
                    new_score=?,
                    score_delta=?,
                    grade_error=?
                WHERE id=?
                """,
                (
                    final_status,
                    ended_at,
                    ended_at,
                    elapsed,
                    new_score,
                    score_delta,
                    grade_error,
                    int(job["id"]),
                ),
            )
            db.commit()

            append_jsonl(
                progress_path,
                {
                    "event": "job_completed",
                    "at": ended_at,
                    "job_id": int(job["id"]),
                    "agent": job["agent"],
                    "task_id": job["task_id"],
                    "status": final_status,
                    "worker_status": worker_status,
                    "return_code": return_code,
                    "elapsed_seconds": round(elapsed, 3),
                    "new_score": new_score,
                    "score_delta": score_delta,
                    "grade_error": grade_error,
                },
            )
            write_json_atomic(
                heartbeat_path,
                heartbeat_payload(
                    run_dir,
                    db,
                    None,
                    (
                        "stopping"
                        if stop_path.exists()
                        else "running"
                    ),
                ),
            )
            write_partial_exports(run_dir, db)

            print(
                f"[{final_status.upper()}] "
                f"{job['agent']} / {job['task_id']} "
                f"score={new_score} elapsed={elapsed:.1f}s",
                flush=True,
            )

            if stop_path.exists():
                write_json_atomic(
                    heartbeat_path,
                    heartbeat_payload(
                        run_dir,
                        db,
                        None,
                        "stopped",
                    ),
                )
                print(
                    "Stop requested. Current job was saved; "
                    "no new job will start."
                )
                return 130

        counts = run_counts(db)
        failed_count = int(counts.get("failed", 0))
        final_state = (
            "complete"
            if failed_count == 0
            else "complete_with_failures"
        )
        db_set_meta(db, "run_completed_at", utc_now())
        db_set_meta(db, "run_final_state", final_state)
        db.commit()
        write_json_atomic(
            heartbeat_path,
            heartbeat_payload(
                run_dir,
                db,
                None,
                final_state,
            ),
        )
        write_partial_exports(run_dir, db)
    return 0 if failed_count == 0 else 4


def result_records(run_dir: Path, db: sqlite3.Connection) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in db.execute("SELECT * FROM jobs ORDER BY job_order"):
        result_path = Path(row["worker_result_path"])
        payload: dict[str, Any] = {}
        if result_path.is_file():
            try:
                payload = read_json(result_path)
            except Exception:
                payload = {}
        record = dict(row)
        record.update(payload)
        records.append(record)
    return records


def summarize_records(records: list[dict[str, Any]]) -> dict[str, Any]:
    completed = [
        row for row in records if row.get("status") == "completed"
    ]
    new_scores = [
        float(row["new_score"])
        for row in completed
        if row.get("new_score") is not None
    ]
    original_scores = [
        float(row["original_score"])
        for row in records
        if row.get("original_score") is not None
    ]
    deltas = [
        float(row["score_delta"])
        for row in completed
        if row.get("score_delta") is not None
    ]
    return {
        "job_count": len(records),
        "completed": len(completed),
        "failed": sum(
            1 for row in records if row.get("status") == "failed"
        ),
        "pending": sum(
            1 for row in records if row.get("status") == "pending"
        ),
        "running": sum(
            1 for row in records if row.get("status") == "running"
        ),
        "judge_required_jobs": sum(
            1 for row in records if row.get("judge_required")
        ),
        "new_scored_count": len(new_scores),
        "new_mean_score": (
            round(sum(new_scores) / len(new_scores), 6)
            if new_scores
            else None
        ),
        "original_scored_count": len(original_scores),
        "original_mean_score": (
            round(
                sum(original_scores) / len(original_scores),
                6,
            )
            if original_scores
            else None
        ),
        "mean_score_delta": (
            round(sum(deltas) / len(deltas), 6)
            if deltas
            else None
        ),
        "absolute_delta_ge_0_15": sum(
            1 for value in deltas if abs(value) >= 0.15
        ),
        "grade_error_count": sum(
            1
            for row in records
            if str(row.get("grade_error") or "").strip()
        ),
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    fields: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen:
                fields.append(key)
                seen.add(key)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=fields,
            extrasaction="ignore",
        )
        writer.writeheader()
        for row in rows:
            flattened = {}
            for key, value in row.items():
                if isinstance(value, (dict, list)):
                    flattened[key] = json.dumps(
                        value,
                        ensure_ascii=False,
                    )
                else:
                    flattened[key] = value
            writer.writerow(flattened)


def write_partial_exports(
    run_dir: Path,
    db: sqlite3.Connection,
) -> None:
    records = result_records(run_dir, db)
    summary = summarize_records(records)
    write_json_atomic(
        run_dir / "results.partial.json",
        {
            "summary": summary,
            "results": records,
        },
    )
    write_json_atomic(run_dir / "summary.json", summary)


def agent_summary(records: list[dict[str, Any]]) -> dict[str, Any]:
    summary = summarize_records(records)
    by_category: dict[str, list[float]] = defaultdict(list)
    by_type: dict[str, list[float]] = defaultdict(list)
    wins = 0
    losses = 0
    ties = 0
    for row in records:
        if row.get("new_score") is not None:
            by_category[str(row.get("category") or "")].append(
                float(row["new_score"])
            )
            by_type[str(row.get("grading_type") or "")].append(
                float(row["new_score"])
            )
        delta = row.get("score_delta")
        if delta is None:
            continue
        number = float(delta)
        if number > 0.00005:
            wins += 1
        elif number < -0.00005:
            losses += 1
        else:
            ties += 1
    summary["score_increased_tasks"] = wins
    summary["score_decreased_tasks"] = losses
    summary["score_unchanged_tasks"] = ties
    summary["category_means"] = {
        key: round(sum(values) / len(values), 6)
        for key, values in sorted(by_category.items())
        if values
    }
    summary["grading_type_means"] = {
        key: round(sum(values) / len(values), 6)
        for key, values in sorted(by_type.items())
        if values
    }
    return summary


def save_xlsx(
    path: Path,
    title: str,
    records: list[dict[str, Any]],
    summary: dict[str, Any],
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.formatting.rule import ColorScaleRule
    except ImportError:
        return

    workbook = openpyxl.Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    light_fill = PatternFill("solid", fgColor="D9EAF7")

    dashboard["A1"] = title
    dashboard["A1"].font = Font(size=16, bold=True)
    dashboard.merge_cells("A1:D1")
    dashboard["A3"] = "Metric"
    dashboard["B3"] = "Value"
    for cell in dashboard[3]:
        cell.fill = header_fill
        cell.font = header_font

    row_index = 4
    for key, value in summary.items():
        if isinstance(value, (dict, list)):
            continue
        dashboard.cell(row=row_index, column=1, value=key)
        dashboard.cell(row=row_index, column=2, value=value)
        row_index += 1
    dashboard.column_dimensions["A"].width = 32
    dashboard.column_dimensions["B"].width = 24

    details = workbook.create_sheet("All Tasks")
    columns = [
        "agent",
        "task_id",
        "task_name",
        "category",
        "grading_type",
        "judge_required",
        "original_status",
        "status",
        "original_score",
        "new_score",
        "score_delta",
        "new_grade_error",
        "grading_elapsed_seconds",
        "source_workspace_sha256",
        "transcript_event_count",
    ]
    for col, name in enumerate(columns, start=1):
        cell = details.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_number, record in enumerate(records, start=2):
        for col, name in enumerate(columns, start=1):
            value = record.get(name)
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            details.cell(row=row_number, column=col, value=value)

    details.freeze_panes = "A2"
    details.auto_filter.ref = details.dimensions
    widths = {
        "A": 16,
        "B": 38,
        "C": 28,
        "D": 20,
        "E": 16,
        "F": 14,
        "G": 16,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 48,
        "M": 18,
        "N": 68,
        "O": 18,
    }
    for column, width in widths.items():
        details.column_dimensions[column].width = width
    if len(records) >= 1:
        details.conditional_formatting.add(
            f"K2:K{len(records)+1}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    changes = workbook.create_sheet("Largest Changes")
    changes_columns = [
        "task_id",
        "category",
        "grading_type",
        "original_score",
        "new_score",
        "score_delta",
        "new_grade_error",
    ]
    sorted_changes = sorted(
        records,
        key=lambda item: abs(float(item.get("score_delta") or 0.0)),
        reverse=True,
    )
    for col, name in enumerate(changes_columns, start=1):
        cell = changes.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
    for row_number, record in enumerate(sorted_changes, start=2):
        for col, name in enumerate(changes_columns, start=1):
            changes.cell(
                row=row_number,
                column=col,
                value=record.get(name),
            )
    changes.freeze_panes = "A2"
    changes.auto_filter.ref = changes.dimensions
    for column, width in {
        "A": 38,
        "B": 20,
        "C": 16,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 48,
    }.items():
        changes.column_dimensions[column].width = width

    errors = workbook.create_sheet("Errors")
    error_rows = [
        row
        for row in records
        if str(row.get("new_grade_error") or row.get("grade_error") or "").strip()
        or row.get("status") == "failed"
    ]
    error_columns = [
        "agent",
        "task_id",
        "grading_type",
        "status",
        "new_score",
        "new_grade_error",
        "error",
        "stderr_path",
    ]
    for col, name in enumerate(error_columns, start=1):
        cell = errors.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
    for row_number, record in enumerate(error_rows, start=2):
        for col, name in enumerate(error_columns, start=1):
            errors.cell(
                row=row_number,
                column=col,
                value=str(record.get(name) or ""),
            )
    errors.freeze_panes = "A2"
    errors.auto_filter.ref = errors.dimensions
    for column in ("A", "B", "C", "D", "E"):
        errors.column_dimensions[column].width = 20
    errors.column_dimensions["F"].width = 60
    errors.column_dimensions["G"].width = 60
    errors.column_dimensions["H"].width = 50

    category_sheet = workbook.create_sheet("Category Summary")
    category_sheet.append(["Category", "Mean New Score"])
    for cell in category_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for category, mean in (
        summary.get("category_means") or {}
    ).items():
        category_sheet.append([category, mean])
    category_sheet.column_dimensions["A"].width = 28
    category_sheet.column_dimensions["B"].width = 20

    grading_sheet = workbook.create_sheet("Grading Types")
    grading_sheet.append(["Grading Type", "Mean New Score"])
    for cell in grading_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for grading_type, mean in (
        summary.get("grading_type_means") or {}
    ).items():
        grading_sheet.append([grading_type, mean])
    grading_sheet.column_dimensions["A"].width = 24
    grading_sheet.column_dimensions["B"].width = 20

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def save_comparison_xlsx(
    path: Path,
    metadata: dict[str, Any],
    agent_summaries: dict[str, Any],
    matrix_rows: list[dict[str, Any]],
    records: list[dict[str, Any]],
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.formatting.rule import ColorScaleRule
    except ImportError:
        return

    workbook = openpyxl.Workbook()
    overall = workbook.active
    overall.title = "Overall Comparison"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    overall["A1"] = "PinchBench Four-Agent Regrade Comparison"
    overall["A1"].font = Font(size=16, bold=True)
    overall.merge_cells("A1:H1")
    overall["A2"] = "Judge"
    overall["B2"] = metadata.get("judge_model")
    headers = [
        "Agent",
        "Original Mean",
        "New Mean",
        "Mean Delta",
        "Completed",
        "Failed",
        "Grade Errors",
        "|Delta| >= 0.15",
    ]
    for col, header in enumerate(headers, start=1):
        cell = overall.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    for row_number, agent in enumerate(
        sorted(agent_summaries),
        start=5,
    ):
        summary = agent_summaries[agent]
        values = [
            agent,
            summary.get("original_mean_score"),
            summary.get("new_mean_score"),
            summary.get("mean_score_delta"),
            summary.get("completed"),
            summary.get("failed"),
            summary.get("grade_error_count"),
            summary.get("absolute_delta_ge_0_15"),
        ]
        for col, value in enumerate(values, start=1):
            overall.cell(row=row_number, column=col, value=value)
    for column, width in {
        "A": 20,
        "B": 18,
        "C": 18,
        "D": 16,
        "E": 14,
        "F": 12,
        "G": 16,
        "H": 18,
    }.items():
        overall.column_dimensions[column].width = width

    matrix = workbook.create_sheet("Task Matrix")
    if matrix_rows:
        columns = list(matrix_rows[0].keys())
        for col, header in enumerate(columns, start=1):
            cell = matrix.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row_number, row in enumerate(matrix_rows, start=2):
            for col, header in enumerate(columns, start=1):
                matrix.cell(
                    row=row_number,
                    column=col,
                    value=row.get(header),
                )
        matrix.freeze_panes = "A2"
        matrix.auto_filter.ref = matrix.dimensions
        matrix.column_dimensions["A"].width = 42
        matrix.column_dimensions["B"].width = 20
        matrix.column_dimensions["C"].width = 18
        for index in range(4, len(columns) + 1):
            matrix.column_dimensions[
                openpyxl.utils.get_column_letter(index)
            ].width = 18
        for index, header in enumerate(columns, start=1):
            if header.endswith("_delta"):
                letter = openpyxl.utils.get_column_letter(index)
                matrix.conditional_formatting.add(
                    f"{letter}2:{letter}{len(matrix_rows)+1}",
                    ColorScaleRule(
                        start_type="min",
                        start_color="F8696B",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFEB84",
                        end_type="max",
                        end_color="63BE7B",
                    ),
                )

    category_sheet = workbook.create_sheet("Category Comparison")
    agents = sorted(agent_summaries)
    categories = sorted(
        {
            category
            for summary in agent_summaries.values()
            for category in (
                summary.get("category_means") or {}
            )
        }
    )
    category_sheet.append(["Category", *agents])
    for cell in category_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for category in categories:
        category_sheet.append(
            [
                category,
                *[
                    (
                        agent_summaries[agent]
                        .get("category_means", {})
                        .get(category)
                    )
                    for agent in agents
                ],
            ]
        )
    category_sheet.column_dimensions["A"].width = 28
    for index in range(2, len(agents) + 2):
        category_sheet.column_dimensions[
            openpyxl.utils.get_column_letter(index)
        ].width = 18

    type_sheet = workbook.create_sheet("Grading Type Comparison")
    grading_types = sorted(
        {
            grading_type
            for summary in agent_summaries.values()
            for grading_type in (
                summary.get("grading_type_means") or {}
            )
        }
    )
    type_sheet.append(["Grading Type", *agents])
    for cell in type_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for grading_type in grading_types:
        type_sheet.append(
            [
                grading_type,
                *[
                    (
                        agent_summaries[agent]
                        .get("grading_type_means", {})
                        .get(grading_type)
                    )
                    for agent in agents
                ],
            ]
        )
    type_sheet.column_dimensions["A"].width = 28
    for index in range(2, len(agents) + 2):
        type_sheet.column_dimensions[
            openpyxl.utils.get_column_letter(index)
        ].width = 18

    audit = workbook.create_sheet("Audit")
    audit.append(["Key", "Value"])
    for cell in audit[1]:
        cell.fill = header_fill
        cell.font = header_font
    for key, value in sorted(metadata.items()):
        audit.append(
            [
                key,
                (
                    json.dumps(value, ensure_ascii=False)
                    if isinstance(value, (dict, list))
                    else value
                ),
            ]
        )
    audit.column_dimensions["A"].width = 34
    audit.column_dimensions["B"].width = 100

    errors = workbook.create_sheet("Errors")
    errors.append(
        [
            "Agent",
            "Task ID",
            "Status",
            "New Score",
            "Grade Error",
        ]
    )
    for cell in errors[1]:
        cell.fill = header_fill
        cell.font = header_font
    for record in records:
        error = str(
            record.get("new_grade_error")
            or record.get("grade_error")
            or ""
        ).strip()
        if not error and record.get("status") != "failed":
            continue
        errors.append(
            [
                record.get("agent"),
                record.get("task_id"),
                record.get("status"),
                record.get("new_score"),
                error,
            ]
        )
    errors.column_dimensions["A"].width = 20
    errors.column_dimensions["B"].width = 42
    errors.column_dimensions["C"].width = 16
    errors.column_dimensions["D"].width = 14
    errors.column_dimensions["E"].width = 80

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def finalize(run_dir: Path) -> int:
    run_dir = run_dir.resolve()
    export_dir = run_dir / "exports"
    export_dir.mkdir(exist_ok=True)

    with db_connect(run_dir / "state.sqlite") as db:
        records = result_records(run_dir, db)
        overall_summary = summarize_records(records)
        metadata = {
            key: json.loads(row["value"])
            for row in db.execute(
                "SELECT key, value FROM metadata"
            )
            for key in [row["key"]]
        }

    metadata["initial_script_revision"] = metadata.get(
        "script_revision"
    )
    metadata["script_revision"] = SCRIPT_REVISION
    metadata["finalized_at"] = utc_now()

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        grouped[str(record.get("agent"))].append(record)

    agent_summaries: dict[str, Any] = {}
    for agent, rows in grouped.items():
        summary = agent_summary(rows)
        agent_summaries[agent] = summary
        slug = safe_slug(agent)
        payload = {
            "metadata": metadata,
            "agent": agent,
            "summary": summary,
            "results": rows,
        }
        write_json_atomic(
            export_dir / f"{slug}_regraded_results.json",
            payload,
        )
        write_csv(
            export_dir / f"{slug}_regraded_results.csv",
            rows,
        )
        save_xlsx(
            export_dir / f"{slug}_regraded_results.xlsx",
            f"{agent} PinchBench Regrade",
            rows,
            summary,
        )

    agents = sorted(grouped)
    task_ids = sorted(
        {
            str(row.get("task_id"))
            for rows in grouped.values()
            for row in rows
        }
    )
    matrix_rows: list[dict[str, Any]] = []
    lookup = {
        (str(row.get("agent")), str(row.get("task_id"))): row
        for row in records
    }
    for task_id in task_ids:
        base = next(
            (
                lookup[(agent, task_id)]
                for agent in agents
                if (agent, task_id) in lookup
            ),
            {},
        )
        row: dict[str, Any] = {
            "task_id": task_id,
            "category": base.get("category"),
            "grading_type": base.get("grading_type"),
        }
        for agent in agents:
            item = lookup.get((agent, task_id), {})
            prefix = safe_slug(agent)
            row[f"{prefix}_original_score"] = item.get(
                "original_score"
            )
            row[f"{prefix}_new_score"] = item.get("new_score")
            row[f"{prefix}_delta"] = item.get("score_delta")
            row[f"{prefix}_status"] = item.get("status")
        matrix_rows.append(row)

    comparison = {
        "metadata": metadata,
        "overall_summary": overall_summary,
        "agent_summaries": agent_summaries,
        "task_matrix": matrix_rows,
    }
    write_json_atomic(
        export_dir / "four_agent_comparison.json",
        comparison,
    )
    write_csv(
        export_dir / "four_agent_task_matrix.csv",
        matrix_rows,
    )
    summary_rows = []
    for agent, summary in agent_summaries.items():
        row = {"agent": agent}
        row.update(
            {
                key: value
                for key, value in summary.items()
                if not isinstance(value, (dict, list))
            }
        )
        summary_rows.append(row)
    write_csv(
        export_dir / "four_agent_summary.csv",
        summary_rows,
    )
    save_comparison_xlsx(
        export_dir / "four_agent_regrade_comparison.xlsx",
        metadata,
        agent_summaries,
        matrix_rows,
        records,
    )

    write_json_atomic(
        run_dir / "results.json",
        {
            "metadata": metadata,
            "summary": overall_summary,
            "agent_summaries": agent_summaries,
            "results": records,
        },
    )
    write_csv(run_dir / "results.csv", records)
    write_json_atomic(
        run_dir / "summary.json",
        {
            "overall": overall_summary,
            "agents": agent_summaries,
        },
    )
    print(f"Exports written to: {export_dir}")
    return 0


def preflight(config_path: Path, judge_model_override: Optional[str]) -> int:
    config = load_config(config_path)
    judge_model = (
        judge_model_override.strip()
        if judge_model_override
        else str(config["judge_model"])
    )
    errors: list[str] = []
    warnings: list[str] = []
    checks: list[dict[str, Any]] = []
    skill_dir = Path(config["skill_dir"])

    def add(name: str, ok: bool, detail: str) -> None:
        checks.append({"name": name, "ok": ok, "detail": detail})
        if not ok:
            errors.append(f"{name}: {detail}")

    add(
        "OPENROUTER_API_KEY",
        bool(os.environ.get("OPENROUTER_API_KEY")),
        "set" if os.environ.get("OPENROUTER_API_KEY") else "missing",
    )
    add("skill_dir", skill_dir.is_dir(), str(skill_dir))

    manifest = skill_dir / "tasks" / "manifest.yaml"
    manifest_hash = sha256_file(manifest) if manifest.is_file() else ""
    add(
        "manifest_sha256",
        manifest_hash == config["expected_manifest_sha256"],
        f"actual={manifest_hash}; expected={config['expected_manifest_sha256']}",
    )

    try:
        reference = load_reference_runner()
        tasks, _ = reference.load_tasks(skill_dir / "tasks")
        add("task_count", len(tasks) == 147, f"loaded={len(tasks)}")
        selected_ids = [
            task.task_id
            for task in tasks
            if task.task_id not in reference.DEFAULT_SKIPPED_TASKS
        ]
        add(
            "formal_task_count",
            len(selected_ids) == 143,
            f"formal={len(selected_ids)}",
        )
    except Exception as exc:
        tasks = []
        selected_ids = []
        add("task_loader", False, str(exc))

    try:
        grader = load_grader(skill_dir)
        signature = inspect.signature(grader.grade_task)
        needed = {
            "task",
            "execution_result",
            "skill_dir",
            "judge_model",
            "judge_timeout_seconds",
            "judge_backend",
        }
        add(
            "grade_task_signature",
            needed.issubset(signature.parameters),
            str(signature),
        )
        add(
            "grader_default_judge",
            True,
            str(getattr(grader, "DEFAULT_JUDGE_MODEL", "")),
        )
    except Exception as exc:
        add("grading_engine", False, str(exc))

    task_id_set = set(selected_ids)
    for entry in config["runs"]:
        agent = str(entry["agent"])
        run_path = Path(entry["path"])
        add(f"{agent}.run_path", run_path.is_dir(), str(run_path))
        if not run_path.is_dir():
            continue
        try:
            run_config = read_json(run_path / "run_config.json")
            add(
                f"{agent}.commit",
                run_config.get("pinchbench_commit")
                == config["expected_pinchbench_commit"],
                str(run_config.get("pinchbench_commit")),
            )
            old_judge = str(run_config.get("judge_model") or "")
            if old_judge and old_judge != config["old_judge_model"]:
                warnings.append(
                    f"{agent}: original judge is {old_judge}, "
                    f"expected {config['old_judge_model']}"
                )
            mapped = result_by_task(run_path / "results.json")
            add(
                f"{agent}.results_count",
                len(mapped) == 143,
                f"count={len(mapped)}",
            )
            add(
                f"{agent}.task_set",
                set(mapped) == task_id_set,
                (
                    f"missing={len(task_id_set-set(mapped))}; "
                    f"extra={len(set(mapped)-task_id_set)}"
                ),
            )
            missing_workspaces = 0
            for task_id, row in mapped.items():
                try:
                    source_workspace(run_path, row, task_id)
                except FileNotFoundError:
                    missing_workspaces += 1
            add(
                f"{agent}.workspaces",
                missing_workspaces == 0,
                f"missing={missing_workspaces}",
            )
        except Exception as exc:
            add(f"{agent}.inspection", False, str(exc))

    output_root = Path(config["output_root"])
    output_root.mkdir(parents=True, exist_ok=True)
    report = {
        "script_revision": SCRIPT_REVISION,
        "generated_at": utc_now(),
        "config": str(config_path.resolve()),
        "judge_model": judge_model,
        "checks": checks,
        "warnings": warnings,
        "errors": errors,
        "passed": not errors,
        "no_api_call_made": True,
        "source_mutation": "none",
    }
    report_path = output_root / (
        f"preflight_{dt.datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
    )
    write_json_atomic(report_path, report)

    for check in checks:
        marker = "PASS" if check["ok"] else "FAIL"
        print(f"[{marker}] {check['name']}: {check['detail']}")
    for warning in warnings:
        print(f"[WARN] {warning}")
    print(f"Report: {report_path}")
    if errors:
        print(f"Preflight failed: {len(errors)} error(s).")
        return 3
    print("PASS: regrade preflight completed.")
    return 0


def status_snapshot(run_dir: Path) -> dict[str, Any]:
    with db_connect(run_dir / "state.sqlite") as db:
        heartbeat = heartbeat_payload(
            run_dir,
            db,
            None,
            "idle",
        )
        metadata = {
            row["key"]: json.loads(row["value"])
            for row in db.execute(
                "SELECT key, value FROM metadata"
            )
        }
        by_agent = []
        for row in db.execute(
            """
            SELECT agent,
                   COUNT(*) AS total,
                   SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) AS completed,
                   SUM(CASE WHEN status='failed' THEN 1 ELSE 0 END) AS failed,
                   SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END) AS pending,
                   AVG(new_score) AS mean_new_score,
                   AVG(original_score) AS mean_original_score,
                   SUM(CASE WHEN grade_error IS NOT NULL AND grade_error<>'' THEN 1 ELSE 0 END) AS grade_errors
            FROM jobs
            GROUP BY agent
            ORDER BY agent
            """
        ):
            by_agent.append(dict(row))
        recent = [
            dict(row)
            for row in db.execute(
                """
                SELECT agent, task_id, status, new_score,
                       score_delta, elapsed_seconds, grade_error
                FROM jobs
                WHERE status IN ('completed', 'failed')
                ORDER BY COALESCE(ended_at, '') DESC
                LIMIT 10
                """
            )
        ]
        current = db.execute(
            """
            SELECT agent, task_id, grading_type, worker_pid,
                   started_at, heartbeat_at, elapsed_seconds
            FROM jobs WHERE status='running'
            LIMIT 1
            """
        ).fetchone()

    actual_heartbeat = {}
    heartbeat_path = run_dir / "heartbeat.json"
    if heartbeat_path.is_file():
        try:
            actual_heartbeat = read_json(heartbeat_path)
        except Exception:
            pass

    return {
        "run_dir": str(run_dir),
        "metadata": metadata,
        "counts": heartbeat["counts"],
        "by_agent": by_agent,
        "current": dict(current) if current else None,
        "recent": recent,
        "heartbeat": actual_heartbeat,
        "stop_requested": (run_dir / "STOP_REQUESTED").exists(),
    }


def print_status(snapshot: dict[str, Any]) -> None:
    metadata = snapshot["metadata"]
    counts = snapshot["counts"]
    total = int(counts.get("total", 0))
    completed = int(counts.get("completed", 0))
    failed = int(counts.get("failed", 0))
    pending = int(counts.get("pending", 0))
    running = int(counts.get("running", 0))
    print("=" * 96)
    print("PinchBench regrade monitor")
    print(f"Run       : {snapshot['run_dir']}")
    print(f"Suite     : {metadata.get('suite')}")
    print(f"Judge     : {metadata.get('judge_model')}")
    print(
        f"Progress  : {completed + failed}/{total} | "
        f"completed={completed} failed={failed} "
        f"running={running} pending={pending}"
    )
    heartbeat = snapshot.get("heartbeat") or {}
    if heartbeat:
        mean = heartbeat.get("mean_new_score")
        eta = heartbeat.get("eta_seconds")
        print(
            "Mean score: "
            + (f"{float(mean):.4f}" if mean is not None else "N/A")
            + " | ETA: "
            + (
                f"{float(eta)/3600:.2f}h"
                if eta is not None
                else "N/A"
            )
        )
        print(
            f"Heartbeat : {heartbeat.get('updated_at')} "
            f"state={heartbeat.get('state')}"
        )
    if snapshot.get("current"):
        current = snapshot["current"]
        print(
            f"Current   : {current.get('agent')} / "
            f"{current.get('task_id')} "
            f"elapsed={float(current.get('elapsed_seconds') or 0):.1f}s"
        )
    print("-" * 96)
    for item in snapshot["by_agent"]:
        print(
            f"{item['agent']:<12} "
            f"done={int(item['completed'] or 0):3d}/"
            f"{int(item['total'] or 0):3d} "
            f"failed={int(item['failed'] or 0):2d} "
            f"pending={int(item['pending'] or 0):3d} "
            f"mean="
            + (
                f"{float(item['mean_new_score']):.4f}"
                if item["mean_new_score"] is not None
                else "N/A"
            )
        )
    if snapshot["recent"]:
        print("-" * 96)
        print("Recent:")
        for item in snapshot["recent"]:
            print(
                f"  {item['agent']:<12} "
                f"{item['task_id']:<42} "
                f"{item['status']:<9} "
                f"score={item['new_score']} "
                f"delta={item['score_delta']}"
            )
    if snapshot.get("stop_requested"):
        print("STOP_REQUESTED is present.")
    print("=" * 96)


def resolve_run_dir(
    config: dict[str, Any],
    value: Optional[str],
    kind: Optional[str] = None,
) -> Path:
    if value:
        return Path(value).resolve()
    return latest_run_dir(Path(config["output_root"]), kind=kind)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="PinchBench frozen-output regrader."
    )
    parser.add_argument(
        "--config",
        default=str(DEFAULT_CONFIG),
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    preflight_parser = subparsers.add_parser("preflight")
    preflight_parser.add_argument("--judge-model", default="")

    run_parser = subparsers.add_parser("run")
    run_parser.add_argument(
        "--suite",
        choices=("smoke", "full"),
        required=True,
    )
    run_parser.add_argument("--judge-model", default="")
    run_parser.add_argument("--verbose", action="store_true")

    resume_parser = subparsers.add_parser("resume")
    resume_parser.add_argument("--run-dir", default="")
    resume_parser.add_argument(
        "--retry-failed",
        action="store_true",
    )
    resume_parser.add_argument("--verbose", action="store_true")

    status_parser = subparsers.add_parser("status")
    status_parser.add_argument("--run-dir", default="")
    status_parser.add_argument("--watch", type=float, default=0.0)

    stop_parser = subparsers.add_parser("stop")
    stop_parser.add_argument("--run-dir", default="")

    finalize_parser = subparsers.add_parser("finalize")
    finalize_parser.add_argument("--run-dir", default="")

    worker_parser = subparsers.add_parser("worker")
    worker_parser.add_argument("--job-file", required=True)

    return parser


def main() -> int:
    args = build_parser().parse_args()
    config_path = Path(args.config).resolve()

    if args.command == "worker":
        return worker_job(Path(args.job_file).resolve())

    config = load_config(config_path)

    if args.command == "preflight":
        return preflight(
            config_path,
            args.judge_model or None,
        )

    if args.command == "run":
        run_dir = create_run(
            config_path,
            args.suite,
            args.judge_model or None,
        )
        print(f"Run directory: {run_dir}")
        code = run_queue(
            config_path,
            run_dir,
            retry_failed=False,
            verbose=args.verbose,
        )
        if code == 0:
            finalize(run_dir)
        return code

    if args.command == "resume":
        run_dir = resolve_run_dir(
            config,
            args.run_dir or None,
            kind="regrade",
        )
        code = run_queue(
            config_path,
            run_dir,
            retry_failed=args.retry_failed,
            verbose=args.verbose,
        )
        if code == 0:
            finalize(run_dir)
        return code

    if args.command == "status":
        run_dir = resolve_run_dir(
            config,
            args.run_dir or None,
        )
        while True:
            if os.name == "nt":
                os.system("cls")
            print_status(status_snapshot(run_dir))
            if args.watch <= 0:
                break
            state = (
                status_snapshot(run_dir)
                .get("heartbeat", {})
                .get("state")
            )
            if state in {"complete", "complete_with_failures"}:
                break
            time.sleep(max(1.0, args.watch))
        return 0

    if args.command == "stop":
        run_dir = resolve_run_dir(
            config,
            args.run_dir or None,
        )
        stop_path = run_dir / "STOP_REQUESTED"
        stop_path.write_text(
            f"requested_at={utc_now()}\n",
            encoding="ascii",
        )
        print(f"Stop requested: {stop_path}")
        print(
            "The current grading job will finish and be saved; "
            "no next job will start."
        )
        return 0

    if args.command == "finalize":
        run_dir = resolve_run_dir(
            config,
            args.run_dir or None,
        )
        return finalize(run_dir)

    raise AssertionError(args.command)


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("Interrupted by user.", file=sys.stderr)
        raise SystemExit(130)
    except Exception:
        traceback.print_exc()
        raise SystemExit(1)
