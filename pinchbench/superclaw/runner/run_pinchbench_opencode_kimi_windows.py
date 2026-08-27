#!/usr/bin/env python3
# Runner revision: 2026-08-10-opencode-kimi-k3-token-time-v1
"""
Run the local PinchBench task set with OpenCode + Kimi K3 on native Windows.

Target evaluation chain:
    PinchBench local tasks -> OpenCode default agent/tools -> OpenRouter
    -> moonshotai/kimi-k3

Design goals:
- Native Windows support (no WSL required).
- One worker / strictly serial task execution.
- Fresh OpenCode session per normal task.
- Multi-session task support via OpenCode --session.
- Prompt delivered through UTF-8 stdin files, never through Windows .cmd argv.
- Raw OpenCode JSONL retained for audit.
- OpenCode events normalized to the transcript shape expected by PinchBench graders.
- Grading through the checked-out PinchBench scripts/lib_grading.py engine.
- PinchBench grading engine from the checked-out commit.
- Hybrid/llm_judge use that commit's DEFAULT_JUDGE_MODEL through direct API
  (judge_backend="api") and OPENROUTER_API_KEY.
- JSON, CSV, and XLSX reports.
- No custom OpenCode agent, no custom tool allow-list, no reasoning variant,
  and no --auto.

Recommended location:
    C:\\pinchbench-opencode-kimi\\runner\\run_pinchbench_opencode_kimi_windows.py

Examples:
    # Validate local files, CLI, selected tasks and prerequisites; does not call the model.
    python run_pinchbench_opencode_kimi_windows.py --preflight --suite all

    # Canary: execute and grade one local automated task.
    python run_pinchbench_opencode_kimi_windows.py --suite automated-only --limit 1 --keep-workspaces --verbose

    # Run the core suite.
    python run_pinchbench_opencode_kimi_windows.py --suite core --keep-workspaces

    # Run the complete task set from the checked-out manifest.
    python run_pinchbench_opencode_kimi_windows.py --suite all --keep-workspaces

Scoring:
- Automated tasks use PinchBench's checked-out automated grading code.
- Hybrid and llm_judge tasks use PinchBench's checked-out DEFAULT_JUDGE_MODEL
  through judge_backend="api".  The current commit uses
  openrouter/anthropic/claude-opus-5 and reads OPENROUTER_API_KEY.
- The judge runs after task execution and is separate from the tested
  OpenCode + Kimi K3 agent.
"""
from __future__ import annotations

import argparse
import csv
import dataclasses
import datetime as dt
import hashlib
import json
import logging
import os
import platform
import queue
import re
import shutil
import signal
import subprocess
import sys
import threading
import time
import traceback
from collections import Counter
from pathlib import Path
from typing import Any, Optional

os.environ.setdefault("PYTHONUTF8", "1")
os.environ.setdefault("OPENCODE_DISABLE_AUTOUPDATE", "true")

try:
    import yaml  # type: ignore
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "缺少 PyYAML。请使用项目虚拟环境运行：\n"
        r"C:\pinchbench-opencode\.venv\Scripts\python.exe -m pip install pyyaml"
    ) from exc

LOGGER = logging.getLogger("pinchbench-opencode-kimi")

RUNNER_REVISION = "2026-08-10-opencode-kimi-k3-opus5-judge-v2"
DEFAULT_MODEL = "openrouter/moonshotai/kimi-k3"
DEFAULT_JUDGE_MODEL = "openrouter/anthropic/claude-opus-5"

# Agreed exclusions for this independent project. They require external
# GitHub/Google Workspace integration tooling and credentials that are not
# part of the current Windows OpenCode test environment.
DEFAULT_SKIPPED_TASKS = {
    "task_gh_issue_triage",
    "task_gws_email_triage",
    "task_gws_cross_service",
    "task_gws_task_management",
}

# Explicitly known internet-dependent tasks from the historical runner.
# Newer task files can also mark network_required: true.
NETWORK_TASKS = {
    "task_earnings_analysis",
    "task_stock",
    "task_market_research",
    "task_polymarket_briefing",
    "task_executive_lookup",
    "task_events",
    "task_deep_research",
    "task_competitive_research",
    "task_oss_alternative_research",
    "task_pricing_research",
    "task_it_procurement",
    "task_eu_regulation_research",
    "task_byok_best_practices",
}

TEXT_FILE_SUFFIXES = {
    ".txt", ".md", ".py", ".js", ".ts", ".tsx", ".jsx", ".json", ".yaml", ".yml",
    ".csv", ".tsv", ".html", ".css", ".xml", ".toml", ".ini", ".cfg", ".log",
    ".ics", ".sh", ".ps1", ".bat", ".cmd", ".sql", ".java", ".go", ".rs", ".rb", ".php",
}

SKIP_WORKSPACE_NAMES = {
    "BOOTSTRAP.md", "SOUL.md", "USER.md", "IDENTITY.md", "HEARTBEAT.md",
    "TOOLS.md", "AGENTS.md", "CLAUDE.md",
}
SKIP_WORKSPACE_DIRS = {
    ".git", ".openclaw", ".opencode", "__pycache__", "node_modules", "skills",
}


@dataclasses.dataclass
class Task:
    task_id: str
    name: str
    category: str
    grading_type: str
    timeout_seconds: int
    workspace_files: list[dict[str, Any]]
    prompt: str
    expected_behavior: str
    grading_criteria: list[str]
    automated_checks: str
    llm_judge_rubric: str
    grading_weights: dict[str, float]
    metadata: dict[str, Any]
    file_path: Path
    multi_session: bool = False
    sessions: list[dict[str, Any]] = dataclasses.field(default_factory=list)


@dataclasses.dataclass
class GradeResult:
    score: Optional[float]
    grading_type: str
    breakdown: dict[str, float]
    notes: str = ""
    error: Optional[str] = None


# ---------------------------------------------------------------------------
# Task loading
# ---------------------------------------------------------------------------

def parse_sections(body: str) -> dict[str, str]:
    sections: dict[str, list[str]] = {}
    current: Optional[str] = None
    for line in body.splitlines():
        match = re.match(r"^##\s+(.+?)\s*$", line)
        if match:
            current = match.group(1).strip()
            sections.setdefault(current, [])
        elif current:
            sections[current].append(line)
    return {key: "\n".join(value).strip() for key, value in sections.items()}


def extract_grading_criteria(text: str) -> list[str]:
    criteria: list[str] = []
    for line in text.splitlines():
        match = re.match(r"^-\s+\[[ xX]\]\s+(.+)$", line.strip())
        if match:
            criteria.append(match.group(1).strip())
    return criteria


def load_task_file(path: Path, category_override: Optional[str] = None) -> Task:
    raw = path.read_text(encoding="utf-8")
    match = re.match(r"^---\s*\r?\n(.*?)\r?\n---\s*\r?\n?(.*)$", raw, re.DOTALL)
    if not match:
        raise ValueError(f"{path} 没有 YAML frontmatter")

    metadata = yaml.safe_load(match.group(1)) or {}
    if not isinstance(metadata, dict):
        raise ValueError(f"{path} 的 YAML frontmatter 不是对象")

    body = match.group(2)
    sections = parse_sections(body)
    task_id = str(metadata.get("id") or path.stem)

    weights = metadata.get("grading_weights") or {}
    if not isinstance(weights, dict):
        weights = {}
    normalized_weights: dict[str, float] = {}
    for key, value in weights.items():
        try:
            normalized_weights[str(key)] = float(value)
        except (TypeError, ValueError):
            pass

    raw_sessions = metadata.get("sessions") or []
    sessions = [item for item in raw_sessions if isinstance(item, dict)] if isinstance(raw_sessions, list) else []

    return Task(
        task_id=task_id,
        name=str(metadata.get("name") or task_id),
        category=str(category_override or metadata.get("category") or ""),
        grading_type=str(metadata.get("grading_type") or "automated"),
        timeout_seconds=int(metadata.get("timeout_seconds") or 120),
        workspace_files=list(metadata.get("workspace_files") or []),
        prompt=sections.get("Prompt", "").strip(),
        expected_behavior=sections.get("Expected Behavior", "").strip(),
        grading_criteria=extract_grading_criteria(sections.get("Grading Criteria", "")),
        automated_checks=sections.get("Automated Checks", "") or "",
        llm_judge_rubric=sections.get("LLM Judge Rubric", "") or "",
        grading_weights=normalized_weights,
        metadata=metadata,
        file_path=path,
        multi_session=bool(metadata.get("multi_session") or sessions),
        sessions=sessions,
    )


def parse_manifest(tasks_dir: Path) -> tuple[list[str], dict[str, str], list[str]]:
    manifest_path = tasks_dir / "manifest.yaml"
    if not manifest_path.exists():
        ordered = [
            p.stem for p in sorted(tasks_dir.glob("task_*.md"))
            if p.stem != "task_XX_name"
        ]
        return ordered, {}, []

    manifest = yaml.safe_load(manifest_path.read_text(encoding="utf-8")) or {}
    if not isinstance(manifest, dict):
        raise ValueError(f"{manifest_path} 不是有效 YAML 对象")

    category_map: dict[str, str] = {}
    core_tasks = [str(x) for x in (manifest.get("core") or [])]

    if "categories" in manifest:
        run_first = [str(x) for x in (manifest.get("run_first") or [])]
        categories = manifest.get("categories") or {}
        all_task_ids: list[str] = []
        if isinstance(categories, dict):
            for category, ids in categories.items():
                for task_id in ids or []:
                    task_id = str(task_id)
                    category_map[task_id] = str(category)
                    all_task_ids.append(task_id)

        seen: set[str] = set()
        ordered: list[str] = []
        for task_id in run_first + all_task_ids:
            if task_id not in seen:
                seen.add(task_id)
                ordered.append(task_id)
        return ordered, category_map, core_tasks

    ordered = [str(x) for x in (manifest.get("tasks") or [])]
    return ordered, category_map, core_tasks


def load_tasks(tasks_dir: Path) -> tuple[list[Task], list[str]]:
    """Load exactly the checked-out manifest when it exists.

    PinchBench ships template/example Markdown files alongside real tasks.
    Falling back to every task_*.md while a manifest exists can accidentally
    count a template as a benchmark task.  The manifest is therefore the
    source of truth for normal runs.
    """
    manifest_exists = (tasks_dir / "manifest.yaml").exists()
    ordered_ids, category_map, core_tasks = parse_manifest(tasks_dir)
    tasks: list[Task] = []

    for task_id in ordered_ids:
        if task_id == "task_XX_name":
            continue
        task_path = tasks_dir / f"{task_id}.md"
        if not task_path.exists():
            LOGGER.warning("manifest 引用了不存在的任务文件: %s", task_path)
            continue
        try:
            task = load_task_file(
                task_path,
                category_override=category_map.get(task_id),
            )
            if (
                task.task_id == "task_XX_name"
                or task.category.strip().lower() == "category_name"
            ):
                LOGGER.debug("忽略 PinchBench 模板任务: %s", task_path)
                continue
            tasks.append(task)
        except Exception as exc:
            LOGGER.warning("加载任务失败 %s: %s", task_path, exc)

    if manifest_exists:
        return tasks, core_tasks

    # Only use file discovery when no manifest is available.
    loaded_ids = {task.task_id for task in tasks}
    for task_path in sorted(tasks_dir.glob("task_*.md")):
        if task_path.stem in loaded_ids or task_path.stem == "task_XX_name":
            continue
        try:
            task = load_task_file(task_path)
            if (
                task.task_id == "task_XX_name"
                or task.category.strip().lower() == "category_name"
            ):
                continue
            tasks.append(task)
        except Exception as exc:
            LOGGER.warning("加载任务失败 %s: %s", task_path, exc)

    return tasks, core_tasks


def filter_tasks(
    tasks: list[Task],
    core_tasks: list[str],
    suite: str,
    limit: Optional[int],
) -> list[Task]:
    suite = suite.strip()
    task_by_id = {task.task_id: task for task in tasks}

    if suite == "all":
        selected = tasks
    elif suite == "core":
        core_set = set(core_tasks)
        selected = [task for task in tasks if task.task_id in core_set] if core_set else tasks[:25]
    elif suite == "automated-only":
        selected = [task for task in tasks if task.grading_type == "automated"]
    elif suite == "llm-judge-only":
        selected = [task for task in tasks if task.grading_type == "llm_judge"]
    elif suite == "hybrid-only":
        selected = [task for task in tasks if task.grading_type == "hybrid"]
    elif suite == "judge-required-only":
        selected = [task for task in tasks if task.grading_type in {"llm_judge", "hybrid"}]
    else:
        ids = [item.strip() for item in suite.split(",") if item.strip()]
        missing = [task_id for task_id in ids if task_id not in task_by_id]
        if missing:
            raise SystemExit(f"suite 中有未知任务: {', '.join(missing)}")
        selected = [task_by_id[task_id] for task_id in ids]

    if limit is not None:
        selected = selected[:limit]
    return selected


def is_network_task(task: Task) -> bool:
    return bool(
        task.task_id in NETWORK_TASKS
        or task.metadata.get("network_required") is True
        or task.metadata.get("requires_network") is True
    )


# ---------------------------------------------------------------------------
# Workspace and prerequisites
# ---------------------------------------------------------------------------

def safe_workspace_dest(dest: str) -> Path:
    rel = Path(dest)
    if rel.is_absolute() or ".." in rel.parts:
        raise ValueError(f"非法 workspace 文件路径: {dest}")
    return rel


def stage_workspace_files(
    task: Task,
    workspace: Path,
    skill_dir: Path,
) -> tuple[list[str], list[str]]:
    staged: list[str] = []
    missing: list[str] = []

    for wf in task.workspace_files:
        if not isinstance(wf, dict):
            missing.append(f"无法识别的 workspace_files 项: {wf!r}")
            continue

        dest_value = wf.get("dest") or wf.get("path") or wf.get("name")
        source_value = wf.get("source")
        content_value = wf.get("content")

        if not dest_value and source_value:
            dest_value = Path(str(source_value)).name
        if not dest_value:
            missing.append(f"workspace_files 项缺少 dest/path: {wf!r}")
            continue

        try:
            dest_rel = safe_workspace_dest(str(dest_value))
        except ValueError as exc:
            missing.append(str(exc))
            continue

        dest_path = workspace / dest_rel
        dest_path.parent.mkdir(parents=True, exist_ok=True)

        if content_value is not None:
            dest_path.write_text(str(content_value), encoding="utf-8")
            staged.append(str(dest_rel))
            continue

        if not source_value:
            missing.append(f"workspace_files 项缺少 source/content: {wf!r}")
            continue

        src_rel = Path(str(source_value))
        candidates = [
            skill_dir / "assets" / src_rel,
            skill_dir / src_rel,
            task.file_path.parent / src_rel,
            task.file_path.parent.parent / "assets" / src_rel,
        ]
        src_path = next((candidate for candidate in candidates if candidate.exists()), None)
        if src_path is None:
            missing.append(
                f"找不到预置文件 {source_value}; 尝试过: "
                + ", ".join(str(candidate) for candidate in candidates)
            )
            continue

        if src_path.is_dir():
            if dest_path.exists():
                shutil.rmtree(dest_path)
            shutil.copytree(src_path, dest_path)
        else:
            shutil.copy2(src_path, dest_path)
        staged.append(str(dest_rel))

    return staged, missing


def resolve_command(name: str) -> Optional[str]:
    candidates = [name]
    if sys.platform == "win32" and not Path(name).suffix:
        candidates = [f"{name}.cmd", f"{name}.exe", name]
    for candidate in candidates:
        resolved = shutil.which(candidate)
        if resolved:
            return resolved
    return None


def check_prerequisites(prereqs: list[str]) -> list[str]:
    missing: list[str] = []
    npm_command = resolve_command("npm") or "npm"

    for req_raw in prereqs:
        req = str(req_raw)
        if req.startswith("cli:"):
            command = req.split(":", 1)[1].strip()
            if not resolve_command(command):
                missing.append(req)
        elif req.startswith("npm:"):
            package = req.split(":", 1)[1].strip()
            try:
                result = subprocess.run(
                    [npm_command, "list", "-g", package, "--depth=0"],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=30,
                    check=False,
                )
                if result.returncode != 0:
                    missing.append(req)
            except Exception:
                missing.append(req)

    return missing


def choose_opencode_command(preferred: str) -> str:
    if preferred and preferred != "auto":
        preferred_path = Path(preferred).expanduser()
        if preferred_path.exists():
            return str(preferred_path.resolve())
        resolved = resolve_command(preferred)
        if resolved:
            return resolved
        raise SystemExit(f"找不到 OpenCode 命令 {preferred!r}")

    for command in (("opencode.cmd", "opencode.exe", "opencode") if sys.platform == "win32" else ("opencode",)):
        resolved = shutil.which(command)
        if resolved:
            return resolved

    raise SystemExit(
        "找不到 OpenCode。请确认已安装 opencode-ai，且 opencode.cmd 在 PATH 中。"
    )


def kill_proc_tree(proc: subprocess.Popen[str]) -> None:
    if proc.poll() is not None:
        return
    try:
        if sys.platform == "win32":
            subprocess.run(
                ["taskkill", "/F", "/T", "/PID", str(proc.pid)],
                capture_output=True,
                check=False,
            )
        else:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
    except Exception:
        try:
            proc.kill()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# OpenCode event handling
# ---------------------------------------------------------------------------

def event_text(event: dict[str, Any]) -> str:
    event_type = str(event.get("type") or "")
    part = event.get("part")

    if event_type == "text" and isinstance(part, dict):
        return str(part.get("text") or "")

    # Compatibility fallbacks for older/newer event shapes.
    if isinstance(part, dict) and part.get("type") == "text":
        return str(part.get("text") or "")
    if isinstance(event.get("text"), str):
        return str(event.get("text") or "")

    message = event.get("message")
    if isinstance(message, dict):
        content = message.get("content")
        if isinstance(content, str):
            return content
        if isinstance(content, list):
            chunks: list[str] = []
            for item in content:
                if isinstance(item, dict) and item.get("type") in {"text", "output_text"}:
                    chunks.append(str(item.get("text") or ""))
            return "".join(chunks)

    return ""


def extract_event_error(event: dict[str, Any]) -> str:
    if event.get("type") != "error":
        return ""
    error = event.get("error")
    if isinstance(error, str):
        return error
    if isinstance(error, dict):
        data = error.get("data")
        if isinstance(data, dict) and data.get("message"):
            return str(data.get("message"))
        if error.get("message"):
            return str(error.get("message"))
        if error.get("name"):
            return str(error.get("name"))
        return json.dumps(error, ensure_ascii=False)
    return str(error or "OpenCode emitted an error event")


def parse_step_finish(event: dict[str, Any]) -> Optional[dict[str, Any]]:
    if event.get("type") != "step_finish":
        return None
    part = event.get("part")
    if not isinstance(part, dict):
        return None

    tokens = part.get("tokens") or {}
    if not isinstance(tokens, dict):
        tokens = {}
    cache = tokens.get("cache") or {}
    if not isinstance(cache, dict):
        cache = {}

    def to_int(value: Any) -> int:
        try:
            return int(value or 0)
        except (TypeError, ValueError):
            return 0

    def to_float(value: Any) -> float:
        try:
            return float(value or 0.0)
        except (TypeError, ValueError):
            return 0.0

    return {
        "input_tokens": to_int(tokens.get("input")),
        "output_tokens": to_int(tokens.get("output")),
        "reasoning_tokens": to_int(tokens.get("reasoning")),
        "cache_read_tokens": to_int(cache.get("read")),
        "cache_write_tokens": to_int(cache.get("write")),
        "cost_usd": to_float(part.get("cost")),
        "finish_reason": str(part.get("reason") or ""),
    }


def normalize_opencode_event(event: dict[str, Any]) -> list[dict[str, Any]]:
    """
    Convert OpenCode raw events to the generic PinchBench transcript shape.

    PinchBench graders expect events shaped like:
      {"type": "message", "message": {"role": ..., "content": [...]}}
    """
    event_type = str(event.get("type") or "")
    normalized: list[dict[str, Any]] = []

    if event_type == "text":
        text = event_text(event)
        if text:
            normalized.append({
                "type": "message",
                "message": {
                    "role": "assistant",
                    "content": [{"type": "text", "text": text}],
                },
                "_adapter": {"source": "opencode", "event_type": event_type},
            })
        return normalized

    if event_type == "tool_use":
        part = event.get("part")
        if not isinstance(part, dict):
            return normalized

        state = part.get("state")
        if not isinstance(state, dict):
            state = {}

        tool_name = str(part.get("tool") or "")
        call_id = str(part.get("callID") or part.get("id") or "")
        tool_input = state.get("input")
        if not isinstance(tool_input, dict):
            tool_input = {}

        normalized.append({
            "type": "message",
            "message": {
                "role": "assistant",
                "content": [{
                    "type": "toolCall",
                    "id": call_id,
                    "name": tool_name,
                    "arguments": tool_input,
                    "input": tool_input,
                }],
            },
            "_adapter": {
                "source": "opencode",
                "event_type": event_type,
                "status": state.get("status"),
            },
        })

        status = str(state.get("status") or "")
        output = state.get("output")
        error = state.get("error")
        result_text = str(output if output is not None else error if error is not None else "")
        normalized.append({
            "type": "message",
            "message": {
                "role": "toolResult",
                "content": [{
                    "type": "toolResult",
                    "toolCallId": call_id,
                    "name": tool_name,
                    "content": result_text,
                    "isError": status == "error",
                }],
            },
            "_adapter": {
                "source": "opencode",
                "event_type": event_type,
                "status": status,
            },
        })
        return normalized

    if event_type == "error":
        error_text = extract_event_error(event)
        normalized.append({
            "type": "message",
            "message": {
                "role": "assistant",
                "content": [{"type": "text", "text": f"[OpenCode error] {error_text}"}],
            },
            "_adapter": {"source": "opencode", "event_type": event_type},
        })

    return normalized


def make_user_transcript_event(prompt: str, turn_index: int, session_label: str) -> dict[str, Any]:
    return {
        "type": "message",
        "message": {
            "role": "user",
            "content": [{"type": "text", "text": prompt}],
        },
        "_adapter": {
            "source": "pinchbench-opencode-runner",
            "turn": turn_index,
            "session_label": session_label,
        },
    }


def _stream_reader(
    stream: Any,
    kind: str,
    output_queue: "queue.Queue[tuple[str, Optional[str]]]",
) -> None:
    try:
        for line in iter(stream.readline, ""):
            output_queue.put((kind, line))
    finally:
        output_queue.put((kind, None))
        try:
            stream.close()
        except Exception:
            pass


def run_opencode_streaming(
    cmd: list[str],
    cwd: Path,
    timeout: float,
    raw_stdout_path: Path,
    stderr_path: Path,
    stdin_path: Path,
) -> dict[str, Any]:
    """
    Run one OpenCode turn.

    The prompt is supplied through ``stdin_path`` rather than as a positional
    command-line argument.  This is intentional: on native Windows an npm
    ``opencode.cmd`` shim is parsed by ``cmd.exe`` and multiline prompt
    arguments can be truncated or altered.  OpenCode 1.18.4 reads non-TTY
    stdin for ``opencode run``, so a UTF-8 file handle preserves the prompt
    exactly and also leaves an auditable copy beside the raw transcript.

    Uses reader threads rather than selectors because Windows selectors do not
    reliably support ordinary subprocess pipes.
    """
    monotonic_start = time.monotonic()
    wall_start_ms = time.time() * 1000.0
    deadline = monotonic_start + timeout

    raw_stdout_path.parent.mkdir(parents=True, exist_ok=True)
    stderr_path.parent.mkdir(parents=True, exist_ok=True)

    stdout_lines: list[str] = []
    stderr_lines: list[str] = []
    raw_events: list[dict[str, Any]] = []
    normalized_events: list[dict[str, Any]] = []
    output_chunks: list[str] = []
    event_counts: Counter[str] = Counter()

    session_id: Optional[str] = None
    ttft: Optional[float] = None
    error_messages: list[str] = []
    status = "success"
    timed_out = False

    usage_seen = False
    input_tokens = 0
    output_tokens = 0
    reasoning_tokens = 0
    cache_read_tokens = 0
    cache_write_tokens = 0
    cost_usd = 0.0
    step_count = 0
    tool_errors = 0
    finish_reasons: list[str] = []

    env = os.environ.copy()
    env.setdefault("PYTHONUTF8", "1")
    env.setdefault("OPENCODE_DISABLE_AUTOUPDATE", "true")
    env.setdefault("NO_COLOR", "1")

    creationflags = 0
    if sys.platform == "win32":
        creationflags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)

    try:
        with stdin_path.open("rb") as stdin_file:
            proc = subprocess.Popen(
                cmd,
                cwd=str(cwd),
                stdin=stdin_file,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace",
                bufsize=1,
                env=env,
                start_new_session=(sys.platform != "win32"),
                creationflags=creationflags,
            )
    except Exception as exc:
        return {
            "status": "error",
            "success": False,
            "returncode": None,
            "output": "",
            "error": f"无法启动 OpenCode: {exc}",
            "stderr": "",
            "elapsed": time.monotonic() - monotonic_start,
            "ttft": None,
            "input_tokens": None,
            "output_tokens": None,
            "reasoning_tokens": None,
            "cache_read_tokens": None,
            "cache_write_tokens": None,
            "cost_usd": None,
            "total_tokens": None,
            "token_source": "opencode_step_finish",
            "token_coverage_complete": False,
            "token_verified_against_openrouter": False,
            "usage_complete": False,
            "step_count": 0,
            "tool_errors": 0,
            "finish_reasons": [],
            "session_id": None,
            "raw_events": [],
            "normalized_events": [],
            "event_counts": {},
            "prompt_transport": "stdin_file",
            "prompt_path": str(stdin_path),
        }

    if proc.stdout is None or proc.stderr is None:
        kill_proc_tree(proc)
        raise RuntimeError("OpenCode stdout/stderr pipe creation failed")

    output_queue: "queue.Queue[tuple[str, Optional[str]]]" = queue.Queue()
    threads = [
        threading.Thread(
            target=_stream_reader,
            args=(proc.stdout, "stdout", output_queue),
            daemon=True,
        ),
        threading.Thread(
            target=_stream_reader,
            args=(proc.stderr, "stderr", output_queue),
            daemon=True,
        ),
    ]
    for thread in threads:
        thread.start()

    open_streams = {"stdout", "stderr"}

    with (
        raw_stdout_path.open("w", encoding="utf-8", newline="") as raw_file,
        stderr_path.open("w", encoding="utf-8", newline="") as err_file,
    ):
        while open_streams or proc.poll() is None:
            if time.monotonic() > deadline:
                timed_out = True
                status = "timeout"
                error_messages.append(f"超时 ({timeout:.0f}s)")
                kill_proc_tree(proc)
                break

            try:
                kind, line = output_queue.get(timeout=0.2)
            except queue.Empty:
                continue

            if line is None:
                open_streams.discard(kind)
                continue

            if kind == "stderr":
                stderr_lines.append(line)
                err_file.write(line)
                err_file.flush()
                continue

            stdout_lines.append(line)
            raw_file.write(line)
            raw_file.flush()

            stripped = line.strip()
            if not stripped:
                continue

            try:
                event = json.loads(stripped)
            except json.JSONDecodeError:
                if ttft is None:
                    ttft = time.monotonic() - monotonic_start
                output_chunks.append(stripped)
                event_counts["non_json_stdout"] += 1
                continue

            if not isinstance(event, dict):
                event_counts["non_object_json"] += 1
                continue

            raw_events.append(event)
            event_type = str(event.get("type") or "unknown")
            event_counts[event_type] += 1

            if not session_id and event.get("sessionID"):
                session_id = str(event.get("sessionID"))

            text = event_text(event)
            if text:
                if ttft is None:
                    part = event.get("part")
                    candidate: Optional[float] = None
                    if isinstance(part, dict):
                        part_time = part.get("time")
                        if isinstance(part_time, dict):
                            try:
                                start_ms = float(part_time.get("start"))
                                estimated = (start_ms - wall_start_ms) / 1000.0
                                if 0.0 <= estimated <= (time.monotonic() - monotonic_start + 5.0):
                                    candidate = estimated
                            except (TypeError, ValueError):
                                pass
                    ttft = candidate if candidate is not None else time.monotonic() - monotonic_start
                output_chunks.append(text)

            normalized_events.extend(normalize_opencode_event(event))

            usage = parse_step_finish(event)
            if usage is not None:
                usage_seen = True
                step_count += 1
                input_tokens += usage["input_tokens"]
                output_tokens += usage["output_tokens"]
                reasoning_tokens += usage["reasoning_tokens"]
                cache_read_tokens += usage["cache_read_tokens"]
                cache_write_tokens += usage["cache_write_tokens"]
                cost_usd += usage["cost_usd"]
                if usage["finish_reason"]:
                    finish_reasons.append(usage["finish_reason"])

            if event_type == "tool_use":
                part = event.get("part")
                if isinstance(part, dict):
                    state = part.get("state")
                    if isinstance(state, dict) and state.get("status") == "error":
                        tool_errors += 1

            event_error = extract_event_error(event)
            if event_error:
                error_messages.append(event_error)
                status = "error"

    try:
        proc.wait(timeout=15)
    except subprocess.TimeoutExpired:
        kill_proc_tree(proc)
        status = "timeout"
        timed_out = True
        error_messages.append("进程未正常退出，已强制终止")

    # Drain anything already queued after process exit.
    drain_deadline = time.monotonic() + 2.0
    while time.monotonic() < drain_deadline:
        try:
            kind, line = output_queue.get_nowait()
        except queue.Empty:
            break
        if line is None:
            continue
        if kind == "stderr":
            stderr_lines.append(line)
        else:
            stdout_lines.append(line)

    for thread in threads:
        thread.join(timeout=1)

    returncode = proc.returncode
    stderr = "".join(stderr_lines).strip()

    if returncode not in (0, None) and status == "success":
        status = "error"
        error_messages.append(stderr or f"OpenCode 退出码 {returncode}")

    # A non-zero exit or timeout is failure. Missing final text alone is not.
    success = status == "success" and not timed_out and returncode in (0, None)

    output = "\n".join(chunk.strip() for chunk in output_chunks if chunk.strip()).strip()
    elapsed = time.monotonic() - monotonic_start

    return {
        "status": status,
        "success": success,
        "returncode": returncode,
        "output": output,
        "error": " | ".join(dict.fromkeys(message for message in error_messages if message)),
        "stderr": stderr,
        "elapsed": elapsed,
        "ttft": ttft,
        "input_tokens": input_tokens if usage_seen else None,
        "output_tokens": output_tokens if usage_seen else None,
        "reasoning_tokens": reasoning_tokens if usage_seen else None,
        "cache_read_tokens": cache_read_tokens if usage_seen else None,
        "cache_write_tokens": cache_write_tokens if usage_seen else None,
        "cost_usd": cost_usd if usage_seen else None,
        "usage_complete": usage_seen,
        "step_count": step_count,
        "tool_errors": tool_errors,
        "finish_reasons": finish_reasons,
        "session_id": session_id,
        "raw_events": raw_events,
        "normalized_events": normalized_events,
        "event_counts": dict(event_counts),
        "stdout_lines": stdout_lines,
        "prompt_transport": "stdin_file",
        "prompt_path": str(stdin_path),
    }


# ---------------------------------------------------------------------------
# PinchBench official/default grading
# ---------------------------------------------------------------------------

def load_pinchbench_grading(skill_dir: Path) -> tuple[Optional[Any], Optional[str]]:
    """Load the grading engine from the exact checked-out PinchBench commit."""
    scripts_dir = (skill_dir / "scripts").resolve()
    grading_path = scripts_dir / "lib_grading.py"
    if not grading_path.exists():
        return None, f"找不到 PinchBench grading engine: {grading_path}"

    scripts_value = str(scripts_dir)
    if scripts_value not in sys.path:
        sys.path.insert(0, scripts_value)

    try:
        import lib_grading  # type: ignore
    except Exception as exc:
        return None, (
            "无法导入 PinchBench scripts/lib_grading.py。"
            "请安装本地仓库依赖，例如："
            r"C:\pinchbench-opencode\.venv\Scripts\python.exe -m pip install -e "
            f'"{skill_dir}"。原始错误: {exc}'
        )

    return lib_grading, None


def grade_with_pinchbench_default(
    *,
    grader_module: Any,
    task: Task,
    execution_result: dict[str, Any],
    workspace: Path,
    skill_dir: Path,
    judge_timeout: float,
    judge_model: str,
    verbose: bool,
) -> GradeResult:
    """Use PinchBench's own grader with its explicitly pinned judge model.

    The judge is called directly through the API backend.  For the current
    commit, DEFAULT_JUDGE_MODEL is openrouter/anthropic/claude-opus-5 and
    lib_agent.py reads OPENROUTER_API_KEY.
    """
    grading_execution = dict(execution_result)
    grading_execution["workspace"] = str(workspace)

    try:
        result = grader_module.grade_task(
            task=task,
            execution_result=grading_execution,
            skill_dir=skill_dir,
            judge_model=judge_model,
            judge_timeout_seconds=judge_timeout,
            judge_backend="api",
            verbose=verbose,
        )

        raw_score = getattr(result, "score", None)
        max_score = getattr(result, "max_score", 1.0)
        score: Optional[float]
        if raw_score is None:
            score = None
        else:
            score = float(raw_score)
            max_score_float = float(max_score or 1.0)
            if max_score_float != 1.0:
                score = score / max_score_float
            score = max(0.0, min(1.0, score))

        notes = str(getattr(result, "notes", "") or "")
        grade_error: Optional[str] = None
        if task.grading_type in {"llm_judge", "hybrid"}:
            lower_notes = notes.lower()
            failure_markers = (
                "llm judge failed",
                "no parseable response",
                "response parsed but no score",
                "openrouter_api_key not set",
                "judge api call failed",
            )
            if any(marker in lower_notes for marker in failure_markers):
                grade_error = notes or "LLM judge failed"

        return GradeResult(
            score=score,
            grading_type=str(getattr(result, "grading_type", task.grading_type)),
            breakdown=dict(getattr(result, "breakdown", {}) or {}),
            notes=notes,
            error=grade_error,
        )
    except Exception as exc:
        return GradeResult(
            score=None,
            grading_type=task.grading_type,
            breakdown={},
            notes="",
            error=(
                "PinchBench 官方评分失败: "
                f"{exc}\n{traceback.format_exc(limit=8)}"
            ),
        )


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def truncate_excel(value: Any, limit: int = 32000) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    return text if len(text) <= limit else text[:limit] + "...[truncated]"


def save_csv(results: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "task_id", "name", "category", "grading_type", "model", "network_task",
        "multi_session", "session_count", "success", "status", "returncode",
        "score", "elapsed", "agent_elapsed", "grading_elapsed", "end_to_end_elapsed",
        "ttft", "input_tokens", "output_tokens", "reasoning_tokens",
        "cache_read_tokens", "cache_write_tokens", "total_tokens", "cost_usd",
        "token_source", "token_coverage_complete", "token_verified_against_openrouter",
        "usage_complete", "step_count", "tool_errors",
        "workspace", "transcript", "error", "grade_error", "grade_notes",
    ]

    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for row in results:
            writer.writerow({
                key: truncate_excel(row.get(key), 10000)
                for key in fields
            })


def save_xlsx(
    results: list[dict[str, Any]],
    summary: dict[str, Any],
    output_path: Path,
) -> None:
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
    except ImportError:
        LOGGER.warning(
            "未安装 openpyxl，跳过 XLSX。可运行: "
            "python -m pip install openpyxl"
        )
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "详细结果"

    headers = [
        "任务ID", "名称", "类别", "打分类型", "模型", "联网题", "多轮任务",
        "会话调用数", "成功", "状态", "退出码", "分数", "旧兼容耗时(s)",
        "Agent耗时(s)", "Grading耗时(s)", "端到端耗时(s)", "TTFT估计(s)",
        "输入Token", "输出Token", "推理Token", "缓存读取Token", "缓存写入Token",
        "派生总Token", "费用USD(OpenCode)", "Token来源", "Token覆盖完整",
        "已与OpenRouter核验", "Usage完整", "Step数", "工具错误数",
        "Workspace", "Transcript", "运行错误", "打分错误", "打分备注",
    ]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(results, start=2):
        values = [
            row.get("task_id"),
            row.get("name"),
            row.get("category"),
            row.get("grading_type"),
            row.get("model") or "",
            "是" if row.get("network_task") else "否",
            "是" if row.get("multi_session") else "否",
            row.get("session_count"),
            "✓" if row.get("success") else "✗",
            row.get("status"),
            row.get("returncode") if row.get("returncode") is not None else "",
            row.get("score") if row.get("score") is not None else "",
            row.get("elapsed"),
            row.get("agent_elapsed") if row.get("agent_elapsed") is not None else "",
            row.get("grading_elapsed") if row.get("grading_elapsed") is not None else "",
            row.get("end_to_end_elapsed") if row.get("end_to_end_elapsed") is not None else "",
            row.get("ttft") if row.get("ttft") is not None else "",
            row.get("input_tokens") if row.get("input_tokens") is not None else "",
            row.get("output_tokens") if row.get("output_tokens") is not None else "",
            row.get("reasoning_tokens") if row.get("reasoning_tokens") is not None else "",
            row.get("cache_read_tokens") if row.get("cache_read_tokens") is not None else "",
            row.get("cache_write_tokens") if row.get("cache_write_tokens") is not None else "",
            row.get("total_tokens") if row.get("total_tokens") is not None else "",
            row.get("cost_usd") if row.get("cost_usd") is not None else "",
            row.get("token_source") or "",
            "是" if row.get("token_coverage_complete") else "否",
            "是" if row.get("token_verified_against_openrouter") else "否",
            "是" if row.get("usage_complete") else "否",
            row.get("step_count"),
            row.get("tool_errors"),
            row.get("workspace"),
            row.get("transcript"),
            row.get("error") or "",
            row.get("grade_error") or "",
            row.get("grade_notes") or "",
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(
                row=row_index,
                column=column,
                value=truncate_excel(value),
            )

    for cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in cells)
        worksheet.column_dimensions[cells[0].column_letter].width = min(
            max_length + 3,
            60,
        )

    summary_sheet = workbook.create_sheet("汇总")
    for row_index, (key, value) in enumerate(summary.items(), start=1):
        summary_sheet.cell(row=row_index, column=1, value=key).font = Font(bold=True)
        summary_sheet.cell(row=row_index, column=2, value=value)
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 30

    workbook.save(output_path)


def optional_sum(results: list[dict[str, Any]], field: str) -> Optional[float]:
    values = [row.get(field) for row in results if row.get(field) is not None]
    if not values:
        return None
    return float(sum(float(value) for value in values))


def print_summary(
    results: list[dict[str, Any]],
    total_elapsed: float,
) -> dict[str, Any]:
    succeeded = sum(1 for row in results if row.get("success"))
    failed = len(results) - succeeded
    scores = [
        float(row["score"])
        for row in results
        if row.get("score") is not None
    ]
    ttfts = [
        float(row["ttft"])
        for row in results
        if row.get("ttft") is not None
    ]
    average_elapsed = (
        sum(float(row.get("agent_elapsed") or row.get("elapsed") or 0.0) for row in results) / len(results)
        if results else 0.0
    )
    total_agent_elapsed = sum(float(row.get("agent_elapsed") or row.get("elapsed") or 0.0) for row in results)
    total_grading_elapsed = sum(float(row.get("grading_elapsed") or 0.0) for row in results)
    total_end_to_end_elapsed = sum(float(row.get("end_to_end_elapsed") or 0.0) for row in results)

    summary = {
        "任务总数": len(results),
        "成功": succeeded,
        "失败": failed,
        "成功率": round(succeeded / len(results), 4) if results else 0.0,
        "有分数任务数": len(scores),
        "打分失败任务数": sum(1 for row in results if row.get("grade_error")),
        "平均分数": round(sum(scores) / len(scores), 4) if scores else None,
        "整批墙钟总耗时(s)": round(total_elapsed, 2),
        "Agent累计耗时(s)": round(total_agent_elapsed, 2),
        "Grading累计耗时(s)": round(total_grading_elapsed, 2),
        "任务端到端累计耗时(s)": round(total_end_to_end_elapsed, 2),
        "平均Agent耗时/任务(s)": round(average_elapsed, 2),
        "平均TTFT估计(s)": round(sum(ttfts) / len(ttfts), 4) if ttfts else None,
        "总输入Token": int(optional_sum(results, "input_tokens") or 0)
            if any(row.get("input_tokens") is not None for row in results) else None,
        "总输出Token": int(optional_sum(results, "output_tokens") or 0)
            if any(row.get("output_tokens") is not None for row in results) else None,
        "总推理Token": int(optional_sum(results, "reasoning_tokens") or 0)
            if any(row.get("reasoning_tokens") is not None for row in results) else None,
        "总缓存读取Token": int(optional_sum(results, "cache_read_tokens") or 0)
            if any(row.get("cache_read_tokens") is not None for row in results) else None,
        "总缓存写入Token": int(optional_sum(results, "cache_write_tokens") or 0)
            if any(row.get("cache_write_tokens") is not None for row in results) else None,
        "派生总Token": int(optional_sum(results, "total_tokens") or 0)
            if any(row.get("total_tokens") is not None for row in results) else None,
        "Token来源": "OpenCode step_finish JSONL (逐 step 累加)",
        "已与OpenRouter逐请求核验": False,
        "总费用USD(OpenCode上报)": round(optional_sum(results, "cost_usd") or 0.0, 6)
            if any(row.get("cost_usd") is not None for row in results) else None,
        "Usage缺失任务数": sum(
            1 for row in results if not row.get("usage_complete")
        ),
        "联网任务数": sum(1 for row in results if row.get("network_task")),
        "多轮任务数": sum(1 for row in results if row.get("multi_session")),
    }

    print("\n" + "=" * 100)
    print("汇总")
    print("=" * 100)
    for key, value in summary.items():
        print(f"{key:<22}: {value}")

    print("-" * 100)
    for row in results:
        marker = "✓" if row.get("success") else "✗"
        score = (
            f"{row['score']:.3f}"
            if row.get("score") is not None
            else "N/A"
        )
        ttft = (
            f"{row['ttft']:.2f}s"
            if row.get("ttft") is not None
            else "N/A"
        )
        error = row.get("error") or row.get("grade_error") or ""
        error_suffix = f"  [{str(error)[:70]}]" if error else ""
        print(
            f"{marker} {row['task_id']:<42} "
            f"score={score:<6} elapsed={row['elapsed']:.1f}s "
            f"ttft={ttft:<8}{error_suffix}"
        )
    print("=" * 100)
    return summary


# ---------------------------------------------------------------------------
# Preflight and run metadata
# ---------------------------------------------------------------------------

def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def command_output(command: list[str], cwd: Optional[Path] = None) -> str:
    try:
        result = subprocess.run(
            command,
            cwd=str(cwd) if cwd else None,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
            check=False,
        )
        return (result.stdout or result.stderr).strip()
    except Exception as exc:
        return f"unavailable: {exc}"


def write_run_config(
    output_path: Path,
    args: argparse.Namespace,
    skill_dir: Path,
    tasks_dir: Path,
    opencode_command: str,
    selected: list[Task],
    judge_model: str,
) -> None:
    manifest_path = tasks_dir / "manifest.yaml"
    git_command = resolve_command("git")

    commit = ""
    if git_command:
        commit = command_output(
            [git_command, "rev-parse", "HEAD"],
            cwd=skill_dir,
        )

    config = {
        "runner_revision": RUNNER_REVISION,
        "created_at": dt.datetime.now(dt.timezone.utc).astimezone().isoformat(),
        "platform": platform.platform(),
        "python": sys.version,
        "opencode_command": opencode_command,
        "opencode_version": command_output([opencode_command, "--version"]),
        "model": args.model,
        "agent": "OpenCode default (not explicitly set)",
        "variant": "default (not explicitly set)",
        "auto_approval": False,
        "prompt_transport": "stdin_file_utf8",
        "prompt_in_argv": False,
        "openrouter_routing": "default",
        "worker_count": 1,
        "task_concurrency": 1,
        "skill_dir": str(skill_dir),
        "tasks_dir": str(tasks_dir),
        "pinchbench_commit": commit,
        "manifest_sha256": sha256_file(manifest_path) if manifest_path.exists() else None,
        "suite": args.suite,
        "limit": args.limit,
        "default_skipped_tasks": sorted(DEFAULT_SKIPPED_TASKS),
        "additional_skip": args.skip,
        "skip_network": args.skip_network,
        "task_count": len(selected),
        "task_ids": [task.task_id for task in selected],
        "timeout_multiplier": args.timeout_multiplier,
        "network_timeout": args.network_timeout,
        "workspace_instruction": not args.no_workspace_instruction,
        "grading_enabled": not args.no_grade,
        "grading_engine": str(skill_dir / "scripts" / "lib_grading.py"),
        "judge_backend": "api",
        "judge_model": judge_model,
        "judge_key_env": "OPENROUTER_API_KEY",
        "judge_key_present": bool(os.environ.get("OPENROUTER_API_KEY")),
        "judge_concurrency": 1,
        "judge_is_separate_from_tested_model": True,
        "token_accounting": {
            "source": "OpenCode --format json step_finish events",
            "aggregation": "sum every step_finish across all turns",
            "derived_total_formula": "input + output + reasoning + cache_read + cache_write",
            "openrouter_reconciled": False,
            "cost_source": "OpenCode step_finish cost",
        },
        "environment_proxy_present": {
            key: bool(os.environ.get(key))
            for key in ("HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "NO_PROXY")
        },
    }
    output_path.write_text(
        json.dumps(config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def print_preflight(
    selected: list[Task],
    skill_dir: Path,
    tasks_dir: Path,
    opencode_command: str,
    grader_module: Optional[Any],
    grader_error: Optional[str],
    args: argparse.Namespace,
) -> tuple[dict[str, list[str]], list[str]]:
    prerequisite_failures: dict[str, list[str]] = {}
    fixture_failures: list[str] = []

    needs_judge = (
        not args.no_grade
        and any(task.grading_type in {"llm_judge", "hybrid"} for task in selected)
    )
    judge_model = str(args.judge_model)
    judge_key_present = bool(os.environ.get("OPENROUTER_API_KEY"))

    print("=" * 100)
    print("PinchBench OpenCode + Kimi K3 Windows runner preflight")
    print("=" * 100)
    print(f"Runner revision      : {RUNNER_REVISION}")
    print(f"Skill dir            : {skill_dir}")
    print(f"Tasks dir            : {tasks_dir}")
    print(f"OpenCode command     : {opencode_command}")
    print(f"OpenCode version     : {command_output([opencode_command, '--version'])}")
    print(f"Model                : {args.model}")
    print(f"Suite                : {args.suite}")
    print(f"Selected tasks       : {len(selected)}")
    print(f"Default skipped      : {len(DEFAULT_SKIPPED_TASKS)}")
    for task_id in sorted(DEFAULT_SKIPPED_TASKS):
        print(f"  - {task_id}")
    print("Worker/concurrency   : 1 / 1")
    print("Judge concurrency    : 1 (synchronous)")
    print("Default agent/tools  : yes (no --agent, no tool list)")
    print("Auto approval        : no")
    print("Prompt transport     : UTF-8 stdin file (never argv)")
    print("Reasoning variant    : default")
    print(f"Grading engine       : {skill_dir / 'scripts' / 'lib_grading.py'}")
    print("Judge backend        : api")
    print(f"Judge model          : {judge_model}")
    print(f"OPENROUTER_API_KEY   : {'set' if judge_key_present else 'missing'}")
    print("OpenClaw required    : no")
    print(f"Grader import        : {'ok' if grader_error is None else 'FAILED'}")
    if grader_error:
        print(f"  {grader_error}")
    print()

    category_counts = Counter(task.category or "(uncategorized)" for task in selected)
    grading_counts = Counter(task.grading_type for task in selected)
    print("Grading types:")
    for key, value in sorted(grading_counts.items()):
        print(f"  {key:<24} {value}")
    print("Categories:")
    for key, value in sorted(category_counts.items()):
        print(f"  {key:<24} {value}")
    print(f"Network-marked tasks : {sum(1 for task in selected if is_network_task(task))}")
    print(f"Multi-session tasks  : {sum(1 for task in selected if task.multi_session)}")
    print(f"Default judge needed : {'yes' if needs_judge else 'no'}")
    if needs_judge and not judge_key_present:
        print("Judge readiness      : FAILED (OPENROUTER_API_KEY missing)")
    elif needs_judge:
        print("Judge readiness      : ok")
    else:
        print("Judge readiness      : not required")
    print()

    for task in selected:
        prerequisites = task.metadata.get("prerequisites") or []
        missing = check_prerequisites(list(prerequisites)) if prerequisites else []
        if missing:
            prerequisite_failures[task.task_id] = missing

        for workspace_file in task.workspace_files:
            if not isinstance(workspace_file, dict):
                fixture_failures.append(f"{task.task_id}: invalid workspace file entry")
                continue
            if workspace_file.get("content") is not None:
                continue
            source = workspace_file.get("source")
            if not source:
                continue
            src_rel = Path(str(source))
            candidates = [
                skill_dir / "assets" / src_rel,
                skill_dir / src_rel,
                task.file_path.parent / src_rel,
                task.file_path.parent.parent / "assets" / src_rel,
            ]
            if not any(candidate.exists() for candidate in candidates):
                fixture_failures.append(f"{task.task_id}: missing fixture {source}")

    if prerequisite_failures:
        print("Missing declared prerequisites:")
        for task_id, missing in prerequisite_failures.items():
            print(f"  {task_id}: {', '.join(missing)}")
    else:
        print("Missing declared prerequisites: none")

    if fixture_failures:
        print("Missing workspace fixtures:")
        for item in fixture_failures:
            print(f"  {item}")
    else:
        print("Missing workspace fixtures: none")

    print("=" * 100)
    return prerequisite_failures, fixture_failures


# ---------------------------------------------------------------------------
# Main execution
# ---------------------------------------------------------------------------

def build_turns(task: Task) -> list[dict[str, Any]]:
    if task.multi_session and task.sessions:
        turns: list[dict[str, Any]] = []
        for index, item in enumerate(task.sessions, start=1):
            prompt = str(item.get("prompt") or "").strip()
            if not prompt:
                continue
            turns.append({
                "id": str(item.get("id") or f"turn_{index}"),
                "prompt": prompt,
                "new_session": bool(item.get("new_session")),
            })
        if turns:
            return turns

    return [{
        "id": "single",
        "prompt": task.prompt,
        "new_session": True,
    }]


def aggregate_optional_int(
    current: Optional[int],
    value: Optional[int],
) -> Optional[int]:
    if value is None:
        return current
    return (current or 0) + int(value)


def aggregate_optional_float(
    current: Optional[float],
    value: Optional[float],
) -> Optional[float]:
    if value is None:
        return current
    return (current or 0.0) + float(value)


def derive_total_tokens(
    input_tokens: Optional[int],
    output_tokens: Optional[int],
    reasoning_tokens: Optional[int],
    cache_read_tokens: Optional[int],
    cache_write_tokens: Optional[int],
) -> Optional[int]:
    """Sum OpenCode step-finish token categories for benchmark accounting.

    This value is derived from OpenCode JSON events and is not claimed to be
    independently reconciled against OpenRouter billing. Raw JSONL is retained.
    """
    values = (input_tokens, output_tokens, reasoning_tokens, cache_read_tokens, cache_write_tokens)
    if all(value is None for value in values):
        return None
    return int(sum(int(value or 0) for value in values))


def execute_task(
    task: Task,
    task_index: int,
    task_count: int,
    skill_dir: Path,
    workspace: Path,
    transcript_dir: Path,
    opencode_command: str,
    args: argparse.Namespace,
) -> dict[str, Any]:
    print(f"[{task_index}/{task_count}] {task.task_id} — {task.name}")
    prompt_preview = re.sub(r"\s+", " ", task.prompt)[:150]
    print(
        f"  prompt: {prompt_preview}"
        f"{'...' if len(task.prompt) > 150 else ''}"
    )

    prerequisites = task.metadata.get("prerequisites") or []
    missing_prerequisites = (
        check_prerequisites(list(prerequisites))
        if prerequisites else []
    )
    if missing_prerequisites:
        error = f"缺少依赖: {', '.join(missing_prerequisites)}"
        print(f"  跳过: {error}\n")
        return {
            "task_id": task.task_id,
            "name": task.name,
            "category": task.category,
            "grading_type": task.grading_type,
            "model": args.model,
            "network_task": is_network_task(task),
            "multi_session": task.multi_session,
            "session_count": 0,
            "success": False,
            "status": "missing_prerequisite",
            "returncode": None,
            "elapsed": 0.0,
            "ttft": None,
            "input_tokens": None,
            "output_tokens": None,
            "reasoning_tokens": None,
            "cache_read_tokens": None,
            "cache_write_tokens": None,
            "cost_usd": None,
            "total_tokens": None,
            "token_source": "opencode_step_finish",
            "token_coverage_complete": False,
            "token_verified_against_openrouter": False,
            "usage_complete": False,
            "step_count": 0,
            "tool_errors": 0,
            "output": "",
            "error": error,
            "stderr": "",
            "score": None,
            "breakdown": {},
            "grade_notes": "",
            "grade_error": error,
            "workspace": str(workspace),
            "transcript": str(transcript_dir),
            "normalized_transcript": [],
            "turn_results": [],
        }

    if workspace.exists():
        shutil.rmtree(workspace)
    workspace.mkdir(parents=True, exist_ok=True)
    transcript_dir.mkdir(parents=True, exist_ok=True)

    staged, fixture_errors = stage_workspace_files(task, workspace, skill_dir)
    if staged:
        LOGGER.info(
            "  staged files: %s",
            ", ".join(staged[:8]) + (" ..." if len(staged) > 8 else ""),
        )
    if fixture_errors:
        error = "缺少或无效的 workspace fixture: " + " | ".join(fixture_errors)
        print(f"  失败: {error}\n")
        return {
            "task_id": task.task_id,
            "name": task.name,
            "category": task.category,
            "grading_type": task.grading_type,
            "model": args.model,
            "network_task": is_network_task(task),
            "multi_session": task.multi_session,
            "session_count": 0,
            "success": False,
            "status": "missing_fixture",
            "returncode": None,
            "elapsed": 0.0,
            "ttft": None,
            "input_tokens": None,
            "output_tokens": None,
            "reasoning_tokens": None,
            "cache_read_tokens": None,
            "cache_write_tokens": None,
            "cost_usd": None,
            "usage_complete": False,
            "step_count": 0,
            "tool_errors": 0,
            "output": "",
            "error": error,
            "stderr": "",
            "score": None,
            "breakdown": {},
            "grade_notes": "",
            "grade_error": error,
            "workspace": str(workspace),
            "transcript": str(transcript_dir),
            "normalized_transcript": [],
            "turn_results": [],
        }

    turns = build_turns(task)
    total_timeout = (
        args.network_timeout
        if is_network_task(task)
        else max(1.0, task.timeout_seconds * args.timeout_multiplier)
    )
    task_deadline = time.monotonic() + total_timeout

    normalized_transcript: list[dict[str, Any]] = []
    outputs: list[str] = []
    errors: list[str] = []
    stderr_chunks: list[str] = []
    turn_results: list[dict[str, Any]] = []

    current_session_id: Optional[str] = None
    task_status = "success"
    task_success = True
    returncode: Optional[int] = 0
    total_elapsed = 0.0
    task_ttft: Optional[float] = None
    input_tokens: Optional[int] = None
    output_tokens: Optional[int] = None
    reasoning_tokens: Optional[int] = None
    cache_read_tokens: Optional[int] = None
    cache_write_tokens: Optional[int] = None
    cost_usd: Optional[float] = None
    usage_complete = True
    step_count = 0
    tool_errors = 0
    event_counts: Counter[str] = Counter()

    for turn_index, turn in enumerate(turns, start=1):
        remaining = task_deadline - time.monotonic()
        if remaining <= 0:
            task_success = False
            task_status = "timeout"
            errors.append(f"任务总超时 ({total_timeout:.0f}s)")
            break

        turn_id = str(turn["id"])
        turn_prompt = str(turn["prompt"])
        new_session = bool(turn.get("new_session"))

        normalized_transcript.append(
            make_user_transcript_event(turn_prompt, turn_index, turn_id)
        )

        prompt = turn_prompt
        if not args.no_workspace_instruction:
            prompt += (
                "\n\nIMPORTANT: You are running in an isolated workspace. "
                "Read, write, and edit files only in the current working directory."
            )

        command = [
            opencode_command,
            "run",
            "--model",
            args.model,
            "--format",
            "json",
            "--dir",
            str(workspace),
        ]

        if not new_session and current_session_id:
            command.extend(["--session", current_session_id])

        # Never append the prompt to ``command`` on native Windows.
        # ``opencode.cmd`` is an npm batch shim and cmd.exe can corrupt
        # multiline positional arguments.  OpenCode 1.18.4 supports reading
        # the run message from non-TTY stdin, so persist the exact UTF-8 prompt
        # and redirect that file into the child process.
        raw_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.jsonl"
        stderr_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.stderr.txt"
        prompt_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.prompt.txt"

        prompt_payload = prompt if prompt.endswith("\n") else prompt + "\n"
        prompt_path.write_text(
            prompt_payload,
            encoding="utf-8",
            newline="\n",
        )
        prompt_bytes = prompt_payload.encode("utf-8")
        prompt_sha256 = hashlib.sha256(prompt_bytes).hexdigest()

        turn_result = run_opencode_streaming(
            command,
            cwd=workspace,
            timeout=max(1.0, remaining),
            raw_stdout_path=raw_path,
            stderr_path=stderr_path,
            stdin_path=prompt_path,
        )

        normalized_transcript.extend(turn_result["normalized_events"])
        if turn_result.get("output"):
            outputs.append(
                f"## Turn {turn_index} ({turn_id})\n{turn_result['output']}"
            )
        if turn_result.get("error"):
            errors.append(
                f"turn {turn_index} ({turn_id}): {turn_result['error']}"
            )
        if turn_result.get("stderr"):
            stderr_chunks.append(
                f"## Turn {turn_index} ({turn_id})\n{turn_result['stderr']}"
            )

        total_elapsed += float(turn_result.get("elapsed") or 0.0)
        if task_ttft is None and turn_result.get("ttft") is not None:
            task_ttft = float(turn_result["ttft"])

        input_tokens = aggregate_optional_int(
            input_tokens,
            turn_result.get("input_tokens"),
        )
        output_tokens = aggregate_optional_int(
            output_tokens,
            turn_result.get("output_tokens"),
        )
        reasoning_tokens = aggregate_optional_int(
            reasoning_tokens,
            turn_result.get("reasoning_tokens"),
        )
        cache_read_tokens = aggregate_optional_int(
            cache_read_tokens,
            turn_result.get("cache_read_tokens"),
        )
        cache_write_tokens = aggregate_optional_int(
            cache_write_tokens,
            turn_result.get("cache_write_tokens"),
        )
        cost_usd = aggregate_optional_float(
            cost_usd,
            turn_result.get("cost_usd"),
        )

        usage_complete = usage_complete and bool(turn_result.get("usage_complete"))
        step_count += int(turn_result.get("step_count") or 0)
        tool_errors += int(turn_result.get("tool_errors") or 0)
        event_counts.update(turn_result.get("event_counts") or {})

        returned_session_id = turn_result.get("session_id")
        if returned_session_id:
            current_session_id = str(returned_session_id)

        turn_results.append({
            "turn": turn_index,
            "turn_id": turn_id,
            "new_session": new_session,
            "session_id": current_session_id,
            "success": turn_result.get("success"),
            "status": turn_result.get("status"),
            "returncode": turn_result.get("returncode"),
            "elapsed": turn_result.get("elapsed"),
            "ttft": turn_result.get("ttft"),
            "input_tokens": turn_result.get("input_tokens"),
            "output_tokens": turn_result.get("output_tokens"),
            "reasoning_tokens": turn_result.get("reasoning_tokens"),
            "cache_read_tokens": turn_result.get("cache_read_tokens"),
            "cache_write_tokens": turn_result.get("cache_write_tokens"),
            "total_tokens": derive_total_tokens(
                turn_result.get("input_tokens"),
                turn_result.get("output_tokens"),
                turn_result.get("reasoning_tokens"),
                turn_result.get("cache_read_tokens"),
                turn_result.get("cache_write_tokens"),
            ),
            "cost_usd": turn_result.get("cost_usd"),
            "token_source": "opencode_step_finish",
            "token_coverage_complete": bool(turn_result.get("usage_complete")),
            "token_verified_against_openrouter": False,
            "usage_complete": turn_result.get("usage_complete"),
            "step_count": turn_result.get("step_count"),
            "tool_errors": turn_result.get("tool_errors"),
            "event_counts": turn_result.get("event_counts"),
            "raw_transcript": str(raw_path),
            "stderr_path": str(stderr_path),
            "prompt_transport": turn_result.get("prompt_transport"),
            "prompt_path": str(prompt_path),
            "prompt_chars": len(prompt_payload),
            "prompt_bytes": len(prompt_bytes),
            "prompt_sha256": prompt_sha256,
            "error": turn_result.get("error"),
        })

        if not turn_result.get("success"):
            task_success = False
            task_status = str(turn_result.get("status") or "error")
            returncode = turn_result.get("returncode")
            break

        returncode = turn_result.get("returncode")

    normalized_path = transcript_dir / "normalized.jsonl"
    with normalized_path.open("w", encoding="utf-8") as file:
        for event in normalized_transcript:
            file.write(json.dumps(event, ensure_ascii=False) + "\n")

    turn_results_path = transcript_dir / "turn_results.json"
    turn_results_path.write_text(
        json.dumps(turn_results, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return {
        "task_id": task.task_id,
        "name": task.name,
        "category": task.category,
        "grading_type": task.grading_type,
        "model": args.model,
        "network_task": is_network_task(task),
        "multi_session": task.multi_session,
        "session_count": len(turn_results),
        "success": task_success,
        "status": task_status,
        "returncode": returncode,
        "elapsed": total_elapsed,
        "ttft": task_ttft,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "reasoning_tokens": reasoning_tokens,
        "cache_read_tokens": cache_read_tokens,
        "cache_write_tokens": cache_write_tokens,
        "total_tokens": derive_total_tokens(
            input_tokens, output_tokens, reasoning_tokens, cache_read_tokens, cache_write_tokens
        ),
        "cost_usd": cost_usd,
        "token_source": "opencode_step_finish",
        "token_coverage_complete": usage_complete and bool(turn_results),
        "token_verified_against_openrouter": False,
        "usage_complete": usage_complete and bool(turn_results),
        "step_count": step_count,
        "tool_errors": tool_errors,
        "event_counts": dict(event_counts),
        "output": "\n\n".join(outputs),
        "error": " | ".join(errors),
        "stderr": "\n\n".join(stderr_chunks),
        "score": None,
        "breakdown": {},
        "grade_notes": "",
        "grade_error": None,
        "workspace": str(workspace),
        "transcript": str(transcript_dir),
        "normalized_transcript_path": str(normalized_path),
        "transcript_data": normalized_transcript,
        "turn_results": turn_results,
    }


def build_args() -> argparse.Namespace:
    script_path = Path(__file__).resolve()
    inferred_root = script_path.parent.parent
    inferred_skill = inferred_root / "skill"

    parser = argparse.ArgumentParser(
        description=(
            "Run local PinchBench tasks with OpenCode + Kimi K3 on native Windows, "
            "one task at a time."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--skill-dir",
        default=str(inferred_skill if inferred_skill.exists() else Path.cwd()),
        help="PinchBench skill 仓库根目录",
    )
    parser.add_argument(
        "--tasks-dir",
        default=None,
        help="任务目录；默认是 <skill-dir>/tasks",
    )
    parser.add_argument(
        "--suite",
        default="core",
        help="all、core、automated-only、llm-judge-only、hybrid-only、judge-required-only，或逗号分隔任务 ID",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="最多运行多少个任务，仅用于测试 runner",
    )
    parser.add_argument(
        "--skip",
        default="",
        help="额外跳过的任务 ID；四个已约定 integration 任务默认跳过",
    )
    parser.add_argument(
        "--skip-network",
        action="store_true",
        help="跳过明确标记的联网任务；正式全量测试通常不应使用",
    )
    parser.add_argument(
        "--opencode",
        default="auto",
        help="OpenCode 命令：auto、opencode.cmd 或绝对路径",
    )
    parser.add_argument(
        "--model",
        default=DEFAULT_MODEL,
        help="OpenCode provider/model ID",
    )
    parser.add_argument(
        "--timeout-multiplier",
        type=float,
        default=3.0,
        help="非联网任务 timeout_seconds 的倍数",
    )
    parser.add_argument(
        "--network-timeout",
        type=float,
        default=300.0,
        help="明确标记的联网任务总超时秒数",
    )
    parser.add_argument(
        "--judge-timeout",
        type=float,
        default=300.0,
        help="PinchBench 默认 LLM judge 单次请求超时秒数",
    )
    parser.add_argument(
        "--judge-model",
        default=DEFAULT_JUDGE_MODEL,
        help="PinchBench LLM judge model；正式口径固定 Claude Opus 5",
    )
    parser.add_argument(
        "--results-dir",
        default=None,
        help="运行结果根目录；默认是 <skill-dir>/../runs",
    )
    parser.add_argument(
        "--keep-workspaces",
        action="store_true",
        help="保留成功任务 workspace；正式评测建议启用",
    )
    parser.add_argument(
        "--no-grade",
        action="store_true",
        help="只执行任务，不打分",
    )
    parser.add_argument(
        "--no-xlsx",
        action="store_true",
        help="不输出 XLSX",
    )
    parser.add_argument(
        "--no-workspace-instruction",
        action="store_true",
        help="不在 prompt 后追加隔离 workspace 提醒",
    )
    parser.add_argument(
        "--preflight",
        action="store_true",
        help="只做本地预检，不调用模型",
    )
    parser.add_argument(
        "--no-judge-cache",
        action="store_true",
        help="禁用 PinchBench Judge 缓存",
    )
    parser.add_argument(
        "--clear-judge-cache",
        action="store_true",
        help="运行前清空 PinchBench Judge 缓存",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="打印更详细日志",
    )
    return parser.parse_args()


def main() -> int:
    args = build_args()
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-8s %(message)s",
        datefmt="%H:%M:%S",
    )

    skill_dir = Path(args.skill_dir).expanduser().resolve()
    tasks_dir = (
        Path(args.tasks_dir).expanduser().resolve()
        if args.tasks_dir
        else skill_dir / "tasks"
    )

    if not skill_dir.exists():
        raise SystemExit(f"找不到 PinchBench 仓库目录: {skill_dir}")
    if not tasks_dir.exists():
        raise SystemExit(
            f"找不到 tasks 目录: {tasks_dir}\n"
            "请确认 --skill-dir 指向本地 pinchbench/skill 仓库。"
        )
    if not (tasks_dir / "manifest.yaml").exists():
        LOGGER.warning("tasks/manifest.yaml 不存在，将按 task_*.md 文件回退加载")

    opencode_command = choose_opencode_command(args.opencode)
    grader_module, grader_error = load_pinchbench_grading(skill_dir)
    judge_key_present = bool(os.environ.get("OPENROUTER_API_KEY"))

    all_tasks, core_tasks = load_tasks(tasks_dir)
    selected = filter_tasks(
        all_tasks,
        core_tasks,
        args.suite,
        args.limit,
    )

    skip_ids = set(DEFAULT_SKIPPED_TASKS)
    skip_ids.update(
        item.strip()
        for item in args.skip.split(",")
        if item.strip()
    )
    if args.skip_network:
        skip_ids.update(
            task.task_id for task in selected if is_network_task(task)
        )
    selected = [
        task for task in selected
        if task.task_id not in skip_ids
    ]

    if not selected:
        raise SystemExit("没有可运行任务。请检查 --suite/--limit/--skip。")

    needs_judge = (
        not args.no_grade
        and any(
            task.grading_type in {"llm_judge", "hybrid"}
            for task in selected
        )
    )

    prerequisite_failures, fixture_failures = print_preflight(
        selected,
        skill_dir,
        tasks_dir,
        opencode_command,
        grader_module,
        grader_error,
        args,
    )

    if args.preflight:
        if fixture_failures:
            return 2
        if prerequisite_failures:
            return 3
        if grader_error:
            return 4
        if needs_judge and not judge_key_present:
            return 5
        return 0

    if grader_error and not args.no_grade:
        raise SystemExit(grader_error)
    if needs_judge and not judge_key_present:
        raise SystemExit(
            "所选任务包含 hybrid/llm_judge，但当前 PowerShell 未设置 "
            "OPENROUTER_API_KEY。为避免跑完后无法评分，已在正式执行前停止。"
        )

    results_root = (
        Path(args.results_dir).expanduser().resolve()
        if args.results_dir
        else skill_dir.parent / "runs"
    )
    run_id = dt.datetime.now().strftime("opencode_kimi_k3_%Y%m%d_%H%M%S")
    results_dir = results_root / run_id
    workspaces_dir = results_dir / "workspaces"
    transcripts_dir = results_dir / "transcripts"

    results_dir.mkdir(parents=True, exist_ok=True)
    workspaces_dir.mkdir(parents=True, exist_ok=True)
    transcripts_dir.mkdir(parents=True, exist_ok=True)
    progress_path = results_dir / "progress.jsonl"
    partial_json_path = results_dir / "results.partial.json"

    judge_model = str(args.judge_model)

    write_run_config(
        results_dir / "run_config.json",
        args,
        skill_dir,
        tasks_dir,
        opencode_command,
        selected,
        judge_model,
    )

    judge_cache_dir = results_root / ".judge_cache"
    if (
        grader_module is not None
        and hasattr(grader_module, "set_judge_cache_dir")
        and not args.no_judge_cache
    ):
        grader_module.set_judge_cache_dir(judge_cache_dir)
        if args.clear_judge_cache and hasattr(grader_module, "clear_judge_cache"):
            grader_module.clear_judge_cache()

    print()
    print(f"Results dir          : {results_dir}")
    print("Grading engine       : PinchBench scripts/lib_grading.py")
    print("Judge backend        : api")
    print(f"Judge model          : {judge_model}")
    print(f"Judge key            : {'set' if judge_key_present else 'not required'}")
    print(f"Judge cache          : {'disabled' if args.no_judge_cache else judge_cache_dir}")
    print()

    results: list[dict[str, Any]] = []
    total_start = time.monotonic()

    for index, task in enumerate(selected, start=1):
        task_end_to_end_start = time.monotonic()
        workspace = workspaces_dir / task.task_id
        transcript_dir = transcripts_dir / task.task_id

        execution = execute_task(
            task,
            index,
            len(selected),
            skill_dir,
            workspace,
            transcript_dir,
            opencode_command,
            args,
        )

        grade_result = GradeResult(
            score=None,
            grading_type=task.grading_type,
            breakdown={},
        )

        grading_attempted = (
            execution["status"] not in {"missing_prerequisite", "missing_fixture"}
            and not args.no_grade
        )
        grading_start = time.monotonic()
        if grading_attempted:
            grading_execution = dict(execution)
            grading_execution["transcript"] = execution.get(
                "transcript_data",
                [],
            )
            grade_result = grade_with_pinchbench_default(
                grader_module=grader_module,
                task=task,
                execution_result=grading_execution,
                workspace=workspace,
                skill_dir=skill_dir,
                judge_timeout=args.judge_timeout,
                judge_model=judge_model,
                verbose=args.verbose,
            )
        grading_elapsed = time.monotonic() - grading_start if grading_attempted else 0.0
        execution["model"] = args.model
        execution["agent_elapsed"] = float(execution.get("elapsed") or 0.0)
        execution["grading_elapsed"] = grading_elapsed
        execution["end_to_end_elapsed"] = time.monotonic() - task_end_to_end_start

        execution["score"] = (
            round(float(grade_result.score), 4)
            if grade_result.score is not None
            else None
        )
        execution["breakdown"] = grade_result.breakdown
        execution["grade_notes"] = grade_result.notes
        execution["grade_error"] = grade_result.error

        # The in-memory normalized transcript is already saved to disk.
        execution.pop("transcript_data", None)
        results.append(execution)

        with progress_path.open("a", encoding="utf-8") as progress_file:
            progress_file.write(json.dumps(execution, ensure_ascii=False) + "\n")
            progress_file.flush()
        partial_json_path.write_text(
            json.dumps({"completed": len(results), "results": results}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        marker = "✓" if execution["success"] else "✗"
        score_text = (
            f"score={execution['score']:.3f}"
            if execution["score"] is not None
            else "score=N/A"
        )
        ttft_text = (
            f"ttft={execution['ttft']:.2f}s"
            if execution["ttft"] is not None
            else "ttft=N/A"
        )
        token_text = (
            f"{execution['input_tokens']}+{execution['output_tokens']}"
            if execution["usage_complete"]
            else "N/A"
        )

        total_token_text = (
            str(execution.get("total_tokens"))
            if execution.get("total_tokens") is not None
            else "N/A"
        )
        print(
            f"  {marker} agent={execution['agent_elapsed']:.1f}s "
            f"grade={execution['grading_elapsed']:.1f}s "
            f"e2e={execution['end_to_end_elapsed']:.1f}s "
            f"{ttft_text} {score_text} tokens(in+out)={token_text} total={total_token_text}"
        )
        if execution.get("error"):
            print(f"  运行错误: {str(execution['error'])[:500]}")
        if execution.get("grade_error"):
            print(f"  打分错误: {str(execution['grade_error'])[:500]}")
        if execution.get("output"):
            preview = re.sub(r"\s+", " ", str(execution["output"]))[:260]
            print(
                f"  输出预览: {preview}"
                f"{'...' if len(str(execution['output'])) > 260 else ''}"
            )
        print()

        if not args.keep_workspaces:
            if execution["success"] and not execution.get("grade_error"):
                shutil.rmtree(workspace, ignore_errors=True)
                execution["workspace"] = (
                    "已清理；使用 --keep-workspaces 可保留"
                )

    total_elapsed = time.monotonic() - total_start
    summary = print_summary(results, total_elapsed)

    payload = {
        "summary": summary,
        "results": results,
    }

    json_path = results_dir / "results.json"
    json_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    csv_path = results_dir / "results.csv"
    save_csv(results, csv_path)

    xlsx_path = results_dir / "results.xlsx"
    if not args.no_xlsx:
        save_xlsx(results, summary, xlsx_path)

    if (
        grader_module is not None
        and hasattr(grader_module, "get_judge_cache_stats")
        and not args.no_judge_cache
    ):
        try:
            stats = grader_module.get_judge_cache_stats()
            print(
                "Judge cache stats   : "
                f"entries={stats.get('entries', 0)} "
                f"hits={stats.get('hits', 0)} "
                f"misses={stats.get('misses', 0)}"
            )
        except Exception as exc:
            LOGGER.warning("读取 Judge cache stats 失败: %s", exc)

    print("\n结果文件:")
    print(f"  Config : {results_dir / 'run_config.json'}")
    print(f"  JSON   : {json_path}")
    print(f"  CSV    : {csv_path}")
    if not args.no_xlsx:
        print(f"  XLSX   : {xlsx_path}")
    print(f"  Logs   : {transcripts_dir}")
    print(f"  Progress JSONL : {progress_path}")
    print(f"  Partial JSON   : {partial_json_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
