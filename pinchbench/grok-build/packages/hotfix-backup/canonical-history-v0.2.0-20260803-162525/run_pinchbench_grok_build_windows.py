#!/usr/bin/env python3
# Runner revision: 2026-08-03-pinchbench-grok-build-windows-v1
r"""
Run the local PinchBench task set with Grok Build on native Windows.

Target evaluation chain:
    PinchBench tasks -> Grok Build built-in agent/tools -> local thin Responses
    compatibility adapter -> OpenRouter -> deepseek/deepseek-v4-pro

Comparability contract:
- Same checked-out PinchBench manifest order, fixtures, four integration exclusions and commit.
- Same serial execution, task timeouts, workspace staging, official grading code and Judge model.
- Prompt bytes travel through Grok Build --prompt-file; benchmark prompts are never put in argv.
- Grok Build output uses streaming-messages-json, retained raw as UTF-8 JSONL and normalized for grading.
- Fresh session per normal task; explicit --resume by returned session ID for multi-session turns.
- Model mismatch is based on non-zero model usage/assistant evidence, avoiding zero-token false positives.
- OpenRouter Judge compatibility retries 2048/4096/8192/16384 output budgets and saves raw responses.
- If a hybrid Judge fails, the official automated component is retained in the result breakdown.
- Run directories are append-only; --resume-run continues only tasks missing from progress.jsonl.

Recommended location:
    C:\pinchbench-grok-build\runner\run_pinchbench_grok_build_windows.py
"""
from __future__ import annotations

import argparse
import csv
import dataclasses
import datetime as dt
import hashlib
import http.client as http_client
import json
import logging
import os
import platform
import queue
import re
import shutil
import signal
import subprocess
import sys
import threading
import time
import traceback
import urllib.error
import urllib.parse
import urllib.request
import tomllib
from collections import Counter
from pathlib import Path
from typing import Any, Optional

os.environ.setdefault("PYTHONUTF8", "1")
os.environ.setdefault("NO_COLOR", "1")
os.environ.setdefault("FORCE_COLOR", "0")
os.environ.setdefault("GROK_BUILD_CLI_TRUST_WORKSPACE", "true")

try:
    import yaml  # type: ignore
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "缺少 PyYAML。请使用项目虚拟环境运行：\n"
        r"C:\pinchbench-grok-build\.venv\Scripts\python.exe -m pip install pyyaml"
    ) from exc

LOGGER = logging.getLogger("pinchbench-grok-build")

RUNNER_REVISION = "2026-08-03-grok-build-windows-v1.0.2-search-adapter-0.1.4-reasoning-summary"
DEFAULT_MODEL_ID = "deepseek/deepseek-v4-pro"
DEFAULT_MODEL_ALIAS = "deepseek-v4-pro-openrouter"
DEFAULT_GROK_BUILD_VERSION = "0.2.118"
DEFAULT_GROK_BUILD_BINARY_SHA256 = "8b365d13ba0956bd8015069a7230370dd11496cd18d03b5eb148a329a8d96f7c"
DEFAULT_ADAPTER_VERSION = "0.1.4"
DEFAULT_ADAPTER_URL = "http://127.0.0.1:8767"
DEFAULT_APPROVAL_MODE = "yolo"
DEFAULT_PINCHBENCH_COMMIT = "819384ae830492365b8363fc26bc2602e73f216d"
DEFAULT_MANIFEST_SHA256 = "38d7cd1bddfa5e9fefc7b6945c91955f36dc5c88c32c994bf8676344b1069a7b"
PINCHBENCH_DEFAULT_JUDGE_MODEL = os.environ.get(
    "PINCHBENCH_JUDGE_MODEL",
    "openrouter/anthropic/claude-opus-5",
)

# Agreed exclusions for this independent project. They require external
# GitHub/Google Workspace integration tooling and credentials that are not
# part of the current Windows OpenCode test environment.
DEFAULT_SKIPPED_TASKS = {
    "task_gh_issue_triage",
    "task_gws_email_triage",
    "task_gws_cross_service",
    "task_gws_task_management",
}

# Explicitly known internet-dependent tasks from the historical runner.
# Newer task files can also mark network_required: true.
NETWORK_TASKS = {
    "task_earnings_analysis",
    "task_stock",
    "task_market_research",
    "task_polymarket_briefing",
    "task_executive_lookup",
    "task_events",
    "task_deep_research",
    "task_competitive_research",
    "task_oss_alternative_research",
    "task_pricing_research",
    "task_it_procurement",
    "task_eu_regulation_research",
    "task_byok_best_practices",
}

TEXT_FILE_SUFFIXES = {
    ".txt", ".md", ".py", ".js", ".ts", ".tsx", ".jsx", ".json", ".yaml", ".yml",
    ".csv", ".tsv", ".html", ".css", ".xml", ".toml", ".ini", ".cfg", ".log",
    ".ics", ".sh", ".ps1", ".bat", ".cmd", ".sql", ".java", ".go", ".rs", ".rb", ".php",
}

SKIP_WORKSPACE_NAMES = {
    "BOOTSTRAP.md", "SOUL.md", "USER.md", "IDENTITY.md", "HEARTBEAT.md",
    "TOOLS.md", "AGENTS.md", "CLAUDE.md", "GEMINI.md", "GROK.md",
}
SKIP_WORKSPACE_DIRS = {
    ".git", ".openclaw", ".opencode", ".gemini", ".grok", ".claude", "__pycache__", "node_modules", "skills",
}


@dataclasses.dataclass
class Task:
    task_id: str
    name: str
    category: str
    grading_type: str
    timeout_seconds: int
    workspace_files: list[dict[str, Any]]
    prompt: str
    expected_behavior: str
    grading_criteria: list[str]
    automated_checks: str
    llm_judge_rubric: str
    grading_weights: dict[str, float]
    metadata: dict[str, Any]
    file_path: Path
    multi_session: bool = False
    sessions: list[dict[str, Any]] = dataclasses.field(default_factory=list)


@dataclasses.dataclass
class GradeResult:
    score: Optional[float]
    grading_type: str
    breakdown: dict[str, float]
    notes: str = ""
    error: Optional[str] = None


# ---------------------------------------------------------------------------
# Task loading
# ---------------------------------------------------------------------------

def parse_sections(body: str) -> dict[str, str]:
    sections: dict[str, list[str]] = {}
    current: Optional[str] = None
    for line in body.splitlines():
        match = re.match(r"^##\s+(.+?)\s*$", line)
        if match:
            current = match.group(1).strip()
            sections.setdefault(current, [])
        elif current:
            sections[current].append(line)
    return {key: "\n".join(value).strip() for key, value in sections.items()}


def extract_grading_criteria(text: str) -> list[str]:
    criteria: list[str] = []
    for line in text.splitlines():
        match = re.match(r"^-\s+\[[ xX]\]\s+(.+)$", line.strip())
        if match:
            criteria.append(match.group(1).strip())
    return criteria


def load_task_file(path: Path, category_override: Optional[str] = None) -> Task:
    raw = path.read_text(encoding="utf-8")
    match = re.match(r"^---\s*\r?\n(.*?)\r?\n---\s*\r?\n?(.*)$", raw, re.DOTALL)
    if not match:
        raise ValueError(f"{path} 没有 YAML frontmatter")

    metadata = yaml.safe_load(match.group(1)) or {}
    if not isinstance(metadata, dict):
        raise ValueError(f"{path} 的 YAML frontmatter 不是对象")

    body = match.group(2)
    sections = parse_sections(body)
    task_id = str(metadata.get("id") or path.stem)

    weights = metadata.get("grading_weights") or {}
    if not isinstance(weights, dict):
        weights = {}
    normalized_weights: dict[str, float] = {}
    for key, value in weights.items():
        try:
            normalized_weights[str(key)] = float(value)
        except (TypeError, ValueError):
            pass

    raw_sessions = metadata.get("sessions") or []
    sessions = [item for item in raw_sessions if isinstance(item, dict)] if isinstance(raw_sessions, list) else []

    return Task(
        task_id=task_id,
        name=str(metadata.get("name") or task_id),
        category=str(category_override or metadata.get("category") or ""),
        grading_type=str(metadata.get("grading_type") or "automated"),
        timeout_seconds=int(metadata.get("timeout_seconds") or 120),
        workspace_files=list(metadata.get("workspace_files") or []),
        prompt=sections.get("Prompt", "").strip(),
        expected_behavior=sections.get("Expected Behavior", "").strip(),
        grading_criteria=extract_grading_criteria(sections.get("Grading Criteria", "")),
        automated_checks=sections.get("Automated Checks", "") or "",
        llm_judge_rubric=sections.get("LLM Judge Rubric", "") or "",
        grading_weights=normalized_weights,
        metadata=metadata,
        file_path=path,
        multi_session=bool(metadata.get("multi_session") or sessions),
        sessions=sessions,
    )


def parse_manifest(tasks_dir: Path) -> tuple[list[str], dict[str, str], list[str]]:
    manifest_path = tasks_dir / "manifest.yaml"
    if not manifest_path.exists():
        ordered = [
            p.stem for p in sorted(tasks_dir.glob("task_*.md"))
            if p.stem != "task_XX_name"
        ]
        return ordered, {}, []

    manifest = yaml.safe_load(manifest_path.read_text(encoding="utf-8")) or {}
    if not isinstance(manifest, dict):
        raise ValueError(f"{manifest_path} 不是有效 YAML 对象")

    category_map: dict[str, str] = {}
    core_tasks = [str(x) for x in (manifest.get("core") or [])]

    if "categories" in manifest:
        run_first = [str(x) for x in (manifest.get("run_first") or [])]
        categories = manifest.get("categories") or {}
        all_task_ids: list[str] = []
        if isinstance(categories, dict):
            for category, ids in categories.items():
                for task_id in ids or []:
                    task_id = str(task_id)
                    category_map[task_id] = str(category)
                    all_task_ids.append(task_id)

        seen: set[str] = set()
        ordered: list[str] = []
        for task_id in run_first + all_task_ids:
            if task_id not in seen:
                seen.add(task_id)
                ordered.append(task_id)
        return ordered, category_map, core_tasks

    ordered = [str(x) for x in (manifest.get("tasks") or [])]
    return ordered, category_map, core_tasks


def load_tasks(tasks_dir: Path) -> tuple[list[Task], list[str]]:
    """Load exactly the checked-out manifest when it exists.

    PinchBench ships template/example Markdown files alongside real tasks.
    Falling back to every task_*.md while a manifest exists can accidentally
    count a template as a benchmark task.  The manifest is therefore the
    source of truth for normal runs.
    """
    manifest_exists = (tasks_dir / "manifest.yaml").exists()
    ordered_ids, category_map, core_tasks = parse_manifest(tasks_dir)
    tasks: list[Task] = []

    for task_id in ordered_ids:
        if task_id == "task_XX_name":
            continue
        task_path = tasks_dir / f"{task_id}.md"
        if not task_path.exists():
            LOGGER.warning("manifest 引用了不存在的任务文件: %s", task_path)
            continue
        try:
            task = load_task_file(
                task_path,
                category_override=category_map.get(task_id),
            )
            if (
                task.task_id == "task_XX_name"
                or task.category.strip().lower() == "category_name"
            ):
                LOGGER.debug("忽略 PinchBench 模板任务: %s", task_path)
                continue
            tasks.append(task)
        except Exception as exc:
            LOGGER.warning("加载任务失败 %s: %s", task_path, exc)

    if manifest_exists:
        return tasks, core_tasks

    # Only use file discovery when no manifest is available.
    loaded_ids = {task.task_id for task in tasks}
    for task_path in sorted(tasks_dir.glob("task_*.md")):
        if task_path.stem in loaded_ids or task_path.stem == "task_XX_name":
            continue
        try:
            task = load_task_file(task_path)
            if (
                task.task_id == "task_XX_name"
                or task.category.strip().lower() == "category_name"
            ):
                continue
            tasks.append(task)
        except Exception as exc:
            LOGGER.warning("加载任务失败 %s: %s", task_path, exc)

    return tasks, core_tasks


def filter_tasks(
    tasks: list[Task],
    core_tasks: list[str],
    suite: str,
    limit: Optional[int],
) -> list[Task]:
    suite = suite.strip()
    task_by_id = {task.task_id: task for task in tasks}

    if suite == "all":
        selected = tasks
    elif suite == "core":
        core_set = set(core_tasks)
        selected = [task for task in tasks if task.task_id in core_set] if core_set else tasks[:25]
    elif suite == "automated-only":
        selected = [task for task in tasks if task.grading_type == "automated"]
    elif suite == "llm-judge-only":
        selected = [task for task in tasks if task.grading_type == "llm_judge"]
    elif suite == "hybrid-only":
        selected = [task for task in tasks if task.grading_type == "hybrid"]
    elif suite == "judge-required-only":
        selected = [task for task in tasks if task.grading_type in {"llm_judge", "hybrid"}]
    else:
        ids = [item.strip() for item in suite.split(",") if item.strip()]
        missing = [task_id for task_id in ids if task_id not in task_by_id]
        if missing:
            raise SystemExit(f"suite 中有未知任务: {', '.join(missing)}")
        selected = [task_by_id[task_id] for task_id in ids]

    if limit is not None:
        selected = selected[:limit]
    return selected


def is_network_task(task: Task) -> bool:
    return bool(
        task.task_id in NETWORK_TASKS
        or task.metadata.get("network_required") is True
        or task.metadata.get("requires_network") is True
    )


# ---------------------------------------------------------------------------
# Workspace and prerequisites
# ---------------------------------------------------------------------------

def safe_workspace_dest(dest: str) -> Path:
    rel = Path(dest)
    if rel.is_absolute() or ".." in rel.parts:
        raise ValueError(f"非法 workspace 文件路径: {dest}")
    return rel


def stage_workspace_files(
    task: Task,
    workspace: Path,
    skill_dir: Path,
) -> tuple[list[str], list[str]]:
    staged: list[str] = []
    missing: list[str] = []

    for wf in task.workspace_files:
        if not isinstance(wf, dict):
            missing.append(f"无法识别的 workspace_files 项: {wf!r}")
            continue

        dest_value = wf.get("dest") or wf.get("path") or wf.get("name")
        source_value = wf.get("source")
        content_value = wf.get("content")

        if not dest_value and source_value:
            dest_value = Path(str(source_value)).name
        if not dest_value:
            missing.append(f"workspace_files 项缺少 dest/path: {wf!r}")
            continue

        try:
            dest_rel = safe_workspace_dest(str(dest_value))
        except ValueError as exc:
            missing.append(str(exc))
            continue

        dest_path = workspace / dest_rel
        dest_path.parent.mkdir(parents=True, exist_ok=True)

        if content_value is not None:
            dest_path.write_text(str(content_value), encoding="utf-8")
            staged.append(str(dest_rel))
            continue

        if not source_value:
            missing.append(f"workspace_files 项缺少 source/content: {wf!r}")
            continue

        src_rel = Path(str(source_value))
        candidates = [
            skill_dir / "assets" / src_rel,
            skill_dir / src_rel,
            task.file_path.parent / src_rel,
            task.file_path.parent.parent / "assets" / src_rel,
        ]
        src_path = next((candidate for candidate in candidates if candidate.exists()), None)
        if src_path is None:
            missing.append(
                f"找不到预置文件 {source_value}; 尝试过: "
                + ", ".join(str(candidate) for candidate in candidates)
            )
            continue

        if src_path.is_dir():
            if dest_path.exists():
                shutil.rmtree(dest_path)
            shutil.copytree(src_path, dest_path)
        else:
            shutil.copy2(src_path, dest_path)
        staged.append(str(dest_rel))

    return staged, missing


def resolve_command(name: str) -> Optional[str]:
    candidates = [name]
    if sys.platform == "win32" and not Path(name).suffix:
        candidates = [f"{name}.cmd", f"{name}.exe", name]
    for candidate in candidates:
        resolved = shutil.which(candidate)
        if resolved:
            return resolved
    return None


def check_prerequisites(prereqs: list[str]) -> list[str]:
    missing: list[str] = []
    npm_command = resolve_command("npm") or "npm"

    for req_raw in prereqs:
        req = str(req_raw)
        if req.startswith("cli:"):
            command = req.split(":", 1)[1].strip()
            if not resolve_command(command):
                missing.append(req)
        elif req.startswith("npm:"):
            package = req.split(":", 1)[1].strip()
            try:
                result = subprocess.run(
                    [npm_command, "list", "-g", package, "--depth=0"],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=30,
                    check=False,
                )
                if result.returncode != 0:
                    missing.append(req)
            except Exception:
                missing.append(req)

    return missing


def choose_grok_build_command(preferred: str, inferred_root: Optional[Path] = None) -> str:
    if preferred and preferred != "auto":
        preferred_path = Path(preferred).expanduser()
        if preferred_path.exists():
            return str(preferred_path.resolve())
        resolved = resolve_command(preferred)
        if resolved:
            return resolved
        raise SystemExit(f"找不到 Grok Build 命令 {preferred!r}")

    candidates: list[Path] = []
    if inferred_root is not None:
        candidates.append(inferred_root / "bin" / "grok.exe")
    candidates.append(Path(r"C:\pinchbench-grok-build\bin\grok.exe"))
    for candidate in candidates:
        if candidate.exists():
            return str(candidate.resolve())

    for command in (("grok.exe", "grok") if sys.platform == "win32" else ("grok",)):
        resolved = shutil.which(command)
        if resolved:
            return resolved

    raise SystemExit(
        "找不到 Grok Build。预期路径: "
        r"C:\pinchbench-grok-build\bin\grok.exe"
    )


def kill_proc_tree(proc: subprocess.Popen[str]) -> None:
    if proc.poll() is not None:
        return
    try:
        if sys.platform == "win32":
            subprocess.run(
                ["taskkill", "/F", "/T", "/PID", str(proc.pid)],
                capture_output=True,
                check=False,
            )
        else:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
    except Exception:
        try:
            proc.kill()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Grok Build CLI event handling
# ---------------------------------------------------------------------------

CLIENT_VISIBLE_ROLES = {"assistant"}


def _to_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _to_float(value: Any) -> Optional[float]:
    try:
        if value is None:
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _stringify_tool_content(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        chunks: list[str] = []
        for item in value:
            if isinstance(item, str):
                chunks.append(item)
            elif isinstance(item, dict):
                text = item.get("text")
                chunks.append(str(text) if isinstance(text, str) else json.dumps(item, ensure_ascii=False))
            else:
                chunks.append(str(item))
        return "\n".join(chunk for chunk in chunks if chunk)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def make_user_transcript_event(prompt: str, turn_index: int, session_label: str) -> dict[str, Any]:
    return {
        "type": "message",
        "message": {"role": "user", "content": [{"type": "text", "text": prompt}]},
        "_adapter": {
            "source": "pinchbench-grok-build-runner",
            "turn": turn_index,
            "session_label": session_label,
        },
    }


def normalize_grok_build_event(event: dict[str, Any]) -> list[dict[str, Any]]:
    """Convert Grok streaming-messages-json frames to PinchBench transcript shape."""
    event_type = str(event.get("type") or "")
    normalized: list[dict[str, Any]] = []

    if event_type == "assistant":
        message = event.get("message") if isinstance(event.get("message"), dict) else {}
        content = message.get("content") if isinstance(message.get("content"), list) else []
        assistant_blocks: list[dict[str, Any]] = []
        tool_results: list[dict[str, Any]] = []
        for block in content:
            if not isinstance(block, dict):
                continue
            block_type = str(block.get("type") or "")
            if block_type == "text":
                text = block.get("text")
                if isinstance(text, str) and text:
                    assistant_blocks.append({"type": "text", "text": text})
            elif block_type in {"tool_use", "server_tool_use"}:
                tool_input = block.get("input") if isinstance(block.get("input"), dict) else {}
                assistant_blocks.append({
                    "type": "toolCall",
                    "id": str(block.get("id") or ""),
                    "name": str(block.get("name") or ""),
                    "arguments": tool_input,
                    "input": tool_input,
                })
            elif block_type == "web_search_tool_result":
                content_value = block.get("content")
                is_error = isinstance(content_value, dict) and str(content_value.get("type") or "").endswith("_error")
                tool_results.append({
                    "type": "message",
                    "message": {
                        "role": "toolResult",
                        "content": [{
                            "type": "toolResult",
                            "toolCallId": str(block.get("tool_use_id") or ""),
                            "name": "web_search",
                            "content": _stringify_tool_content(content_value),
                            "isError": is_error,
                        }],
                    },
                    "_adapter": {"source": "grok-build", "event_type": event_type},
                })
        if assistant_blocks:
            normalized.append({
                "type": "message",
                "message": {"role": "assistant", "content": assistant_blocks},
                "_adapter": {"source": "grok-build", "event_type": event_type},
            })
        normalized.extend(tool_results)
        return normalized

    if event_type == "user":
        message = event.get("message") if isinstance(event.get("message"), dict) else {}
        content = message.get("content") if isinstance(message.get("content"), list) else []
        for block in content:
            if not isinstance(block, dict) or block.get("type") != "tool_result":
                continue
            normalized.append({
                "type": "message",
                "message": {
                    "role": "toolResult",
                    "content": [{
                        "type": "toolResult",
                        "toolCallId": str(block.get("tool_use_id") or ""),
                        "name": str(block.get("name") or ""),
                        "content": _stringify_tool_content(block.get("content")),
                        "isError": bool(block.get("is_error")),
                    }],
                },
                "_adapter": {"source": "grok-build", "event_type": event_type},
            })
        return normalized

    if event_type == "result" and (bool(event.get("is_error")) or str(event.get("subtype") or "").startswith("error")):
        errors = event.get("errors")
        error_text = _stringify_tool_content(errors) or str(event.get("result") or "Grok Build error")
        normalized.append({
            "type": "message",
            "message": {"role": "assistant", "content": [{"type": "text", "text": f"[Grok Build error] {error_text}"}]},
            "_adapter": {"source": "grok-build", "event_type": event_type},
        })
    return normalized


def _stream_reader(stream: Any, kind: str, output_queue: "queue.Queue[tuple[str, Optional[str]]]") -> None:
    try:
        for line in iter(stream.readline, ""):
            output_queue.put((kind, line))
    finally:
        output_queue.put((kind, None))
        try:
            stream.close()
        except Exception:
            pass


def merge_no_proxy(existing: str, required: list[str]) -> str:
    parts = [item.strip() for item in existing.split(",") if item.strip()]
    seen = {item.lower() for item in parts}
    for item in required:
        if item.lower() not in seen:
            parts.append(item)
            seen.add(item.lower())
    return ",".join(parts)


def run_grok_build_streaming(
    cmd: list[str],
    cwd: Path,
    timeout: float,
    raw_stdout_path: Path,
    stderr_path: Path,
    prompt_path: Path,
    requested_alias: str,
    expected_model_id: str,
    grok_build_home: Path,
) -> dict[str, Any]:
    """Run one Grok Build headless turn and parse streaming-messages-json."""
    monotonic_start = time.monotonic()
    deadline = monotonic_start + timeout
    raw_stdout_path.parent.mkdir(parents=True, exist_ok=True)
    stderr_path.parent.mkdir(parents=True, exist_ok=True)

    stdout_lines: list[str] = []
    stderr_lines: list[str] = []
    raw_events: list[dict[str, Any]] = []
    normalized_events: list[dict[str, Any]] = []
    assistant_texts: list[str] = []
    event_counts: Counter[str] = Counter()
    finish_reasons: list[str] = []
    error_messages: list[str] = []
    available_tools: set[str] = set()
    observed_models: set[str] = set()
    assistant_models: set[str] = set()
    positive_usage_models: set[str] = set()
    model_usage: dict[str, Any] = {}
    permission_modes: set[str] = set()
    permission_denials: list[Any] = []

    session_id: Optional[str] = None
    request_id: Optional[str] = None
    ttft: Optional[float] = None
    status = "success"
    timed_out = False
    result_seen = False
    result_success = False
    final_output = ""
    usage_seen = False
    input_tokens = output_tokens = cache_read_tokens = cache_write_tokens = total_tokens = 0
    step_count = tool_errors = tool_call_count = 0
    cost_usd: Optional[float] = None

    env = os.environ.copy()
    env["PYTHONUTF8"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    env["NO_COLOR"] = "1"
    env["FORCE_COLOR"] = "0"
    env["RUST_LOG"] = "error"
    env["LANG"] = "en_US.UTF-8"
    env["LC_ALL"] = "en_US.UTF-8"
    env["LC_MESSAGES"] = "en_US.UTF-8"
    env["LANGUAGE"] = "en_US:en"
    env["GROK_HOME"] = str(grok_build_home)
    no_proxy = merge_no_proxy(str(env.get("NO_PROXY") or env.get("no_proxy") or ""), ["localhost", "127.0.0.1", "::1"])
    env["NO_PROXY"] = no_proxy
    env["no_proxy"] = no_proxy

    creationflags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0) if sys.platform == "win32" else 0
    try:
        proc = subprocess.Popen(
            cmd,
            cwd=str(cwd),
            stdin=subprocess.DEVNULL,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
            errors="replace",
            bufsize=1,
            env=env,
            start_new_session=(sys.platform != "win32"),
            creationflags=creationflags,
        )
    except Exception as exc:
        return {
            "status": "error", "success": False, "returncode": None,
            "output": "", "error": f"无法启动 Grok Build: {exc}", "stderr": "",
            "elapsed": time.monotonic() - monotonic_start, "ttft": None,
            "input_tokens": None, "output_tokens": None, "reasoning_tokens": None,
            "cache_read_tokens": None, "cache_write_tokens": None, "cost_usd": None,
            "total_tokens": None, "usage_complete": False, "step_count": 0,
            "tool_errors": 0, "tool_call_count": 0, "finish_reasons": [],
            "session_id": None, "request_id": None, "raw_events": [],
            "normalized_events": [], "event_counts": {}, "prompt_transport": "prompt_file_utf8",
            "prompt_path": str(prompt_path), "observed_models": [], "positive_usage_models": [],
            "unexpected_models": [], "model_usage": {}, "permission_mode": None,
            "permission_denials": [], "available_tools": [],
        }

    if proc.stdout is None or proc.stderr is None:
        kill_proc_tree(proc)
        raise RuntimeError("Grok Build stdout/stderr pipe creation failed")

    output_queue: "queue.Queue[tuple[str, Optional[str]]]" = queue.Queue()
    threads = [
        threading.Thread(target=_stream_reader, args=(proc.stdout, "stdout", output_queue), daemon=True),
        threading.Thread(target=_stream_reader, args=(proc.stderr, "stderr", output_queue), daemon=True),
    ]
    for thread in threads:
        thread.start()
    open_streams = {"stdout", "stderr"}

    with raw_stdout_path.open("w", encoding="utf-8", newline="") as raw_file, stderr_path.open("w", encoding="utf-8", newline="") as err_file:
        while open_streams or proc.poll() is None:
            if time.monotonic() > deadline:
                timed_out = True
                status = "timeout"
                error_messages.append(f"超时 ({timeout:.0f}s)")
                kill_proc_tree(proc)
                break
            try:
                kind, line = output_queue.get(timeout=0.2)
            except queue.Empty:
                continue
            if line is None:
                open_streams.discard(kind)
                continue
            if kind == "stderr":
                stderr_lines.append(line)
                err_file.write(line); err_file.flush()
                continue
            stdout_lines.append(line)
            raw_file.write(line); raw_file.flush()
            stripped = line.strip()
            if not stripped:
                continue
            try:
                event = json.loads(stripped)
            except json.JSONDecodeError:
                event_counts["non_json_stdout"] += 1
                continue
            if not isinstance(event, dict):
                event_counts["non_object_json"] += 1
                continue
            raw_events.append(event)
            event_type = str(event.get("type") or "unknown")
            event_counts[event_type] += 1
            normalized_events.extend(normalize_grok_build_event(event))

            candidate_session = event.get("session_id") or event.get("sessionId")
            if candidate_session:
                session_id = str(candidate_session)

            if event_type == "system" and str(event.get("subtype") or "") == "init":
                model = str(event.get("model") or "")
                if model and model != "unknown": observed_models.add(model)
                tools = event.get("tools")
                if isinstance(tools, list): available_tools.update(str(x) for x in tools if x)
                permission = str(event.get("permissionMode") or event.get("permission_mode") or "")
                if permission: permission_modes.add(permission)

            elif event_type == "assistant":
                message = event.get("message") if isinstance(event.get("message"), dict) else {}
                model = str(message.get("model") or event.get("model") or "")
                if model and model != "unknown":
                    observed_models.add(model); assistant_models.add(model)
                stop_reason = str(message.get("stop_reason") or "")
                if stop_reason: finish_reasons.append(stop_reason)
                content = message.get("content") if isinstance(message.get("content"), list) else []
                if content and ttft is None: ttft = time.monotonic() - monotonic_start
                for block in content:
                    if not isinstance(block, dict): continue
                    block_type = str(block.get("type") or "")
                    if block_type == "text" and isinstance(block.get("text"), str):
                        assistant_texts.append(str(block.get("text")))
                    elif block_type in {"tool_use", "server_tool_use"}:
                        tool_call_count += 1
                        name = str(block.get("name") or "")
                        if name: available_tools.add(name)
                    elif block_type == "web_search_tool_result":
                        content_value = block.get("content")
                        if isinstance(content_value, dict) and str(content_value.get("type") or "").endswith("_error"):
                            tool_errors += 1

            elif event_type == "user":
                message = event.get("message") if isinstance(event.get("message"), dict) else {}
                content = message.get("content") if isinstance(message.get("content"), list) else []
                for block in content:
                    if not isinstance(block, dict) or block.get("type") != "tool_result": continue
                    if bool(block.get("is_error")):
                        tool_errors += 1
                        text = _stringify_tool_content(block.get("content"))
                        if re.search(r"(?i)permission|denied|not approved|approval", text):
                            permission_denials.append(text)

            elif event_type == "result":
                result_seen = True
                subtype = str(event.get("subtype") or "")
                is_error = bool(event.get("is_error")) or subtype.startswith("error")
                result_success = subtype == "success" and not is_error
                if not result_success: status = "error"
                if isinstance(event.get("result"), str): final_output = str(event.get("result"))
                request = event.get("request_id") or event.get("requestId")
                if request: request_id = str(request)
                stop_reason = str(event.get("stop_reason") or "")
                if stop_reason: finish_reasons.append(stop_reason)
                step_count = _to_int(event.get("num_turns"))
                usage = event.get("usage") if isinstance(event.get("usage"), dict) else {}
                input_tokens = _to_int(usage.get("input_tokens"))
                output_tokens = _to_int(usage.get("output_tokens"))
                cache_read_tokens = _to_int(usage.get("cache_read_input_tokens"))
                cache_write_tokens = _to_int(usage.get("cache_creation_input_tokens"))
                total_tokens = input_tokens + output_tokens + cache_read_tokens + cache_write_tokens
                model_usage = event.get("modelUsage") if isinstance(event.get("modelUsage"), dict) else {}
                for model_name, model_row in model_usage.items():
                    if not model_name: continue
                    observed_models.add(str(model_name))
                    if isinstance(model_row, dict):
                        evidence = sum(_to_int(model_row.get(key)) for key in (
                            "inputTokens", "outputTokens", "cacheReadInputTokens",
                            "cacheCreationInputTokens", "modelCalls"
                        ))
                        if evidence > 0: positive_usage_models.add(str(model_name))
                usage_seen = bool(model_usage) or total_tokens > 0
                cost_value = _to_float(event.get("total_cost_usd"))
                cost_usd = cost_value if cost_value is not None and cost_value > 0 else None
                if is_error:
                    error_messages.append(_stringify_tool_content(event.get("errors")) or final_output or subtype)

            elif event_type == "error":
                status = "error"
                error_messages.append(str(event.get("message") or event.get("error") or "Grok Build error"))

    try:
        proc.wait(timeout=15)
    except subprocess.TimeoutExpired:
        kill_proc_tree(proc)
        status = "timeout"; timed_out = True
        error_messages.append("进程未正常退出，已强制终止")
    for thread in threads: thread.join(timeout=1)

    returncode = proc.returncode
    stderr = "".join(stderr_lines).strip()
    if returncode not in (0, None) and status == "success":
        status = "error"; error_messages.append(stderr or f"Grok Build 退出码 {returncode}")
    if returncode in (0, None) and status == "success" and not result_seen:
        status = "error"; error_messages.append("Grok Build 未输出最终 result 事件")

    accepted_models = {requested_alias, expected_model_id}
    # modelUsage is the authoritative billing/usage evidence. Assistant frame model
    # labels can be friendly display names, so use them only as a fallback when
    # the terminal result provides no positive per-model usage at all.
    evidence_models = set(positive_usage_models) if positive_usage_models else set(assistant_models)
    unexpected_models = sorted(model for model in evidence_models if model and model not in accepted_models and model != "unknown")
    if unexpected_models and status == "success":
        status = "model_mismatch"
        error_messages.append("检测到有实际调用证据的非目标模型: " + ", ".join(unexpected_models))

    success = status == "success" and not timed_out and returncode in (0, None) and result_seen and result_success and not unexpected_models
    output = final_output.strip() if final_output.strip() else "\n".join(x.strip() for x in assistant_texts if x.strip()).strip()
    return {
        "status": status, "success": success, "returncode": returncode,
        "output": output, "error": " | ".join(dict.fromkeys(x for x in error_messages if x)),
        "stderr": stderr, "elapsed": time.monotonic() - monotonic_start, "ttft": ttft,
        "input_tokens": input_tokens if usage_seen else None,
        "output_tokens": output_tokens if usage_seen else None,
        "reasoning_tokens": None,
        "cache_read_tokens": cache_read_tokens if usage_seen else None,
        "cache_write_tokens": cache_write_tokens if usage_seen else None,
        "cost_usd": cost_usd,
        "total_tokens": total_tokens if usage_seen else None,
        "usage_complete": usage_seen,
        "step_count": step_count, "tool_errors": tool_errors,
        "tool_call_count": tool_call_count, "finish_reasons": list(dict.fromkeys(finish_reasons)),
        "session_id": session_id, "request_id": request_id,
        "raw_events": raw_events, "normalized_events": normalized_events,
        "event_counts": dict(event_counts), "stdout_lines": stdout_lines,
        "prompt_transport": "prompt_file_utf8", "prompt_path": str(prompt_path),
        "observed_models": sorted(observed_models),
        "positive_usage_models": sorted(positive_usage_models),
        "unexpected_models": unexpected_models, "model_usage": model_usage,
        "permission_mode": ",".join(sorted(permission_modes)) if permission_modes else "yolo",
        "permission_denials": permission_denials, "available_tools": sorted(available_tools),
    }


# ---------------------------------------------------------------------------
# PinchBench official/default grading
# ---------------------------------------------------------------------------

INVALID_FILENAME = '<>:"/\\|?*'


def utc_now() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat()


def safe_slug(value: str) -> str:
    text = str(value).strip()
    for char in INVALID_FILENAME:
        text = text.replace(char, "_")
    text = "_".join(text.split())
    return text.strip("._") or "item"


def write_json_atomic(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(path.name + f".{os.getpid()}.{threading.get_ident()}.tmp")
    payload = json.dumps(value, ensure_ascii=False, indent=2)
    with temp.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write(payload); handle.flush(); os.fsync(handle.fileno())
    delays = (0.05, 0.10, 0.20, 0.40, 0.80, 1.0, 1.0, 1.0, 1.0, 1.0)
    try:
        for index, delay in enumerate(delays, start=1):
            try:
                os.replace(temp, path); return
            except PermissionError:
                if index == len(delays): raise
                time.sleep(delay)
    finally:
        try:
            if temp.exists(): temp.unlink()
        except OSError:
            pass


def _openrouter_message_text(message: Any) -> str:
    if not isinstance(message, dict): return ""
    content = message.get("content")
    if isinstance(content, str): return content
    if isinstance(content, list):
        pieces: list[str] = []
        for block in content:
            if isinstance(block, str): pieces.append(block)
            elif isinstance(block, dict):
                text = block.get("text")
                nested = block.get("content")
                if isinstance(text, str): pieces.append(text)
                elif isinstance(nested, str): pieces.append(nested)
        return "".join(pieces)
    return ""


def install_openrouter_judge_compatibility(grader: Any, task_id: str, raw_root: Path) -> dict[str, Any]:
    """Patch only the Judge HTTP transport; rubric, weights and grading code remain official."""
    context: dict[str, Any] = {
        "enabled": False, "api_attempts": 0, "empty_content_retries": 0,
        "length_retries": 0, "network_failures": 0, "raw_response_files": [],
        "completion_budgets": [], "last_finish_reason": None, "last_usage": {},
    }
    original = getattr(grader, "_grok_runner_original_call_judge_api", None)
    if not callable(original):
        candidate = getattr(grader, "call_judge_api", None)
        if not callable(candidate): return context
        original = candidate
        setattr(grader, "_grok_runner_original_call_judge_api", original)
    module = sys.modules.get(getattr(original, "__module__", ""))
    system_message = getattr(module, "_JUDGE_SYSTEM_MSG", None)
    if not isinstance(system_message, str) or not system_message: return context
    task_raw_root = raw_root / safe_slug(task_id)
    task_raw_root.mkdir(parents=True, exist_ok=True)
    context["enabled"] = True

    def save_raw(attempt: int, payload: dict[str, Any], status_code: Optional[int], response_data: Any, error_text: str = "") -> None:
        path = task_raw_root / f"attempt_{attempt:02d}.json"
        write_json_atomic(path, {
            "schema_version": 1, "at": utc_now(), "task_id": task_id,
            "attempt": attempt, "model": payload.get("model"),
            "max_completion_tokens": payload.get("max_completion_tokens"),
            "temperature": payload.get("temperature"),
            "prompt_sha256": hashlib.sha256(str(payload["messages"][1]["content"]).encode("utf-8")).hexdigest(),
            "http_status": status_code, "response": response_data, "error": error_text,
        })
        context["raw_response_files"].append(str(path))

    def compatible_call(*, prompt: str, model: str, timeout_seconds: float = 120.0) -> dict[str, Any]:
        if not model.startswith("openrouter/"):
            return original(prompt=prompt, model=model, timeout_seconds=timeout_seconds)
        api_key = os.environ.get("OPENROUTER_API_KEY")
        if not api_key:
            return {"status": "error", "text": "", "error": "OPENROUTER_API_KEY not set"}
        bare_model = model.removeprefix("openrouter/")
        budgets = (2048, 4096, 8192, 16384)
        last_description = "empty content"
        for local_index, budget in enumerate(budgets, start=1):
            context["api_attempts"] += 1
            attempt = int(context["api_attempts"])
            context["completion_budgets"].append(budget)
            payload_obj = {
                "model": bare_model,
                "messages": [{"role": "system", "content": system_message}, {"role": "user", "content": prompt}],
                "temperature": 0.0, "max_completion_tokens": budget,
            }
            request = urllib.request.Request(
                "https://openrouter.ai/api/v1/chat/completions",
                data=json.dumps(payload_obj, ensure_ascii=False).encode("utf-8"),
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json", "HTTP-Referer": "https://pinchbench.com", "X-Title": "PinchBench-Judge"},
                method="POST",
            )
            try:
                with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
                    status_code = getattr(response, "status", 200)
                    response_data = json.loads(response.read().decode("utf-8"))
            except urllib.error.HTTPError as exc:
                body = ""
                try: body = exc.read().decode("utf-8", errors="replace")
                except Exception: pass
                save_raw(attempt, payload_obj, exc.code, None, body[:4000])
                return {"status": "error", "text": "", "error": f"HTTP {exc.code}: {body[:500]}"}
            except (http_client.IncompleteRead, http_client.RemoteDisconnected, ConnectionResetError, BrokenPipeError, urllib.error.URLError, TimeoutError, OSError) as exc:
                context["network_failures"] += 1
                save_raw(attempt, payload_obj, None, None, repr(exc))
                if local_index < len(budgets):
                    time.sleep(min(4, 2 ** (local_index - 1))); continue
                return {"status": "timeout", "text": "", "error": repr(exc)}
            save_raw(attempt, payload_obj, status_code, response_data)
            choices = response_data.get("choices", []) if isinstance(response_data, dict) else []
            if not choices:
                last_description = "No choices in response"
                if local_index < len(budgets): time.sleep(min(4, 2 ** (local_index - 1))); continue
                break
            choice = choices[0] if isinstance(choices[0], dict) else {}
            text = _openrouter_message_text(choice.get("message", {}))
            finish_reason = choice.get("finish_reason")
            usage = response_data.get("usage", {}) if isinstance(response_data, dict) else {}
            context["last_finish_reason"] = finish_reason; context["last_usage"] = usage
            truncated = str(finish_reason or "").lower() in {"length", "max_tokens"}
            if isinstance(text, str) and text.strip() and not truncated:
                return {"status": "success", "text": text}
            if truncated:
                context["length_retries"] += 1
                last_description = f"Judge response truncated; finish_reason={finish_reason!r}; visible_text_length={len(text)}; budget={budget}; usage={usage}"
            else:
                context["empty_content_retries"] += 1
                last_description = f"Judge returned null/empty content; finish_reason={finish_reason!r}; budget={budget}; usage={usage}"
            if local_index < len(budgets): time.sleep(min(4, 2 ** (local_index - 1)))
        raise RuntimeError(last_description + "; retries exhausted; raw responses: " + " | ".join(context["raw_response_files"]))

    grader.call_judge_api = compatible_call
    return context


def load_pinchbench_grading(skill_dir: Path) -> tuple[Optional[Any], Optional[str]]:
    """Load the grading engine from the exact checked-out PinchBench commit."""
    scripts_dir = (skill_dir / "scripts").resolve()
    grading_path = scripts_dir / "lib_grading.py"
    if not grading_path.exists():
        return None, f"找不到 PinchBench grading engine: {grading_path}"

    scripts_value = str(scripts_dir)
    if scripts_value not in sys.path:
        sys.path.insert(0, scripts_value)

    try:
        import lib_grading  # type: ignore
    except Exception as exc:
        return None, (
            "无法导入 PinchBench scripts/lib_grading.py。"
            "请安装本地仓库依赖，例如："
            r"C:\pinchbench-opencode\.venv\Scripts\python.exe -m pip install -e "
            f'"{skill_dir}"。原始错误: {exc}'
        )

    return lib_grading, None


def grade_with_pinchbench_default(
    *, grader_module: Any, task: Task, execution_result: dict[str, Any],
    workspace: Path, skill_dir: Path, judge_timeout: float, judge_model: str,
    verbose: bool, judge_raw_root: Path,
) -> GradeResult:
    """Use official grading logic with a transport-only Judge compatibility shim."""
    grading_execution = dict(execution_result)
    grading_execution["workspace"] = str(workspace)
    judge_context: dict[str, Any] = {}
    if task.grading_type in {"llm_judge", "hybrid"}:
        judge_context = install_openrouter_judge_compatibility(grader_module, task.task_id, judge_raw_root)
    try:
        result = grader_module.grade_task(
            task=task, execution_result=grading_execution, skill_dir=skill_dir,
            judge_model=judge_model, judge_timeout_seconds=judge_timeout,
            judge_backend="api", verbose=verbose,
        )
        raw_score = getattr(result, "score", None)
        max_score = getattr(result, "max_score", 1.0)
        score = None if raw_score is None else max(0.0, min(1.0, float(raw_score) / float(max_score or 1.0)))
        notes = str(getattr(result, "notes", "") or "")
        grade_error: Optional[str] = None
        if task.grading_type in {"llm_judge", "hybrid"}:
            lower = notes.lower()
            markers = ("llm judge failed", "no parseable response", "response parsed but no score", "openrouter_api_key not set", "judge api call failed")
            if any(marker in lower for marker in markers): grade_error = notes or "LLM judge failed"
        if judge_context:
            notes = (notes + " | " if notes else "") + "judge_transport=" + json.dumps(judge_context, ensure_ascii=False)
        return GradeResult(score=score, grading_type=str(getattr(result, "grading_type", task.grading_type)), breakdown=dict(getattr(result, "breakdown", {}) or {}), notes=notes, error=grade_error)
    except Exception as exc:
        breakdown: dict[str, float] = {}
        recovery_note = ""
        if task.grading_type == "hybrid" and hasattr(grader_module, "_grade_automated"):
            try:
                auto = grader_module._grade_automated(task, grading_execution, skill_dir=skill_dir, verbose=verbose)
                breakdown = {f"automated.{k}": float(v) for k, v in dict(getattr(auto, "breakdown", {}) or {}).items()}
                recovery_note = "Hybrid automated component preserved after Judge failure."
            except Exception as auto_exc:
                recovery_note = f"Hybrid automated recovery also failed: {auto_exc}"
        context_text = json.dumps(judge_context, ensure_ascii=False) if judge_context else "{}"
        return GradeResult(
            score=None, grading_type=task.grading_type, breakdown=breakdown,
            notes=recovery_note + (" | judge_transport=" + context_text if judge_context else ""),
            error="PinchBench 官方评分失败: " + f"{exc}\n{traceback.format_exc(limit=8)}",
        )


# ---------------------------------------------------------------------------
# Reporting

# ---------------------------------------------------------------------------

def truncate_excel(value: Any, limit: int = 32000) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    return text if len(text) <= limit else text[:limit] + "...[truncated]"


def save_csv(results: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "task_id", "name", "category", "grading_type", "network_task",
        "multi_session", "session_count", "success", "status", "returncode",
        "score", "elapsed", "ttft", "input_tokens", "output_tokens",
        "reasoning_tokens", "cache_read_tokens", "cache_write_tokens",
        "cost_usd", "total_tokens", "usage_complete", "step_count", "tool_errors",
        "tool_call_count", "observed_models", "positive_usage_models", "unexpected_models", "model_usage",
        "grok_build_cli_versions", "permission_modes", "permission_denials",
        "workspace", "transcript", "error", "grade_error", "grade_notes",
    ]

    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for row in results:
            writer.writerow({
                key: truncate_excel(row.get(key), 10000)
                for key in fields
            })


def save_xlsx(
    results: list[dict[str, Any]],
    summary: dict[str, Any],
    output_path: Path,
) -> None:
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
    except ImportError:
        LOGGER.warning(
            "未安装 openpyxl，跳过 XLSX。可运行: "
            "python -m pip install openpyxl"
        )
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "详细结果"

    headers = [
        "任务ID", "名称", "类别", "打分类型", "联网题", "多轮任务",
        "会话调用数", "成功", "状态", "退出码", "分数", "耗时(s)",
        "TTFT估计(s)", "输入Token", "输出Token", "推理Token",
        "缓存读取Token", "缓存写入Token", "费用USD", "总Token", "Usage完整",
        "Step数", "工具错误数", "工具调用数", "观测模型", "正Usage模型",
        "非目标模型", "模型Usage", "Grok Build版本", "权限模式", "权限拒绝",
        "Workspace", "Transcript", "运行错误",
        "打分错误", "打分备注",
    ]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(results, start=2):
        values = [
            row.get("task_id"),
            row.get("name"),
            row.get("category"),
            row.get("grading_type"),
            "是" if row.get("network_task") else "否",
            "是" if row.get("multi_session") else "否",
            row.get("session_count"),
            "✓" if row.get("success") else "✗",
            row.get("status"),
            row.get("returncode") if row.get("returncode") is not None else "",
            row.get("score") if row.get("score") is not None else "",
            row.get("elapsed"),
            row.get("ttft") if row.get("ttft") is not None else "",
            row.get("input_tokens") if row.get("input_tokens") is not None else "",
            row.get("output_tokens") if row.get("output_tokens") is not None else "",
            row.get("reasoning_tokens") if row.get("reasoning_tokens") is not None else "",
            row.get("cache_read_tokens") if row.get("cache_read_tokens") is not None else "",
            row.get("cache_write_tokens") if row.get("cache_write_tokens") is not None else "",
            row.get("cost_usd") if row.get("cost_usd") is not None else "",
            row.get("total_tokens") if row.get("total_tokens") is not None else "",
            "是" if row.get("usage_complete") else "否",
            row.get("step_count"),
            row.get("tool_errors"),
            row.get("tool_call_count"),
            json.dumps(row.get("observed_models") or [], ensure_ascii=False),
            json.dumps(row.get("positive_usage_models") or [], ensure_ascii=False),
            json.dumps(row.get("unexpected_models") or [], ensure_ascii=False),
            json.dumps(row.get("model_usage") or {}, ensure_ascii=False),
            json.dumps(row.get("grok_build_cli_versions") or [], ensure_ascii=False),
            json.dumps(row.get("permission_modes") or [], ensure_ascii=False),
            json.dumps(row.get("permission_denials") or [], ensure_ascii=False),
            row.get("workspace"),
            row.get("transcript"),
            row.get("error") or "",
            row.get("grade_error") or "",
            row.get("grade_notes") or "",
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(
                row=row_index,
                column=column,
                value=truncate_excel(value),
            )

    for cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in cells)
        worksheet.column_dimensions[cells[0].column_letter].width = min(
            max_length + 3,
            60,
        )

    summary_sheet = workbook.create_sheet("汇总")
    for row_index, (key, value) in enumerate(summary.items(), start=1):
        summary_sheet.cell(row=row_index, column=1, value=key).font = Font(bold=True)
        summary_sheet.cell(row=row_index, column=2, value=value)
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 30

    workbook.save(output_path)


def optional_sum(results: list[dict[str, Any]], field: str) -> Optional[float]:
    values = [row.get(field) for row in results if row.get(field) is not None]
    if not values:
        return None
    return float(sum(float(value) for value in values))


def print_summary(
    results: list[dict[str, Any]],
    total_elapsed: float,
) -> dict[str, Any]:
    succeeded = sum(1 for row in results if row.get("success"))
    failed = len(results) - succeeded
    scores = [
        float(row["score"])
        for row in results
        if row.get("score") is not None
    ]
    ttfts = [
        float(row["ttft"])
        for row in results
        if row.get("ttft") is not None
    ]
    average_elapsed = (
        sum(float(row.get("elapsed") or 0.0) for row in results) / len(results)
        if results else 0.0
    )

    summary = {
        "任务总数": len(results),
        "成功": succeeded,
        "失败": failed,
        "成功率": round(succeeded / len(results), 4) if results else 0.0,
        "有分数任务数": len(scores),
        "打分失败任务数": sum(1 for row in results if row.get("grade_error")),
        "平均分数": round(sum(scores) / len(scores), 4) if scores else None,
        "总耗时(s)": round(total_elapsed, 2),
        "平均耗时/任务(s)": round(average_elapsed, 2),
        "平均TTFT估计(s)": round(sum(ttfts) / len(ttfts), 4) if ttfts else None,
        "总输入Token": int(optional_sum(results, "input_tokens") or 0)
            if any(row.get("input_tokens") is not None for row in results) else None,
        "总输出Token": int(optional_sum(results, "output_tokens") or 0)
            if any(row.get("output_tokens") is not None for row in results) else None,
        "总推理Token": int(optional_sum(results, "reasoning_tokens") or 0)
            if any(row.get("reasoning_tokens") is not None for row in results) else None,
        "总费用USD": round(optional_sum(results, "cost_usd") or 0.0, 6)
            if any(row.get("cost_usd") is not None for row in results) else None,
        "Usage缺失任务数": sum(
            1 for row in results if not row.get("usage_complete")
        ),
        "模型不一致任务数": sum(
            1 for row in results if row.get("unexpected_models")
        ),
        "权限拒绝任务数": sum(
            1 for row in results if row.get("permission_denials")
        ),
        "联网任务数": sum(1 for row in results if row.get("network_task")),
        "多轮任务数": sum(1 for row in results if row.get("multi_session")),
    }

    print("\n" + "=" * 100)
    print("汇总")
    print("=" * 100)
    for key, value in summary.items():
        print(f"{key:<22}: {value}")

    print("-" * 100)
    for row in results:
        marker = "✓" if row.get("success") else "✗"
        score = (
            f"{row['score']:.3f}"
            if row.get("score") is not None
            else "N/A"
        )
        ttft = (
            f"{row['ttft']:.2f}s"
            if row.get("ttft") is not None
            else "N/A"
        )
        error = row.get("error") or row.get("grade_error") or ""
        error_suffix = f"  [{str(error)[:70]}]" if error else ""
        print(
            f"{marker} {row['task_id']:<42} "
            f"score={score:<6} elapsed={row['elapsed']:.1f}s "
            f"ttft={ttft:<8}{error_suffix}"
        )
    print("=" * 100)
    return summary


# ---------------------------------------------------------------------------
# Preflight and run metadata
# ---------------------------------------------------------------------------

def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def command_output(command: list[str], cwd: Optional[Path] = None) -> str:
    try:
        result = subprocess.run(
            command,
            cwd=str(cwd) if cwd else None,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
            check=False,
        )
        return (result.stdout or result.stderr).strip()
    except Exception as exc:
        return f"unavailable: {exc}"


def read_json_file(path: Path) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    try:
        value = json.loads(path.read_text(encoding="utf-8-sig"))
        if not isinstance(value, dict):
            return None, f"JSON 根节点不是对象: {path}"
        return value, None
    except Exception as exc:
        return None, f"无法解析 {path}: {exc}"


def write_run_config(
    output_path: Path, args: argparse.Namespace, skill_dir: Path, tasks_dir: Path,
    grok_build_command: str, selected: list[Task], judge_model: str,
    adapter_health: dict[str, Any], resume_history: Optional[list[dict[str, Any]]] = None,
) -> None:
    manifest_path = tasks_dir / "manifest.yaml"
    git_command = resolve_command("git")
    commit = command_output([git_command, "rev-parse", "HEAD"], cwd=skill_dir) if git_command else ""
    grok_version = command_output([grok_build_command, "--version"])
    config_path = Path(args.grok_build_home).expanduser().resolve() / "config.toml"
    config = {
        "runner_revision": RUNNER_REVISION,
        "created_at": dt.datetime.now(dt.timezone.utc).astimezone().isoformat(),
        "platform": platform.platform(), "python": sys.version,
        "grok_build_command": grok_build_command,
        "grok_build_cli_version": grok_version,
        "expected_grok_build_version": args.expected_grok_build_version,
        "grok_build_binary_sha256": sha256_file(Path(grok_build_command)),
        "expected_grok_build_binary_sha256": args.expected_grok_build_binary_sha256,
        "model": args.model, "model_alias": args.model_alias,
        "agent": "Grok Build built-in default agent/tools",
        "variant": "default; no custom rules or tool allow-list",
        "approval_mode": args.approval_mode, "auto_approval": args.approval_mode == "yolo",
        "sandbox": False, "prompt_transport": "prompt_file_utf8", "prompt_in_argv": False,
        "output_format": "streaming-messages-json", "include_partial_messages": False,
        "ttft_method": "first complete assistant message frame",
        "grok_build_home": str(Path(args.grok_build_home).expanduser().resolve()),
        "grok_build_config": str(config_path),
        "grok_build_config_sha256": sha256_file(config_path) if config_path.exists() else None,
        "adapter_url": args.adapter_url, "adapter_expected_version": args.expected_adapter_version,
        "adapter_health": adapter_health,
        "adapter_chain": "Grok Build Responses API -> localhost thin compatibility adapter -> OpenRouter Responses API -> deepseek/deepseek-v4-pro",
        "adapter_scope": "force benchmark model; normalize web-search events; suppress legacy SSE DONE; repair missing reasoning summary arrays; stringify structured tool payloads",
        "worker_count": 1, "task_concurrency": 1,
        "skill_dir": str(skill_dir), "tasks_dir": str(tasks_dir),
        "pinchbench_commit": commit, "expected_pinchbench_commit": args.expected_pinchbench_commit,
        "expected_manifest_sha256": args.expected_manifest_sha256,
        "manifest_sha256": sha256_file(manifest_path) if manifest_path.exists() else None,
        "suite": args.suite, "limit": args.limit,
        "default_skipped_tasks": sorted(DEFAULT_SKIPPED_TASKS), "additional_skip": args.skip,
        "skip_network": args.skip_network, "task_count": len(selected),
        "task_ids": [task.task_id for task in selected],
        "timeout_multiplier": args.timeout_multiplier, "network_timeout": args.network_timeout,
        "workspace_instruction": not args.no_workspace_instruction,
        "grading_enabled": not args.no_grade,
        "grading_engine": str(skill_dir / "scripts" / "lib_grading.py"),
        "judge_backend": "api", "judge_model": judge_model,
        "judge_key_env": "OPENROUTER_API_KEY", "judge_key_present": bool(os.environ.get("OPENROUTER_API_KEY")),
        "judge_concurrency": 1, "judge_is_separate_from_tested_model": True,
        "judge_transport_compatibility": {"budgets": [2048, 4096, 8192, 16384], "raw_response_retention": True, "hybrid_automated_recovery": True},
        "environment_proxy_present": {key: bool(os.environ.get(key)) for key in ("HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "NO_PROXY")},
        "resume_history": resume_history or [],
        "comparability_note": "Task loading, exclusions, serial execution, timeouts, workspace staging, transcript normalization, official grader, hybrid weights, exports and Judge model match the validated Windows runners. Only Grok Build invocation/event parsing and the documented thin search compatibility layer differ.",
    }
    write_json_atomic(output_path, config)


def read_adapter_health(adapter_url: str, timeout: float = 5.0) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    try:
        parsed = urllib.parse.urlsplit(adapter_url)
        if parsed.scheme != "http" or parsed.hostname not in {"127.0.0.1", "localhost", "::1"}:
            return None, f"Adapter URL must be localhost HTTP: {adapter_url}"
        request = urllib.request.Request(adapter_url.rstrip("/") + "/healthz", headers={"Accept": "application/json"}, method="GET")
        with urllib.request.urlopen(request, timeout=timeout) as response:
            payload = json.loads(response.read().decode("utf-8"))
        if not isinstance(payload, dict): return None, "Adapter health response is not a JSON object."
        return payload, None
    except Exception as exc:
        return None, f"Adapter health check failed: {exc}"


def read_toml_file(path: Path) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    try:
        raw = path.read_bytes()
        if raw.startswith(b"\xef\xbb\xbf"): raw = raw[3:]
        value = tomllib.loads(raw.decode("utf-8"))
        return value if isinstance(value, dict) else None, None
    except Exception as exc:
        return None, f"无法解析 TOML {path}: {exc}"


def find_grok_build_context_files(inferred_root: Path, grok_build_home: Path, skill_dir: Path) -> list[str]:
    del skill_dir  # The PinchBench checkout is not an ancestor of task workspaces.
    candidates = [
        inferred_root / "AGENTS.md", inferred_root / "CLAUDE.md", inferred_root / "Claude.md",
        inferred_root / "CLAUDE.local.md", inferred_root / "GROK.md",
        grok_build_home / "AGENTS.md", grok_build_home / "CLAUDE.md",
        grok_build_home / "Claude.md", grok_build_home / "CLAUDE.local.md",
        grok_build_home / "GROK.md",
    ]
    found: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate.resolve(strict=False)).lower()
        if key in seen:
            continue
        seen.add(key)
        if candidate.exists():
            found.append(str(candidate.resolve()))
    for base, directories in (
        (grok_build_home, ("rules", "skills", "plugins", "memory", "mcp", ".claude", ".cursor")),
        (inferred_root, (".grok", ".claude", ".cursor")),
    ):
        for directory in directories:
            candidate = base / directory
            if candidate.exists():
                found.append(str(candidate.resolve()))
    return found


def print_preflight(
    selected: list[Task], skill_dir: Path, tasks_dir: Path, grok_build_command: str,
    grader_module: Optional[Any], grader_error: Optional[str], args: argparse.Namespace,
    inferred_root: Path,
) -> tuple[dict[str, list[str]], list[str], list[str], dict[str, Any]]:
    prerequisite_failures: dict[str, list[str]] = {}
    fixture_failures: list[str] = []
    environment_failures: list[str] = []
    needs_judge = not args.no_grade and any(task.grading_type in {"llm_judge", "hybrid"} for task in selected)
    judge_key_present = bool(os.environ.get("OPENROUTER_API_KEY"))
    git_command = resolve_command("git")
    actual_commit = command_output([git_command, "rev-parse", "HEAD"], cwd=skill_dir).strip() if git_command else ""
    if not actual_commit: environment_failures.append("Cannot determine PinchBench git commit from skill dir.")
    elif args.expected_pinchbench_commit and actual_commit != args.expected_pinchbench_commit:
        environment_failures.append(f"PinchBench commit mismatch: expected {args.expected_pinchbench_commit}, actual {actual_commit}")
    manifest_path = tasks_dir / "manifest.yaml"
    actual_manifest_sha256 = sha256_file(manifest_path) if manifest_path.exists() else ""
    if args.expected_manifest_sha256 and actual_manifest_sha256.lower() != args.expected_manifest_sha256.lower():
        environment_failures.append(f"PinchBench manifest SHA-256 mismatch: expected {args.expected_manifest_sha256}, actual {actual_manifest_sha256 or 'missing'}")

    version_output = command_output([grok_build_command, "--version"])
    version_match = re.search(r"\d+\.\d+\.\d+(?:[-+][^\s]+)?", version_output)
    actual_version = version_match.group(0) if version_match else version_output.strip()
    if args.expected_grok_build_version and actual_version != args.expected_grok_build_version:
        environment_failures.append(f"Grok Build version mismatch: expected {args.expected_grok_build_version}, actual {version_output}")
    actual_binary_hash = sha256_file(Path(grok_build_command))
    if args.expected_grok_build_binary_sha256 and actual_binary_hash.lower() != args.expected_grok_build_binary_sha256.lower():
        environment_failures.append(f"Grok Build binary SHA-256 mismatch: expected {args.expected_grok_build_binary_sha256}, actual {actual_binary_hash}")

    grok_home = Path(args.grok_build_home).expanduser().resolve()
    config_path = grok_home / "config.toml"
    if not config_path.exists(): environment_failures.append(f"找不到 Grok Build config.toml: {config_path}")
    else:
        config, config_error = read_toml_file(config_path)
        if config_error: environment_failures.append(config_error)
        elif config is not None:
            models = config.get("models") if isinstance(config.get("models"), dict) else {}
            model_table = config.get("model") if isinstance(config.get("model"), dict) else {}
            model_config = model_table.get(args.model_alias) if isinstance(model_table.get(args.model_alias), dict) else {}
            expected_base = args.adapter_url.rstrip("/") + "/v1"
            checks = {
                "models.default": (models.get("default"), args.model_alias),
                "models.web_search": (models.get("web_search"), args.model_alias),
                "model.model": (model_config.get("model"), args.model),
                "model.base_url": (str(model_config.get("base_url") or "").rstrip("/"), expected_base),
                "model.env_key": (model_config.get("env_key"), "OPENROUTER_API_KEY"),
                "model.api_backend": (model_config.get("api_backend"), "responses"),
            }
            compat = config.get("compat") if isinstance(config.get("compat"), dict) else {}
            for vendor in ("cursor", "claude"):
                vendor_config = compat.get(vendor) if isinstance(compat.get(vendor), dict) else {}
                for key in ("skills", "rules", "agents", "mcps", "hooks", "sessions"):
                    checks[f"compat.{vendor}.{key}"] = (vendor_config.get(key), False)
            codex_config = compat.get("codex") if isinstance(compat.get("codex"), dict) else {}
            checks["compat.codex.sessions"] = (codex_config.get("sessions"), False)
            for label, (actual, expected) in checks.items():
                if actual != expected: environment_failures.append(f"Grok config {label} mismatch: expected {expected!r}, actual {actual!r}")
            if model_config.get("api_key"): environment_failures.append("Grok config contains inline api_key; env_key is required.")
    context_files = find_grok_build_context_files(inferred_root, grok_home, skill_dir)
    if context_files: environment_failures.append("检测到可能注入自定义指令/扩展的文件或目录: " + " | ".join(context_files))

    adapter_health, adapter_error = read_adapter_health(args.adapter_url)
    if adapter_error:
        environment_failures.append(adapter_error); adapter_health = {}
    else:
        assert adapter_health is not None
        if adapter_health.get("ok") is not True: environment_failures.append("Adapter health ok != true")
        if str(adapter_health.get("version") or "") != args.expected_adapter_version:
            environment_failures.append(f"Adapter version mismatch: expected {args.expected_adapter_version}, actual {adapter_health.get('version')}")
        if str(adapter_health.get("target_model") or "") != args.model:
            environment_failures.append(f"Adapter target_model mismatch: expected {args.model}, actual {adapter_health.get('target_model')}")

    models_output = command_output([grok_build_command, "models"])
    if args.model_alias not in models_output: environment_failures.append(f"grok models does not list alias {args.model_alias}")
    if "Default model: " + args.model_alias not in models_output: environment_failures.append(f"Grok default model is not {args.model_alias}")

    print("=" * 100)
    print("PinchBench Grok Build Windows runner preflight")
    print("=" * 100)
    print(f"Runner revision      : {RUNNER_REVISION}")
    print(f"Skill dir            : {skill_dir}")
    print(f"PinchBench commit    : {actual_commit or 'unavailable'}")
    print(f"Manifest SHA-256     : {actual_manifest_sha256 or 'missing'}")
    print(f"Grok command         : {grok_build_command}")
    print(f"Grok version         : {version_output}")
    print(f"Binary SHA-256       : {actual_binary_hash}")
    print(f"GROK_HOME            : {grok_home}")
    print(f"Model alias          : {args.model_alias}")
    print(f"Actual model         : {args.model}")
    print(f"Adapter              : {args.adapter_url} v{adapter_health.get('version', 'unavailable')}")
    print(f"Suite / tasks        : {args.suite} / {len(selected)}")
    print("Worker/concurrency   : 1 / 1")
    print("Prompt transport     : UTF-8 --prompt-file (never prompt argv)")
    print("Output transport     : streaming-messages-json; raw JSONL/stderr separate")
    print(f"Judge model          : {args.judge_model}")
    print(f"OPENROUTER_API_KEY   : {'set' if judge_key_present else 'missing'}")
    print(f"Grader import        : {'ok' if grader_error is None else 'FAILED'}")
    print(f"Custom context       : {'none' if not context_files else 'FOUND'}")
    print()

    for task in selected:
        prerequisites = task.metadata.get("prerequisites") or []
        missing = check_prerequisites(list(prerequisites)) if prerequisites else []
        if missing: prerequisite_failures[task.task_id] = missing
        for wf in task.workspace_files:
            if not isinstance(wf, dict): fixture_failures.append(f"{task.task_id}: invalid workspace file entry"); continue
            if wf.get("content") is not None: continue
            source = wf.get("source")
            if not source: continue
            rel = Path(str(source))
            candidates = [skill_dir / "assets" / rel, skill_dir / rel, task.file_path.parent / rel, task.file_path.parent.parent / "assets" / rel]
            if not any(x.exists() for x in candidates): fixture_failures.append(f"{task.task_id}: missing fixture {source}")

    print(f"Missing prerequisites: {sum(len(x) for x in prerequisite_failures.values())}")
    print(f"Missing fixtures     : {len(fixture_failures)}")
    if environment_failures:
        print("Environment failures:")
        for item in environment_failures: print(f"  - {item}")
    else: print("Environment failures: none")
    if needs_judge and not judge_key_present: print("Judge readiness      : FAILED")
    else: print("Judge readiness      : ok" if needs_judge else "Judge readiness      : not required")
    print("=" * 100)
    return prerequisite_failures, fixture_failures, environment_failures, dict(adapter_health)


# ---------------------------------------------------------------------------
# Main execution

# ---------------------------------------------------------------------------

def build_turns(task: Task) -> list[dict[str, Any]]:
    """Build the PinchBench turn sequence with cross-Agent-compatible semantics.

    PinchBench multi-session task metadata commonly omits ``new_session`` on
    the first item. The first executable turn must always start a fresh Grok Build
    session. Later turns resume the previous session unless their metadata
    explicitly requests ``new_session: true``.
    """
    if task.multi_session and task.sessions:
        turns: list[dict[str, Any]] = []
        for index, item in enumerate(task.sessions, start=1):
            prompt = str(item.get("prompt") or "").strip()
            if not prompt:
                continue

            is_first_executable_turn = not turns
            new_session = (
                True
                if is_first_executable_turn
                else bool(item.get("new_session"))
            )

            turns.append({
                "id": str(item.get("id") or f"turn_{index}"),
                "prompt": prompt,
                "new_session": new_session,
            })
        if turns:
            return turns

    return [{
        "id": "single",
        "prompt": task.prompt,
        "new_session": True,
    }]


def aggregate_optional_int(
    current: Optional[int],
    value: Optional[int],
) -> Optional[int]:
    if value is None:
        return current
    return (current or 0) + int(value)


def aggregate_optional_float(
    current: Optional[float],
    value: Optional[float],
) -> Optional[float]:
    if value is None:
        return current
    return (current or 0.0) + float(value)


def merge_model_usage_rows(target: dict[str, Any], source: dict[str, Any]) -> None:
    """Accumulate per-model usage across resumed/multi-turn Grok calls."""
    additive_ints = {
        "inputTokens", "outputTokens", "cacheReadInputTokens",
        "cacheCreationInputTokens", "modelCalls", "webSearchRequests",
    }
    additive_floats = {"costUSD"}
    for model_name, raw_row in source.items():
        if not isinstance(raw_row, dict):
            continue
        row = target.setdefault(str(model_name), {})
        if not isinstance(row, dict):
            row = {}
            target[str(model_name)] = row
        for key, value in raw_row.items():
            if key in additive_ints:
                row[key] = _to_int(row.get(key)) + _to_int(value)
            elif key in additive_floats:
                row[key] = float(row.get(key) or 0.0) + float(value or 0.0)
            else:
                # Context-window and descriptive fields are snapshots, not sums.
                row[key] = value


def execute_task(
    task: Task, task_index: int, task_count: int, skill_dir: Path,
    workspace: Path, transcript_dir: Path, grok_build_command: str,
    args: argparse.Namespace,
) -> dict[str, Any]:
    print(f"[{task_index}/{task_count}] {task.task_id} — {task.name}")
    preview = re.sub(r"\s+", " ", task.prompt)[:150]
    print(f"  prompt: {preview}{'...' if len(task.prompt) > 150 else ''}")

    def failure_result(status: str, error: str) -> dict[str, Any]:
        return {
            "task_id": task.task_id, "name": task.name, "category": task.category,
            "grading_type": task.grading_type, "network_task": is_network_task(task),
            "multi_session": task.multi_session, "session_count": 0, "success": False,
            "status": status, "returncode": None, "elapsed": 0.0, "ttft": None,
            "input_tokens": None, "output_tokens": None, "reasoning_tokens": None,
            "cache_read_tokens": None, "cache_write_tokens": None, "cost_usd": None,
            "total_tokens": None, "usage_complete": False, "step_count": 0,
            "tool_errors": 0, "tool_call_count": 0, "observed_models": [],
            "positive_usage_models": [], "unexpected_models": [], "model_usage": {},
            "grok_build_cli_versions": [args.expected_grok_build_version],
            "permission_modes": [], "permission_denials": [], "output": "", "error": error,
            "stderr": "", "score": None, "breakdown": {}, "grade_notes": "",
            "grade_error": error, "workspace": str(workspace), "transcript": str(transcript_dir),
            "normalized_transcript": [], "turn_results": [],
        }

    prerequisites = task.metadata.get("prerequisites") or []
    missing = check_prerequisites(list(prerequisites)) if prerequisites else []
    if missing:
        error = f"缺少依赖: {', '.join(missing)}"; print(f"  跳过: {error}\n")
        return failure_result("missing_prerequisite", error)

    if workspace.exists(): shutil.rmtree(workspace)
    workspace.mkdir(parents=True, exist_ok=True)
    transcript_dir.mkdir(parents=True, exist_ok=True)
    staged, fixture_errors = stage_workspace_files(task, workspace, skill_dir)
    if fixture_errors:
        error = "缺少或无效的 workspace fixture: " + " | ".join(fixture_errors)
        print(f"  失败: {error}\n"); return failure_result("missing_fixture", error)

    turns = build_turns(task)
    total_timeout = args.network_timeout if is_network_task(task) else max(1.0, task.timeout_seconds * args.timeout_multiplier)
    deadline = time.monotonic() + total_timeout
    normalized_transcript: list[dict[str, Any]] = []
    outputs: list[str] = []; errors: list[str] = []; stderr_chunks: list[str] = []; turn_results: list[dict[str, Any]] = []
    current_session_id: Optional[str] = None
    task_status = "success"; task_success = True; returncode: Optional[int] = 0
    total_elapsed = 0.0; task_ttft: Optional[float] = None
    input_tokens = output_tokens = cache_read_tokens = cache_write_tokens = total_tokens = None
    cost_usd: Optional[float] = None
    usage_complete = True; step_count = tool_errors = tool_call_count = 0
    event_counts: Counter[str] = Counter(); observed_models: set[str] = set(); positive_usage_models: set[str] = set(); unexpected_models: set[str] = set()
    model_usage: dict[str, Any] = {}; permission_modes: set[str] = set(); permission_denials: list[Any] = []

    for turn_index, turn in enumerate(turns, start=1):
        remaining = deadline - time.monotonic()
        if remaining <= 0:
            task_success = False; task_status = "timeout"; errors.append(f"任务总超时 ({total_timeout:.0f}s)"); break
        turn_id = str(turn["id"]); turn_prompt = str(turn["prompt"]); new_session = bool(turn.get("new_session"))
        resume_required = turn_index > 1 and not new_session
        if resume_required and not current_session_id:
            task_success = False; task_status = "session_error"; errors.append(f"turn {turn_index} ({turn_id}) 需要恢复会话，但前一轮没有 session_id"); break
        normalized_transcript.append(make_user_transcript_event(turn_prompt, turn_index, turn_id))
        prompt = turn_prompt
        if not args.no_workspace_instruction:
            prompt += "\n\nIMPORTANT: You are running in an isolated workspace. Read, write, and edit files only in the current working directory."
        raw_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.jsonl"
        stderr_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.stderr.txt"
        prompt_path = transcript_dir / f"turn_{turn_index:02d}_{turn_id}.prompt.txt"
        prompt_payload = prompt if prompt.endswith("\n") else prompt + "\n"
        prompt_path.write_text(prompt_payload, encoding="utf-8", newline="\n")
        prompt_bytes = prompt_payload.encode("utf-8")
        command = [
            grok_build_command, "--prompt-file", str(prompt_path), "-m", args.model_alias,
            "--cwd", str(workspace), "--output-format", "streaming-messages-json",
            "--yolo", "--no-auto-update", "--verbatim",
        ]
        if resume_required and current_session_id: command.extend(["--resume", current_session_id])
        result = run_grok_build_streaming(
            command, cwd=workspace, timeout=max(1.0, remaining), raw_stdout_path=raw_path,
            stderr_path=stderr_path, prompt_path=prompt_path, requested_alias=args.model_alias,
            expected_model_id=args.model, grok_build_home=Path(args.grok_build_home).expanduser().resolve(),
        )
        normalized_transcript.extend(result["normalized_events"])
        if result.get("output"): outputs.append(f"## Turn {turn_index} ({turn_id})\n{result['output']}")
        if result.get("error"): errors.append(f"turn {turn_index} ({turn_id}): {result['error']}")
        if result.get("stderr"): stderr_chunks.append(f"## Turn {turn_index} ({turn_id})\n{result['stderr']}")
        total_elapsed += float(result.get("elapsed") or 0.0)
        if task_ttft is None and result.get("ttft") is not None: task_ttft = float(result["ttft"])
        input_tokens = aggregate_optional_int(input_tokens, result.get("input_tokens"))
        output_tokens = aggregate_optional_int(output_tokens, result.get("output_tokens"))
        cache_read_tokens = aggregate_optional_int(cache_read_tokens, result.get("cache_read_tokens"))
        cache_write_tokens = aggregate_optional_int(cache_write_tokens, result.get("cache_write_tokens"))
        total_tokens = aggregate_optional_int(total_tokens, result.get("total_tokens"))
        cost_usd = aggregate_optional_float(cost_usd, result.get("cost_usd"))
        usage_complete = usage_complete and bool(result.get("usage_complete"))
        step_count += int(result.get("step_count") or 0); tool_errors += int(result.get("tool_errors") or 0); tool_call_count += int(result.get("tool_call_count") or 0)
        event_counts.update(result.get("event_counts") or {})
        observed_models.update(str(x) for x in result.get("observed_models") or [])
        positive_usage_models.update(str(x) for x in result.get("positive_usage_models") or [])
        unexpected_models.update(str(x) for x in result.get("unexpected_models") or [])
        if isinstance(result.get("model_usage"), dict):
            merge_model_usage_rows(model_usage, result["model_usage"])
        if result.get("permission_mode"): permission_modes.add(str(result.get("permission_mode")))
        permission_denials.extend(result.get("permission_denials") or [])
        if result.get("session_id"): current_session_id = str(result.get("session_id"))
        turn_results.append({
            "turn": turn_index, "turn_id": turn_id, "new_session": new_session,
            "session_id": current_session_id, "request_id": result.get("request_id"),
            "success": result.get("success"), "status": result.get("status"),
            "returncode": result.get("returncode"), "elapsed": result.get("elapsed"),
            "ttft": result.get("ttft"), "input_tokens": result.get("input_tokens"),
            "output_tokens": result.get("output_tokens"), "cache_read_tokens": result.get("cache_read_tokens"),
            "cache_write_tokens": result.get("cache_write_tokens"), "total_tokens": result.get("total_tokens"),
            "cost_usd": result.get("cost_usd"), "usage_complete": result.get("usage_complete"),
            "step_count": result.get("step_count"), "tool_errors": result.get("tool_errors"),
            "tool_call_count": result.get("tool_call_count"), "event_counts": result.get("event_counts"),
            "observed_models": result.get("observed_models"), "positive_usage_models": result.get("positive_usage_models"),
            "unexpected_models": result.get("unexpected_models"), "model_usage": result.get("model_usage"),
            "permission_mode": result.get("permission_mode"), "permission_denials": result.get("permission_denials"),
            "available_tools": result.get("available_tools"), "raw_transcript": str(raw_path),
            "stderr_path": str(stderr_path), "prompt_transport": result.get("prompt_transport"),
            "prompt_path": str(prompt_path), "prompt_chars": len(prompt_payload), "prompt_bytes": len(prompt_bytes),
            "prompt_sha256": hashlib.sha256(prompt_bytes).hexdigest(), "error": result.get("error"),
        })
        if not result.get("success"):
            task_success = False; task_status = str(result.get("status") or "error"); returncode = result.get("returncode"); break
        returncode = result.get("returncode")

    normalized_path = transcript_dir / "normalized.jsonl"
    with normalized_path.open("w", encoding="utf-8", newline="\n") as handle:
        for event in normalized_transcript: handle.write(json.dumps(event, ensure_ascii=False) + "\n")
    write_json_atomic(transcript_dir / "turn_results.json", turn_results)
    return {
        "task_id": task.task_id, "name": task.name, "category": task.category,
        "grading_type": task.grading_type, "network_task": is_network_task(task),
        "multi_session": task.multi_session, "session_count": len(turn_results),
        "success": task_success, "status": task_status, "returncode": returncode,
        "elapsed": total_elapsed, "ttft": task_ttft, "input_tokens": input_tokens,
        "output_tokens": output_tokens, "reasoning_tokens": None, "cache_read_tokens": cache_read_tokens,
        "cache_write_tokens": cache_write_tokens, "cost_usd": cost_usd, "total_tokens": total_tokens,
        "usage_complete": usage_complete and bool(turn_results), "step_count": step_count,
        "tool_errors": tool_errors, "tool_call_count": tool_call_count, "event_counts": dict(event_counts),
        "observed_models": sorted(observed_models), "positive_usage_models": sorted(positive_usage_models),
        "unexpected_models": sorted(unexpected_models), "model_usage": model_usage,
        "grok_build_cli_versions": [args.expected_grok_build_version],
        "permission_modes": sorted(permission_modes), "permission_denials": permission_denials,
        "output": "\n\n".join(outputs), "error": " | ".join(errors), "stderr": "\n\n".join(stderr_chunks),
        "score": None, "breakdown": {}, "grade_notes": "", "grade_error": None,
        "workspace": str(workspace), "transcript": str(transcript_dir),
        "normalized_transcript_path": str(normalized_path), "transcript_data": normalized_transcript,
        "turn_results": turn_results,
    }


def build_args() -> argparse.Namespace:
    script_path = Path(__file__).resolve()
    root = script_path.parent.parent
    parser = argparse.ArgumentParser(description="Run PinchBench with Grok Build on native Windows, serially.", formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("--skill-dir", default=str(root / "skill" if (root / "skill").exists() else Path.cwd()))
    parser.add_argument("--tasks-dir", default=None)
    parser.add_argument("--expected-pinchbench-commit", default=DEFAULT_PINCHBENCH_COMMIT)
    parser.add_argument("--expected-manifest-sha256", default=DEFAULT_MANIFEST_SHA256)
    parser.add_argument("--suite", default="core", help="all/core/automated-only/llm-judge-only/hybrid-only/judge-required-only or comma-separated task IDs")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--skip", default="")
    parser.add_argument("--skip-network", action="store_true")
    parser.add_argument("--grok-build-cli", default=str(root / "bin" / "grok.exe"))
    parser.add_argument("--expected-grok-build-version", default=DEFAULT_GROK_BUILD_VERSION)
    parser.add_argument("--expected-grok-build-binary-sha256", default=DEFAULT_GROK_BUILD_BINARY_SHA256)
    parser.add_argument("--grok-build-home", default=str(root / "benchmark-home"))
    parser.add_argument("--adapter-url", default=DEFAULT_ADAPTER_URL)
    parser.add_argument("--expected-adapter-version", default=DEFAULT_ADAPTER_VERSION)
    parser.add_argument("--model", default=DEFAULT_MODEL_ID, help="Actual provider model ID")
    parser.add_argument("--model-alias", default=DEFAULT_MODEL_ALIAS, help="Grok Build custom-model alias")
    parser.add_argument("--approval-mode", choices=["yolo"], default=DEFAULT_APPROVAL_MODE)
    parser.add_argument("--timeout-multiplier", type=float, default=3.0)
    parser.add_argument("--network-timeout", type=float, default=300.0)
    parser.add_argument("--judge-timeout", type=float, default=300.0)
    parser.add_argument("--judge-model", default=PINCHBENCH_DEFAULT_JUDGE_MODEL)
    parser.add_argument("--results-dir", default=None)
    parser.add_argument("--resume-run", default=None, help="Existing run directory; continue task IDs absent from progress.jsonl")
    parser.add_argument("--keep-workspaces", action="store_true")
    parser.add_argument("--no-grade", action="store_true")
    parser.add_argument("--no-xlsx", action="store_true")
    parser.add_argument("--no-workspace-instruction", action="store_true")
    parser.add_argument("--preflight", action="store_true")
    parser.add_argument("--no-judge-cache", action="store_true")
    parser.add_argument("--clear-judge-cache", action="store_true")
    parser.add_argument("--verbose", action="store_true")
    return parser.parse_args()


def load_progress_rows(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not path.exists(): return rows
    for line in path.read_text(encoding="utf-8-sig", errors="replace").splitlines():
        if not line.strip(): continue
        try:
            row = json.loads(line)
            if isinstance(row, dict): rows.append(row)
        except json.JSONDecodeError:
            LOGGER.warning("Ignoring incomplete progress line in %s", path)
    # Last row wins if a historical file contains duplicate task IDs.
    mapping: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for row in rows:
        task_id = str(row.get("task_id") or "")
        if not task_id: continue
        if task_id not in mapping: order.append(task_id)
        mapping[task_id] = row
    return [mapping[task_id] for task_id in order]


def main() -> int:
    args = build_args()
    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO, format="%(asctime)s %(levelname)-8s %(message)s", datefmt="%H:%M:%S")
    root = Path(__file__).resolve().parent.parent
    grok_home = Path(args.grok_build_home).expanduser().resolve()
    os.environ["GROK_HOME"] = str(grok_home)
    os.environ.setdefault("PYTHONUTF8", "1"); os.environ.setdefault("PYTHONIOENCODING", "utf-8")
    os.environ.setdefault("NO_COLOR", "1"); os.environ.setdefault("FORCE_COLOR", "0"); os.environ.setdefault("RUST_LOG", "error")
    grok_home.mkdir(parents=True, exist_ok=True)

    skill_dir = Path(args.skill_dir).expanduser().resolve()
    tasks_dir = Path(args.tasks_dir).expanduser().resolve() if args.tasks_dir else skill_dir / "tasks"
    if not skill_dir.exists(): raise SystemExit(f"找不到 PinchBench 仓库目录: {skill_dir}")
    if not tasks_dir.exists(): raise SystemExit(f"找不到 tasks 目录: {tasks_dir}")
    grok_command = choose_grok_build_command(args.grok_build_cli, root)
    grader_module, grader_error = load_pinchbench_grading(skill_dir)
    judge_key_present = bool(os.environ.get("OPENROUTER_API_KEY"))
    all_tasks, core_tasks = load_tasks(tasks_dir)
    selected = filter_tasks(all_tasks, core_tasks, args.suite, args.limit)
    skip_ids = set(DEFAULT_SKIPPED_TASKS) | {x.strip() for x in args.skip.split(",") if x.strip()}
    if args.skip_network: skip_ids.update(task.task_id for task in selected if is_network_task(task))
    selected = [task for task in selected if task.task_id not in skip_ids]
    if not selected: raise SystemExit("没有可运行任务。")
    needs_judge = not args.no_grade and any(task.grading_type in {"llm_judge", "hybrid"} for task in selected)
    prereq_failures, fixture_failures, env_failures, adapter_health = print_preflight(selected, skill_dir, tasks_dir, grok_command, grader_module, grader_error, args, root)
    if not args.no_xlsx:
        try:
            import openpyxl  # type: ignore  # noqa: F401
        except ImportError:
            env_failures.append("openpyxl is required for the formal XLSX export.")
    if args.preflight:
        if env_failures: return 2
        if fixture_failures: return 3
        if prereq_failures: return 4
        if grader_error: return 5
        if needs_judge and not judge_key_present: return 6
        return 0
    if env_failures: raise SystemExit("Grok Build 环境预检失败:\n- " + "\n- ".join(env_failures))
    if fixture_failures: raise SystemExit("Workspace fixture 预检失败:\n- " + "\n- ".join(fixture_failures))
    if prereq_failures:
        details = [f"{task_id}: {', '.join(items)}" for task_id, items in prereq_failures.items()]
        raise SystemExit("Task prerequisite 预检失败:\n- " + "\n- ".join(details))
    if grader_error and not args.no_grade: raise SystemExit(grader_error)
    if needs_judge and not judge_key_present: raise SystemExit("所选任务需要 Judge，但当前 PowerShell 没有 OPENROUTER_API_KEY。")

    results_root = Path(args.results_dir).expanduser().resolve() if args.results_dir else root / "runs"
    resume_history: list[dict[str, Any]] = []
    if args.resume_run:
        results_dir = Path(args.resume_run).expanduser().resolve()
        if not results_dir.exists(): raise SystemExit(f"Resume run does not exist: {results_dir}")
        existing_config_path = results_dir / "run_config.json"
        if not existing_config_path.exists(): raise SystemExit(f"Missing run_config.json: {existing_config_path}")
        existing_config = json.loads(existing_config_path.read_text(encoding="utf-8-sig"))
        expected_ids = [task.task_id for task in selected]
        if list(existing_config.get("task_ids") or []) != expected_ids:
            raise SystemExit("Resume task list differs from the original run_config.json; refusing to mix runs.")
        if existing_config.get("model") != args.model or existing_config.get("model_alias") != args.model_alias:
            raise SystemExit("Resume model differs from the original run; refusing to mix runs.")
        resume_history = list(existing_config.get("resume_history") or [])
        resume_history.append({"at": utc_now(), "pid": os.getpid()})
    else:
        results_dir = results_root / dt.datetime.now().strftime("grok_build_%Y%m%d_%H%M%S")
    workspaces_dir = results_dir / "workspaces"; transcripts_dir = results_dir / "transcripts"
    results_dir.mkdir(parents=True, exist_ok=True); workspaces_dir.mkdir(parents=True, exist_ok=True); transcripts_dir.mkdir(parents=True, exist_ok=True)
    progress_path = results_dir / "progress.jsonl"; partial_path = results_dir / "results.partial.json"; heartbeat_path = results_dir / "heartbeat.json"
    progress_path.touch(exist_ok=True)
    results = load_progress_rows(progress_path)
    completed_ids = {str(row.get("task_id") or "") for row in results}
    write_json_atomic(partial_path, {"completed": len(results), "results": results})
    write_run_config(results_dir / "run_config.json", args, skill_dir, tasks_dir, grok_command, selected, str(args.judge_model), adapter_health, resume_history)

    judge_cache_slug = re.sub(r"[^A-Za-z0-9._-]+", "_", str(args.judge_model)).strip("_") or "judge"
    judge_cache_dir = results_root / ".judge_cache" / judge_cache_slug
    if grader_module is not None and hasattr(grader_module, "set_judge_cache_dir") and not args.no_judge_cache:
        grader_module.set_judge_cache_dir(judge_cache_dir)
        if args.clear_judge_cache and not args.resume_run and hasattr(grader_module, "clear_judge_cache"): grader_module.clear_judge_cache()

    print(); print(f"Results dir          : {results_dir}")
    print(f"Grok Build           : {command_output([grok_command, '--version'])}")
    print(f"Model alias / ID     : {args.model_alias} / {args.model}")
    print(f"Judge model          : {args.judge_model}")
    print(f"Already completed    : {len(results)}"); print()
    total_start = time.monotonic()
    for index, task in enumerate(selected, start=1):
        if task.task_id in completed_ids:
            print(f"[{index}/{len(selected)}] {task.task_id} — already completed; skipping")
            continue
        write_json_atomic(heartbeat_path, {"at": utc_now(), "state": "running", "task_id": task.task_id, "task_index": index, "task_count": len(selected), "pid": os.getpid()})
        workspace = workspaces_dir / task.task_id; transcript_dir = transcripts_dir / task.task_id
        execution = execute_task(task, index, len(selected), skill_dir, workspace, transcript_dir, grok_command, args)
        grade_result = GradeResult(score=None, grading_type=task.grading_type, breakdown={})
        if execution["status"] not in {"missing_prerequisite", "missing_fixture"} and not args.no_grade:
            grading_execution = dict(execution); grading_execution["transcript"] = execution.get("transcript_data", [])
            grade_result = grade_with_pinchbench_default(
                grader_module=grader_module, task=task, execution_result=grading_execution,
                workspace=workspace, skill_dir=skill_dir, judge_timeout=args.judge_timeout,
                judge_model=str(args.judge_model), verbose=args.verbose,
                judge_raw_root=results_dir / "judge_raw_responses",
            )
        execution["score"] = round(float(grade_result.score), 4) if grade_result.score is not None else None
        execution["breakdown"] = grade_result.breakdown; execution["grade_notes"] = grade_result.notes; execution["grade_error"] = grade_result.error
        execution.pop("transcript_data", None)
        results.append(execution); completed_ids.add(task.task_id)
        with progress_path.open("a", encoding="utf-8", newline="\n") as handle:
            handle.write(json.dumps(execution, ensure_ascii=False) + "\n"); handle.flush(); os.fsync(handle.fileno())
        write_json_atomic(partial_path, {"completed": len(results), "results": results})
        write_json_atomic(heartbeat_path, {"at": utc_now(), "state": "completed_task", "task_id": task.task_id, "completed": len(results), "task_count": len(selected), "pid": os.getpid()})
        marker = "✓" if execution["success"] else "✗"
        score_text = f"score={execution['score']:.3f}" if execution["score"] is not None else "score=N/A"
        token_text = f"{execution['input_tokens']}+{execution['output_tokens']}" if execution["usage_complete"] else "N/A"
        print(f"  {marker} elapsed={execution['elapsed']:.1f}s {score_text} tokens={token_text}")
        if execution.get("unexpected_models"): print("  非目标模型: " + ", ".join(execution["unexpected_models"]))
        if execution.get("error"): print("  运行错误: " + str(execution["error"])[:500])
        if execution.get("grade_error"): print("  打分错误: " + str(execution["grade_error"])[:500])
        print()
        if not args.keep_workspaces and execution["success"] and not execution.get("grade_error"):
            shutil.rmtree(workspace, ignore_errors=True); execution["workspace"] = "已清理；使用 --keep-workspaces 可保留"

    total_elapsed = time.monotonic() - total_start
    summary = print_summary(results, total_elapsed)
    payload = {"summary": summary, "results": results}
    json_path = results_dir / "results.json"; write_json_atomic(json_path, payload)
    csv_path = results_dir / "results.csv"; save_csv(results, csv_path)
    xlsx_path = results_dir / "results.xlsx"
    if not args.no_xlsx: save_xlsx(results, summary, xlsx_path)
    write_json_atomic(heartbeat_path, {"at": utc_now(), "state": "finished", "completed": len(results), "task_count": len(selected), "pid": os.getpid()})
    print("\n结果文件:"); print(f"  Config : {results_dir / 'run_config.json'}"); print(f"  JSON   : {json_path}"); print(f"  CSV    : {csv_path}")
    if not args.no_xlsx: print(f"  XLSX   : {xlsx_path}")
    print(f"  Progress JSONL : {progress_path}"); print(f"  Partial JSON   : {partial_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
